import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import RepresentativeContact


NAME_HEADERS = ("тп", "торговый представитель", "representative", "name")
WORK_PHONE_HEADERS = ("раб номер", "рабочий номер", "work phone", "work_phone")
PERSONAL_PHONE_HEADERS = ("лич номер", "личный номер", "personal phone", "personal_phone")
WORK_ZONE_HEADERS = ("раб зона", "рабочая зона", "work zone", "work_zone")
REPRESENTATIVE_NOISE_PARTS = {"угли", "огли", "o", "g", "li", "qizi", "кизи", "қизи"}


def normalize_representative_name(value: Any) -> str:
    text = normalize_text(value).lower().replace("ё", "е")
    text = re.sub(r"[\u2010-\u2015]", "-", text)
    text = re.sub(r"[^0-9a-zа-я]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\bт\s*п\b", "тп", text)
    text = re.sub(r"\bтп\s*[- ]?\s*(\d+)\b", r"тп-\1", text)
    return text


def representative_aliases(value: Any) -> set[str]:
    normalized = normalize_representative_name(value)
    if not normalized:
        return set()
    aliases = {normalized, normalized.replace("-", ""), normalized.replace(" ", "")}
    match = re.search(r"\bтп-?(\d+)\b", normalized)
    if match:
        number = match.group(1)
        aliases.update({f"тп-{number}", f"тп{number}", f"тп {number}"})
    name_parts = representative_name_parts(normalized)
    if name_parts:
        aliases.add(" ".join(name_parts))
        aliases.add(name_parts[-1])
        for part in name_parts:
            aliases.update(representative_token_aliases(part))
    return {alias for alias in aliases if alias}


def representative_name_parts(normalized: str) -> list[str]:
    return [
        part
        for part in normalized.split()
        if len(part) > 1
        and part not in REPRESENTATIVE_NOISE_PARTS
        and part != "тп"
        and not re.fullmatch(r"тп-?\d+|\d+", part)
    ]


def representative_token_aliases(part: str) -> set[str]:
    aliases = {part}
    suffixes = ("жон", "jon")
    for suffix in suffixes:
        if part.endswith(suffix) and len(part) > len(suffix) + 2:
            aliases.add(part[: -len(suffix)])
    return aliases


def representative_tp_code(value: Any) -> str:
    match = re.search(r"\bтп-?(\d+)\b", normalize_representative_name(value))
    return f"ТП{match.group(1)}" if match else ""


def representative_name_without_tp_code(value: Any) -> str:
    text = normalize_text(value)
    return re.sub(r"^\s*т\s*п\s*[- ]?\s*\d+\s*", "", text, flags=re.I).strip()


def display_representative_name(
    representative: Any = "",
    contact: RepresentativeContact | None = None,
) -> str:
    representative_text = normalize_text(representative)
    contact_name = normalize_text(getattr(contact, "name", ""))
    representative_code = representative_tp_code(representative_text)
    contact_code = representative_tp_code(contact_name)
    tp_code = representative_code or contact_code
    if not tp_code:
        return representative_text or contact_name

    representative_parts = representative_name_parts(normalize_representative_name(representative_text))
    if representative_code:
        source_tail = representative_name_without_tp_code(representative_text)
        return f"{tp_code} {source_tail}" if source_tail else tp_code
    if representative_text:
        if len(representative_parts) >= 2:
            return f"{tp_code} {representative_text}"

    source_name = contact_name or representative_text
    source_tail = representative_name_without_tp_code(source_name)
    if source_tail:
        return f"{tp_code} {source_tail}"
    return tp_code


def normalize_phone(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D+", "", text)
    if len(digits) == 9:
        digits = f"998{digits}"
    if len(digits) == 12 and digits.startswith("998"):
        return f"+{digits[:3]} {digits[3:5]} {digits[5:8]} {digits[8:10]} {digits[10:12]}"
    if text.startswith("+"):
        return text
    return f"+{digits}" if digits and text.replace(" ", "").isdigit() else text


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def find_representative_contact(db: Session, representative: Any) -> RepresentativeContact | None:
    wanted = representative_aliases(representative)
    if not wanted:
        return None
    contacts = db.execute(
        select(RepresentativeContact).where(RepresentativeContact.is_active.is_(True))
    ).scalars().all()
    for contact in contacts:
        contact_aliases = representative_aliases(contact.name)
        contact_aliases.add(normalize_representative_name(contact.normalized_name))
        if wanted.intersection(contact_aliases):
            return contact
    return None


def build_representative_comment(
    payment_type: Any,
    representative: Any = "",
    contact: RepresentativeContact | None = None,
) -> str:
    lines = []
    payment = normalize_text(payment_type)
    rep_name = display_representative_name(representative, contact)
    if payment:
        lines.append(payment)
    if rep_name:
        lines.append(rep_name)
    if contact and normalize_text(contact.work_phone):
        lines.append(f"Рабочий номер: {normalize_phone(contact.work_phone)}")
    if contact and normalize_text(contact.personal_phone):
        lines.append(f"Личный номер: {normalize_phone(contact.personal_phone)}")
    return "\n".join(lines)


def import_representative_contacts_from_xlsx(db: Session, xlsx_path: str | Path) -> dict[str, Any]:
    path = Path(xlsx_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    summary = {"created": 0, "updated": 0, "skipped": 0, "rows": 0}
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue
        headers = {normalize_header(value): index for index, value in enumerate(header_row)}
        name_index = first_header_index(headers, NAME_HEADERS)
        if name_index is None:
            summary["skipped"] += 1
            continue
        work_phone_index = first_header_index(headers, WORK_PHONE_HEADERS)
        personal_phone_index = first_header_index(headers, PERSONAL_PHONE_HEADERS)
        work_zone_index = first_header_index(headers, WORK_ZONE_HEADERS)
        for row_number, row in enumerate(rows, start=2):
            summary["rows"] += 1
            name = cell(row, name_index)
            normalized_name = normalize_representative_name(name)
            if not normalized_name:
                summary["skipped"] += 1
                continue
            contact = db.execute(
                select(RepresentativeContact).where(RepresentativeContact.normalized_name == normalized_name)
            ).scalars().one_or_none()
            created = contact is None
            if contact is None:
                contact = RepresentativeContact(name=name, normalized_name=normalized_name)
            contact.name = name
            contact.work_phone = normalize_phone(cell(row, work_phone_index)) or None
            contact.personal_phone = normalize_phone(cell(row, personal_phone_index)) or None
            contact.work_zone = cell(row, work_zone_index) or None
            contact.is_active = True
            contact.raw_payload = {
                "source_file": path.name,
                "source_sheet": sheet.title,
                "source_row": row_number,
                "imported_at": datetime.now(timezone.utc).isoformat(),
            }
            db.add(contact)
            summary["created" if created else "updated"] += 1
    return summary


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", normalize_text(value).lower().replace("ё", "е")).strip()


def first_header_index(headers: dict[str, int], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        index = headers.get(normalize_header(alias))
        if index is not None:
            return index
    return None


def cell(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return normalize_text(row[index])
