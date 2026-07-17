import hashlib
import hmac
import json
import logging
import math
import os
import re
import time
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time as datetime_time, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from .observability_context import current_correlation_id
from .imports_service import create_import, preview_import
from .excel_importer import reverse_geocode_yandex
from .logistics_calendar_service import is_logistics_non_working_day, resolve_effective_delivery_date
from .logistics_service import build_logistics_report_xlsx, list_logistics_dates
from .models import AuditLog, ImportJob, Order, PendingEvent, SmartupFulfillment
from .schemas import ImportCreate
from .redaction import redact_secrets
from .skladbot_request_dry_run import (
    SKLADBOT_REQUEST_CREATE_EVENT_TYPE,
    create_skladbot_dry_run_for_import,
    create_skladbot_dry_run_for_orders,
    process_pending_skladbot_request_creates,
)
from .smartup_saga import (
    execute_status_sagas,
    imports_from_sagas,
    load_recoverable_fulfillment_events,
    load_slot_sagas,
    mark_skladbot_results,
    normalize_saga_mode,
    prepare_deal_sagas,
    record_shadow_results,
    saga_report,
    smartup_saga_fault,
)
from .spreadsheet_safety import force_workbook_text_literals
from .telegram_routing_contract import (
    TelegramMessageKind,
    load_telegram_routing_contract,
    validate_route_values,
)
from .telegram_output_contract import (
    logistics_report_caption,
    smartup_export_caption,
    smartup_export_filename as export_filename,
)


logger = logging.getLogger(__name__)

SMARTUP_AUTO_IMPORT_EVENT_TYPE = "smartup_auto_import_run"
SMARTUP_LOGISTICS_REPORT_EVENT_TYPE = "smartup_logistics_report"
SMARTUP_CLIENT_EXPORT_EVENT_TYPE = "smartup_client_export"
SMARTUP_AUTO_IMPORT_SOURCE = "smartup_auto"
SMARTUP_EXPORT_REQUEST_PATH = "/b/trade/txs/tdeal/order$export"
SMARTUP_CHANGE_STATUS_PATH = "/b/trade/txs/tdeal/order$change_status"
DEFAULT_SMARTUP_BASE_URL = "https://smartup.online"
DEFAULT_SCHEDULE_TIMES = ("12:00", "15:00", "17:50")
DEFAULT_FINAL_TIME = "17:50"
DEFAULT_TIMEZONE = "Asia/Tashkent"
DEFAULT_DISABLED_WEEKDAYS = (5, 6)
STALE_SMARTUP_SLOT_TIMEOUT = timedelta(minutes=30)
DEFAULT_LOGISTICS_RETRY_BASE_SECONDS = 60
DEFAULT_LOGISTICS_MAX_ATTEMPTS = 5
DEFAULT_LOGISTICS_CATCHUP_DAYS = 2
DEFAULT_LOGISTICS_CLAIM_TIMEOUT_MINUTES = 10
TERMINAL_PAYMENT_CODE = "PYMT:2"
SMARTUP_NEW_STATUS = "B#N"
SMARTUP_WAITING_STATUS = "B#W"
TERMINAL_PAYMENT_NAME = "Терминал"
EXPORT_WORKBOOK_HEADERS = [
    "Дата заказа",
    "Дата отгрузки",
    "Тип оплаты",
    "Клиент",
    "Адрес",
    "Координаты",
    "Торговый представитель",
    "Товары",
    "Кол-во ШТ",
    "Кол-во блок",
    "_pieces_per_block",
    "Цена из файла",
    "Сумма из файла",
    "Цена за блок",
    "Сумма позиции",
    "Статус",
    "ID заказа",
    "ID импорта",
    "Smartup deal_id",
    "Smartup product_id",
    "Smartup status",
    "Smartup delivery_date original",
    "Smartup delivery_date adjusted",
    "Smartup delivery_date adjustment_reason",
    "Smartup delivery_date skipped_dates",
    "Источник файла",
    "Строка файла",
]


class SmartupAutoImportError(Exception):
    pass


class TelegramDeliveryAmbiguousError(SmartupAutoImportError):
    pass


@dataclass(frozen=True)
class SmartupAutoImportConfig:
    enabled: bool = False
    backend_import_enabled: bool = False
    change_status_enabled: bool = False
    process_skladbot_now: bool = False
    saga_mode: str = "disabled"
    schedule_times: tuple[str, ...] = DEFAULT_SCHEDULE_TIMES
    disabled_weekdays: tuple[int, ...] = DEFAULT_DISABLED_WEEKDAYS
    final_time: str = DEFAULT_FINAL_TIME
    logistics_due_time: str = ""
    logistics_retry_base_seconds: int = DEFAULT_LOGISTICS_RETRY_BASE_SECONDS
    logistics_max_attempts: int = DEFAULT_LOGISTICS_MAX_ATTEMPTS
    logistics_catchup_days: int = DEFAULT_LOGISTICS_CATCHUP_DAYS
    logistics_claim_timeout_minutes: int = DEFAULT_LOGISTICS_CLAIM_TIMEOUT_MINUTES
    logistics_route_recovery_export_date: str = ""
    slot_grace_minutes: int = 10
    poll_seconds: int = 30
    timezone_name: str = DEFAULT_TIMEZONE
    output_dir: Path = Path("outputs/smartup_exports")
    smartup_base_url: str = DEFAULT_SMARTUP_BASE_URL
    smartup_username: str = ""
    smartup_password: str = ""
    smartup_project_code: str = ""
    smartup_filial_id: str = ""
    smartup_filial_code: str = ""
    smartup_timeout_seconds: int = 30
    new_status_code: str = SMARTUP_NEW_STATUS
    waiting_status_code: str = SMARTUP_WAITING_STATUS
    terminal_payment_code: str = TERMINAL_PAYMENT_CODE
    pieces_per_block: int = 10
    default_block_price: int = 240000
    client_chat_id: str = ""
    logistics_chat_id: str = ""
    alert_chat_id: str = ""
    legacy_alert_chat_id: str = ""
    admin_chat_ids: tuple[str, ...] = ()
    route_fingerprint_key: str = ""
    environment_name: str = ""
    skladbot_create_requests_mode: str = "dry_run"
    telegram_bot_token: str = ""
    telegram_timeout_seconds: int = 120

    def validate_for_run(self) -> None:
        if not self.smartup_username or not self.smartup_password:
            raise SmartupAutoImportError("SMARTUP_USERNAME и SMARTUP_PASSWORD обязательны для Smartup automation")
        if self.backend_import_enabled and not self.change_status_enabled:
            raise SmartupAutoImportError(
                "SMARTUP_AUTO_IMPORT_BACKEND_IMPORT_ENABLED=true требует "
                "SMARTUP_AUTO_IMPORT_CHANGE_STATUS_ENABLED=true"
            )
        if self.change_status_enabled and not self.waiting_status_code:
            raise SmartupAutoImportError("Не задан Smartup статус ожидания")
        if self.pieces_per_block <= 0:
            raise SmartupAutoImportError("TAKSKLAD_DEFAULT_PIECES_PER_BLOCK должен быть больше нуля")
        if self.default_block_price <= 0:
            raise SmartupAutoImportError("TAKSKLAD_DEFAULT_BLOCK_PRICE должен быть больше нуля")
        if normalize_saga_mode(self.saga_mode) != self.saga_mode:
            raise SmartupAutoImportError(
                "SMARTUP_AUTO_IMPORT_SAGA_MODE должен быть disabled, shadow или enforced"
            )
        route_errors = smartup_production_route_errors(self)
        production_runtime_errors = smartup_production_runtime_errors(self)
        if route_errors or production_runtime_errors:
            raise SmartupAutoImportError("; ".join([*route_errors, *production_runtime_errors]))
        if self.logistics_route_recovery_export_date and not is_iso_date(
            self.logistics_route_recovery_export_date
        ):
            raise SmartupAutoImportError(
                "SMARTUP_AUTO_IMPORT_LOGISTICS_ROUTE_RECOVERY_EXPORT_DATE должен быть YYYY-MM-DD"
            )

    @property
    def effective_logistics_due_time(self) -> str:
        return normalize_text(self.logistics_due_time) or normalize_text(self.final_time) or DEFAULT_FINAL_TIME

    @property
    def production(self) -> bool:
        return normalize_text(self.environment_name).casefold() == "production"

    @property
    def timezone(self) -> ZoneInfo:
        return ZoneInfo(self.timezone_name)


class SmartupClient:
    def __init__(self, config: SmartupAutoImportConfig):
        self.config = config
        self.base_url = config.smartup_base_url.rstrip("/")

    def export_orders(self, export_date: date, *, target_delivery_date: date | None = None) -> dict[str, Any]:
        display_date = format_display_date(export_date)
        delivery_date = format_display_date(target_delivery_date) if target_delivery_date else ""
        deal_date = "" if target_delivery_date else display_date
        payload = {
            "filial_code": self.config.smartup_filial_code,
            "external_id": "",
            "deal_id": "",
            "begin_deal_date": deal_date,
            "end_deal_date": deal_date,
            "delivery_date": delivery_date,
            "begin_created_on": "",
            "end_created_on": "",
            "begin_modified_on": "",
            "end_modified_on": "",
            "statuses": [self.config.new_status_code],
        }
        if self.config.smartup_filial_code:
            payload["filial_codes"] = [{"filial_code": self.config.smartup_filial_code}]
        return self._post(SMARTUP_EXPORT_REQUEST_PATH, payload)

    def change_status(self, deal_ids: list[str], status_code: str) -> dict[str, Any]:
        unique_deal_ids = []
        seen = set()
        for deal_id in deal_ids:
            normalized = normalize_text(deal_id)
            if normalized and normalized not in seen:
                unique_deal_ids.append(normalized)
                seen.add(normalized)
        if not unique_deal_ids:
            return {"successes": [], "errors": [], "submitted": 0}
        payload = {
            "order": [
                {
                    "deal_id": deal_id,
                    "status": status_code,
                }
                for deal_id in unique_deal_ids
            ]
        }
        response = self._post(SMARTUP_CHANGE_STATUS_PATH, payload)
        errors = status_change_errors(response)
        return {
            **response,
            "errors": errors,
            "submitted": len(unique_deal_ids),
            "deal_ids": unique_deal_ids,
            "successful_deal_ids": status_change_successful_deal_ids(response, unique_deal_ids),
            "failed_deal_ids": status_change_failed_deal_ids(response),
            "status": status_code,
        }

    def get_deal_statuses(self, deal_ids: list[str]) -> dict[str, str]:
        """Read back exact Smartup deals before deciding whether a write may be retried."""
        unique_deal_ids: list[str] = []
        seen: set[str] = set()
        for deal_id in deal_ids:
            normalized = normalize_text(deal_id)
            if normalized and normalized not in seen:
                unique_deal_ids.append(normalized)
                seen.add(normalized)

        statuses: dict[str, str] = {}
        readable_statuses = list(
            dict.fromkeys(
                status
                for status in (self.config.new_status_code, self.config.waiting_status_code)
                if normalize_text(status)
            )
        )
        for deal_id in unique_deal_ids:
            payload = {
                "filial_code": self.config.smartup_filial_code,
                "external_id": "",
                "deal_id": deal_id,
                "begin_deal_date": "",
                "end_deal_date": "",
                "delivery_date": "",
                "begin_created_on": "",
                "end_created_on": "",
                "begin_modified_on": "",
                "end_modified_on": "",
                "statuses": readable_statuses,
            }
            if self.config.smartup_filial_code:
                payload["filial_codes"] = [{"filial_code": self.config.smartup_filial_code}]
            response = self._post(SMARTUP_EXPORT_REQUEST_PATH, payload)
            exact_orders = [
                order
                for order in extract_smartup_orders(response)
                if normalize_text(order.get("deal_id")) == deal_id
            ]
            if exact_orders:
                statuses[deal_id] = smartup_status(exact_orders[0])
        return statuses

    def _post(self, path: str, payload: dict[str, Any]) -> dict[str, Any]:
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json; charset=utf-8",
            "X-Correlation-ID": current_correlation_id(),
        }
        if self.config.smartup_project_code:
            headers["project_code"] = self.config.smartup_project_code
        if self.config.smartup_filial_id:
            headers["filial_id"] = self.config.smartup_filial_id
        url = f"{self.base_url}{path}"
        with httpx.Client(timeout=self.config.smartup_timeout_seconds) as client:
            response = client.post(
                url,
                headers=headers,
                json=payload,
                auth=(self.config.smartup_username, self.config.smartup_password),
            )
            try:
                response.raise_for_status()
            except httpx.HTTPStatusError as exc:
                detail = response.text[:500] if response is not None else ""
                raise SmartupAutoImportError(
                    f"Smartup API {path} failed: HTTP {exc.response.status_code} {detail}"
                ) from None
        try:
            data = response.json()
        except ValueError as exc:
            raise SmartupAutoImportError(f"Smartup API {path} вернул не JSON") from exc
        if not isinstance(data, dict):
            raise SmartupAutoImportError(f"Smartup API {path} вернул неожиданный формат")
        return data


def smartup_source_scope(config: SmartupAutoImportConfig) -> str:
    """Stable non-secret identity for a Smartup project/filial boundary."""
    identity = "|".join((
        normalize_text(config.smartup_base_url).rstrip("/").casefold(),
        normalize_text(config.smartup_project_code).casefold(),
        normalize_text(config.smartup_filial_id).casefold(),
        normalize_text(config.smartup_filial_code).casefold(),
    ))
    return f"smartup:{hashlib.sha256(identity.encode('utf-8')).hexdigest()}"


STATUS_CHANGE_DEAL_ID_KEYS = ("deal_id", "dealId", "external_id", "externalId", "order_id", "id", "code")


def status_change_errors(response: Any) -> list[Any]:
    if not isinstance(response, dict):
        return []
    errors = response.get("errors")
    if not errors:
        errors = response.get("error")
    if not errors:
        return []
    return errors if isinstance(errors, list) else [errors]


def status_change_has_errors(response: Any) -> bool:
    return bool(status_change_errors(response))


def status_change_successful_deal_ids(response: Any, submitted_deal_ids: list[str]) -> list[str]:
    submitted = unique_values(submitted_deal_ids)
    submitted_set = set(submitted)
    if not status_change_has_errors(response):
        return submitted
    if not isinstance(response, dict):
        return []
    explicit = response.get("successful_deal_ids")
    if isinstance(explicit, list):
        return [deal_id for deal_id in unique_values(explicit) if deal_id in submitted_set]
    successful = []
    for item in status_change_success_items(response):
        deal_id = status_change_item_deal_id(item)
        if deal_id and deal_id in submitted_set:
            successful.append(deal_id)
    return unique_values(successful)


def status_change_failed_deal_ids(response: Any) -> list[str]:
    return unique_values(
        status_change_item_deal_id(item)
        for item in status_change_errors(response)
        if status_change_item_deal_id(item)
    )


def status_change_success_items(response: dict[str, Any]) -> list[Any]:
    for key in ("successes", "success", "updated", "orders"):
        value = response.get(key)
        if isinstance(value, list):
            return value
    return []


def status_change_item_deal_id(item: Any) -> str:
    if isinstance(item, dict):
        for key in STATUS_CHANGE_DEAL_ID_KEYS:
            value = normalize_text(item.get(key))
            if value:
                return value
        return ""
    return normalize_text(item)


def status_change_error_message(response: dict[str, Any]) -> str:
    return (
        "Smartup status change failed: "
        f"errors={redact_json(status_change_errors(response))}; "
        f"successful_deal_ids={redact_json(response.get('successful_deal_ids') or [])}; "
        f"failed_deal_ids={redact_json(response.get('failed_deal_ids') or [])}"
    )


class TelegramDocumentSender:
    def __init__(self, token: str, timeout_seconds: int = 120):
        self.token = normalize_text(token)
        self.timeout_seconds = timeout_seconds

    @property
    def configured(self) -> bool:
        return bool(self.token)

    def send_document(self, chat_id: str, content: bytes, filename: str, caption: str = "") -> dict[str, Any]:
        if not self.configured:
            raise SmartupAutoImportError("TELEGRAM_BOT_TOKEN не задан")
        with httpx.Client(timeout=self.timeout_seconds) as client:
            response = client.post(
                f"https://api.telegram.org/bot{self.token}/sendDocument",
                data={"chat_id": chat_id, "caption": caption[:1000]},
                files={"document": (filename, content)},
            )
            response.raise_for_status()
        payload = response.json()
        if not payload.get("ok"):
            raise SmartupAutoImportError(f"Telegram sendDocument failed: {redact_json(payload)}")
        return payload

    def send_message(self, chat_id: str, text_value: str) -> dict[str, Any]:
        if not self.configured:
            raise SmartupAutoImportError("TELEGRAM_BOT_TOKEN не задан")
        with httpx.Client(timeout=self.timeout_seconds) as client:
            response = client.post(
                f"https://api.telegram.org/bot{self.token}/sendMessage",
                data={
                    "chat_id": chat_id,
                    "text": text_value[:4096],
                    "disable_web_page_preview": "true",
                },
            )
            response.raise_for_status()
        payload = response.json()
        if not payload.get("ok"):
            raise SmartupAutoImportError(f"Telegram sendMessage failed: {redact_json(payload)}")
        return payload


def load_smartup_auto_import_config(environ: dict[str, str] | None = None) -> SmartupAutoImportConfig:
    environ = environ or os.environ
    return SmartupAutoImportConfig(
        enabled=parse_bool(environ.get("SMARTUP_AUTO_IMPORT_ENABLED"), default=False),
        backend_import_enabled=parse_bool(environ.get("SMARTUP_AUTO_IMPORT_BACKEND_IMPORT_ENABLED"), default=False),
        change_status_enabled=parse_bool(environ.get("SMARTUP_AUTO_IMPORT_CHANGE_STATUS_ENABLED"), default=False),
        process_skladbot_now=parse_bool(environ.get("SMARTUP_AUTO_IMPORT_PROCESS_SKLADBOT_NOW"), default=False),
        saga_mode=normalize_saga_mode(environ.get("SMARTUP_AUTO_IMPORT_SAGA_MODE")),
        schedule_times=parse_schedule_times(environ.get("SMARTUP_AUTO_IMPORT_TIMES")),
        disabled_weekdays=parse_disabled_weekdays(environ.get("SMARTUP_AUTO_IMPORT_DISABLED_WEEKDAYS")),
        final_time=normalize_text(environ.get("SMARTUP_AUTO_IMPORT_FINAL_TIME")) or DEFAULT_FINAL_TIME,
        logistics_due_time=normalize_text(environ.get("SMARTUP_AUTO_IMPORT_LOGISTICS_DUE_TIME")),
        logistics_retry_base_seconds=max(
            1,
            parse_int(
                environ.get("SMARTUP_AUTO_IMPORT_LOGISTICS_RETRY_BASE_SECONDS"),
                DEFAULT_LOGISTICS_RETRY_BASE_SECONDS,
            ),
        ),
        logistics_max_attempts=max(
            1,
            parse_int(
                environ.get("SMARTUP_AUTO_IMPORT_LOGISTICS_MAX_ATTEMPTS"),
                DEFAULT_LOGISTICS_MAX_ATTEMPTS,
            ),
        ),
        logistics_catchup_days=max(
            0,
            parse_int(
                environ.get("SMARTUP_AUTO_IMPORT_LOGISTICS_CATCHUP_DAYS"),
                DEFAULT_LOGISTICS_CATCHUP_DAYS,
            ),
        ),
        logistics_claim_timeout_minutes=max(
            1,
            parse_int(
                environ.get("SMARTUP_AUTO_IMPORT_LOGISTICS_CLAIM_TIMEOUT_MINUTES"),
                DEFAULT_LOGISTICS_CLAIM_TIMEOUT_MINUTES,
            ),
        ),
        logistics_route_recovery_export_date=normalize_text(
            environ.get("SMARTUP_AUTO_IMPORT_LOGISTICS_ROUTE_RECOVERY_EXPORT_DATE")
        ),
        slot_grace_minutes=max(1, parse_int(environ.get("SMARTUP_AUTO_IMPORT_SLOT_GRACE_MINUTES"), 10)),
        poll_seconds=max(30, parse_int(environ.get("SMARTUP_AUTO_IMPORT_POLL_SECONDS"), 30)),
        timezone_name=normalize_text(environ.get("TAKSKLAD_TIMEZONE")) or DEFAULT_TIMEZONE,
        output_dir=Path(normalize_text(environ.get("SMARTUP_AUTO_IMPORT_OUTPUT_DIR")) or "outputs/smartup_exports"),
        smartup_base_url=normalize_text(environ.get("SMARTUP_BASE_URL")) or DEFAULT_SMARTUP_BASE_URL,
        smartup_username=normalize_text(environ.get("SMARTUP_USERNAME")),
        smartup_password=normalize_text(environ.get("SMARTUP_PASSWORD")),
        smartup_project_code=normalize_text(environ.get("SMARTUP_PROJECT_CODE")),
        smartup_filial_id=normalize_text(environ.get("SMARTUP_FILIAL_ID")),
        smartup_filial_code=normalize_text(environ.get("SMARTUP_FILIAL_CODE")),
        smartup_timeout_seconds=max(5, parse_int(environ.get("SMARTUP_TIMEOUT_SECONDS"), 30)),
        new_status_code=normalize_text(environ.get("SMARTUP_AUTO_IMPORT_NEW_STATUS_CODE")) or SMARTUP_NEW_STATUS,
        waiting_status_code=normalize_text(environ.get("SMARTUP_AUTO_IMPORT_WAITING_STATUS_CODE"))
        or SMARTUP_WAITING_STATUS,
        terminal_payment_code=normalize_text(environ.get("SMARTUP_AUTO_IMPORT_TERMINAL_PAYMENT_CODE"))
        or TERMINAL_PAYMENT_CODE,
        pieces_per_block=max(1, parse_int(environ.get("TAKSKLAD_DEFAULT_PIECES_PER_BLOCK"), 10)),
        default_block_price=max(1, parse_int(environ.get("TAKSKLAD_DEFAULT_BLOCK_PRICE"), 240000)),
        client_chat_id=normalize_text(environ.get("SMARTUP_AUTO_IMPORT_CLIENT_CHAT_ID")),
        logistics_chat_id=normalize_text(environ.get("SMARTUP_AUTO_IMPORT_LOGISTICS_CHAT_ID")),
        alert_chat_id=normalize_text(environ.get("TAKSKLAD_AUTOMATION_ALERT_CHAT_ID")),
        legacy_alert_chat_id=normalize_text(environ.get("SMARTUP_AUTO_IMPORT_ALERT_CHAT_ID")),
        admin_chat_ids=parse_csv_values(environ.get("TELEGRAM_ADMIN_CHAT_IDS")),
        route_fingerprint_key=normalize_text(environ.get("SMARTUP_AUTO_IMPORT_ROUTE_FINGERPRINT_KEY")),
        environment_name=normalize_text(environ.get("TAKSKLAD_ENV")),
        skladbot_create_requests_mode=(
            normalize_text(environ.get("SKLADBOT_CREATE_REQUESTS_MODE")).casefold() or "dry_run"
        ),
        telegram_bot_token=normalize_text(environ.get("TELEGRAM_BOT_TOKEN")),
        telegram_timeout_seconds=max(5, parse_int(environ.get("TELEGRAM_WORKER_FILE_TIMEOUT_SECONDS"), 120)),
    )


def run_due_smartup_auto_imports(
    db: Session,
    config: SmartupAutoImportConfig | None = None,
    *,
    now: datetime | None = None,
    smartup_client: Any | None = None,
    telegram_sender: Any | None = None,
) -> list[dict[str, Any]]:
    config = config or load_smartup_auto_import_config()
    if not config.enabled:
        return [{"status": "disabled"}]
    config.validate_for_run()
    local_now = normalize_local_now(now, config.timezone)
    results: list[dict[str, Any]] = []
    slot_failure_count = 0
    if local_now.weekday() in config.disabled_weekdays:
        results.append({
            "status": "idle",
            "reason": "weekday_disabled",
            "weekday": local_now.weekday(),
            "now": local_now.isoformat(),
        })
    else:
        target_delivery_dates = scheduled_smartup_target_delivery_dates(db, local_now.date(), config)
        for slot in config.schedule_times:
            if not is_slot_due(local_now, slot, config.slot_grace_minutes):
                continue
            for target_delivery_date in target_delivery_dates:
                try:
                    results.append(run_scheduled_smartup_auto_import_slot(
                        db,
                        config,
                        slot_label=slot,
                        now=local_now,
                        target_delivery_date=target_delivery_date,
                        smartup_client=smartup_client,
                        telegram_sender=telegram_sender,
                    ))
                except Exception as exc:
                    slot_failure_count += 1
                    logger.error(
                        "Smartup scheduled slot failed; logistics dependency gate remains closed "
                        "slot=%s error_class=%s",
                        slot,
                        type(exc).__name__,
                    )
                    results.append({
                        "status": "failed",
                        "reason": "smartup_slot_failed",
                        "export_date": local_now.date().isoformat(),
                        "slot": slot,
                        "target_delivery_date": target_delivery_date.isoformat(),
                        "error": sanitize_automation_error_text(exc, limit=500),
                    })

    results.extend(run_due_smartup_logistics_reports(
        db,
        config,
        now=local_now,
        telegram_sender=telegram_sender,
    ))
    if slot_failure_count:
        raise SmartupAutoImportError(
            f"Smartup scheduled slots failed: {slot_failure_count}; "
            "logistics dependency gate was evaluated fail-closed"
        )
    return results or [{"status": "idle", "now": local_now.isoformat()}]


def run_due_smartup_logistics_reports(
    db: Session,
    config: SmartupAutoImportConfig,
    *,
    now: datetime | None = None,
    telegram_sender: Any | None = None,
) -> list[dict[str, Any]]:
    """Deliver only after durable proof that the full Smartup cycle is terminal."""
    if not config.backend_import_enabled:
        return []
    local_now = normalize_local_now(now, config.timezone)
    due_time = parse_slot_time(config.effective_logistics_due_time)
    results: list[dict[str, Any]] = []
    for days_ago in range(config.logistics_catchup_days, -1, -1):
        export_date = local_now.date() - timedelta(days=days_ago)
        due_at = datetime.combine(export_date, due_time, tzinfo=config.timezone)
        if local_now < due_at:
            continue
        dependency = smartup_logistics_dependency_proof(db, config, export_date)
        if dependency["status"] != "ready":
            if dependency["reason"] != "final_cycle_event_failed":
                queue_smartup_logistics_dependency_alert(db, export_date, dependency["reason"])
            results.append({
                "status": "logistics_blocked",
                "provenance": "auto_smartup",
                "export_date": export_date.isoformat(),
                "due_time": config.effective_logistics_due_time,
                "reason": dependency["reason"],
                "dependency": dependency,
                "logistics_reports": [],
            })
            continue
        delivery_dates = delivery_dates_for_auto_logistics(db, export_date, config)
        if not delivery_dates:
            continue
        results.append({
            "status": "logistics_due",
            "provenance": "auto_smartup",
            "export_date": export_date.isoformat(),
            "due_time": config.effective_logistics_due_time,
            "delivery_dates": delivery_dates,
            "logistics_reports": send_final_logistics_reports(
                db,
                config,
                export_date=export_date,
                telegram_sender=telegram_sender,
                extra_delivery_dates=delivery_dates,
                now=local_now,
            ),
        })
    return results


def smartup_logistics_dependency_proof(
    db: Session,
    config: SmartupAutoImportConfig,
    export_date: date,
) -> dict[str, Any]:
    expected_targets = scheduled_smartup_target_delivery_dates(db, export_date, config)
    expected_slots = load_telegram_routing_contract().route_for(
        TelegramMessageKind.SMARTUP_CLIENT_EXPORT
    ).schedules
    completed_cycles = 0
    order_ids: set[str] = set()
    delivery_dates: set[str] = set()

    for slot in expected_slots:
        for target_delivery_date in expected_targets:
            key = smartup_slot_idempotency_key(
                export_date,
                slot,
                target_delivery_date=target_delivery_date,
            )
            event = db.execute(
                select(PendingEvent).where(PendingEvent.idempotency_key == key)
            ).scalar_one_or_none()
            if event is None:
                return _blocked_dependency("final_cycle_event_missing", completed_cycles, len(order_ids))
            if event.status != "completed":
                reason = (
                    "final_cycle_event_failed"
                    if event.status in {"failed", "blocked", "dead", "error", "cancelled"}
                    else "final_cycle_event_not_terminal"
                )
                return _blocked_dependency(reason, completed_cycles, len(order_ids))
            result = (event.payload or {}).get("result")
            if not isinstance(result, dict):
                return _blocked_dependency("final_cycle_result_missing", completed_cycles, len(order_ids))
            run_status = normalize_text(result.get("status"))
            if run_status == "no_orders":
                completed_cycles += 1
                continue
            if run_status != "completed":
                return _blocked_dependency("final_cycle_result_not_successful", completed_cycles, len(order_ids))
            client_status = normalize_text((result.get("client_export") or {}).get("status"))
            if client_status not in {"sent", "already_sent"}:
                return _blocked_dependency("client_export_not_terminal", completed_cycles, len(order_ids))
            imports = result.get("imports")
            if not isinstance(imports, list) or not imports:
                return _blocked_dependency("imports_missing", completed_cycles, len(order_ids))
            for imported in imports:
                if not isinstance(imported, dict):
                    return _blocked_dependency("import_result_invalid", completed_cycles, len(order_ids))
                if normalize_text(imported.get("status")) not in {"completed", "deduplicated"}:
                    return _blocked_dependency("import_not_completed", completed_cycles, len(order_ids))
                if int(imported.get("invalid_rows") or 0) or imported.get("errors"):
                    return _blocked_dependency("import_partial_or_invalid", completed_cycles, len(order_ids))
                resolved = {
                    normalize_text(value)
                    for value in imported.get("resolved_order_ids") or []
                    if normalize_text(value)
                }
                if not resolved:
                    return _blocked_dependency("import_order_proof_missing", completed_cycles, len(order_ids))
                order_ids.update(resolved)
            delivery_dates.update(
                normalize_text(value)
                for value in result.get("delivery_dates") or []
                if normalize_text(value)
            )
            completed_cycles += 1

    for order_id in sorted(order_ids):
        try:
            order = db.get(Order, uuid.UUID(order_id))
        except ValueError:
            order = None
        if order is None:
            return _blocked_dependency("canonical_order_missing", completed_cycles, len(order_ids))
        create_event_id = normalize_text((order.raw_payload or {}).get("skladbot_create_event_id"))
        try:
            create_event = db.get(PendingEvent, uuid.UUID(create_event_id)) if create_event_id else None
        except ValueError:
            create_event = None
        if create_event is None or create_event.event_type != SKLADBOT_REQUEST_CREATE_EVENT_TYPE:
            return _blocked_dependency("fulfillment_event_missing", completed_cycles, len(order_ids))
        if create_event.status != "completed":
            reason = (
                "fulfillment_failed_or_ambiguous"
                if create_event.status in {"failed", "blocked", "dead", "error", "cancelled"}
                else "fulfillment_not_terminal"
            )
            return _blocked_dependency(reason, completed_cycles, len(order_ids))
        create_status = normalize_text(
            ((create_event.payload or {}).get("last_result") or {}).get("status")
        )
        if create_status not in {"created", "created_recovered", "already_linked"}:
            return _blocked_dependency("fulfillment_terminal_result_invalid", completed_cycles, len(order_ids))

    return {
        "status": "ready",
        "reason": "all_smartup_slots_imports_client_exports_and_fulfillments_terminal",
        "expected_cycles": len(expected_slots) * len(expected_targets),
        "completed_cycles": completed_cycles,
        "orders_proven": len(order_ids),
        "delivery_dates": sorted(delivery_dates),
    }


def _blocked_dependency(reason: str, completed_cycles: int, orders_proven: int) -> dict[str, Any]:
    return {
        "status": "blocked",
        "reason": reason,
        "completed_cycles": completed_cycles,
        "orders_proven": orders_proven,
    }


def queue_smartup_logistics_dependency_alert(
    db: Session,
    export_date: date,
    reason: str,
) -> PendingEvent:
    """Queue at most one sanitized admin-only alert for a blocked daily cycle."""
    key = f"telegram:notification:v1:smartup_logistics_dependency:{export_date.isoformat()}"
    existing = db.execute(
        select(PendingEvent).where(PendingEvent.idempotency_key == key)
    ).scalar_one_or_none()
    if existing is not None:
        return existing
    event = PendingEvent(
        event_type="telegram_notification",
        status="pending",
        idempotency_key=key,
        payload={
            "kind": "smartup_logistics_dependency_alert",
            "route_role": "admin",
            "export_date": export_date.isoformat(),
            "reason_code": normalize_text(reason),
            "text": "\n".join([
                "TakSklad: логистический отчёт заблокирован",
                f"Дата цикла: {format_display_date(export_date)}",
                f"Причина: {normalize_text(reason)}",
                "Отправка в logistics не выполнялась.",
            ]),
        },
    )
    db.add(event)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        existing = db.execute(
            select(PendingEvent).where(PendingEvent.idempotency_key == key)
        ).scalar_one()
        return existing
    db.refresh(event)
    return event


def run_scheduled_smartup_auto_import_slot(
    db: Session,
    config: SmartupAutoImportConfig,
    *,
    slot_label: str,
    now: datetime | None = None,
    target_delivery_date: date | None = None,
    smartup_client: Any | None = None,
    telegram_sender: Any | None = None,
) -> dict[str, Any]:
    local_now = normalize_local_now(now, config.timezone)
    export_date = local_now.date()
    if config.production:
        allowed_slots = load_telegram_routing_contract().route_for(
            TelegramMessageKind.SMARTUP_CLIENT_EXPORT
        ).schedules
        if slot_label not in allowed_slots:
            raise SmartupAutoImportConfigError("Smartup slot is absent from Telegram routing contract")
    lock_acquired, lock_connection = acquire_smartup_slot_advisory_lock(
        db,
        export_date,
        slot_label,
        target_delivery_date=target_delivery_date,
    )
    if not lock_acquired:
        return {
            "status": "skipped",
            "reason": "slot_locked",
            "export_date": export_date.isoformat(),
            "slot": slot_label,
            "target_delivery_date": target_delivery_date.isoformat() if target_delivery_date else "",
        }
    try:
        event, skipped = claim_smartup_slot(
            db,
            export_date,
            slot_label,
            local_now,
            target_delivery_date=target_delivery_date,
        )
        if skipped:
            return skipped
        try:
            result = run_smartup_auto_import_once(
                db,
                config,
                now=local_now,
                slot_label=slot_label,
                target_delivery_date=target_delivery_date,
                smartup_client=smartup_client,
                telegram_sender=telegram_sender,
            )
        except Exception as exc:
            db.rollback()
            if isinstance(exc, TelegramDeliveryAmbiguousError):
                mark_smartup_slot_blocked(db, event.id, exc)
            else:
                mark_smartup_slot_failed(db, event.id, exc)
            notify_smartup_automation_error(
                db,
                config,
                export_date=export_date,
                slot_label=slot_label,
                exc=exc,
                telegram_sender=telegram_sender,
            )
            raise
        mark_smartup_slot_completed(db, event.id, result)
        return result
    finally:
        release_smartup_slot_advisory_lock(
            lock_connection,
            export_date,
            slot_label,
            target_delivery_date=target_delivery_date,
        )


def run_smartup_auto_import_once(
    db: Session,
    config: SmartupAutoImportConfig,
    *,
    now: datetime | None = None,
    slot_label: str = "",
    target_delivery_date: date | None = None,
    smartup_client: Any | None = None,
    telegram_sender: Any | None = None,
) -> dict[str, Any]:
    config.validate_for_run()
    local_now = normalize_local_now(now, config.timezone)
    export_date = local_now.date()
    export_date_display = format_display_date(export_date)
    client = smartup_client or SmartupClient(config)
    require_fake_saga_clients(config, client, telegram_sender)

    if config.saga_mode == "enforced":
        prepare_orphaned_smartup_sagas(
            db,
            config,
            export_date=export_date,
            slot_label=slot_label,
            target_delivery_date=target_delivery_date,
        )
        recovery = recover_smartup_slot_sagas(
            db,
            config,
            client=client,
            export_date=export_date,
            slot_label=slot_label,
            target_delivery_date=target_delivery_date,
        )
        if recovery is not None:
            return recovery

    raw_response = client.export_orders(export_date, target_delivery_date=target_delivery_date)
    raw_orders = extract_smartup_orders(raw_response)
    selected_orders = filter_smartup_orders(
        raw_orders,
        export_date,
        config,
        target_delivery_date=target_delivery_date,
    )
    part = next_export_part(config.output_dir, export_date)

    if not selected_orders:
        result = {
            "status": "no_orders",
            "slot": slot_label,
            "export_date": export_date.isoformat(),
            "target_delivery_date": target_delivery_date.isoformat() if target_delivery_date else "",
            "raw_orders": len(raw_orders),
            "selected_orders": 0,
            "part": part,
            "skladbot_processing": {"status": "skipped"},
            "logistics_reports": [],
            "smartup_saga": saga_report([], config.saga_mode),
        }
        if is_final_slot(slot_label, config):
            if config.process_skladbot_now:
                result["skladbot_processing"] = process_pending_skladbot_request_creates(db)
        record_smartup_audit(db, "smartup_auto_import_no_orders", result)
        return result

    filename = export_filename(export_date, part)
    import_rows = build_import_rows(selected_orders, export_date, filename, config, db=db)
    if not import_rows:
        raise SmartupAutoImportError("Smartup export не дал строк для импорта после фильтра")
    grouped_rows = group_rows_by_delivery_date(import_rows)
    export_path, workbook_sha256 = write_export_workbook(config.output_dir, export_date, filename, import_rows)
    source_batch_key = smartup_source_batch_key(export_date, part, workbook_sha256)
    delivery_dates = sorted(grouped_rows)
    audit_payload = {
        "version": 1,
        "slot": slot_label,
        "export_date": export_date.isoformat(),
        "export_date_display": export_date_display,
        "target_delivery_date": target_delivery_date.isoformat() if target_delivery_date else "",
        "part": part,
        "filename": filename,
        "source_batch_key": source_batch_key,
        "export_path": str(export_path),
        "sha256": workbook_sha256,
        "raw_orders": len(raw_orders),
        "selected_orders": len(selected_orders),
        "rows": len(import_rows),
        "deal_ids": unique_deal_ids(selected_orders),
        "delivery_dates": delivery_dates,
        "delivery_date_adjustments": delivery_date_adjustments(import_rows),
        "previews": [],
        "backend_import_enabled": config.backend_import_enabled,
        "change_status_enabled": config.change_status_enabled,
        "process_skladbot_now": config.process_skladbot_now,
        "saga_mode": config.saga_mode,
    }
    audit_path = write_export_audit(config.output_dir, export_date, filename, audit_payload)
    audit_payload["audit_path"] = str(audit_path)
    update_export_audit(audit_path, audit_payload)
    try:
        previews = preview_delivery_groups(
            db,
            grouped_rows,
            filename,
            workbook_sha256,
            source_batch_key=source_batch_key,
        )
        audit_payload["previews"] = previews
        update_export_audit(audit_path, audit_payload)
        assert_previews_safe(previews)
    except Exception as exc:
        failed_preview = {
            **audit_payload,
            "status": "failed_preview",
            "error": str(exc)[:1000],
        }
        update_export_audit(audit_path, failed_preview)
        record_smartup_audit(db, "smartup_auto_import_preview_failed", failed_preview)
        raise

    if not config.backend_import_enabled:
        result = {
            **audit_payload,
            "status": "shadow_preview",
            "imports": [],
            "status_change": {"status": "skipped"},
            "skladbot_processing": {"status": "skipped"},
            "logistics_reports": [],
            "smartup_saga": saga_report([], config.saga_mode),
        }
        record_smartup_audit(db, "smartup_auto_import_shadow_preview", result)
        return result

    imports = create_delivery_group_imports(
        db,
        grouped_rows,
        filename,
        workbook_sha256,
        source_batch_key=source_batch_key,
        export_date=export_date,
        part=part,
        slot_label=slot_label,
        source_chat_id=config.client_chat_id,
        target_delivery_date=target_delivery_date,
        saga_mode=config.saga_mode,
    )
    saga_events = []
    if config.saga_mode in {"shadow", "enforced"}:
        saga_events = prepare_deal_sagas(
            db,
            imports,
            source_batch_key=source_batch_key,
            target_status=config.waiting_status_code,
            export_date=export_date,
            slot_label=slot_label,
            target_delivery_date=target_delivery_date,
            mode=config.saga_mode,
            source_scope=smartup_source_scope(config),
        )
    if config.saga_mode == "enforced":
        status_change = execute_status_sagas(
            db,
            client,
            saga_events,
            target_status=config.waiting_status_code,
            successful_ids=status_change_successful_deal_ids,
            failed_ids=status_change_failed_deal_ids,
        )
    else:
        status_change = client.change_status(unique_deal_ids(selected_orders), config.waiting_status_code)
        if config.saga_mode == "shadow":
            record_shadow_results(
                db,
                saga_events,
                status_change,
                status_change_successful_deal_ids,
                status_change_failed_deal_ids,
            )
    if config.saga_mode == "enforced":
        smartup_saga_fault("local_state_to_skladbot", "batch")
    imports = queue_skladbot_after_smartup_status(
        db,
        imports,
        successful_deal_ids=status_change_successful_deal_ids(status_change, unique_deal_ids(selected_orders)),
        commit=not bool(saga_events),
    )
    if saga_events:
        mark_skladbot_results(db, saga_events)
        assert_enforced_sagas_complete(saga_events, config.saga_mode)
    if status_change_has_errors(status_change):
        result = {
            **audit_payload,
            "status": "failed_status_change",
            "imports": imports,
            "status_change": status_change,
            "client_export": {"status": "skipped"},
            "skladbot_processing": {"status": "skipped"},
            "logistics_reports": [],
            "smartup_saga": saga_report(saga_events, config.saga_mode),
        }
        update_export_audit(audit_path, result)
        record_smartup_audit(db, "smartup_auto_import_status_change_failed", result)
        raise SmartupAutoImportError(status_change_error_message(status_change))
    client_export = send_smartup_export_to_client(
        db,
        config,
        export_path=export_path,
        filename=filename,
        export_date=export_date,
        slot_label=slot_label,
        part=part,
        selected_orders=len(selected_orders),
        rows=len(import_rows),
        delivery_dates=delivery_dates,
        target_delivery_date=target_delivery_date,
        telegram_sender=telegram_sender,
    )

    skladbot_processing = {"status": "skipped"}
    if config.process_skladbot_now:
        skladbot_processing = process_pending_skladbot_request_creates(db)

    result = {
        **audit_payload,
        "status": "completed",
        "imports": imports,
        "status_change": status_change,
        "client_export": client_export,
        "skladbot_processing": skladbot_processing,
        "logistics_reports": [],
        "smartup_saga": saga_report(saga_events, config.saga_mode),
    }
    update_export_audit(audit_path, result)
    record_smartup_audit(db, "smartup_auto_import_completed", result)
    return result


def require_fake_saga_clients(
    config: SmartupAutoImportConfig,
    smartup_client: Any | None,
    telegram_sender: Any | None,
) -> None:
    del telegram_sender  # Telegram delivery is independent from saga safety.
    if config.saga_mode == "disabled":
        return
    if smartup_client is None or not callable(getattr(smartup_client, "change_status", None)):
        raise SmartupAutoImportError("Smartup saga требует client.change_status")
    if config.saga_mode == "enforced" and not callable(getattr(smartup_client, "get_deal_statuses", None)):
        raise SmartupAutoImportError("Smartup saga enforced требует client.get_deal_statuses")
    if config.process_skladbot_now:
        raise SmartupAutoImportError(
            "Smartup saga в Phase 10 запрещает немедленную обработку SkladBot; разрешена только durable queue"
        )


def recover_smartup_slot_sagas(
    db: Session,
    config: SmartupAutoImportConfig,
    *,
    client: Any,
    export_date: date,
    slot_label: str,
    target_delivery_date: date | None,
) -> dict[str, Any] | None:
    saga_events = load_slot_sagas(
        db,
        export_date=export_date,
        slot_label=slot_label,
        target_delivery_date=target_delivery_date,
    )
    if not saga_events:
        return None
    status_change = execute_status_sagas(
        db,
        client,
        saga_events,
        target_status=config.waiting_status_code,
        successful_ids=status_change_successful_deal_ids,
        failed_ids=status_change_failed_deal_ids,
    )
    imports = imports_from_sagas(saga_events)
    smartup_saga_fault("local_state_to_skladbot", "batch")
    imports = queue_skladbot_after_smartup_status(
        db,
        imports,
        successful_deal_ids=status_change_successful_deal_ids(
            status_change,
            [event.aggregate_id for event in saga_events],
        ),
        commit=False,
    )
    mark_skladbot_results(db, saga_events)
    assert_enforced_sagas_complete(saga_events, config.saga_mode)
    result = {
        "status": "completed_recovery" if not status_change_has_errors(status_change) else "failed_status_change",
        "slot": slot_label,
        "export_date": export_date.isoformat(),
        "target_delivery_date": target_delivery_date.isoformat() if target_delivery_date else "",
        "raw_orders": 0,
        "selected_orders": 0,
        "imports": imports,
        "status_change": status_change,
        "client_export": {"status": "skipped", "reason": "saga_recovery"},
        "skladbot_processing": {"status": "skipped"},
        "logistics_reports": [],
        "smartup_saga": saga_report(saga_events, config.saga_mode),
    }
    record_smartup_audit(db, "smartup_auto_import_saga_recovery", result)
    if status_change_has_errors(status_change):
        raise SmartupAutoImportError(status_change_error_message(status_change))
    return result


def sweep_incomplete_smartup_fulfillments(
    db: Session,
    config: SmartupAutoImportConfig,
    *,
    client: Any | None = None,
    limit: int = 50,
) -> dict[str, Any]:
    """Recover durable Smartup/SkladBot gaps without depending on the original slot."""
    if config.saga_mode != "enforced":
        return {"status": "skipped", "reason": "saga_not_enforced", "workflows": 0}
    lock_acquired, lock_connection = acquire_smartup_fulfillment_sweep_lock(db)
    if not lock_acquired:
        return {"status": "skipped", "reason": "sweep_locked", "workflows": 0}
    try:
        smartup_client = client or SmartupClient(config)
        events = load_recoverable_fulfillment_events(
            db,
            source_scope=smartup_source_scope(config),
            limit=limit,
        )
        if not events:
            return {"status": "idle", "workflows": 0, "queued": 0, "unresolved": 0}
        status_change = execute_status_sagas(
            db,
            smartup_client,
            events,
            target_status=config.waiting_status_code,
            successful_ids=status_change_successful_deal_ids,
            failed_ids=status_change_failed_deal_ids,
        )
        successful_deal_ids = status_change_successful_deal_ids(
            status_change,
            [event.aggregate_id for event in events],
        )
        imports = queue_skladbot_after_smartup_status(
            db,
            imports_from_sagas(events),
            successful_deal_ids=successful_deal_ids,
            commit=False,
        )
        mark_skladbot_results(db, events)
        unresolved = [
            event
            for event in events
            if normalize_text((event.payload or {}).get("saga_state")) != "skladbot_queued"
        ]
        result = {
            "status": "completed" if not unresolved else "incomplete",
            "workflows": len(events),
            "queued": len(events) - len(unresolved),
            "unresolved": len(unresolved),
            "imports": imports,
            "status_change": status_change,
        }
        record_smartup_audit(db, "smartup_fulfillment_sweep", result)
        return result
    finally:
        release_smartup_fulfillment_sweep_lock(lock_connection)


def acquire_smartup_fulfillment_sweep_lock(db: Session) -> tuple[bool, Any | None]:
    bind = db.get_bind()
    if getattr(getattr(bind, "dialect", None), "name", "") != "postgresql":
        return True, None
    connection = bind.connect()
    try:
        acquired = bool(connection.execute(
            text("SELECT pg_try_advisory_lock(hashtextextended(:identity, 0))"),
            {"identity": "taksklad:smartup:fulfillment-sweep:v1"},
        ).scalar())
        connection.commit()
    except Exception:
        connection.close()
        raise
    if not acquired:
        connection.close()
        return False, None
    return True, connection


def release_smartup_fulfillment_sweep_lock(lock_connection: Any | None) -> None:
    if lock_connection is None:
        return
    try:
        lock_connection.execute(
            text("SELECT pg_advisory_unlock(hashtextextended(:identity, 0))"),
            {"identity": "taksklad:smartup:fulfillment-sweep:v1"},
        )
        lock_connection.commit()
    finally:
        lock_connection.close()


def assert_enforced_sagas_complete(events: list[PendingEvent], mode: str) -> None:
    if mode != "enforced":
        return
    incomplete = []
    for event in events:
        payload = event.payload or {}
        expected_count = int(payload.get("skladbot_expected_order_count") or 0)
        covered_count = int(payload.get("skladbot_covered_order_count") or 0)
        legacy_complete = expected_count == 0 and int(payload.get("skladbot_event_count") or 0) == 1
        coverage_complete = expected_count > 0 and covered_count == expected_count
        if normalize_text(payload.get("saga_state")) != "skladbot_queued" or not (
            coverage_complete or legacy_complete
        ):
            incomplete.append(event)
    if incomplete:
        raise SmartupAutoImportError(
            f"Smartup saga SkladBot intent incomplete: {len(incomplete)} deal(s) remain retryable"
        )


def prepare_orphaned_smartup_sagas(
    db: Session,
    config: SmartupAutoImportConfig,
    *,
    export_date: date,
    slot_label: str,
    target_delivery_date: date | None,
) -> list[PendingEvent]:
    export_iso = export_date.isoformat()
    normalized_slot = normalize_text(slot_label)
    expected_target = target_delivery_date.isoformat() if target_delivery_date else ""
    metadata_expr = ImportJob.raw_payload["smartup_auto"]
    imports = db.execute(
        select(ImportJob)
        .where(ImportJob.source == SMARTUP_AUTO_IMPORT_SOURCE)
        .where(ImportJob.status.in_(("completed", "completed_with_errors")))
        .where(ImportJob.rows_imported > 0)
        .where(metadata_expr["export_date"].as_string() == export_iso)
        .where(metadata_expr["slot"].as_string() == normalized_slot)
        .where(metadata_expr["target_delivery_date"].as_string() == expected_target)
        .where(metadata_expr["saga_mode"].as_string() == "enforced")
        .order_by(ImportJob.created_at, ImportJob.id)
    ).scalars().all()
    candidates = []
    seen_deals = set()
    for import_job in imports:
        raw_payload = import_job.raw_payload or {}
        metadata = raw_payload.get("smartup_auto") if isinstance(raw_payload.get("smartup_auto"), dict) else {}
        for deal_id in sorted({normalize_text(value) for value in metadata.get("deal_ids") or [] if normalize_text(value)}):
            if deal_id in seen_deals:
                continue
            seen_deals.add(deal_id)
            candidates.append({
                "delivery_date": normalize_text((metadata.get("delivery_dates") or [""])[0]),
                "import_id": str(import_job.id),
                "status": import_job.status,
                "rows_total": int(import_job.rows_total or 0),
                "rows_imported": int(import_job.rows_imported or 0),
                "orders_created": int(raw_payload.get("orders_created") or 0),
                "items_created": int(raw_payload.get("items_created") or 0),
                "duplicate_rows": int(raw_payload.get("duplicate_rows") or 0),
                "invalid_rows": int(raw_payload.get("invalid_rows") or 0),
                "deal_ids": [deal_id],
                "resolved_order_ids": list(raw_payload.get("resolved_order_ids") or []),
            })
    if not candidates:
        return []
    return prepare_deal_sagas(
        db,
        candidates,
        source_batch_key=f"orphan-recovery:{export_date.isoformat()}:{slot_label}",
        target_status=config.waiting_status_code,
        export_date=export_date,
        slot_label=slot_label,
        target_delivery_date=target_delivery_date,
        mode="enforced",
        source_scope=smartup_source_scope(config),
    )


def create_delivery_group_imports(
    db: Session,
    grouped_rows: dict[str, list[dict[str, Any]]],
    filename: str,
    sha256: str,
    *,
    source_batch_key: str,
    export_date: date,
    part: int,
    slot_label: str,
    source_chat_id: str,
    target_delivery_date: date | None = None,
    saga_mode: str = "disabled",
) -> list[dict[str, Any]]:
    results = []
    for delivery_date, rows in sorted(grouped_rows.items()):
        rows_by_deal_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            rows_by_deal_id[normalize_text(row.get("Smartup deal_id"))].append(row)
        for deal_id, deal_rows in sorted(rows_by_deal_id.items()):
            result = create_smartup_import(
                db,
                delivery_date,
                deal_id,
                deal_rows,
                filename,
                sha256,
                source_batch_key,
                export_date=export_date,
                part=part,
                slot_label=slot_label,
                source_chat_id=source_chat_id,
                target_delivery_date=target_delivery_date,
                saga_mode=saga_mode,
            )
            results.append(result)
    return results


def create_smartup_import(
    db: Session,
    delivery_date: str,
    deal_id: str,
    rows: list[dict[str, Any]],
    filename: str,
    sha256: str,
    source_batch_key: str,
    *,
    export_date: date,
    part: int,
    slot_label: str,
    source_chat_id: str,
    target_delivery_date: date | None = None,
    saga_mode: str = "disabled",
) -> dict[str, Any]:
    import_rows = with_source_batch_key(rows, source_batch_key)
    payload = ImportCreate(
        source=SMARTUP_AUTO_IMPORT_SOURCE,
        filename=filename,
        sha256=sha256,
        telegram_chat_id=source_chat_id,
        telegram_event_id=(
            f"smartup-auto:{export_date.isoformat()}:{slot_label}:"
            f"{delivery_date}:{deal_id}:part-{part}"
        ),
        rows=import_rows,
    )
    import_result = create_import(db, payload, skladbot_create_mode="dry_run")
    metadata = {
        "version": 1,
        "export_date": export_date.isoformat(),
        "export_date_display": format_display_date(export_date),
        "delivery_dates": [delivery_date],
        "part": part,
        "slot": slot_label,
        "target_delivery_date": target_delivery_date.isoformat() if target_delivery_date else "",
        "saga_mode": normalize_saga_mode(saga_mode),
        "filename": filename,
        "source_batch_key": source_batch_key,
        "rows": len(import_rows),
        "deal_ids": unique_values(row.get("Smartup deal_id") for row in import_rows),
        "delivery_date_adjustments": delivery_date_adjustments(import_rows),
    }
    attach_smartup_metadata_to_import(db, import_result.id, metadata)
    return {
        "delivery_date": delivery_date,
        "import_id": import_result.id,
        "status": import_result.status,
        "rows_total": import_result.rows_total,
        "rows_imported": import_result.rows_imported,
        "orders_created": import_result.orders_created,
        "items_created": import_result.items_created,
        "duplicate_rows": import_result.duplicate_rows,
        "invalid_rows": import_result.invalid_rows,
        "resolved_order_ids": list(import_result.resolved_order_ids),
        "skladbot_dry_run_status": import_result.skladbot_dry_run_status,
        "skladbot_dry_run_ready": import_result.skladbot_dry_run_ready,
        "skladbot_dry_run_blocked": import_result.skladbot_dry_run_blocked,
        "skladbot_dry_run_event_id": import_result.skladbot_dry_run_event_id,
        "deal_ids": metadata["deal_ids"],
    }


def queue_skladbot_after_smartup_status(
    db: Session,
    imports: list[dict[str, Any]],
    *,
    successful_deal_ids: list[str] | None = None,
    commit: bool = True,
) -> list[dict[str, Any]]:
    successful_deal_id_set = {normalize_text(value) for value in successful_deal_ids or [] if normalize_text(value)}
    results = []
    for item in imports:
        import_deal_ids = {normalize_text(value) for value in item.get("deal_ids") or [] if normalize_text(value)}
        if successful_deal_ids is not None and (not import_deal_ids or not import_deal_ids <= successful_deal_id_set):
            results.append({
                **item,
                "skladbot_after_status": {
                    "status": "skipped",
                    "reason": "smartup_status_not_confirmed",
                    "deal_ids": sorted(import_deal_ids),
                },
            })
            continue
        import_id = normalize_text(item.get("import_id"))
        if not import_id:
            results.append(item)
            continue
        try:
            resolved_order_ids = [
                normalize_text(value)
                for value in item.get("resolved_order_ids") or []
                if normalize_text(value)
            ]
            if resolved_order_ids:
                queue_result = create_skladbot_dry_run_for_orders(
                    db,
                    resolved_order_ids,
                    import_id=import_id,
                )
            else:
                queue_result = create_skladbot_dry_run_for_import(db, import_id)
            import_job = db.get(ImportJob, uuid.UUID(import_id))
            if import_job is not None:
                import_job.raw_payload = {
                    **(import_job.raw_payload or {}),
                    "skladbot_dry_run": queue_result,
                }
                db.add(import_job)
            if commit:
                db.commit()
            updated = {
                **item,
                "skladbot_dry_run_status": queue_result.get("status", ""),
                "skladbot_dry_run_ready": queue_result.get("ready", 0),
                "skladbot_dry_run_blocked": queue_result.get("blocked", 0),
                "skladbot_dry_run_event_id": queue_result.get("event_id", ""),
                "skladbot_after_status": queue_result,
            }
        except Exception as exc:
            db.rollback()
            logger.exception("SkladBot after-status queue failed for Smartup import %s", import_id)
            raise SmartupAutoImportError(
                f"SkladBot after-status queue failed for Smartup import {import_id}: {str(exc)[:500]}"
            ) from exc
        results.append(updated)
    return results


def attach_smartup_metadata_to_import(db: Session, import_id: str, metadata: dict[str, Any]) -> None:
    try:
        import_uuid = uuid.UUID(import_id)
    except ValueError:
        return
    import_job = db.get(ImportJob, import_uuid)
    if import_job is None:
        return
    import_job.raw_payload = {
        **(import_job.raw_payload or {}),
        "smartup_auto": metadata,
    }
    db.add(AuditLog(
        action="smartup_auto_import_metadata_attached",
        entity_type="import",
        entity_id=import_id,
        payload=metadata,
    ))
    db.commit()


def send_smartup_export_to_client(
    db: Session,
    config: SmartupAutoImportConfig,
    *,
    export_path: Path,
    filename: str,
    export_date: date,
    slot_label: str,
    part: int,
    selected_orders: int,
    rows: int,
    delivery_dates: list[str],
    target_delivery_date: date | None = None,
    telegram_sender: Any | None = None,
) -> dict[str, Any]:
    if not config.client_chat_id:
        result = {"status": "skipped", "reason": "SMARTUP_AUTO_IMPORT_CLIENT_CHAT_ID is empty"}
        record_smartup_client_export_audit(db, export_date, filename, result)
        return result
    sender = telegram_sender or TelegramDocumentSender(config.telegram_bot_token, config.telegram_timeout_seconds)
    if not getattr(sender, "configured", True):
        result = {"status": "skipped", "reason": "TELEGRAM_BOT_TOKEN is empty"}
        record_smartup_client_export_audit(db, export_date, filename, result)
        return result
    route_fingerprint = smartup_route_fingerprint(config, "client")
    delivery_scope = (
        target_delivery_date.isoformat() if target_delivery_date is not None else "all_targets"
    )
    idempotency_key = (
        f"smartup:client_export:v1:{export_date.isoformat()}:"
        f"{normalize_text(slot_label)}:{delivery_scope}"
    )
    existing_delivery = db.execute(
        select(PendingEvent).where(PendingEvent.idempotency_key == idempotency_key)
    ).scalar_one_or_none()
    if existing_delivery is not None:
        if existing_delivery.status == "completed":
            result = {
                "status": "already_sent",
                "provenance": "auto_smartup",
                "route_role": "client",
                "route_fingerprint": route_fingerprint,
                "filename": filename,
                "export_date": export_date.isoformat(),
                "part": part,
            }
            record_smartup_client_export_audit(db, export_date, filename, result)
            return result
        raise TelegramDeliveryAmbiguousError(
            "Smartup client export has a non-terminal or ambiguous durable delivery claim"
        )
    delivery_event = PendingEvent(
        event_type=SMARTUP_CLIENT_EXPORT_EVENT_TYPE,
        status="processing",
        attempts=1,
        idempotency_key=idempotency_key,
        payload={
            "version": 1,
            "export_date": export_date.isoformat(),
            "slot": normalize_text(slot_label),
            "target_delivery_date": delivery_scope,
            "route_role": "client",
            "route_fingerprint": route_fingerprint,
            "delivery_started": False,
        },
    )
    db.add(delivery_event)
    db.commit()
    db.refresh(delivery_event)
    try:
        content = Path(export_path).read_bytes()
        delivery_event.payload = {**(delivery_event.payload or {}), "delivery_started": True}
        db.add(delivery_event)
        db.commit()
        sender.send_document(
            config.client_chat_id,
            content,
            filename,
            caption=smartup_export_caption(export_date, slot_label, part, selected_orders, rows, delivery_dates),
        )
        result = {
            "status": "sent",
            "provenance": "auto_smartup",
            "route_role": "client",
            "route_fingerprint": route_fingerprint,
            "target_type": telegram_chat_target_type(config.client_chat_id),
            "filename": filename,
            "export_date": export_date.isoformat(),
            "part": part,
        }
        delivery_event.status = "completed"
        delivery_event.last_error = ""
        delivery_event.completed_at = datetime.now(timezone.utc)
        delivery_event.payload = {
            **(delivery_event.payload or {}),
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "result": sanitize_audit_payload(result),
        }
        db.add(delivery_event)
        db.commit()
    except Exception as exc:
        result = {
            "status": "ambiguous",
            "provenance": "auto_smartup",
            "route_role": "client",
            "route_fingerprint": route_fingerprint,
            "target_type": telegram_chat_target_type(config.client_chat_id),
            "filename": filename,
            "error": sanitize_automation_error_text(exc, limit=500),
            "manual_recovery_required": True,
        }
        delivery_event.status = "blocked"
        delivery_event.last_error = "SMARTUP_CLIENT_EXPORT_DELIVERY_AMBIGUOUS"
        delivery_event.payload = {
            **(delivery_event.payload or {}),
            "manual_recovery_required": True,
            "result": sanitize_audit_payload(result),
        }
        db.add(delivery_event)
        db.commit()
        record_smartup_client_export_audit(db, export_date, filename, result)
        raise TelegramDeliveryAmbiguousError(
            "Smartup client export delivery result is ambiguous; automatic retry blocked"
        ) from exc
    record_smartup_client_export_audit(db, export_date, filename, result)
    return result


def record_smartup_client_export_audit(
    db: Session,
    export_date: date,
    filename: str,
    result: dict[str, Any],
) -> None:
    db.add(AuditLog(
        action="smartup_auto_import_client_export_file",
        entity_type="smartup_export",
        entity_id=f"{export_date.isoformat()}:{filename}",
        payload=result,
    ))
    db.commit()


def preview_delivery_groups(
    db: Session,
    grouped_rows: dict[str, list[dict[str, Any]]],
    filename: str,
    sha256: str,
    *,
    source_batch_key: str,
) -> list[dict[str, Any]]:
    previews = []
    for delivery_date, rows in sorted(grouped_rows.items()):
        preview_rows = with_source_batch_key(rows, source_batch_key)
        preview = preview_import(
            db,
            ImportCreate(
                source=SMARTUP_AUTO_IMPORT_SOURCE,
                filename=filename,
                sha256=sha256,
                rows=preview_rows,
            ),
        )
        previews.append({
            "delivery_date": delivery_date,
            "status": preview.status,
            "rows_total": preview.rows_total,
            "rows_importable": preview.rows_importable,
            "orders_new": preview.orders_new,
            "items_new": preview.items_new,
            "duplicate_rows": preview.duplicate_rows,
            "invalid_rows": preview.invalid_rows,
            "errors": preview.errors,
        })
    return previews


def with_source_batch_key(rows: list[dict[str, Any]], source_batch_key: str) -> list[dict[str, Any]]:
    return [
        {
            **row,
            "source_batch_key": source_batch_key,
        }
        for row in rows
    ]


def assert_previews_safe(previews: list[dict[str, Any]]) -> None:
    invalid = [item for item in previews if item.get("invalid_rows")]
    failed = [item for item in previews if item.get("status") == "failed"]
    if invalid or failed:
        raise SmartupAutoImportError(f"Smartup import preview failed: {redact_json(previews)}")


def send_final_logistics_reports(
    db: Session,
    config: SmartupAutoImportConfig,
    *,
    export_date: date,
    telegram_sender: Any | None = None,
    extra_delivery_dates: list[str] | None = None,
    now: datetime | None = None,
) -> list[dict[str, Any]]:
    delivery_dates = set(extra_delivery_dates or [])
    delivery_dates.update(delivery_dates_for_export_date(db, export_date))
    if not delivery_dates:
        return [{"status": "skipped", "reason": "no_delivery_dates"}]
    if not config.logistics_chat_id:
        return [{"status": "skipped", "reason": "SMARTUP_AUTO_IMPORT_LOGISTICS_CHAT_ID is empty"}]
    route_fingerprint = smartup_route_fingerprint(config, "logistics")
    current_time = smartup_datetime_utc(
        normalize_local_now(now, config.timezone) if now is not None else datetime.now(timezone.utc)
    )

    sender = telegram_sender or TelegramDocumentSender(config.telegram_bot_token, config.telegram_timeout_seconds)
    if not getattr(sender, "configured", True):
        return [{"status": "skipped", "reason": "TELEGRAM_BOT_TOKEN is empty"}]

    results = []
    for delivery_date in sorted(delivery_dates):
        parsed_delivery_date = parse_smartup_date(delivery_date)
        normalized_delivery_date = (
            parsed_delivery_date.isoformat() if parsed_delivery_date else normalize_text(delivery_date)
        )
        event, skipped = claim_smartup_logistics_report(
            db,
            config,
            export_date,
            normalized_delivery_date,
            route_fingerprint=route_fingerprint,
            now=current_time,
        )
        if skipped is not None:
            results.append(skipped)
            continue
        if parsed_delivery_date and is_logistics_non_working_day(
            db,
            parsed_delivery_date,
            default_non_working_weekdays=config.disabled_weekdays,
        ):
            result = {
                "status": "skipped",
                "provenance": "auto_smartup",
                "route_role": "logistics",
                "route_fingerprint": route_fingerprint,
                "delivery_date": parsed_delivery_date.isoformat(),
                "reason": "non_working_logistics_day",
            }
            db.add(AuditLog(
                action="smartup_auto_import_logistics_report",
                entity_type="delivery_date",
                entity_id=parsed_delivery_date.isoformat(),
                payload=sanitize_audit_payload(result),
            ))
            mark_smartup_logistics_report_completed(db, event.id, result)
            results.append(result)
            continue
        try:
            content, filename = build_logistics_report_xlsx(db, delivery_date)
        except Exception as exc:
            result = {
                "status": "failed",
                "provenance": "auto_smartup",
                "route_role": "logistics",
                "route_fingerprint": route_fingerprint,
                "delivery_date": delivery_date,
                "error": sanitize_automation_error_text(exc, limit=500),
                "delivery_started": False,
            }
            mark_smartup_logistics_report_failed(db, config, event.id, result, now=current_time)
            notify_smartup_automation_error(
                db,
                config,
                export_date=export_date,
                slot_label=f"logistics:{normalized_delivery_date}",
                exc=SmartupAutoImportError(
                    f"Logistics report build failed: {sanitize_automation_error_text(exc, limit=500)}"
                ),
                telegram_sender=telegram_sender,
            )
        else:
            try:
                sender.send_document(
                    config.logistics_chat_id,
                    content,
                    filename,
                    caption=logistics_report_caption(delivery_date),
                )
            except Exception as exc:
                result = {
                    "status": "ambiguous",
                    "provenance": "auto_smartup",
                    "route_role": "logistics",
                    "route_fingerprint": route_fingerprint,
                    "delivery_date": delivery_date,
                    "error": sanitize_automation_error_text(exc, limit=500),
                    "delivery_started": True,
                    "manual_recovery_required": True,
                }
                mark_smartup_logistics_report_ambiguous(db, event.id, result)
                notify_smartup_automation_error(
                    db,
                    config,
                    export_date=export_date,
                    slot_label=f"logistics:{normalized_delivery_date}",
                    exc=SmartupAutoImportError(
                        "Logistics Telegram delivery result is ambiguous; automatic retry blocked"
                    ),
                    telegram_sender=telegram_sender,
                )
            else:
                result = {
                    "status": "sent",
                    "provenance": "auto_smartup",
                    "route_role": "logistics",
                    "route_fingerprint": route_fingerprint,
                    "delivery_date": delivery_date,
                    "filename": filename,
                }
                mark_smartup_logistics_report_completed(db, event.id, result)
        results.append(result)
        db.add(AuditLog(
            action="smartup_auto_import_logistics_report",
            entity_type="delivery_date",
            entity_id=delivery_date,
            payload=sanitize_audit_payload(result),
        ))
    db.commit()
    return results


def claim_smartup_logistics_report(
    db: Session,
    config: SmartupAutoImportConfig,
    export_date: date,
    delivery_date: str,
    *,
    route_fingerprint: str,
    now: datetime,
) -> tuple[PendingEvent | None, dict[str, Any] | None]:
    idempotency_key = smartup_logistics_report_idempotency_key(
        export_date,
        delivery_date,
        route_fingerprint,
    )
    existing = db.execute(
        select(PendingEvent).where(PendingEvent.idempotency_key == idempotency_key)
    ).scalar_one_or_none()
    if existing is not None:
        retry_reason = smartup_logistics_retry_reason(existing, config, now)
        if not retry_reason:
            if existing.status == "blocked":
                skipped_reason = "manual_recovery_required"
            elif existing.status == "completed":
                skipped_reason = "already_sent"
            else:
                skipped_reason = "already_processing"
            return None, {
                "status": "skipped",
                "reason": skipped_reason,
                "provenance": "auto_smartup",
                "route_role": "logistics",
                "route_fingerprint": route_fingerprint,
                "delivery_date": delivery_date,
                "event_id": str(existing.id),
            }
        if retry_reason in {"retry_backoff", "retry_exhausted"}:
            return None, {
                "status": "skipped",
                "reason": retry_reason,
                "provenance": "auto_smartup",
                "route_role": "logistics",
                "route_fingerprint": route_fingerprint,
                "delivery_date": delivery_date,
                "attempts": int(existing.attempts or 0),
            }
        existing.status = "processing"
        existing.attempts = int(existing.attempts or 0) + 1
        existing.last_error = ""
        payload = {**(existing.payload or {})}
        payload.pop("next_attempt_at", None)
        payload.update({
            "retry_claimed_at": now.isoformat(),
            "retry_reason": retry_reason,
            "route_fingerprint": route_fingerprint,
        })
        existing.payload = payload
        db.add(existing)
        db.commit()
        db.refresh(existing)
        return existing, None

    sent_audit = find_existing_smartup_logistics_report_audit(
        db,
        delivery_date,
    )
    sent_route_fingerprint = normalize_text((sent_audit.payload or {}).get("route_fingerprint")) \
        if sent_audit is not None else ""
    route_matches = bool(sent_route_fingerprint and sent_route_fingerprint == route_fingerprint)
    if sent_audit is not None:
        legacy_delivery_state = (
            "route_matched" if route_matches else "legacy_unproven_assumed_delivered"
        )
        event = PendingEvent(
            event_type=SMARTUP_LOGISTICS_REPORT_EVENT_TYPE,
            idempotency_key=idempotency_key,
            status="completed",
            attempts=0,
            payload={
                "version": 2,
                "export_date": export_date.isoformat(),
                "delivery_date": delivery_date,
                "provenance": "auto_smartup",
                "route_role": "logistics",
                "route_fingerprint": route_fingerprint,
                "legacy_delivery_state": legacy_delivery_state,
                "completed_at": now.isoformat(),
                "result": sanitize_audit_payload(sent_audit.payload or {}),
                "legacy_audit_log_id": str(sent_audit.id),
            },
        )
        db.add(event)
        try:
            db.commit()
        except IntegrityError:
            db.rollback()
        return None, {
            "status": "skipped",
            "reason": "already_sent" if route_matches else "legacy_assumed_delivered",
            "provenance": "auto_smartup",
            "route_role": "logistics",
            "route_fingerprint": route_fingerprint,
            "delivery_date": delivery_date,
            "audit_log_id": str(sent_audit.id),
        }

    event = PendingEvent(
        event_type=SMARTUP_LOGISTICS_REPORT_EVENT_TYPE,
        idempotency_key=idempotency_key,
        status="processing",
        attempts=1,
        payload={
            "version": 2,
            "export_date": export_date.isoformat(),
            "delivery_date": delivery_date,
            "provenance": "auto_smartup",
            "route_role": "logistics",
            "route_fingerprint": route_fingerprint,
            "claimed_at": now.isoformat(),
        },
    )
    db.add(event)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        return None, {
            "status": "skipped",
            "reason": "already_claimed",
            "delivery_date": delivery_date,
        }
    db.refresh(event)
    return event, None


def mark_smartup_logistics_report_completed(
    db: Session,
    event_id: uuid.UUID,
    result: dict[str, Any],
) -> None:
    event = db.get(PendingEvent, event_id)
    if event is None:
        return
    event.status = "completed"
    event.last_error = ""
    event.payload = {
        **(event.payload or {}),
        "completed_at": datetime.now(timezone.utc).isoformat(),
        "result": sanitize_audit_payload(result),
    }
    db.add(event)
    db.commit()


def mark_smartup_logistics_report_failed(
    db: Session,
    config: SmartupAutoImportConfig,
    event_id: uuid.UUID,
    result: dict[str, Any],
    *,
    now: datetime,
) -> None:
    event = db.get(PendingEvent, event_id)
    if event is None:
        return
    event.status = "failed"
    event.last_error = normalize_text(result.get("error"))[:1000]
    backoff_seconds = min(
        3600,
        config.logistics_retry_base_seconds * (2 ** max(0, int(event.attempts or 1) - 1)),
    )
    event.payload = {
        **(event.payload or {}),
        "failed_at": now.isoformat(),
        "next_attempt_at": (now + timedelta(seconds=backoff_seconds)).isoformat(),
        "backoff_seconds": backoff_seconds,
        "result": sanitize_audit_payload(result),
    }
    db.add(event)
    db.commit()


def mark_smartup_logistics_report_ambiguous(
    db: Session,
    event_id: uuid.UUID,
    result: dict[str, Any],
) -> None:
    event = db.get(PendingEvent, event_id)
    if event is None:
        return
    event.status = "blocked"
    event.last_error = "LOGISTICS_TELEGRAM_DELIVERY_AMBIGUOUS"
    event.payload = {
        **(event.payload or {}),
        "blocked_at": datetime.now(timezone.utc).isoformat(),
        "manual_recovery_required": True,
        "result": sanitize_audit_payload(result),
    }
    db.add(event)
    db.commit()


def smartup_logistics_report_idempotency_key(
    export_date: date,
    delivery_date: str,
    route_fingerprint: str,
) -> str:
    return (
        f"smartup:logistics_report:v2:{export_date.isoformat()}:"
        f"{normalize_text(delivery_date)}:{normalize_text(route_fingerprint)}"
    )


def find_existing_smartup_logistics_report_audit(
    db: Session,
    delivery_date: str,
) -> AuditLog | None:
    return db.execute(
        select(AuditLog)
        .where(AuditLog.action == "smartup_auto_import_logistics_report")
        .where(AuditLog.entity_id == delivery_date)
        .where(AuditLog.payload["status"].as_string() == "sent")
        .order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
        .limit(1)
    ).scalar_one_or_none()


def smartup_logistics_retry_reason(
    event: PendingEvent,
    config: SmartupAutoImportConfig,
    now: datetime,
) -> str:
    if event.status == "completed":
        return ""
    if event.status == "processing":
        payload = event.payload or {}
        claimed_at = parse_iso_datetime(
            payload.get("retry_claimed_at") or payload.get("claimed_at")
        )
        last_seen_at = claimed_at or smartup_event_last_seen_at(event)
        if last_seen_at is None:
            return ""
        if smartup_datetime_utc(now) - smartup_datetime_utc(last_seen_at) >= timedelta(
            minutes=config.logistics_claim_timeout_minutes
        ):
            return "stale_processing"
        return ""
    if event.status == "failed":
        if int(event.attempts or 0) >= config.logistics_max_attempts:
            return "retry_exhausted"
        next_attempt_at = parse_iso_datetime((event.payload or {}).get("next_attempt_at"))
        if next_attempt_at is not None and smartup_datetime_utc(now) < smartup_datetime_utc(next_attempt_at):
            return "retry_backoff"
        return "failed"
    return ""


def smartup_route_fingerprint(config: SmartupAutoImportConfig, role: str) -> str:
    routes = {
        "client": config.client_chat_id,
        "logistics": config.logistics_chat_id,
        "alert": config.alert_chat_id,
    }
    chat_id = normalize_text(routes.get(role))
    key = normalize_text(config.route_fingerprint_key)
    if role not in routes or not chat_id:
        raise SmartupAutoImportError(f"Smartup route {role or 'unknown'} не настроен")
    if not key:
        raise SmartupAutoImportError("SMARTUP_AUTO_IMPORT_ROUTE_FINGERPRINT_KEY не задан")
    digest = hmac.new(
        key.encode("utf-8"),
        f"taksklad-smartup-route:v1:{role}:{chat_id}".encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return f"hmac-sha256:v1:{digest[:24]}"


def parse_iso_datetime(value: Any) -> datetime | None:
    text_value = normalize_text(value)
    if not text_value:
        return None
    try:
        return datetime.fromisoformat(text_value.replace("Z", "+00:00"))
    except ValueError:
        return None


def delivery_dates_for_export_date(db: Session, export_date: date) -> list[str]:
    export_iso = export_date.isoformat()
    imports = db.execute(
        select(ImportJob)
        .where(ImportJob.source == SMARTUP_AUTO_IMPORT_SOURCE)
        .where(ImportJob.raw_payload["smartup_auto"]["export_date"].as_string() == export_iso)
        .order_by(ImportJob.created_at.asc())
    ).scalars().all()
    result = set()
    for import_job in imports:
        metadata = (import_job.raw_payload or {}).get("smartup_auto") or {}
        for value in metadata.get("delivery_dates") or []:
            parsed = parse_smartup_date(value)
            if parsed:
                result.add(parsed.isoformat())
    return sorted(result)


def delivery_dates_for_auto_logistics(
    db: Session,
    export_date: date,
    config: SmartupAutoImportConfig,
) -> list[str]:
    """Use durable Smartup metadata, with order truth as a recovery source."""
    result = set(delivery_dates_for_export_date(db, export_date))
    scheduled_dates = {
        value.isoformat()
        for value in scheduled_smartup_target_delivery_dates(db, export_date, config)
    }
    available_order_dates = {
        normalize_text(value)
        for value in list_logistics_dates(db)
        if normalize_text(value)
    }
    result.update(scheduled_dates & available_order_dates)
    return sorted(result)


def delivery_dates_from_smartup_run_results(results: list[dict[str, Any]]) -> list[str]:
    return sorted(unique_values(
        value
        for result in results
        for value in result.get("delivery_dates", [])
    ))


def build_import_rows(
    orders: list[dict[str, Any]],
    export_date: date,
    filename: str,
    config: SmartupAutoImportConfig,
    *,
    db: Session | None = None,
) -> list[dict[str, Any]]:
    rows = []
    geocode_cache: dict[str, tuple[str, str]] = {}
    for order_index, order in enumerate(orders, start=1):
        deal_id = normalize_text(order.get("deal_id"))
        if not deal_id:
            raise SmartupAutoImportError(f"Smartup order #{order_index}: нет deal_id")
        delivery_date = parse_smartup_date(order.get("delivery_date"))
        if delivery_date is None:
            raise SmartupAutoImportError(f"Smartup deal_id={deal_id}: нет delivery_date")
        delivery_resolution = resolve_effective_delivery_date(
            db,
            delivery_date,
            default_non_working_weekdays=config.disabled_weekdays,
        )
        client = normalize_text(order.get("person_name") or order.get("person_code"))
        if not client:
            raise SmartupAutoImportError(f"Smartup deal_id={deal_id}: нет person_name")
        coordinates = smartup_coordinates(order)
        address = smartup_address(order, coordinates, geocode_cache=geocode_cache)
        representative = normalize_text(
            order.get("sales_manager_name")
            or order.get("sales_manager_code")
            or order.get("manager_code")
            or order.get("expeditor_name")
        )
        products = order.get("order_products") or []
        if not isinstance(products, list) or not products:
            raise SmartupAutoImportError(f"Smartup deal_id={deal_id}: нет order_products")
        for product_index, product in enumerate(products, start=1):
            product_name = clean_product_name(product.get("product_name") or product.get("product_code"))
            if not product_name:
                raise SmartupAutoImportError(f"Smartup deal_id={deal_id}: нет product_name")
            quantity_pieces = parse_int(product.get("order_quant") or product.get("sold_quant"))
            if quantity_pieces <= 0:
                raise SmartupAutoImportError(
                    f"Smartup deal_id={deal_id}, product={product_name}: quantity <= 0"
                )
            quantity_blocks = math.ceil(quantity_pieces / config.pieces_per_block)
            product_price = parse_money(product.get("product_price"))
            sold_amount = parse_money(product.get("sold_amount"))
            line_total = sold_amount or quantity_blocks * config.default_block_price
            product_id = normalize_text(
                product.get("external_id")
                or product.get("product_unit_id")
                or product.get("product_code")
                or product_index
            )
            rows.append({
                "Дата заказа": format_display_date(export_date),
                "Дата отгрузки": format_display_date(delivery_resolution.effective_date),
                "Тип оплаты": TERMINAL_PAYMENT_NAME,
                "Клиент": client,
                "Адрес": address,
                "Координаты": coordinates,
                "Торговый представитель": representative,
                "Товары": product_name,
                "Кол-во ШТ": quantity_pieces,
                "Кол-во блок": quantity_blocks,
                "_pieces_per_block": config.pieces_per_block,
                "Цена из файла": product_price,
                "Сумма из файла": sold_amount,
                "Цена за блок": config.default_block_price,
                "Сумма позиции": line_total,
                "Статус": "not_completed",
                "ID заказа": f"smartup:{deal_id}",
                "ID импорта": f"smartup:{deal_id}:{product_id}:{product_index}",
                "Smartup deal_id": deal_id,
                "Smartup product_id": product_id,
                "Smartup status": smartup_status(order),
                "Smartup delivery_date original": format_display_date(delivery_resolution.original_date),
                "Smartup delivery_date adjusted": "yes" if delivery_resolution.adjusted else "",
                "Smartup delivery_date adjustment_reason": delivery_resolution.reason,
                "Smartup delivery_date skipped_dates": ",".join(delivery_resolution.skipped_dates),
                "Источник файла": filename,
                "Строка файла": len(rows) + 2,
            })
    return rows


def filter_smartup_orders(
    orders: list[dict[str, Any]],
    export_date: date,
    config: SmartupAutoImportConfig,
    *,
    target_delivery_date: date | None = None,
) -> list[dict[str, Any]]:
    result = []
    for order in orders:
        deal_date = smartup_deal_date(order)
        if target_delivery_date is None and deal_date != export_date:
            continue
        if target_delivery_date is not None and parse_smartup_date(order.get("delivery_date")) != target_delivery_date:
            continue
        if smartup_status(order) != config.new_status_code:
            continue
        if smartup_payment_code(order) != config.terminal_payment_code:
            continue
        result.append(order)
    return result


def extract_smartup_orders(response: dict[str, Any]) -> list[dict[str, Any]]:
    orders = response.get("order") or response.get("orders") or []
    if not isinstance(orders, list):
        raise SmartupAutoImportError("Smartup export response: order должен быть массивом")
    return [order for order in orders if isinstance(order, dict)]


def smartup_deal_date(order: dict[str, Any]) -> date | None:
    return parse_smartup_date(
        order.get("deal_date")
        or order.get("deal_time")
        or order.get("created_on")
        or order.get("modified_on")
    )


def smartup_status(order: dict[str, Any]) -> str:
    return normalize_text(order.get("status_code") or order.get("status"))


def smartup_payment_code(order: dict[str, Any]) -> str:
    return normalize_text(order.get("payment_type_code") or order.get("payment_code") or order.get("payment_type"))


def smartup_coordinates(order: dict[str, Any]) -> str:
    latitude = normalize_coordinate_value(order.get("person_latitude") or order.get("latitude"))
    longitude = normalize_coordinate_value(order.get("person_longitude") or order.get("longitude"))
    if not latitude or not longitude:
        return ""
    return f"{latitude},{longitude}"


def smartup_address(
    order: dict[str, Any],
    coordinates: str,
    *,
    geocode_cache: dict[str, tuple[str, str]] | None = None,
) -> str:
    address = normalize_text(order.get("delivery_address_full") or order.get("delivery_address_short"))
    if address:
        return address
    if coordinates:
        try:
            geocoded_address, _ = reverse_geocode_yandex(coordinates, cache=geocode_cache)
        except Exception as exc:
            logger.warning("Smartup reverse geocode failed: %s", exc)
            geocoded_address = ""
        if geocoded_address:
            return geocoded_address
        return f"GPS: {coordinates}"
    return ""


def clean_product_name(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = re.split(r"\s+/\s+", text, maxsplit=1)[0]
    return " ".join(text.split())


def group_rows_by_delivery_date(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        delivery_date = parse_smartup_date(row.get("Дата отгрузки"))
        if delivery_date is None:
            raise SmartupAutoImportError(f"Строка без даты отгрузки: {row}")
        grouped[delivery_date.isoformat()].append(row)
    return dict(grouped)


def delivery_date_adjustments(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    adjustments = []
    seen = set()
    for row in rows:
        if normalize_text(row.get("Smartup delivery_date adjusted")) != "yes":
            continue
        key = (
            normalize_text(row.get("Smartup deal_id")),
            normalize_text(row.get("Smartup delivery_date original")),
            normalize_text(row.get("Дата отгрузки")),
        )
        if key in seen:
            continue
        seen.add(key)
        adjustments.append({
            "deal_id": key[0],
            "original_delivery_date": key[1],
            "effective_delivery_date": key[2],
            "reason": normalize_text(row.get("Smartup delivery_date adjustment_reason")),
            "skipped_dates": normalize_text(row.get("Smartup delivery_date skipped_dates")),
        })
    return adjustments


def write_export_workbook(
    output_dir: Path,
    export_date: date,
    filename: str,
    rows: list[dict[str, Any]],
) -> tuple[Path, str]:
    day_dir = export_day_dir(output_dir, export_date)
    day_dir.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Заказы"
    sheet.append(EXPORT_WORKBOOK_HEADERS)
    apply_export_header_style(sheet)
    for row in rows:
        sheet.append([row.get(header, "") for header in EXPORT_WORKBOOK_HEADERS])
    autosize_columns(sheet)
    buffer = BytesIO()
    force_workbook_text_literals(workbook)
    workbook.save(buffer)
    content = buffer.getvalue()
    path = day_dir / filename
    path.write_bytes(content)
    return path, hashlib.sha256(content).hexdigest()


def write_export_audit(
    output_dir: Path,
    export_date: date,
    filename: str,
    payload: dict[str, Any],
) -> Path:
    day_dir = export_day_dir(output_dir, export_date)
    day_dir.mkdir(parents=True, exist_ok=True)
    path = day_dir / f"{Path(filename).stem}.audit.json"
    update_export_audit(path, payload)
    return path


def update_export_audit(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def export_day_dir(output_dir: Path, export_date: date) -> Path:
    return Path(output_dir) / export_date.isoformat()


def smartup_source_batch_key(export_date: date, part: int, sha256: str) -> str:
    normalized_sha = normalize_text(sha256).lower()
    if normalized_sha:
        return f"smartup:{export_date.isoformat()}:part:{part}:sha256:{normalized_sha}"
    return f"smartup:{export_date.isoformat()}:part:{part}"


def next_export_part(output_dir: Path, export_date: date) -> int:
    day_dir = export_day_dir(output_dir, export_date)
    if not day_dir.exists():
        return 1
    pattern = re.compile(rf"^Терминал {re.escape(format_display_date(export_date))} Часть (\d+)\.xlsx$")
    parts = []
    for path in day_dir.glob("Терминал *.xlsx"):
        match = pattern.match(path.name)
        if match:
            parts.append(int(match.group(1)))
    return (max(parts) + 1) if parts else 1


def apply_export_header_style(sheet) -> None:
    fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
    sheet.freeze_panes = "A2"


def autosize_columns(sheet) -> None:
    for column in sheet.columns:
        width = 8
        for cell in column:
            width = max(width, min(60, len(str(cell.value or "")) + 2))
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def record_smartup_audit(db: Session, action: str, payload: dict[str, Any]) -> None:
    db.add(AuditLog(
        action=action,
        entity_type="smartup_auto_import",
        entity_id=normalize_text(payload.get("export_date")),
        payload=sanitize_audit_payload(payload),
    ))
    db.commit()


def notify_smartup_automation_error(
    db: Session,
    config: SmartupAutoImportConfig,
    *,
    export_date: date,
    slot_label: str,
    exc: Exception,
    telegram_sender: Any | None = None,
) -> dict[str, Any]:
    chat_id = config.alert_chat_id
    route_fingerprint = smartup_route_fingerprint_if_configured(config, "alert")
    result: dict[str, Any]
    if not chat_id:
        result = {"status": "skipped", "reason": "TAKSKLAD_AUTOMATION_ALERT_CHAT_ID is empty"}
    else:
        sender = telegram_sender or TelegramDocumentSender(config.telegram_bot_token, config.telegram_timeout_seconds)
        if not getattr(sender, "configured", True):
            result = {"status": "skipped", "reason": "TELEGRAM_BOT_TOKEN is empty"}
        else:
            message = "\n".join([
                "Smartup automation error",
                f"Дата выгрузки: {format_display_date(export_date)}",
                f"Слот: {slot_label}",
                f"Ошибка: {sanitize_automation_error_text(exc, limit=1200)}",
            ])
            try:
                sender.send_message(chat_id, message)
                result = {
                    "status": "sent",
                    "route_role": "alert",
                    "route_fingerprint": route_fingerprint,
                    "target_type": telegram_chat_target_type(chat_id),
                    "slot": slot_label,
                }
            except Exception as alert_exc:
                logger.error(
                    "Smartup automation error alert failed: %s",
                    sanitize_automation_error_text(alert_exc, limit=500),
                )
                result = {
                    "status": "failed",
                    "route_role": "alert",
                    "route_fingerprint": route_fingerprint,
                    "target_type": telegram_chat_target_type(chat_id),
                    "slot": slot_label,
                    "error": sanitize_automation_error_text(alert_exc, limit=500),
                }

    db.add(AuditLog(
        action="smartup_auto_import_error_alert",
        entity_type="smartup_auto_import",
        entity_id=export_date.isoformat(),
        payload=sanitize_audit_payload({
            **result,
            "export_date": export_date.isoformat(),
            "slot": slot_label,
            "source_error": sanitize_automation_error_text(exc, limit=500),
        }),
    ))
    db.commit()
    return result


def sanitize_automation_error_text(exc: Any, *, limit: int) -> str:
    sanitized = redact_secrets(exc)
    sanitized = re.sub(
        r"(?i)(\bchat[_ -]?id\b\s*[:=]\s*)-?[1-9]\d*",
        r"\1***",
        sanitized,
    )
    return sanitized[:max(1, int(limit))]


def sanitize_audit_payload(payload: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {}

    def sanitize(value: Any, key: str = "") -> Any:
        normalized_key = normalize_text(key).casefold()
        if (
            normalized_key in {"smartup_password", "telegram_bot_token", "token", "password"}
            or "chat_id" in normalized_key
        ):
            return "***"
        if isinstance(value, dict):
            return {item_key: sanitize(item_value, str(item_key)) for item_key, item_value in value.items()}
        if isinstance(value, list):
            return [sanitize(item) for item in value]
        if isinstance(value, tuple):
            return [sanitize(item) for item in value]
        if isinstance(value, str):
            return redact_secrets(value)
        return value

    return sanitize(payload)


def acquire_smartup_slot_advisory_lock(
    db: Session,
    export_date: date,
    slot_label: str,
    *,
    target_delivery_date: date | None = None,
) -> tuple[bool, Any | None]:
    bind = db.get_bind()
    dialect_name = getattr(getattr(bind, "dialect", None), "name", "")
    if dialect_name != "postgresql":
        return True, None
    lock_key = smartup_advisory_lock_key(export_date, slot_label, target_delivery_date=target_delivery_date)
    connection = bind.connect()
    try:
        acquired = bool(connection.execute(
            text("SELECT pg_try_advisory_lock(:lock_key)"),
            {"lock_key": lock_key},
        ).scalar())
        connection.commit()
    except Exception:
        try:
            connection.rollback()
        except Exception:
            pass
        connection.close()
        raise
    if not acquired:
        connection.close()
        return False, None
    return True, connection


def release_smartup_slot_advisory_lock(
    lock_connection: Any | None,
    export_date: date,
    slot_label: str,
    *,
    target_delivery_date: date | None = None,
) -> None:
    if lock_connection is None:
        return
    try:
        lock_connection.execute(
            text("SELECT pg_advisory_unlock(:lock_key)"),
            {"lock_key": smartup_advisory_lock_key(export_date, slot_label, target_delivery_date=target_delivery_date)},
        )
        lock_connection.commit()
    except Exception:
        logger.exception("Smartup auto import advisory lock release failed")
        try:
            lock_connection.rollback()
        except Exception:
            pass
    finally:
        lock_connection.close()


def claim_smartup_slot(
    db: Session,
    export_date: date,
    slot_label: str,
    now: datetime,
    *,
    target_delivery_date: date | None = None,
) -> tuple[PendingEvent | None, dict[str, Any] | None]:
    idempotency_key = smartup_slot_idempotency_key(
        export_date,
        slot_label,
        target_delivery_date=target_delivery_date,
    )
    existing = db.execute(
        select(PendingEvent).where(PendingEvent.idempotency_key == idempotency_key)
    ).scalar_one_or_none()
    if existing is not None:
        retry_reason = smartup_slot_retry_reason(existing, now)
        if retry_reason:
            existing.status = "processing"
            existing.attempts = int(existing.attempts or 0) + 1
            existing.last_error = ""
            payload = {**(existing.payload or {})}
            payload.pop("failed_at", None)
            payload.pop("error", None)
            payload.update({
                "retry_claimed_at": now.isoformat(),
                "retry_attempt": int(existing.attempts or 0),
                "retry_reason": retry_reason,
            })
            existing.payload = {
                **payload,
            }
            db.add(existing)
            if retry_reason == "stale_processing":
                db.add(AuditLog(
                    action="smartup_auto_import_stale_processing_reset",
                    entity_type="pending_event",
                    entity_id=str(existing.id),
                    payload={
                        "event_id": str(existing.id),
                        "export_date": export_date.isoformat(),
                        "slot": slot_label,
                        "attempts": int(existing.attempts or 0),
                        "last_seen_at": smartup_event_last_seen_at(existing).isoformat()
                        if smartup_event_last_seen_at(existing) is not None
                        else "",
                    },
                ))
            db.commit()
            db.refresh(existing)
            return existing, None
        return None, {
            "status": "skipped",
            "reason": "slot_already_claimed",
            "event_id": str(existing.id),
            "event_status": existing.status,
            "export_date": export_date.isoformat(),
            "slot": slot_label,
            "target_delivery_date": target_delivery_date.isoformat() if target_delivery_date else "",
        }
    event = PendingEvent(
        event_type=SMARTUP_AUTO_IMPORT_EVENT_TYPE,
        idempotency_key=idempotency_key,
        status="processing",
        attempts=1,
        payload={
            "version": 1,
            "export_date": export_date.isoformat(),
            "target_delivery_date": target_delivery_date.isoformat() if target_delivery_date else "",
            "slot": slot_label,
            "claimed_at": now.isoformat(),
        },
    )
    db.add(event)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        return None, {
            "status": "skipped",
            "reason": "slot_already_claimed",
            "export_date": export_date.isoformat(),
            "slot": slot_label,
            "target_delivery_date": target_delivery_date.isoformat() if target_delivery_date else "",
        }
    db.refresh(event)
    return event, None


def smartup_slot_retry_reason(event: PendingEvent, now: datetime) -> str:
    if event.status == "failed":
        return "failed"
    if event.status == "processing" and smartup_processing_slot_is_stale(event, now):
        return "stale_processing"
    return ""


def smartup_processing_slot_is_stale(event: PendingEvent, now: datetime) -> bool:
    last_seen_at = smartup_event_last_seen_at(event)
    if last_seen_at is None:
        return False
    return smartup_datetime_utc(now) - smartup_datetime_utc(last_seen_at) >= STALE_SMARTUP_SLOT_TIMEOUT


def smartup_event_last_seen_at(event: PendingEvent) -> datetime | None:
    value = event.updated_at or event.created_at
    if isinstance(value, datetime):
        return value
    return None


def smartup_datetime_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def mark_smartup_slot_completed(db: Session, event_id: uuid.UUID, result: dict[str, Any]) -> None:
    event = db.get(PendingEvent, event_id)
    if event is None:
        return
    event.status = "completed"
    event.last_error = ""
    event.payload = {
        **(event.payload or {}),
        "completed_at": datetime.now(timezone.utc).isoformat(),
        "result": sanitize_audit_payload(result),
    }
    db.add(event)
    db.commit()


def mark_smartup_slot_failed(db: Session, event_id: uuid.UUID, exc: Exception) -> None:
    event = db.get(PendingEvent, event_id)
    if event is None:
        return
    event.status = "failed"
    safe_error = sanitize_automation_error_text(exc, limit=1000)
    event.last_error = safe_error
    event.payload = {
        **(event.payload or {}),
        "failed_at": datetime.now(timezone.utc).isoformat(),
        "error": safe_error,
    }
    db.add(event)
    db.add(AuditLog(
        action="smartup_auto_import_failed",
        entity_type="pending_event",
        entity_id=str(event.id),
        payload={"error": safe_error, "event_id": str(event.id)},
    ))
    db.commit()


def mark_smartup_slot_blocked(db: Session, event_id: uuid.UUID, exc: Exception) -> None:
    event = db.get(PendingEvent, event_id)
    if event is None:
        return
    safe_error = sanitize_automation_error_text(exc, limit=1000)
    event.status = "blocked"
    event.last_error = safe_error
    event.payload = {
        **(event.payload or {}),
        "blocked_at": datetime.now(timezone.utc).isoformat(),
        "manual_recovery_required": True,
        "error": safe_error,
    }
    db.add(event)
    db.commit()


def build_smartup_auto_import_status(
    db: Session,
    config: SmartupAutoImportConfig,
    *,
    limit: int = 5,
) -> dict[str, Any]:
    event_limit = max(1, min(int(limit or 5), 50))
    validation_errors, validation_warnings = smartup_auto_import_status_findings(config)
    last_events = db.execute(
        select(PendingEvent)
        .where(PendingEvent.event_type == SMARTUP_AUTO_IMPORT_EVENT_TYPE)
        .order_by(PendingEvent.created_at.desc(), PendingEvent.id.desc())
        .limit(event_limit)
    ).scalars().all()
    last_logistics_events = db.execute(
        select(PendingEvent)
        .where(PendingEvent.event_type == SMARTUP_LOGISTICS_REPORT_EVENT_TYPE)
        .order_by(PendingEvent.created_at.desc(), PendingEvent.id.desc())
        .limit(event_limit)
    ).scalars().all()
    logistics_state_rows = db.execute(
        select(PendingEvent.status, func.count(PendingEvent.id))
        .where(PendingEvent.event_type == SMARTUP_LOGISTICS_REPORT_EVENT_TYPE)
        .group_by(PendingEvent.status)
    ).all()
    logistics_states = {
        normalize_text(state): int(count or 0)
        for state, count in logistics_state_rows
        if normalize_text(state)
    }
    current_logistics_route_fingerprint = smartup_route_fingerprint_if_configured(config, "logistics")
    logistics_retry_exhausted = 0
    logistics_manual_recovery = 0
    if current_logistics_route_fingerprint:
        logistics_retry_exhausted = int(db.execute(
            select(func.count(PendingEvent.id))
            .where(PendingEvent.event_type == SMARTUP_LOGISTICS_REPORT_EVENT_TYPE)
            .where(PendingEvent.status == "failed")
            .where(PendingEvent.attempts >= config.logistics_max_attempts)
            .where(
                PendingEvent.payload["route_fingerprint"].as_string()
                == current_logistics_route_fingerprint
            )
        ).scalar_one() or 0)
        logistics_manual_recovery = int(db.execute(
            select(func.count(PendingEvent.id))
            .where(PendingEvent.event_type == SMARTUP_LOGISTICS_REPORT_EVENT_TYPE)
            .where(PendingEvent.status == "blocked")
            .where(
                PendingEvent.payload["route_fingerprint"].as_string()
                == current_logistics_route_fingerprint
            )
        ).scalar_one() or 0)
    pending_skladbot_creates = db.execute(
        select(func.count(PendingEvent.id))
        .where(PendingEvent.event_type == SKLADBOT_REQUEST_CREATE_EVENT_TYPE)
        .where(PendingEvent.status.in_(("pending", "processing", "failed")))
    ).scalar_one()
    fulfillment_rows = db.execute(
        select(SmartupFulfillment.state, func.count(SmartupFulfillment.id))
        .group_by(SmartupFulfillment.state)
    ).all()
    fulfillment_states = {
        normalize_text(state): int(count or 0)
        for state, count in fulfillment_rows
        if normalize_text(state)
    }
    manual_review_count = sum(
        fulfillment_states.get(state, 0)
        for state in ("smartup_ambiguous", "skladbot_ambiguous", "blocked_stock", "payload_mismatch", "manual_review")
    )
    return {
        "status": "failed" if (
            validation_errors or manual_review_count or logistics_retry_exhausted or logistics_manual_recovery
        ) else "ok",
        "configuration": {
            "enabled": config.enabled,
            "backend_import_enabled": config.backend_import_enabled,
            "change_status_enabled": config.change_status_enabled,
            "process_skladbot_now": config.process_skladbot_now,
            "saga_mode": config.saga_mode,
            "schedule_times": list(config.schedule_times),
            "disabled_weekdays": list(config.disabled_weekdays),
            "final_time": config.final_time,
            "logistics_due_time": config.effective_logistics_due_time,
            "logistics_retry_base_seconds": config.logistics_retry_base_seconds,
            "logistics_max_attempts": config.logistics_max_attempts,
            "logistics_catchup_days": config.logistics_catchup_days,
            "logistics_claim_timeout_minutes": config.logistics_claim_timeout_minutes,
            "logistics_route_recovery_export_date": config.logistics_route_recovery_export_date,
            "slot_grace_minutes": config.slot_grace_minutes,
            "poll_seconds": config.poll_seconds,
            "timezone": config.timezone_name,
            "output_dir": str(config.output_dir),
            "smartup_base_url_configured": bool(config.smartup_base_url),
            "smartup_auth_configured": bool(config.smartup_username and config.smartup_password),
            "smartup_project_configured": bool(config.smartup_project_code),
            "smartup_filial_configured": bool(config.smartup_filial_id or config.smartup_filial_code),
            "client_chat_configured": bool(config.client_chat_id),
            "logistics_chat_configured": bool(config.logistics_chat_id),
            "alert_chat_configured": bool(config.alert_chat_id),
            "legacy_alert_chat_configured": bool(config.legacy_alert_chat_id),
            "admin_allowlist_configured": bool(config.admin_chat_ids),
            "route_fingerprint_key_configured": bool(config.route_fingerprint_key),
            "production_route_contract_enforced": bool(config.production and config.enabled),
            "skladbot_create_requests_mode": config.skladbot_create_requests_mode,
            "client_route_type": telegram_chat_target_type(config.client_chat_id),
            "logistics_route_type": telegram_chat_target_type(config.logistics_chat_id),
            "alert_route_type": telegram_chat_target_type(config.alert_chat_id),
            "client_route_fingerprint": smartup_route_fingerprint_if_configured(config, "client"),
            "logistics_route_fingerprint": smartup_route_fingerprint_if_configured(config, "logistics"),
            "alert_route_fingerprint": smartup_route_fingerprint_if_configured(config, "alert"),
            "telegram_bot_configured": bool(config.telegram_bot_token),
            "new_status_code_configured": bool(config.new_status_code),
            "waiting_status_code_configured": bool(config.waiting_status_code),
        },
        "validation": {
            "errors": validation_errors,
            "warnings": validation_warnings,
        },
        "queues": {
            "pending_skladbot_request_creates": int(pending_skladbot_creates or 0),
            "fulfillment_states": dict(sorted(fulfillment_states.items())),
            "fulfillment_manual_review": manual_review_count,
            "logistics_delivery_states": dict(sorted(logistics_states.items())),
            "logistics_retry_exhausted": logistics_retry_exhausted,
            "logistics_manual_recovery": logistics_manual_recovery,
        },
        "last_events": [summarize_smartup_auto_import_event(event) for event in last_events],
        "last_logistics_events": [summarize_smartup_logistics_event(event) for event in last_logistics_events],
    }


def smartup_auto_import_status_findings(config: SmartupAutoImportConfig) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    if config.backend_import_enabled and not config.change_status_enabled:
        errors.append(
            "SMARTUP_AUTO_IMPORT_BACKEND_IMPORT_ENABLED=true требует "
            "SMARTUP_AUTO_IMPORT_CHANGE_STATUS_ENABLED=true"
        )
    if config.change_status_enabled and not config.waiting_status_code:
        errors.append("SMARTUP_AUTO_IMPORT_CHANGE_STATUS_ENABLED=true требует статус ожидания")
    if config.enabled and not (config.smartup_username and config.smartup_password):
        errors.append("SMARTUP_AUTO_IMPORT_ENABLED=true требует Smartup auth")
    errors.extend(smartup_production_route_errors(config))
    errors.extend(smartup_production_runtime_errors(config))
    if config.logistics_route_recovery_export_date and not is_iso_date(
        config.logistics_route_recovery_export_date
    ):
        errors.append("SMARTUP_AUTO_IMPORT_LOGISTICS_ROUTE_RECOVERY_EXPORT_DATE должен быть YYYY-MM-DD")
    if (
        config.enabled
        and (config.client_chat_id or config.logistics_chat_id or config.alert_chat_id)
        and not config.route_fingerprint_key
        and "SMARTUP_AUTO_IMPORT_ROUTE_FINGERPRINT_KEY не задан" not in errors
    ):
        errors.append("SMARTUP_AUTO_IMPORT_ROUTE_FINGERPRINT_KEY не задан")
    if (
        (config.enabled or config.backend_import_enabled or config.change_status_enabled)
        and not config.telegram_bot_token
        and not config.production
    ):
        warnings.append("TELEGRAM_BOT_TOKEN не задан: файлы/alerts в Telegram будут пропущены")
    if config.enabled and not config.client_chat_id and not config.production:
        warnings.append("SMARTUP_AUTO_IMPORT_CLIENT_CHAT_ID не задан: export-файл клиенту не отправится")
    if config.enabled and not config.logistics_chat_id and not config.production:
        warnings.append("SMARTUP_AUTO_IMPORT_LOGISTICS_CHAT_ID не задан: финальный отчет логистики не отправится")
    if config.process_skladbot_now:
        warnings.append("SMARTUP_AUTO_IMPORT_PROCESS_SKLADBOT_NOW=true: SkladBot create queue обрабатывается сразу")
    if config.logistics_route_recovery_export_date:
        warnings.append("Smartup logistics route recovery разрешён для одной export-date")
    if config.saga_mode in {"shadow", "enforced"}:
        if config.saga_mode == "shadow":
            warnings.append("SMARTUP_AUTO_IMPORT_SAGA_MODE=shadow не блокирует legacy completion invariant")
    return errors, warnings


def smartup_production_route_errors(config: SmartupAutoImportConfig) -> list[str]:
    if not (config.production and config.enabled):
        return []
    errors: list[str] = []
    route_settings = {
        "SMARTUP_AUTO_IMPORT_CLIENT_CHAT_ID": config.client_chat_id,
        "SMARTUP_AUTO_IMPORT_LOGISTICS_CHAT_ID": config.logistics_chat_id,
        "TAKSKLAD_AUTOMATION_ALERT_CHAT_ID": config.alert_chat_id,
    }
    route_errors = validate_route_values(
        config.client_chat_id, config.logistics_chat_id, config.alert_chat_id
    )
    for setting_name in route_errors:
        if setting_name == "TELEGRAM_ROUTE_ROLE_COLLISION":
            errors.append("client, logistics и admin Telegram routes должны быть попарно различны")
        elif setting_name in route_settings:
            errors.append(f"{setting_name} имеет неверный target type или отсутствует")
    if tuple(config.admin_chat_ids) != (config.alert_chat_id,):
        errors.append("TELEGRAM_ADMIN_CHAT_IDS должен содержать только unified admin route")
    if config.legacy_alert_chat_id:
        errors.append("SMARTUP_AUTO_IMPORT_ALERT_CHAT_ID должен быть пустым")
    if not config.route_fingerprint_key:
        errors.append("SMARTUP_AUTO_IMPORT_ROUTE_FINGERPRINT_KEY не задан")
    if not config.telegram_bot_token:
        errors.append("TELEGRAM_BOT_TOKEN не задан для production Smartup automation")
    return errors


def smartup_production_runtime_errors(config: SmartupAutoImportConfig) -> list[str]:
    if not config.production:
        return []
    errors: list[str] = []
    if not config.enabled:
        errors.append(
            "Production Smartup automation требует SMARTUP_AUTO_IMPORT_ENABLED=true"
        )
        return errors
    if not config.backend_import_enabled:
        errors.append(
            "Production Smartup automation требует "
            "SMARTUP_AUTO_IMPORT_BACKEND_IMPORT_ENABLED=true"
        )
    if normalize_text(config.skladbot_create_requests_mode).casefold() != "enabled":
        errors.append(
            "Production Smartup automation требует SKLADBOT_CREATE_REQUESTS_MODE=enabled"
        )
    contract = load_telegram_routing_contract()
    expected_slots = contract.route_for(TelegramMessageKind.SMARTUP_CLIENT_EXPORT).schedules
    logistics_slot = contract.route_for(TelegramMessageKind.SMARTUP_LOGISTICS_REPORT).schedules[0]
    normalized_slots = tuple(normalize_text(value) for value in config.schedule_times if normalize_text(value))
    if normalized_slots != expected_slots:
        errors.append("Production Smartup automation требует exact contract slot times")
    if normalize_text(config.final_time) != logistics_slot:
        errors.append("SMARTUP_AUTO_IMPORT_FINAL_TIME не совпадает с routing contract")
    if normalize_text(config.effective_logistics_due_time) != logistics_slot:
        errors.append("SMARTUP_AUTO_IMPORT_LOGISTICS_DUE_TIME не совпадает с routing contract")
    return errors


def telegram_chat_target_type(value: Any) -> str:
    normalized = normalize_text(value)
    if not normalized:
        return "missing"
    if not re.fullmatch(r"-?[1-9]\d*", normalized):
        return "invalid"
    return "group" if int(normalized) < 0 else "personal"


def smartup_route_fingerprint_if_configured(config: SmartupAutoImportConfig, role: str) -> str:
    if not config.route_fingerprint_key:
        return ""
    route_value = {
        "client": config.client_chat_id,
        "logistics": config.logistics_chat_id,
        "alert": config.alert_chat_id,
    }.get(role)
    if not route_value:
        return ""
    return smartup_route_fingerprint(config, role)


def summarize_smartup_auto_import_event(event: PendingEvent) -> dict[str, Any]:
    payload = event.payload or {}
    result = payload.get("result") if isinstance(payload.get("result"), dict) else {}
    status_change = result.get("status_change") if isinstance(result.get("status_change"), dict) else {}
    skladbot_processing = (
        result.get("skladbot_processing") if isinstance(result.get("skladbot_processing"), dict) else {}
    )
    imports = result.get("imports") if isinstance(result.get("imports"), list) else []
    logistics_reports = result.get("logistics_reports") if isinstance(result.get("logistics_reports"), list) else []
    return {
        "id": str(event.id),
        "status": event.status,
        "attempts": int(event.attempts or 0),
        "idempotency_key": event.idempotency_key or "",
        "export_date": normalize_text(result.get("export_date") or payload.get("export_date")),
        "target_delivery_date": normalize_text(
            result.get("target_delivery_date") or payload.get("target_delivery_date")
        ),
        "slot": normalize_text(result.get("slot") or payload.get("slot")),
        "raw_orders": int(result.get("raw_orders") or 0),
        "selected_orders": int(result.get("selected_orders") or 0),
        "rows": int(result.get("rows") or 0),
        "delivery_dates": list(result.get("delivery_dates") or []),
        "imports": len(imports),
        "status_change": normalize_text(status_change.get("status")),
        "skladbot_processing": normalize_text(skladbot_processing.get("status")),
        "logistics_reports": len(logistics_reports),
        "last_error": redact_secrets(event.last_error or "")[:500],
        "created_at": event.created_at.isoformat() if isinstance(event.created_at, datetime) else "",
        "updated_at": event.updated_at.isoformat() if isinstance(event.updated_at, datetime) else "",
    }


def summarize_smartup_logistics_event(event: PendingEvent) -> dict[str, Any]:
    payload = event.payload or {}
    return {
        "id": str(event.id),
        "status": event.status,
        "attempts": int(event.attempts or 0),
        "export_date": normalize_text(payload.get("export_date")),
        "delivery_date": normalize_text(payload.get("delivery_date")),
        "provenance": normalize_text(payload.get("provenance")),
        "route_role": normalize_text(payload.get("route_role")),
        "route_fingerprint": normalize_text(payload.get("route_fingerprint")),
        "legacy_delivery_state": normalize_text(payload.get("legacy_delivery_state")),
        "last_error": sanitize_automation_error_text(event.last_error or "", limit=500),
        "created_at": event.created_at.isoformat() if isinstance(event.created_at, datetime) else "",
        "updated_at": event.updated_at.isoformat() if isinstance(event.updated_at, datetime) else "",
    }


def smartup_slot_idempotency_key(
    export_date: date,
    slot_label: str,
    *,
    target_delivery_date: date | None = None,
) -> str:
    key = f"smartup:auto_import:v1:{export_date.isoformat()}:{slot_label}"
    if target_delivery_date is not None:
        key = f"{key}:delivery:{target_delivery_date.isoformat()}"
    return key


def smartup_advisory_lock_key(
    export_date: date,
    slot_label: str,
    *,
    target_delivery_date: date | None = None,
) -> int:
    digest = hashlib.sha256(
        smartup_slot_idempotency_key(
            export_date,
            slot_label,
            target_delivery_date=target_delivery_date,
        ).encode("utf-8")
    ).digest()
    value = int.from_bytes(digest[:8], byteorder="big", signed=False)
    if value >= 2**63:
        value -= 2**64
    return value


def is_final_slot(slot_label: str, config: SmartupAutoImportConfig) -> bool:
    return normalize_text(slot_label) == normalize_text(config.final_time)


def scheduled_smartup_target_delivery_date(export_date: date) -> date:
    return export_date + timedelta(days=1)


def scheduled_smartup_target_delivery_dates(
    db: Session | None,
    export_date: date,
    config: SmartupAutoImportConfig,
) -> list[date]:
    first_delivery_date = scheduled_smartup_target_delivery_date(export_date)
    effective_delivery_date = resolve_effective_delivery_date(
        db,
        first_delivery_date,
        default_non_working_weekdays=config.disabled_weekdays,
    ).effective_date
    delivery_dates = []
    current = first_delivery_date
    while current <= effective_delivery_date:
        delivery_dates.append(current)
        current += timedelta(days=1)
    return delivery_dates


def is_slot_due(now: datetime, slot_label: str, grace_minutes: int) -> bool:
    slot_time = parse_slot_time(slot_label)
    slot_at = datetime.combine(now.date(), slot_time, tzinfo=now.tzinfo)
    return slot_at <= now < slot_at + timedelta(minutes=grace_minutes)


def parse_schedule_times(value: str | None) -> tuple[str, ...]:
    raw = normalize_text(value)
    if not raw:
        return DEFAULT_SCHEDULE_TIMES
    result = []
    for part in raw.split(","):
        label = normalize_text(part)
        if not label:
            continue
        parse_slot_time(label)
        result.append(label)
    return tuple(result) or DEFAULT_SCHEDULE_TIMES


def parse_disabled_weekdays(value: str | None) -> tuple[int, ...]:
    raw = normalize_text(value)
    if not raw:
        return DEFAULT_DISABLED_WEEKDAYS
    result = []
    for part in raw.split(","):
        label = normalize_text(part).casefold()
        if not label:
            continue
        weekday = parse_weekday_label(label)
        if weekday not in result:
            result.append(weekday)
    return tuple(result)


def parse_weekday_label(value: str) -> int:
    labels = {
        "mon": 0,
        "monday": 0,
        "пн": 0,
        "понедельник": 0,
        "tue": 1,
        "tuesday": 1,
        "вт": 1,
        "вторник": 1,
        "wed": 2,
        "wednesday": 2,
        "ср": 2,
        "среда": 2,
        "thu": 3,
        "thursday": 3,
        "чт": 3,
        "четверг": 3,
        "fri": 4,
        "friday": 4,
        "пт": 4,
        "пятница": 4,
        "sat": 5,
        "saturday": 5,
        "сб": 5,
        "суббота": 5,
        "sun": 6,
        "sunday": 6,
        "вс": 6,
        "воскресенье": 6,
    }
    if value in labels:
        return labels[value]
    weekday = parse_int(value, -1)
    if 0 <= weekday <= 6:
        return weekday
    raise SmartupAutoImportError(f"Некорректный день недели Smartup auto import: {value}")


def parse_slot_time(value: str) -> datetime_time:
    text = normalize_text(value)
    try:
        return datetime.strptime(text, "%H:%M").time()
    except ValueError as exc:
        raise SmartupAutoImportError(f"Некорректное время Smartup auto import: {text}") from exc


def normalize_local_now(now: datetime | None, timezone_value: ZoneInfo) -> datetime:
    if now is None:
        return datetime.now(timezone_value)
    if now.tzinfo is None:
        return now.replace(tzinfo=timezone_value)
    return now.astimezone(timezone_value)


def format_display_date(value: date) -> str:
    return value.strftime("%d.%m.%Y")


def display_date_from_any(value: Any) -> str:
    parsed = parse_smartup_date(value)
    return format_display_date(parsed) if parsed else normalize_text(value)


def parse_smartup_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace("T", " ")
    candidates = [
        text,
        text[:10],
        text.split(" ", 1)[0],
    ]
    for candidate in candidates:
        candidate = candidate.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y.%m.%d"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                pass
    match = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", text)
    if match:
        day, month, year = match.groups()
        if len(year) == 2:
            year = f"20{year}"
        try:
            return date(int(year), int(month), int(day))
        except ValueError:
            return None
    return None


def is_iso_date(value: Any) -> bool:
    normalized = normalize_text(value)
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", normalized):
        return False
    try:
        date.fromisoformat(normalized)
    except ValueError:
        return False
    return True


def parse_bool(value: Any, default: bool = False) -> bool:
    text = normalize_text(value).casefold()
    if not text:
        return default
    return text in {"1", "true", "yes", "on", "да"}


def parse_int(value: Any, default: int = 0) -> int:
    text = normalize_text(value).replace(" ", "").replace(",", ".")
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return default


def parse_csv_values(value: Any) -> tuple[str, ...]:
    result: list[str] = []
    seen: set[str] = set()
    for item in normalize_text(value).split(","):
        normalized = normalize_text(item)
        if normalized and normalized not in seen:
            result.append(normalized)
            seen.add(normalized)
    return tuple(result)


def parse_money(value: Any) -> int:
    if isinstance(value, (int, float)):
        return int(value)
    text = normalize_text(value).replace("\xa0", " ").strip()
    if not text:
        return 0
    normalized = text.replace(" ", "").replace(",", ".")
    try:
        return int(float(normalized))
    except ValueError:
        digits = "".join(char for char in text if char.isdigit())
        return int(digits) if digits else 0


def normalize_coordinate_value(value: Any) -> str:
    text = normalize_text(value).replace(",", ".")
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return ""
    return f"{number:.12f}".rstrip("0").rstrip(".")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def unique_values(values) -> list[str]:
    result = []
    seen = set()
    for value in values:
        text = normalize_text(value)
        if text and text not in seen:
            result.append(text)
            seen.add(text)
    return result


def unique_deal_ids(orders: list[dict[str, Any]]) -> list[str]:
    return unique_values(order.get("deal_id") for order in orders)


def redact_json(value: Any) -> str:
    text = json.dumps(value, ensure_ascii=False, default=str)
    return re.sub(r"(?i)(password|token|authorization)(['\"]?\s*[:=]\s*['\"]?)[^,'\"}\]]+", r"\1\2***", text)[:1000]


def worker_sleep(config: SmartupAutoImportConfig) -> None:
    time.sleep(config.poll_seconds)
