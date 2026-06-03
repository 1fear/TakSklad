from io import BytesIO
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .models import Order, OrderItem
from .orders_service import ApiError, COMPLETED_STATUSES
from .reports_service import payment_group

TERMINAL_NO_KIZ_STATUSES = {"archived_no_kiz", "cancelled", "removed_from_google_sheet"}


KIZ_REPORT_HEADERS = [
    "Дата отгрузки",
    "Номер заявки SkladBot",
    "Клиент",
    "Адрес",
    "Координаты",
    "Тип оплаты",
    "Товар",
    "Кол-во блок",
    "КИЗ",
    "Цена заказа",
    "Источник файла",
]

KIZ_SUMMARY_HEADERS = [
    "Дата отгрузки",
    "Номер заявки SkladBot",
    "Клиент",
    "Адрес",
    "Координаты",
    "Тип оплаты",
    "План блоков",
    "Отсканировано блоков",
    "Цена заказа",
    "Источник файла",
]


def list_completed_kiz_source_files(db: Session):
    groups = group_items_by_source_document(load_items(db))
    result = []
    for source_key, items in sorted(groups.items(), key=lambda item: source_group_sort_key(item[1])):
        source_file = source_file_for_items(items)
        if not source_file or not items:
            continue
        planned_blocks = sum(max(0, item.quantity_blocks or 0) for item in items)
        scanned_blocks = sum(max(0, item.scanned_blocks or 0) for item in items)
        completed = all(item_is_completed(item) for item in items)
        if not completed:
            continue
        dates = sorted({item.order.order_date.isoformat() for item in items if item.order and item.order.order_date})
        result.append({
            "source_key": source_key,
            "source_file": source_file,
            "dates": dates,
            "items": len(items),
            "planned_blocks": planned_blocks,
            "scanned_blocks": scanned_blocks,
        })
    return result


def list_completed_kiz_dates(db: Session):
    items = load_items(db)
    grouped = {}
    for item in items:
        order = item.order
        if not order or not order.order_date:
            continue
        grouped.setdefault(order.order_date, []).append(item)

    result = []
    for shipment_date, date_items in sorted(grouped.items()):
        if not date_kiz_is_completed(date_items):
            continue
        report_items = reportable_kiz_items(date_items)
        if not report_items:
            continue
        result.append({
            "date": shipment_date.isoformat(),
            "items": len(report_items),
            "orders": len({item.order.id for item in report_items if item.order}),
            "planned_blocks": sum(max(0, item.quantity_blocks or 0) for item in report_items),
            "scanned_blocks": sum(len(item.scan_codes or []) for item in report_items),
        })
    return result


def build_kiz_source_file_report_xlsx(db: Session, source_file: str, source_key: str = ""):
    source_file = str(source_file or "").strip()
    source_key = str(source_key or "").strip()
    if not source_file:
        raise ApiError(422, "source_file is required")

    items = [
        item
        for item in load_items(db)
        if item_matches_source(item, source_file=source_file, source_key=source_key)
    ]
    if not items:
        raise ApiError(404, f"No rows for source file {source_file}")
    if not all(item_is_completed(item) for item in items):
        raise ApiError(409, f"Source file {source_file} is not fully completed")

    return build_kiz_items_report_xlsx(items, source_file, kiz_source_file_report_filename(source_file))


def build_kiz_date_report_xlsx(db: Session, shipment_date: str):
    target_date = parse_report_date(shipment_date)
    if not target_date:
        raise ApiError(422, "shipment_date is required")

    items = [
        item
        for item in load_items(db)
        if item.order and item.order.order_date == target_date
    ]
    if not items:
        raise ApiError(404, f"No rows for shipment date {target_date.isoformat()}")
    ensure_date_kiz_completed(items, target_date.isoformat())
    report_items = reportable_kiz_items(items)
    if not report_items:
        raise ApiError(404, f"No KIZ scans for shipment date {target_date.isoformat()}")
    display = target_date.strftime("%d.%m.%Y")
    return build_kiz_items_report_xlsx(
        report_items,
        f"Дата {display}",
        f"TakSklad_КИЗ_{display}.xlsx",
    )


def build_kiz_date_range_report_xlsx(db: Session, date_from: str, date_to: str):
    start_date = parse_report_date(date_from)
    end_date = parse_report_date(date_to)
    if not start_date or not end_date:
        raise ApiError(422, "date_from and date_to are required")
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    items_by_date = {}
    for item in load_items(db):
        order = item.order
        if not order or not order.order_date:
            continue
        if start_date <= order.order_date <= end_date:
            items_by_date.setdefault(order.order_date, []).append(item)
    if not items_by_date:
        raise ApiError(404, "No rows for shipment date range")

    for shipment_date, items in sorted(items_by_date.items()):
        ensure_date_kiz_completed(items, shipment_date.isoformat())

    report_items = []
    for items in items_by_date.values():
        report_items.extend(reportable_kiz_items(items))
    if not report_items:
        raise ApiError(404, "No KIZ scans for shipment date range")

    start_display = start_date.strftime("%d.%m.%Y")
    end_display = end_date.strftime("%d.%m.%Y")
    return build_kiz_items_report_xlsx(
        report_items,
        f"Период {start_display}-{end_display}",
        f"TakSklad_КИЗ_{start_display}-{end_display}.xlsx",
    )


def build_kiz_items_report_xlsx(items, source_label, filename):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    summary_sheet.append(KIZ_SUMMARY_HEADERS)
    apply_header_style(summary_sheet)
    for row in build_summary_rows(items, source_label):
        summary_sheet.append(row)
    autosize_columns(summary_sheet)

    grouped = {}
    for item in items:
        grouped.setdefault(payment_group(item.order.payment_type if item.order else ""), []).append(item)

    for group, group_items in sorted(grouped.items()):
        sheet = workbook.create_sheet(payment_sheet_title(group))
        sheet.append(KIZ_REPORT_HEADERS)
        apply_header_style(sheet)
        for item in sorted(group_items, key=item_sort_key):
            order = item.order
            raw_payload = item.raw_payload or {}
            order_raw = order.raw_payload if order else {}
            codes = [scan.code for scan in sorted(item.scan_codes, key=lambda value: (str(value.scanned_at or ""), str(value.id)))]
            for code in codes:
                sheet.append([
                    order.order_date.strftime("%d.%m.%Y") if order and order.order_date else "",
                    (order_raw or {}).get("skladbot_request_number") or "",
                    order.client if order else "",
                    order.address if order else "",
                    (order_raw or {}).get("coordinates") or "",
                    order.payment_type if order else "",
                    item.product,
                    item.quantity_blocks,
                    code,
                    parse_int(raw_payload.get("line_total")),
                    source_file_for_item(item) or source_label,
                ])
        autosize_columns(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), filename


def build_summary_rows(items, source_label):
    grouped = {}
    for item in items:
        if not item.order:
            continue
        grouped.setdefault(item.order.id, {"order": item.order, "items": []})["items"].append(item)

    rows = []
    for group in sorted(grouped.values(), key=lambda value: (str(value["order"].order_date or ""), value["order"].client)):
        order = group["order"]
        order_items = group["items"]
        order_raw = order.raw_payload or {}
        planned_blocks = sum(max(0, item.quantity_blocks or 0) for item in order_items)
        scanned_blocks = sum(max(0, item.scanned_blocks or 0) for item in order_items)
        order_total = sum(parse_int((item.raw_payload or {}).get("line_total")) for item in order_items)
        rows.append([
            order.order_date.strftime("%d.%m.%Y") if order.order_date else "",
            order_raw.get("skladbot_request_number") or "",
            order.client,
            order.address,
            order_raw.get("coordinates") or "",
            order.payment_type,
            planned_blocks,
            scanned_blocks,
            order_total,
            summary_source_for_items(order_items, source_label),
        ])
    return rows


def reportable_kiz_items(items):
    return [
        item
        for item in items
        if item_is_completed(item) and item.scan_codes
    ]


def date_kiz_is_completed(items):
    return not incomplete_kiz_items(items)


def ensure_date_kiz_completed(items, label):
    incomplete = incomplete_kiz_items(items)
    if incomplete:
        raise ApiError(409, f"Shipment date {label} is not fully completed: {len(incomplete)} positions left")


def incomplete_kiz_items(items):
    return [
        item
        for item in items
        if item_requires_kiz_completion(item) and not item_is_completed(item)
    ]


def item_requires_kiz_completion(item):
    status = str(item.status or "").strip()
    if status in TERMINAL_NO_KIZ_STATUSES:
        return False
    if item.order and str(item.order.status or "").strip() in TERMINAL_NO_KIZ_STATUSES:
        return False
    return bool(item.requires_kiz and (item.quantity_blocks or 0) > 0)


def summary_source_for_items(items, source_label):
    source_files = sorted({source_file_for_item(item) for item in items if source_file_for_item(item)})
    if len(source_files) == 1:
        return source_files[0]
    if source_files:
        return ", ".join(source_files[:2]) + (f" +{len(source_files) - 2}" if len(source_files) > 2 else "")
    return source_label


def parse_report_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def load_items(db: Session):
    return db.execute(
        select(OrderItem)
        .options(
            selectinload(OrderItem.order),
            selectinload(OrderItem.scan_codes),
        )
    ).scalars().all()


def group_items_by_source_document(items):
    groups = {}
    for item in items:
        source_file = source_file_for_item(item)
        if not source_file:
            continue
        groups.setdefault(source_document_key(item), []).append(item)
    return groups


def source_document_key(item):
    raw_payload = item.raw_payload or {}
    backend_import_id = str(raw_payload.get("backend_import_id") or "").strip()
    source_file = source_file_for_item(item)
    if backend_import_id:
        return f"import:{backend_import_id}:file:{source_file}"
    return f"file:{source_file}"


def source_file_for_item(item):
    return str((item.raw_payload or {}).get("source_file") or "").strip()


def source_file_for_items(items):
    for item in items:
        source_file = source_file_for_item(item)
        if source_file:
            return source_file
    return ""


def source_group_sort_key(items):
    first_item = items[0] if items else None
    order = first_item.order if first_item else None
    raw_payload = first_item.raw_payload if first_item else {}
    return (
        str(order.order_date or "") if order else "",
        str((raw_payload or {}).get("backend_import_id") or ""),
        source_file_for_item(first_item) if first_item else "",
    )


def item_matches_source(item, source_file="", source_key=""):
    if source_key:
        return source_document_key(item) == source_key
    return source_file_for_item(item) == source_file


def item_is_completed(item):
    if item.status in COMPLETED_STATUSES:
        return True
    return (item.quantity_blocks or 0) > 0 and (item.scanned_blocks or 0) >= (item.quantity_blocks or 0)


def item_sort_key(item):
    order = item.order
    return (
        str(order.order_date or "") if order else "",
        order.client if order else "",
        item.product,
        str(item.id),
    )


def payment_sheet_title(group):
    if group == "terminal":
        return "Терминал"
    if group == "transfer":
        return "Перевод"
    return "Неизвестно"


def kiz_source_file_report_filename(source_file):
    safe_name = safe_filename(source_file.rsplit(".", 1)[0])
    return f"TakSklad_КИЗ_{safe_name}.xlsx"


def safe_filename(value):
    allowed = []
    for char in str(value or ""):
        if char.isalnum() or char in "._- ":
            allowed.append(char)
        else:
            allowed.append("_")
    result = "".join(allowed).strip(" ._")
    return result or "file"


def parse_int(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def apply_header_style(sheet):
    fill = PatternFill("solid", fgColor="F0E68C")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = fill
    sheet.freeze_panes = "A2"


def autosize_columns(sheet):
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)
