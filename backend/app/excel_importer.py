import hashlib
import json
import os
import re
import urllib.parse
from datetime import date, datetime
from pathlib import Path

import httpx
import openpyxl


MAX_HEADER_SCAN_ROWS = 40
SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
SUMMARY_MARKERS = {"итого", "total", "grand total", "всего"}
DATE_PATTERN = re.compile(r"(?<!\d)(\d{1,2})[._/-](\d{1,2})[._/-](\d{2,4})(?!\d)")
COUNTRY_PREFIXES = ("узбекистан", "uzbekistan", "o'zbekiston", "oʻzbekiston")
MISSING_ADDRESS_MARKERS = {
    "адрес не указан",
    "адрес не найден",
    "адреса не найдены",
    "адрес не определен",
    "адрес отсутствует",
}
YANDEX_GEOCODER_ENV_VAR = "YANDEX_GEOCODER_API_KEY"
YANDEX_GEOCODER_URL = "https://geocode-maps.yandex.ru/v1/"

REQUIRED_ALIASES = {
    "client": [
        "ФИО или Наименование торговой точки",
        "Клиент",
        "Юр. лицо",
        "Юр лицо",
        "Наименование",
        "Покупатель",
        "Контрагент",
        "Наименование клиента",
        "Название компании",
        "Название компании/Имя человека",
        "Юридическое лицо",
    ],
    "payment": [
        "Тип оплаты",
        "Оплата",
        "Способ оплаты",
        "Форма оплаты",
        "Комментарий оплаты",
    ],
    "product": [
        "Наименование Товара",
        "Товары",
        "Товар",
        "Номенклатура",
        "ТМЦ",
        "SKU",
        "Артикул",
        "Продукт",
        "Наименование продукции",
    ],
    "quantity": [
        "Кол-во",
        "Количество",
        "Кол-во ШТ",
        "Количество ШТ",
        "Количество заказа",
        "Кол-во заказа",
        "Заказано",
        "В заявке",
        "Штук",
        "ШТ",
    ],
}

OPTIONAL_ALIASES = {
    "date": [
        "Дата доставки",
        "Дата отгрузки",
        "Дата получения заказа",
        "Дата заказа",
        "Дата",
        "Дата выгрузки",
        "Дата поставки",
        "Дата документа",
    ],
    "coordinates": [
        "Координаты",
        "Координаты клиента",
        "GPS",
        "Локация",
    ],
    "address": [
        "Адрес доставки",
        "Адрес",
        "Адрес клиента",
        "Адрес торговой точки",
        "Адрес получателя",
        "Локация",
    ],
    "representative": [
        "Торговый представитель",
        "ТП",
        "Менеджер",
        "Номер телефона",
        "Торговый",
        "Агент",
        "Ответственный",
    ],
    "blocks": [
        "Кол-во блок",
        "Кол-во блоков",
        "Блоков",
        "Количество блоков",
        "План КИЗ",
    ],
    "unit_price": [
        "Цена",
        "Цена за блок",
        "Цена блока",
        "Цена за штуку",
    ],
    "line_total": [
        "Цена заказа",
        "Сумма с переоценкой",
        "Сумма",
        "Итого сумма",
        "Итого",
    ],
    "skladbot_request_number": [
        "Номер заявки SkladBot",
        "Заявка SkladBot",
        "Номер заявки",
    ],
    "skladbot_request_id": [
        "ID заявки SkladBot",
        "SkladBot ID",
        "ID заявки",
    ],
}


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def normalize_lookup_text(value):
    text = normalize_text(value).casefold().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", "", text)


def parse_int(value):
    text = normalize_text(value).replace(" ", "").replace(",", ".")
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def parse_money(value):
    if isinstance(value, (int, float)):
        return int(value)
    text = normalize_text(value)
    if not text:
        return 0
    text = text.replace("\xa0", " ").strip()
    if re.fullmatch(r"\d+([.,]0+)?", text.replace(" ", "")):
        return int(float(text.replace(" ", "").replace(",", ".")))
    digits = re.sub(r"\D+", "", text)
    return int(digits) if digits else 0


def stable_hash(value):
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def file_sha256(path):
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def is_supported_excel_file_name(file_name):
    return Path(normalize_text(file_name)).suffix.lower() in SUPPORTED_EXCEL_EXTENSIONS


def row_has_data(row):
    return bool(row and any(normalize_text(cell) for cell in row if cell is not None))


def get_cell(row, index):
    if index is None or index >= len(row):
        return ""
    return normalize_text(row[index])


def build_header_index(header):
    result = {}
    for index, value in enumerate(header):
        key = normalize_lookup_text(value)
        if key and key not in result:
            result[key] = index
    return result


def build_header_positions(header):
    result = {}
    for index, value in enumerate(header):
        key = normalize_lookup_text(value)
        if key:
            result.setdefault(key, []).append(index)
    return result


def find_column(header_index, aliases):
    for alias in aliases:
        key = normalize_lookup_text(alias)
        if key in header_index:
            return header_index[key]
    return None


def find_columns(header_positions, aliases):
    result = []
    for alias in aliases:
        key = normalize_lookup_text(alias)
        result.extend(header_positions.get(key, []))
    return sorted(set(result))


def build_columns(header):
    header_index = build_header_index(header)
    header_positions = build_header_positions(header)
    columns = {}
    missing = []
    for key, aliases in REQUIRED_ALIASES.items():
        index = find_column(header_index, aliases)
        columns[key] = index
        if index is None:
            missing.append(aliases[0])
    for key, aliases in OPTIONAL_ALIASES.items():
        columns[key] = find_column(header_index, aliases)
    columns["coordinates_candidates"] = find_columns(header_positions, OPTIONAL_ALIASES["coordinates"])
    optional_found = sum(1 for key in OPTIONAL_ALIASES if columns.get(key) is not None)
    required_found = len(REQUIRED_ALIASES) - len(missing)
    return columns, missing, required_found * 10 + optional_found


def detect_header_row(rows):
    best = None
    for row_number, row in enumerate(rows, start=1):
        if not row_has_data(row):
            continue
        columns, missing, score = build_columns(row)
        if missing:
            continue
        candidate = {"header_row": row_number, "columns": columns, "score": score}
        if not best or candidate["score"] > best["score"]:
            best = candidate
    return best


def parse_date_text(value):
    text = normalize_text(value).replace("_", ".").replace("/", ".").replace("-", ".")
    if not text:
        return ""
    for fmt in ("%d.%m.%Y", "%Y.%m.%d", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
        except ValueError:
            pass
    return ""


def extract_dates_from_text(value):
    result = []
    for day, month, year in DATE_PATTERN.findall(normalize_text(value)):
        if len(year) == 2:
            year = "20" + year
        parsed = parse_date_text(f"{day}.{month}.{year}")
        if parsed:
            result.append(parsed)
    return result


def default_date_from_file_name(file_name):
    dates = extract_dates_from_text(Path(normalize_text(file_name)).stem)
    return dates[-1] if dates else ""


def default_date_from_context(rows, header_row):
    dates = []
    for row in rows[: max(header_row - 1, 0)]:
        for cell in row:
            dates.extend(extract_dates_from_text(cell))
    return dates[-1] if dates else ""


def is_summary_row(row, columns):
    if not row_has_data(row):
        return False
    values = [normalize_lookup_text(cell) for cell in row if normalize_lookup_text(cell)]
    if not values:
        return False
    if values[0] in SUMMARY_MARKERS:
        return True
    key_values = [
        normalize_lookup_text(get_cell(row, columns.get("client"))),
        normalize_lookup_text(get_cell(row, columns.get("payment"))),
        normalize_lookup_text(get_cell(row, columns.get("product"))),
    ]
    return key_values.count("итого") >= 2


def detect_excel_source(workbook, file_name):
    preferred_sheets = []
    if "Заявки" in workbook.sheetnames:
        preferred_sheets.append("Заявки")
    preferred_sheets.extend(sheet for sheet in workbook.sheetnames if sheet not in preferred_sheets)

    best = None
    for sheet_name in preferred_sheets:
        worksheet = workbook[sheet_name]
        preview_rows = list(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(MAX_HEADER_SCAN_ROWS, worksheet.max_row or MAX_HEADER_SCAN_ROWS),
                values_only=True,
            )
        )
        detected = detect_header_row(preview_rows)
        if not detected:
            continue
        default_date = default_date_from_file_name(file_name) or default_date_from_context(
            preview_rows,
            detected["header_row"],
        )
        candidate = {
            "sheet_name": sheet_name,
            "first_data_row": detected["header_row"] + 1,
            "columns": detected["columns"],
            "default_date": default_date,
            "score": detected["score"] + (5 if sheet_name == "Заявки" else 0),
        }
        if not best or candidate["score"] > best["score"]:
            best = candidate
    return best


def clean_address_for_display(value):
    text = normalize_text(value)
    lowered = text.casefold()
    if is_missing_address_text(text):
        return ""
    for prefix in COUNTRY_PREFIXES:
        if lowered == prefix:
            return ""
        if lowered.startswith(prefix + ","):
            return text[len(prefix):].lstrip(" ,")
    return text


def is_missing_address_text(value):
    text = normalize_text(value).casefold().replace("ё", "е")
    return not text or text in MISSING_ADDRESS_MARKERS or text.startswith("координаты")


def normalize_coordinates(value):
    text = normalize_text(value)
    if not text:
        return ""
    numbers = re.findall(r"-?\d+(?:[.,]\d+)?", text)
    if len(numbers) < 2:
        return ""
    first = numbers[0].replace(",", ".")
    second = numbers[1].replace(",", ".")
    return f"{first}, {second}"


def parse_coordinate_component(value):
    text = normalize_text(value).replace(",", ".")
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def format_coordinate(value):
    return f"{value:.12f}".rstrip("0").rstrip(".")


def normalize_split_coordinates(latitude_value, longitude_value):
    latitude = parse_coordinate_component(latitude_value)
    longitude = parse_coordinate_component(longitude_value)
    if latitude is None or longitude is None:
        return ""
    if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
        return ""
    return f"{format_coordinate(latitude)}, {format_coordinate(longitude)}"


def normalize_coordinates_from_row(row, columns):
    candidates = list(columns.get("coordinates_candidates") or [])
    primary = columns.get("coordinates")
    if primary is not None and primary not in candidates:
        candidates.insert(0, primary)

    expanded_candidates = []
    for index in candidates:
        for offset in (0, 1, 2):
            expanded = index + offset
            if expanded < len(row) and expanded not in expanded_candidates:
                expanded_candidates.append(expanded)

    for index in expanded_candidates:
        coordinates = normalize_coordinates(get_cell(row, index))
        if coordinates:
            return coordinates

    for index in candidates:
        if index + 1 >= len(row):
            continue
        coordinates = normalize_split_coordinates(get_cell(row, index), get_cell(row, index + 1))
        if coordinates:
            return coordinates
    return ""


def yandex_key():
    return normalize_text(os.environ.get(YANDEX_GEOCODER_ENV_VAR))


def geocode_address_yandex(address, cache=None):
    address = clean_address_for_display(address)
    if not address:
        return "", "адрес не указан"
    if cache is not None and address in cache:
        return cache[address]

    api_key = yandex_key()
    if not api_key:
        result = ("", "не указан ключ Яндекс Геокодера")
        if cache is not None:
            cache[address] = result
        return result

    params = {
        "apikey": api_key,
        "geocode": address,
        "format": "json",
        "lang": "ru_RU",
        "results": "1",
    }
    url = YANDEX_GEOCODER_URL + "?" + urllib.parse.urlencode(params)
    try:
        response = httpx.get(url, timeout=10)
        response.raise_for_status()
        payload = response.json()
        members = payload.get("response", {}).get("GeoObjectCollection", {}).get("featureMember", [])
        if not members:
            result = ("", "координаты не найдены")
        else:
            point = members[0].get("GeoObject", {}).get("Point", {}).get("pos") or ""
            parts = [part for part in point.replace(",", ".").split() if part]
            if len(parts) < 2:
                result = ("", "Яндекс не вернул координаты")
            else:
                longitude, latitude = parts[0], parts[1]
                result = (f"{latitude}, {longitude}", "")
    except Exception as exc:
        result = ("", f"{exc.__class__.__name__}: {exc}")

    if cache is not None:
        cache[address] = result
    return result


def reverse_geocode_yandex(coordinates, cache=None):
    coordinates = normalize_coordinates(coordinates)
    if not coordinates:
        return "", "некорректные координаты"

    cache_key = f"reverse:{coordinates}"
    if cache is not None and cache_key in cache:
        return cache[cache_key]

    api_key = yandex_key()
    if not api_key:
        result = ("", "не указан ключ Яндекс Геокодера")
        if cache is not None:
            cache[cache_key] = result
        return result

    params = {
        "apikey": api_key,
        "geocode": coordinates,
        "format": "json",
        "lang": "ru_RU",
        "sco": "latlong",
        "results": "1",
        "kind": "house",
    }
    url = YANDEX_GEOCODER_URL + "?" + urllib.parse.urlencode(params)
    try:
        response = httpx.get(url, timeout=10)
        response.raise_for_status()
        payload = response.json()
        members = payload.get("response", {}).get("GeoObjectCollection", {}).get("featureMember", [])
        if not members:
            result = ("", "адрес не найден")
        else:
            geo_object = members[0].get("GeoObject", {})
            meta = geo_object.get("metaDataProperty", {}).get("GeocoderMetaData", {})
            address = clean_address_for_display(meta.get("text") or geo_object.get("name"))
            result = (address, "") if address else ("", "Яндекс не вернул адрес")
    except Exception as exc:
        result = ("", f"{exc.__class__.__name__}: {exc}")

    if cache is not None:
        cache[cache_key] = result
    return result


def excel_file_to_import_payload(file_path, file_name=None, source="telegram", shipment_date=None):
    file_name = normalize_text(file_name) or Path(file_path).name
    if not is_supported_excel_file_name(file_name):
        raise ValueError("Поддерживаются только Excel-файлы .xlsx и .xlsm")

    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        source_info = detect_excel_source(workbook, file_name)
        if not source_info:
            raise ValueError("Не найден лист с обязательными колонками: клиент, оплата, товар, количество")

        sheet_name = source_info["sheet_name"]
        worksheet = workbook[sheet_name]
        columns = source_info["columns"]
        shipment_date = parse_date_text(shipment_date)
        default_date = shipment_date or source_info.get("default_date") or datetime.now().strftime("%d.%m.%Y")
        sha256 = file_sha256(file_path)
        default_pieces_per_block = max(1, int(os.environ.get("TAKSKLAD_DEFAULT_PIECES_PER_BLOCK", "10") or "10"))
        default_block_price = max(0, int(os.environ.get("TAKSKLAD_DEFAULT_BLOCK_PRICE", "240000") or "240000"))
        geocode_cache = {}
        rows = []
        warnings = []
        geocoded_count = 0
        geocode_failed_count = 0
        source_rows_count = 0

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=source_info["first_data_row"], values_only=True),
            start=source_info["first_data_row"],
        ):
            if not row_has_data(row) or is_summary_row(row, columns):
                continue
            source_rows_count += 1
            client = get_cell(row, columns["client"])
            payment = get_cell(row, columns["payment"])
            product = get_cell(row, columns["product"])
            quantity = parse_int(get_cell(row, columns["quantity"]))
            blocks = parse_int(get_cell(row, columns.get("blocks")))

            if not client or not payment or not product or quantity <= 0:
                warnings.append(f"{file_name}, строка {row_number}: пропущена, не заполнены клиент/оплата/товар/количество")
                continue

            if blocks <= 0:
                blocks = (quantity + default_pieces_per_block - 1) // default_pieces_per_block

            date_value = shipment_date or parse_date_text(get_cell(row, columns.get("date"))) or default_date
            address = clean_address_for_display(get_cell(row, columns.get("address")))
            coordinates = normalize_coordinates_from_row(row, columns)
            if not address and coordinates:
                geocoded_address, geocode_error = reverse_geocode_yandex(coordinates, cache=geocode_cache)
                if geocoded_address:
                    address = geocoded_address
                    geocoded_count += 1
                else:
                    geocode_failed_count += 1
                    address = f"Координаты: {coordinates}"
                    warnings.append(
                        f"{file_name}, строка {row_number}: адрес по координатам не получен ({geocode_error})"
                    )
            if not address:
                address = "Адрес не указан"
            if not coordinates:
                geocoded_coordinates, geocode_error = geocode_address_yandex(address, cache=geocode_cache)
                if geocoded_coordinates:
                    coordinates = geocoded_coordinates
                    geocoded_count += 1
                else:
                    geocode_failed_count += 1
                    warnings.append(
                        f"{file_name}, строка {row_number}: координаты по адресу не получены ({geocode_error})"
                    )
            representative = get_cell(row, columns.get("representative"))
            imported_unit_price = parse_money(get_cell(row, columns.get("unit_price")))
            imported_line_total = parse_money(get_cell(row, columns.get("line_total")))
            calculated_line_total = blocks * default_block_price
            line_total = imported_line_total or calculated_line_total
            source_id = stable_hash({"sha256": sha256, "sheet": sheet_name, "row": row_number})

            rows.append({
                "Дата отгрузки": date_value,
                "Тип оплаты": payment,
                "Клиент": client,
                "Адрес": address,
                "Координаты": coordinates,
                "Торговый представитель": representative,
                "Товары": product,
                "Кол-во ШТ": quantity,
                "Кол-во блок": blocks,
                "Цена за блок": default_block_price,
                "Цена из файла": imported_unit_price,
                "Сумма из файла": imported_line_total,
                "Сумма позиции": line_total,
                "Сумма рассчитанная": calculated_line_total,
                "Отсканированные коды": "",
                "Статус": "Не выполнено",
                "ID импорта": source_id,
                "ID заказа": stable_hash({
                    "date": date_value,
                    "payment": payment,
                    "client": client,
                    "address": address,
                    "product": product,
                    "row": row_number,
                }),
                "Источник файла": file_name,
                "Строка файла": str(row_number),
                "Дата импорта": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                "_pieces_per_block": default_pieces_per_block,
                "Номер заявки SkladBot": get_cell(row, columns.get("skladbot_request_number")),
                "ID заявки SkladBot": get_cell(row, columns.get("skladbot_request_id")),
            })
    finally:
        workbook.close()

    return {
        "source": source,
        "filename": file_name,
        "sha256": sha256,
        "rows": rows,
        "meta": {
            "sheet_name": sheet_name,
            "source_rows_count": source_rows_count,
            "shipment_date": shipment_date or default_date,
            "geocoded_count": geocoded_count,
            "geocode_failed_count": geocode_failed_count,
            "warnings": warnings,
        },
    }
