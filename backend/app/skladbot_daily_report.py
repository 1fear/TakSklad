import logging
import os
import time
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import exists, func, or_, select
from sqlalchemy.orm import aliased

from .models import Order, OrderItem
from .skladbot_client import SkladBotClient, env_float, env_int, sanitize_skladbot_error
from .skladbot_contracts import (
    business_timezone,
    business_today,
    canonical_remote_request_id,
    canonical_skladbot_request_evidence_link,
    canonical_skladbot_request_number,
    extract_list_items,
    format_internal_smartup_ids,
    normalize_request_payload,
    normalize_text,
    parse_date,
    parse_int,
    request_list_value,
)
from .representative_contacts import display_representative_name
from .spreadsheet_safety import force_workbook_text_literals
from .telegram_output_contract import build_skladbot_daily_report_message, daily_report_filename


DEFAULT_DAILY_REPORT_REQUEST_TYPE_IDS = (3387, 3388, 3389, 3391, 3403)
SKLADBOT_DAILY_REPORT_REQUEST_TYPE_IDS_ENV = "SKLADBOT_DAILY_REPORT_REQUEST_TYPE_IDS"
REQUEST_CATEGORY_SHIPMENT = "Отгрузка"
REQUEST_CATEGORY_DEFECT_SHIPMENT = "Отгрузка в браке"
REQUEST_CATEGORY_RETURN = "Возврат"
REQUEST_CATEGORY_RECEIVING = "Приемка"
REQUEST_CATEGORY_OTHER = "Прочее"
REQUEST_CATEGORIES = (
    REQUEST_CATEGORY_SHIPMENT,
    REQUEST_CATEGORY_DEFECT_SHIPMENT,
    REQUEST_CATEGORY_RETURN,
    REQUEST_CATEGORY_RECEIVING,
    REQUEST_CATEGORY_OTHER,
)

REQUEST_HEADERS = [
    "ID",
    "Номер",
    "Smartup ID",
    "Категория",
    "Тип",
    "Статус",
    "В архиве",
    "Дата создания",
    "Дата обновления",
    "Дата выгрузки",
    "Юрлицо/точка",
    "Торговый представитель",
    "Раб зона",
    "Клиент SkladBot",
    "Адрес",
    "Комментарий",
    "Блоков план",
    "Блоков факт",
    "Отклонение",
    "Товаров",
    "Причина включения",
]

REQUEST_PRODUCT_HEADERS = [
    "Заявка",
    "ID заявки",
    "Smartup ID",
    "Тип",
    "Дата выгрузки",
    "Юрлицо/точка",
    "Торговый представитель",
    "Раб зона",
    "Товар",
    "Артикул",
    "Штрихкод",
    "Блоков план",
    "Принято факт",
    "Блоков факт",
    "Отклонение",
]

MOVEMENT_HEADERS = [
    "Направление",
    "Дата",
    "Заявка/документ",
    "Тип движения",
    "Клиент",
    "Товар",
    "Артикул",
    "Штрихкод",
    "Кол-во",
    "Короб",
    "Ячейка",
]

STOCK_HEADERS = [
    "Клиент",
    "Товар",
    "Артикул",
    "Штрихкод",
    "Остаток",
    "Обычный остаток",
    "Номинальный остаток",
    "Доступно",
]

PRIMARY_DAILY_SCOPE = "Дата создания / дата выгрузки / движение склада"
COVERAGE_STATUS_COMPLETE = "complete"
COVERAGE_STATUS_PARTIAL = "partial"
COVERAGE_STATUS_FAILED = "failed"
READ_STYLE_POST_ENDPOINTS = {
    "/warehouse/transactions",
    "/products",
    "/report/stock",
}
WRITE_POST_ENDPOINTS = {
    "/requests",
}

COVERAGE_FIELDS = [
    "report_date",
    "primary_scope",
    "coverage_status",
    "pages_fetched",
    "list_pages_fetched",
    "page_limit",
    "max_pages",
    "max_pages_per_request_type",
    "list_page_guard_max_total",
    "max_pages_reached",
    "movement_pages_fetched",
    "movements_rows_returned",
    "duplicate_movement_ids",
    "movements_limit",
    "movements_truncation_possible",
    "products_rows_returned",
    "products_limit",
    "products_truncation_possible",
    "stock_rows_returned",
    "stock_limit",
    "stock_truncation_possible",
    "read_style_post_retry_count",
    "read_style_post_error_count",
    "detail_pages_fetched",
    "total_http_pages_fetched",
    "list_rows_total",
    "unique_request_ids",
    "duplicate_request_ids",
    "detail_attempted",
    "detail_attempted_in_scope",
    "detail_attempted_unknown_date",
    "detail_attempted_out_of_scope_sample",
    "detail_success",
    "detail_errors",
    "detail_limit_reached",
    "skipped_due_to_out_of_scope",
    "out_of_scope_skipped_without_detail",
    "in_scope_candidates_not_detailed",
    "included_operational_requests",
    "excluded_diagnostic_requests",
    "out_of_scope_requests",
    "completed_only_count",
    "archived_only_count",
    "neither_count",
    "missing_date_count",
    "conflicting_date_count",
    "included_date_conflict_count",
    "unloading_movement_conflict_count",
    "movement_without_unloading_count",
    "unloading_without_matching_movement_count",
    "api_error_count",
    "warnings",
]

EXCLUDED_REQUEST_HEADERS = [
    "request_id",
    "client",
    "created_at",
    "unloading_date",
    "movement_date",
    "completed_at",
    "archived_at",
    "status",
    "exclusion_reason",
    "diagnostic_reason",
    "error_message",
    "source_page",
    "detail_loaded",
]

DATE_DIAGNOSTIC_HEADERS = [
    "request_id",
    "number",
    "report_date",
    "primary_scope",
    "date_field_used",
    "inclusion_reason",
    "exclusion_reason",
    "diagnostic_reason",
    "date_confidence",
    "created_at",
    "unloading_date",
    "movement_date",
    "completed_at",
    "archived_at",
]

LOGGER = logging.getLogger(__name__)
DEFAULT_DAILY_REPORT_MAX_PAGES = 60
SKLADBOT_DAILY_REPORT_MAX_RUNTIME_SECONDS_ENV = "SKLADBOT_DAILY_REPORT_MAX_RUNTIME_SECONDS"
DEFAULT_DAILY_REPORT_MAX_RUNTIME_SECONDS = 25 * 60
SKLADBOT_DAILY_REPORT_OUT_OF_SCOPE_DETAIL_SAMPLE_LIMIT_ENV = "SKLADBOT_DAILY_REPORT_OUT_OF_SCOPE_DETAIL_SAMPLE_LIMIT"
DEFAULT_DAILY_REPORT_OUT_OF_SCOPE_DETAIL_SAMPLE_LIMIT = 25

REQUEST_DETAIL_BUCKET_IN_SCOPE = "in_scope"
REQUEST_DETAIL_BUCKET_UNKNOWN_DATE = "unknown_date"
REQUEST_DETAIL_BUCKET_CREATED_TODAY_DIAGNOSTIC = "created_today_diagnostic"
REQUEST_DETAIL_BUCKET_DIAGNOSTIC_OUT_OF_SCOPE = "diagnostic_out_of_scope"
REQUEST_DETAIL_BUCKET_KNOWN_OUT_OF_SCOPE = "known_out_of_scope"
REQUEST_DETAIL_BUCKET_ORDER = {
    REQUEST_DETAIL_BUCKET_IN_SCOPE: 0,
    REQUEST_DETAIL_BUCKET_UNKNOWN_DATE: 1,
    REQUEST_DETAIL_BUCKET_CREATED_TODAY_DIAGNOSTIC: 2,
    REQUEST_DETAIL_BUCKET_DIAGNOSTIC_OUT_OF_SCOPE: 3,
    REQUEST_DETAIL_BUCKET_KNOWN_OUT_OF_SCOPE: 4,
}


class SkladBotDailyReportTimeout(RuntimeError):
    pass


class SkladBotReadOnlyClient:
    def __init__(self, client: Any):
        self._client = client

    def __getattr__(self, name: str) -> Any:
        if name in {"create_request", "update_request", "delete_request", "return_request"}:
            raise AttributeError(f"SkladBot daily report client is read-only: {name} is forbidden")
        return getattr(self._client, name)

    def get(self, path: str, params: dict[str, Any] | None = None) -> Any:
        return self._client.get(path, params or {})

    def get_request_detail(self, request_id: int) -> Any:
        return self._client.get_request_detail(request_id)

    def post(self, path: str, payload: dict[str, Any] | None = None) -> Any:
        normalized_path = f"/{normalize_text(path).lstrip('/')}"
        if normalized_path not in READ_STYLE_POST_ENDPOINTS:
            raise RuntimeError(f"SkladBot daily report is read-only; POST {normalized_path} is forbidden")
        return self._client.post(normalized_path, payload or {})


def read_only_client(client: Any) -> SkladBotReadOnlyClient:
    if isinstance(client, SkladBotReadOnlyClient):
        return client
    return SkladBotReadOnlyClient(client)


def daily_report_max_runtime_seconds() -> int:
    return max(1, env_int(SKLADBOT_DAILY_REPORT_MAX_RUNTIME_SECONDS_ENV, DEFAULT_DAILY_REPORT_MAX_RUNTIME_SECONDS))


def ensure_daily_report_runtime_budget(started_at: float, max_runtime_seconds: int, stage: str) -> None:
    if max_runtime_seconds <= 0:
        return
    elapsed = time.monotonic() - started_at
    if elapsed > max_runtime_seconds:
        raise SkladBotDailyReportTimeout(
            f"SkladBot daily report runtime exceeded at {stage}: {elapsed:.1f}s > {max_runtime_seconds}s"
        )


def read_style_post(
    client: Any,
    path: str,
    payload: dict[str, Any] | None = None,
    coverage: dict[str, Any] | None = None,
) -> Any:
    normalized_path = f"/{normalize_text(path).lstrip('/')}"
    if normalized_path not in READ_STYLE_POST_ENDPOINTS:
        raise RuntimeError(f"Read-style POST retry is forbidden for {normalized_path}")
    retries = max(0, env_int("SKLADBOT_DAILY_REPORT_READ_POST_RETRIES", env_int("SKLADBOT_DAILY_REPORT_429_RETRIES", 2)))
    retry_seconds = max(
        0.0,
        env_float(
            "SKLADBOT_DAILY_REPORT_READ_POST_RETRY_SECONDS",
            env_float("SKLADBOT_DAILY_REPORT_429_RETRY_SECONDS", 15.0),
        ),
    )
    for attempt in range(retries + 1):
        try:
            return client.post(normalized_path, payload or {})
        except Exception as exc:
            if attempt >= retries or not is_transient_read_style_post_error(exc):
                if coverage is not None:
                    coverage["read_style_post_error_count"] = parse_int(coverage.get("read_style_post_error_count")) + 1
                raise
            if coverage is not None:
                coverage["read_style_post_retry_count"] = parse_int(coverage.get("read_style_post_retry_count")) + 1
            if retry_seconds:
                time.sleep(retry_seconds)
    raise RuntimeError(f"SkladBot read-style POST failed: {normalized_path}")


def is_transient_read_style_post_error(exc: Exception) -> bool:
    text = normalize_text(exc).lower()
    return any(marker in text for marker in (
        "429",
        "too many requests",
        "timeout",
        "timed out",
        "http 500",
        "http 502",
        "http 503",
        "http 504",
        " 500",
        " 502",
        " 503",
        " 504",
    ))


def configured_request_type_ids(environ: dict[str, str] | None = None) -> list[int]:
    environ = environ or os.environ
    raw = normalize_text(environ.get(SKLADBOT_DAILY_REPORT_REQUEST_TYPE_IDS_ENV))
    if not raw:
        return []
    result = []
    for part in raw.replace(";", ",").split(","):
        value = parse_int(part)
        if value > 0 and value not in result:
            result.append(value)
    return result


def collect_skladbot_daily_report(
    report_date: date | None = None,
    client: Any | None = None,
) -> dict[str, Any]:
    client = read_only_client(client or SkladBotClient())
    report_date = report_date or business_today()
    started_at = time.monotonic()
    max_runtime_seconds = daily_report_max_runtime_seconds()
    LOGGER.info(
        "SkladBot daily report started report_date=%s max_runtime_seconds=%s",
        report_date.isoformat(),
        max_runtime_seconds,
    )
    generated_at = datetime.now(timezone.utc).astimezone(business_timezone())
    result = {
        "report_date": report_date,
        "primary_scope": PRIMARY_DAILY_SCOPE,
        "generated_at": generated_at,
        "customer_id": getattr(client, "customer_id", None),
        "requests": [],
        "excluded_requests": [],
        "date_diagnostics": [],
        "movements": [],
        "stock": {"total": 0, "rows": [], "raw": {}, "error": ""},
        "errors": [],
        "api_errors": [],
        "coverage": default_coverage(report_date),
    }
    if not getattr(client, "configured", False):
        result["errors"].append("SKLADBOT_API_TOKEN is not configured")
        result["coverage"]["coverage_status"] = COVERAGE_STATUS_FAILED
        result["coverage"]["api_error_count"] = 1
        add_coverage_warning(result["coverage"], "skladbot_not_configured")
        finalize_report_coverage(result)
        result["summary"] = summarize_daily_report(result)
        return result

    ensure_daily_report_runtime_budget(started_at, max_runtime_seconds, "request_types")
    request_types = load_request_types(client, result["errors"])
    LOGGER.info("SkladBot daily report request types loaded report_date=%s count=%s", report_date.isoformat(), len(request_types))
    ensure_daily_report_runtime_budget(started_at, max_runtime_seconds, "movements")
    movements = fetch_daily_movements(client, report_date, result["errors"], result["coverage"])
    LOGGER.info("SkladBot daily report movements fetched report_date=%s count=%s", report_date.isoformat(), len(movements))
    result["movements"] = movements
    ensure_daily_report_runtime_budget(started_at, max_runtime_seconds, "requests")
    request_result = fetch_daily_requests(
        client,
        report_date,
        request_types,
        result["errors"],
        movements=movements,
        started_at=started_at,
        max_runtime_seconds=max_runtime_seconds,
        coverage=result["coverage"],
    )
    result["requests"] = request_result["requests"]
    result["excluded_requests"] = request_result["excluded_requests"]
    result["date_diagnostics"] = request_result["date_diagnostics"]
    result["api_errors"] = request_result["api_errors"]
    result["coverage"] = request_result["coverage"]
    ensure_daily_report_runtime_budget(started_at, max_runtime_seconds, "stock")
    result["stock"] = fetch_current_stock(client, result["errors"], result["coverage"])
    finalize_report_coverage(result)
    result["summary"] = summarize_daily_report(result)
    LOGGER.info(
        "SkladBot daily report finished report_date=%s coverage_status=%s pages=%s detail_attempted=%s detail_success=%s detail_errors=%s errors=%s",
        report_date.isoformat(),
        (result.get("coverage") or {}).get("coverage_status"),
        (result.get("coverage") or {}).get("pages_fetched"),
        (result.get("coverage") or {}).get("detail_attempted"),
        (result.get("coverage") or {}).get("detail_success"),
        (result.get("coverage") or {}).get("detail_errors"),
        len(result.get("errors") or []),
    )
    return result


def load_request_types(client: Any, errors: list[str]) -> list[dict[str, Any]]:
    configured_ids = configured_request_type_ids()
    if configured_ids:
        return [{"id": type_id, "name": ""} for type_id in configured_ids]
    try:
        payload = client.get("/requests/filter/fields")
        request_types = extract_request_types(payload)
    except Exception as exc:
        errors.append(f"Не удалось получить типы заявок SkladBot: {sanitize_skladbot_error(exc)}")
        request_types = []
    if request_types:
        return request_types
    return [{"id": type_id, "name": ""} for type_id in DEFAULT_DAILY_REPORT_REQUEST_TYPE_IDS]


def extract_request_types(payload: Any) -> list[dict[str, Any]]:
    rows = []
    for value in find_values_by_key(payload, {"types", "request_types", "requestTypes"}):
        if isinstance(value, dict):
            value = extract_list_items(value)
        if not isinstance(value, list):
            continue
        for item in value:
            if not isinstance(item, dict):
                continue
            type_id = parse_int(item.get("id") or item.get("value"))
            name = normalize_text(item.get("name") or item.get("title") or item.get("label") or item.get("type"))
            if type_id > 0 and (name or "request" not in normalize_text(item.get("group")).lower()):
                rows.append({"id": type_id, "name": name})
    seen = set()
    result = []
    for row in rows:
        if row["id"] in seen:
            continue
        seen.add(row["id"])
        result.append(row)
    return result


def find_values_by_key(value: Any, names: set[str]) -> list[Any]:
    result = []
    if isinstance(value, dict):
        for key, nested in value.items():
            if str(key) in names:
                result.append(nested)
            result.extend(find_values_by_key(nested, names))
    elif isinstance(value, list):
        for item in value:
            result.extend(find_values_by_key(item, names))
    return result


def fetch_daily_requests(
    client: Any,
    report_date: date,
    request_types: list[dict[str, Any]],
    errors: list[str],
    movements: list[dict[str, Any]] | None = None,
    started_at: float | None = None,
    max_runtime_seconds: int | None = None,
    coverage: dict[str, Any] | None = None,
) -> dict[str, Any]:
    started_at = started_at if started_at is not None else time.monotonic()
    max_runtime_seconds = max_runtime_seconds if max_runtime_seconds is not None else daily_report_max_runtime_seconds()
    limit = max(1, env_int("SKLADBOT_DAILY_REPORT_REQUESTS_LIMIT", getattr(client, "limit", 500) or 500))
    max_pages = max(1, env_int("SKLADBOT_DAILY_REPORT_MAX_PAGES", DEFAULT_DAILY_REPORT_MAX_PAGES))
    default_detail_limit = max(250, limit * max(1, len(request_types)))
    detail_limit = max(1, env_int("SKLADBOT_DAILY_REPORT_DETAIL_LIMIT", default_detail_limit))
    out_of_scope_sample_limit = max(
        0,
        env_int(
            SKLADBOT_DAILY_REPORT_OUT_OF_SCOPE_DETAIL_SAMPLE_LIMIT_ENV,
            min(DEFAULT_DAILY_REPORT_OUT_OF_SCOPE_DETAIL_SAMPLE_LIMIT, detail_limit),
        ),
    )
    request_delay = max(0.0, env_float("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", 3.0))
    coverage = coverage or default_coverage(report_date, page_limit=limit, max_pages=max_pages)
    coverage["report_date"] = report_date.isoformat()
    coverage["page_limit"] = limit
    coverage["max_pages"] = max_pages
    coverage["max_pages_per_request_type"] = max_pages
    coverage["list_page_guard_max_total"] = max_pages * max(1, len(request_types))
    result = []
    excluded_requests = []
    date_diagnostics = []
    api_errors = []
    movement_dates = movement_dates_by_request_number(movements or [])
    list_entries = crawl_daily_request_list_pages(
        client=client,
        report_date=report_date,
        request_types=request_types,
        limit=limit,
        max_pages=max_pages,
        request_delay=request_delay,
        coverage=coverage,
        errors=errors,
        api_errors=api_errors,
        started_at=started_at,
        max_runtime_seconds=max_runtime_seconds,
    )
    checked_details = 0
    detail_entries = prioritize_request_detail_entries(list_entries, report_date, movement_dates)
    for entry in detail_entries:
        ensure_daily_report_runtime_budget(started_at, max_runtime_seconds, "request_detail")
        request_id = parse_int(request_list_value(entry.get("list_item") or {}, "id"))
        detail_bucket = request_list_detail_bucket(entry.get("list_item") or {}, report_date, movement_dates)
        if (
            detail_bucket == REQUEST_DETAIL_BUCKET_KNOWN_OUT_OF_SCOPE
            and coverage.get("detail_attempted_out_of_scope_sample", 0) >= out_of_scope_sample_limit
        ):
            coverage["out_of_scope_skipped_without_detail"] += 1
            coverage["out_of_scope_requests"] += 1
            excluded_requests.append(excluded_request_from_list_entry(
                entry,
                "out_of_scope",
                "known_out_of_scope_without_detail",
                "List-level date is outside report scope; detail fetch skipped",
            ))
            continue
        if checked_details >= detail_limit:
            coverage["detail_limit_reached"] = True
            add_coverage_warning(coverage, "detail_limit")
            if detail_bucket in (REQUEST_DETAIL_BUCKET_IN_SCOPE, REQUEST_DETAIL_BUCKET_UNKNOWN_DATE):
                coverage["in_scope_candidates_not_detailed"] += 1
            elif detail_bucket == REQUEST_DETAIL_BUCKET_KNOWN_OUT_OF_SCOPE:
                coverage["out_of_scope_skipped_without_detail"] += 1
                coverage["out_of_scope_requests"] += 1
            excluded_requests.append(excluded_request_from_list_entry(
                entry,
                "detail_limit_reached",
                "detail_not_loaded",
                "Лимит детализации заявок достигнут",
            ))
            continue
        try:
            coverage["detail_attempted"] += 1
            increment_detail_bucket_counter(coverage, detail_bucket)
            checked_details += 1
            LOGGER.info(
                "SkladBot daily report detail fetch started report_date=%s request_index=%s request_id=%s",
                report_date.isoformat(),
                checked_details,
                request_id,
            )
            coverage["detail_pages_fetched"] += 1
            coverage["total_http_pages_fetched"] += 1
            detail = get_daily_request_detail(client, request_id, request_delay)
            coverage["detail_success"] += 1
        except Exception as exc:
            error_message = sanitize_skladbot_error(exc)
            coverage["detail_errors"] += 1
            errors.append(f"Не удалось получить заявку {request_id}: {error_message}")
            api_errors.append({
                "source": "detail",
                "request_id": request_id,
                "message": error_message,
                "source_page": entry.get("source_page") or "",
            })
            excluded_requests.append(excluded_request_from_list_entry(
                entry,
                "detail_error",
                "api_error",
                error_message,
            ))
            continue
        if request_delay:
            time.sleep(request_delay)
        list_item = entry.get("list_item") or {}
        request = normalize_request_payload(list_item, detail)
        request["category"] = categorize_request_type(request.get("type") or entry.get("type_name"))
        request["type_id"] = entry.get("type_id")
        request["source_page"] = entry.get("source_page")
        apply_request_scope(request, report_date, movement_dates)
        attach_request_product_identity_keys(request)
        date_diagnostics.append(date_diagnostic_row(request))
        if request.get("include_operational"):
            request["include_reasons"] = [request.get("inclusion_reason")]
            result.append(request)
            update_coverage_included_counts(coverage, request)
        else:
            excluded_requests.append(excluded_request_from_request(request, detail_loaded=True))
            update_coverage_exclusion_counts(coverage, request)

    result.sort(key=lambda item: (
        category_sort_key(item.get("category")),
        parse_int(item.get("id")),
    ))
    coverage["included_operational_requests"] = len(result)
    coverage["excluded_diagnostic_requests"] = len(excluded_requests)
    coverage["skipped_due_to_out_of_scope"] = coverage.get("out_of_scope_requests", 0)
    coverage["api_error_count"] = len(api_errors)
    finalize_coverage_status(coverage, errors)
    return {
        "requests": result,
        "excluded_requests": excluded_requests,
        "date_diagnostics": date_diagnostics,
        "api_errors": api_errors,
        "coverage": coverage,
    }


def crawl_daily_request_list_pages(
    client: Any,
    report_date: date,
    request_types: list[dict[str, Any]],
    limit: int,
    max_pages: int,
    request_delay: float,
    coverage: dict[str, Any],
    errors: list[str],
    api_errors: list[dict[str, Any]],
    started_at: float,
    max_runtime_seconds: int,
) -> list[dict[str, Any]]:
    result = []
    seen_ids = set()
    for request_type in request_types:
        type_id = parse_int(request_type.get("id"))
        if type_id <= 0:
            continue
        page = 1
        page_signatures = set()
        while True:
            ensure_daily_report_runtime_budget(started_at, max_runtime_seconds, "request_list")
            try:
                list_payload = client.get("/requests", {
                    "customer_id": getattr(client, "customer_id", None),
                    "type_id": type_id,
                    "limit": limit,
                    "page": page,
                })
                list_items = extract_list_items(list_payload)
                coverage["pages_fetched"] += 1
                coverage["list_pages_fetched"] += 1
                coverage["total_http_pages_fetched"] += 1
                coverage["list_rows_total"] += len(list_items)
                LOGGER.info(
                    "SkladBot daily report list page fetched report_date=%s type_id=%s page=%s rows=%s pages_fetched=%s",
                    report_date.isoformat(),
                    type_id,
                    page,
                    len(list_items),
                    coverage["pages_fetched"],
                )
                if request_delay:
                    time.sleep(request_delay)
            except Exception as exc:
                error_message = sanitize_skladbot_error(exc)
                errors.append(f"Не удалось получить список заявок type_id={type_id} page={page}: {error_message}")
                api_errors.append({
                    "source": "list",
                    "type_id": type_id,
                    "page": page,
                    "message": error_message,
                })
                add_coverage_warning(coverage, "list_error")
                break
            page_ids = tuple(
                parse_int(request_list_value(item, "id"))
                for item in list_items
                if isinstance(item, dict) and parse_int(request_list_value(item, "id")) > 0
            )
            repeated_page = bool(page_ids and page_ids in page_signatures)
            if page_ids:
                page_signatures.add(page_ids)
            for list_item in list_items:
                if not isinstance(list_item, dict):
                    continue
                request_id = parse_int(request_list_value(list_item, "id"))
                if request_id <= 0:
                    add_coverage_warning(coverage, "missing_request_id")
                    continue
                if request_id in seen_ids:
                    coverage["duplicate_request_ids"] += 1
                    continue
                seen_ids.add(request_id)
                result.append({
                    "list_item": list_item,
                    "type_id": type_id,
                    "type_name": request_type.get("name") or list_item.get("type") or "",
                    "source_page": page,
                })
            if repeated_page:
                add_coverage_warning(coverage, "repeated_page_ids")
                break
            if not list_items:
                break
            if len(list_items) < limit:
                break
            if page >= max_pages:
                coverage["max_pages_reached"] = True
                add_coverage_warning(coverage, "max_pages")
                break
            page += 1
    coverage["unique_request_ids"] = len(seen_ids)
    return result


def prioritize_request_detail_entries(
    list_entries: list[dict[str, Any]],
    report_date: date,
    movement_dates: dict[str, date],
) -> list[dict[str, Any]]:
    return sorted(list_entries, key=lambda entry: (
        request_list_detail_priority(entry.get("list_item") or {}, report_date, movement_dates),
        parse_int(request_list_value(entry.get("list_item") or {}, "id")),
    ))


def request_list_detail_priority(
    list_item: dict[str, Any],
    report_date: date,
    movement_dates: dict[str, date],
) -> int:
    bucket = request_list_detail_bucket(list_item, report_date, movement_dates)
    return REQUEST_DETAIL_BUCKET_ORDER.get(bucket, 9)


def request_list_detail_bucket(
    list_item: dict[str, Any],
    report_date: date,
    movement_dates: dict[str, date],
) -> str:
    request_number = normalize_text(request_list_value(list_item, "delivery_number", "number"))
    if request_number and movement_dates.get(request_number) == report_date:
        return REQUEST_DETAIL_BUCKET_IN_SCOPE

    unloading_date = parse_date(request_list_value(list_item, "unloading_date", "unloadingDate"))
    created_date = parse_date(request_list_value(list_item, "created_at", "createdAt"))
    completed_date = parse_date(request_list_value(list_item, "completedAt", "completed_at"))
    archived_date = parse_date(request_list_value(list_item, "archivedAt", "archived_at"))

    if unloading_date == report_date:
        return REQUEST_DETAIL_BUCKET_IN_SCOPE
    if not any([created_date, unloading_date, completed_date, archived_date]):
        return REQUEST_DETAIL_BUCKET_UNKNOWN_DATE
    if created_date == report_date:
        return REQUEST_DETAIL_BUCKET_IN_SCOPE
    if completed_date == report_date or archived_date == report_date:
        return REQUEST_DETAIL_BUCKET_DIAGNOSTIC_OUT_OF_SCOPE
    return REQUEST_DETAIL_BUCKET_KNOWN_OUT_OF_SCOPE


def increment_detail_bucket_counter(coverage: dict[str, Any], detail_bucket: str) -> None:
    if detail_bucket == REQUEST_DETAIL_BUCKET_IN_SCOPE:
        coverage["detail_attempted_in_scope"] += 1
    elif detail_bucket == REQUEST_DETAIL_BUCKET_UNKNOWN_DATE:
        coverage["detail_attempted_unknown_date"] += 1
    else:
        coverage["detail_attempted_out_of_scope_sample"] += 1


def default_coverage(
    report_date: date | None,
    page_limit: int = 0,
    max_pages: int = 0,
) -> dict[str, Any]:
    return {
        "report_date": report_date.isoformat() if isinstance(report_date, date) else normalize_text(report_date),
        "primary_scope": PRIMARY_DAILY_SCOPE,
        "coverage_status": COVERAGE_STATUS_COMPLETE,
        "pages_fetched": 0,
        "list_pages_fetched": 0,
        "page_limit": page_limit,
        "max_pages": max_pages,
        "max_pages_per_request_type": max_pages,
        "list_page_guard_max_total": max_pages,
        "max_pages_reached": False,
        "movement_pages_fetched": 0,
        "movements_rows_returned": 0,
        "duplicate_movement_ids": 0,
        "movements_limit": 0,
        "movements_truncation_possible": False,
        "products_rows_returned": 0,
        "products_limit": 0,
        "products_truncation_possible": False,
        "stock_rows_returned": 0,
        "stock_limit": 0,
        "stock_truncation_possible": False,
        "read_style_post_retry_count": 0,
        "read_style_post_error_count": 0,
        "detail_pages_fetched": 0,
        "total_http_pages_fetched": 0,
        "list_rows_total": 0,
        "unique_request_ids": 0,
        "duplicate_request_ids": 0,
        "detail_attempted": 0,
        "detail_attempted_in_scope": 0,
        "detail_attempted_unknown_date": 0,
        "detail_attempted_out_of_scope_sample": 0,
        "detail_success": 0,
        "detail_errors": 0,
        "detail_limit_reached": False,
        "skipped_due_to_out_of_scope": 0,
        "out_of_scope_skipped_without_detail": 0,
        "in_scope_candidates_not_detailed": 0,
        "included_operational_requests": 0,
        "excluded_diagnostic_requests": 0,
        "out_of_scope_requests": 0,
        "completed_only_count": 0,
        "archived_only_count": 0,
        "neither_count": 0,
        "missing_date_count": 0,
        "conflicting_date_count": 0,
        "included_date_conflict_count": 0,
        "unloading_movement_conflict_count": 0,
        "movement_without_unloading_count": 0,
        "unloading_without_matching_movement_count": 0,
        "api_error_count": 0,
        "warnings": "",
        "_warnings": [],
    }


def add_coverage_warning(coverage: dict[str, Any], warning: str) -> None:
    warning = normalize_text(warning)
    if not warning:
        return
    warnings = coverage.setdefault("_warnings", [])
    if warning not in warnings:
        warnings.append(warning)


def mark_possible_truncation(
    coverage: dict[str, Any] | None,
    prefix: str,
    rows_returned: int,
    limit: int,
    warning: str,
) -> None:
    if coverage is None:
        return
    coverage[f"{prefix}_rows_returned"] = parse_int(coverage.get(f"{prefix}_rows_returned")) + max(0, rows_returned)
    coverage[f"{prefix}_limit"] = max(parse_int(coverage.get(f"{prefix}_limit")), max(0, limit))
    if limit > 0 and rows_returned >= limit:
        coverage[f"{prefix}_truncation_possible"] = True
        add_coverage_warning(coverage, warning)


def update_coverage_included_counts(coverage: dict[str, Any], request: dict[str, Any]) -> None:
    unloading_date = parse_date(request.get("unloading_date"))
    movement_date = parse_date(request.get("movement_date"))
    if movement_date and not unloading_date:
        coverage["movement_without_unloading_count"] += 1
    if unloading_date and not movement_date:
        coverage["unloading_without_matching_movement_count"] += 1
    if unloading_date and movement_date and unloading_date != movement_date:
        coverage["included_date_conflict_count"] += 1
        coverage["unloading_movement_conflict_count"] += 1
        add_coverage_warning(coverage, "date_conflict_unloading_vs_movement")


def finalize_coverage_status(coverage: dict[str, Any], errors: list[str] | None = None) -> None:
    errors = errors or []
    if errors:
        add_coverage_warning(coverage, "api_error")
    warnings = list(coverage.get("_warnings") or [])
    if coverage.get("unique_request_ids") == 0 and (
        coverage.get("api_error_count") or "list_error" in warnings
    ):
        coverage["coverage_status"] = COVERAGE_STATUS_FAILED
    elif (
        coverage.get("max_pages_reached")
        or coverage.get("detail_limit_reached")
        or coverage.get("in_scope_candidates_not_detailed")
        or coverage.get("detail_errors")
        or coverage.get("api_error_count")
        or warnings
    ):
        coverage["coverage_status"] = COVERAGE_STATUS_PARTIAL
    else:
        coverage["coverage_status"] = COVERAGE_STATUS_COMPLETE
    coverage["warnings"] = "; ".join(warnings)


def finalize_report_coverage(report: dict[str, Any]) -> None:
    coverage = report.get("coverage") or default_coverage(report.get("report_date"))
    coverage["included_operational_requests"] = len(report.get("requests") or [])
    coverage["excluded_diagnostic_requests"] = len(report.get("excluded_requests") or [])
    coverage["api_error_count"] = max(parse_int(coverage.get("api_error_count")), len(report.get("errors") or []))
    finalize_coverage_status(coverage, report.get("errors") or [])
    report["coverage"] = coverage


def movement_dates_by_request_number(movements: list[dict[str, Any]]) -> dict[str, date]:
    result = {}
    for movement in movements:
        request_number = normalize_text(movement.get("request_number"))
        movement_date = parse_date(movement.get("date"))
        if request_number and movement_date and request_number not in result:
            result[request_number] = movement_date
    return result


def apply_request_scope(
    request: dict[str, Any],
    report_date: date,
    movement_dates: dict[str, date],
) -> None:
    request_number = normalize_text(request.get("number"))
    created_date = parse_date(request.get("created_at"))
    unloading_date = parse_date(request.get("unloading_date"))
    completed_date = parse_date(request.get("completed_at"))
    archived_date = parse_date(request.get("archived_at"))
    movement_date = movement_dates.get(request_number)
    request["report_date"] = report_date.isoformat()
    request["primary_scope"] = PRIMARY_DAILY_SCOPE
    request["movement_date"] = movement_date.isoformat() if movement_date else ""
    request["date_field_used"] = ""
    request["inclusion_reason"] = ""
    request["exclusion_reason"] = ""
    request["diagnostic_reason"] = ""
    request["date_confidence"] = "unknown"
    request["include_operational"] = False

    if unloading_date and movement_date and unloading_date != movement_date:
        request["diagnostic_reason"] = "conflicting_date_fields"

    if unloading_date == report_date:
        request["date_field_used"] = "unloading_date"
        request["inclusion_reason"] = "Дата выгрузки"
        request["date_confidence"] = "high"
    elif movement_date == report_date:
        request["date_field_used"] = "movement_date"
        request["inclusion_reason"] = "Движение склада"
        request["date_confidence"] = "high"
    elif created_date == report_date:
        request["date_field_used"] = "created_at"
        request["inclusion_reason"] = "Дата создания"
        request["date_confidence"] = "high"
    elif completed_date == report_date or archived_date == report_date:
        request["date_field_used"] = "completed_at" if completed_date == report_date else "archived_at"
        request["exclusion_reason"] = "out_of_scope"
        request["diagnostic_reason"] = "completed_or_archived_date_only"
        request["date_confidence"] = "diagnostic"
    elif not any([created_date, unloading_date, movement_date, completed_date, archived_date]):
        request["exclusion_reason"] = "missing_date"
        request["diagnostic_reason"] = "missing_date"
    else:
        request["exclusion_reason"] = "out_of_scope"
        request["diagnostic_reason"] = request.get("diagnostic_reason") or "out_of_scope"

    if request.get("inclusion_reason"):
        if request_is_completed_and_archived(request):
            request["include_operational"] = True
            request["exclusion_reason"] = ""
        else:
            request["exclusion_reason"] = "status_not_completed_archived"
            if not request.get("diagnostic_reason") or request.get("diagnostic_reason") == "conflicting_date_fields":
                request["diagnostic_reason"] = status_diagnostic_reason(request)
    elif not request.get("diagnostic_reason"):
        request["diagnostic_reason"] = status_diagnostic_reason(request) or "out_of_scope"

    if not request_is_completed_and_archived(request) and request.get("exclusion_reason") != "status_not_completed_archived":
        request["exclusion_reason"] = "status_not_completed_archived"
        status_reason = status_diagnostic_reason(request)
        if status_reason:
            request["diagnostic_reason"] = status_reason


def status_diagnostic_reason(request: dict[str, Any]) -> str:
    completed = bool(request.get("is_completed"))
    archived = bool(request.get("archived"))
    if completed and not archived:
        return "completed_only"
    if archived and not completed:
        return "archived_only"
    if not completed and not archived:
        return "neither"
    return ""


def update_coverage_exclusion_counts(coverage: dict[str, Any], request: dict[str, Any]) -> None:
    diagnostic_reason = normalize_text(request.get("diagnostic_reason"))
    exclusion_reason = normalize_text(request.get("exclusion_reason"))
    date_field_used = normalize_text(request.get("date_field_used"))
    if diagnostic_reason == "completed_only":
        coverage["completed_only_count"] += 1
    if diagnostic_reason == "archived_only":
        coverage["archived_only_count"] += 1
    if diagnostic_reason == "neither":
        coverage["neither_count"] += 1
    if diagnostic_reason == "missing_date" or exclusion_reason == "missing_date":
        coverage["missing_date_count"] += 1
    if diagnostic_reason == "conflicting_date_fields":
        coverage["conflicting_date_count"] += 1
    if exclusion_reason == "out_of_scope":
        coverage["out_of_scope_requests"] += 1
    primary_status_fields = {"unloading_date", "movement_date", "created_at"}
    if exclusion_reason == "status_not_completed_archived" and date_field_used in primary_status_fields:
        add_coverage_warning(coverage, "status_not_completed_archived")


def request_status_label(request: dict[str, Any]) -> str:
    completed = bool(request.get("is_completed"))
    archived = bool(request.get("archived"))
    if completed and archived:
        return "completed+archived"
    if completed:
        return "completed_only"
    if archived:
        return "archived_only"
    return "neither"


def excluded_request_from_request(request: dict[str, Any], detail_loaded: bool) -> dict[str, Any]:
    return {
        "request_id": request.get("id") or "",
        "client": request.get("recipient") or request.get("customer_name") or "",
        "created_at": request.get("created_at") or "",
        "unloading_date": request.get("unloading_date") or "",
        "movement_date": request.get("movement_date") or "",
        "completed_at": request.get("completed_at") or "",
        "archived_at": request.get("archived_at") or "",
        "status": request_status_label(request),
        "exclusion_reason": request.get("exclusion_reason") or "",
        "diagnostic_reason": request.get("diagnostic_reason") or "",
        "error_message": "",
        "source_page": request.get("source_page") or "",
        "detail_loaded": bool(detail_loaded),
    }


def excluded_request_from_list_entry(
    entry: dict[str, Any],
    exclusion_reason: str,
    diagnostic_reason: str,
    error_message: str,
) -> dict[str, Any]:
    list_item = entry.get("list_item") or {}
    return {
        "request_id": parse_int(request_list_value(list_item, "id")) or "",
        "client": first_text(list_item, "customer", "client"),
        "created_at": request_list_value(list_item, "created_at", "createdAt"),
        "unloading_date": request_list_value(list_item, "unloading_date", "unloadingDate"),
        "movement_date": "",
        "completed_at": request_list_value(list_item, "completedAt", "completed_at"),
        "archived_at": request_list_value(list_item, "archivedAt", "archived_at"),
        "status": list_item_status_label(list_item),
        "exclusion_reason": exclusion_reason,
        "diagnostic_reason": diagnostic_reason,
        "error_message": error_message,
        "source_page": entry.get("source_page") or "",
        "detail_loaded": False,
    }


def list_item_status_label(list_item: dict[str, Any]) -> str:
    return request_status_label({
        "is_completed": bool(list_item.get("isCompleted") or list_item.get("is_completed")),
        "archived": bool(list_item.get("archived")),
    })


def date_diagnostic_row(request: dict[str, Any]) -> dict[str, Any]:
    return {
        "request_id": request.get("id") or "",
        "number": request.get("number") or "",
        "report_date": request.get("report_date") or "",
        "primary_scope": request.get("primary_scope") or PRIMARY_DAILY_SCOPE,
        "date_field_used": request.get("date_field_used") or "",
        "inclusion_reason": request.get("inclusion_reason") or "",
        "exclusion_reason": request.get("exclusion_reason") or "",
        "diagnostic_reason": request.get("diagnostic_reason") or "",
        "date_confidence": request.get("date_confidence") or "",
        "created_at": request.get("created_at") or "",
        "unloading_date": request.get("unloading_date") or "",
        "movement_date": request.get("movement_date") or "",
        "completed_at": request.get("completed_at") or "",
        "archived_at": request.get("archived_at") or "",
    }


def attach_request_product_identity_keys(request: dict[str, Any]) -> None:
    request_id = normalize_text(request.get("id"))
    for product in request.get("products") or []:
        product_identity = (
            normalize_text(product.get("vendor_code"))
            or normalize_text(product.get("barcode"))
            or normalize_text(product.get("name"))
        )
        product["source_type"] = "request_product"
        product["source_identity_key"] = "|".join([
            "request",
            request_id,
            product_identity,
            normalize_text(product.get("amount")),
            "request_product",
        ])


def get_daily_request_detail(client: Any, request_id: int, request_delay: float) -> Any:
    rate_limit_retries = max(0, env_int("SKLADBOT_DAILY_REPORT_429_RETRIES", 2))
    retry_seconds = max(request_delay, env_float("SKLADBOT_DAILY_REPORT_429_RETRY_SECONDS", 15.0))
    for attempt in range(rate_limit_retries + 1):
        try:
            return client.get_request_detail(request_id)
        except Exception as exc:
            if attempt >= rate_limit_retries or not is_skladbot_rate_limit_error(exc):
                raise
            if retry_seconds:
                time.sleep(retry_seconds)
    raise RuntimeError(f"Не удалось получить заявку {request_id}")


def prioritize_request_list_items(list_items: list[Any], report_date: date) -> list[Any]:
    return sorted(list_items, key=lambda item: (
        0 if date_matches(request_list_value(item, "created_at", "createdAt"), report_date) else 1,
    ))


def report_date_request_list_items(list_items: list[Any], report_date: date) -> list[Any]:
    return [
        item
        for item in prioritize_request_list_items(list_items, report_date)
        if date_matches(request_list_value(item, "created_at", "createdAt"), report_date)
    ]


def is_skladbot_rate_limit_error(exc: Exception) -> bool:
    text = sanitize_skladbot_error(exc).lower()
    return "429" in text or "too many requests" in text


def request_inclusion_reasons(
    request: dict[str, Any],
    report_date: date,
) -> list[str]:
    if not request_is_completed_and_archived(request):
        return []
    if request_created_on_report_date(request, report_date):
        return ["создана"]
    return []


def request_is_completed_and_archived(request: dict[str, Any]) -> bool:
    return bool(request.get("is_completed") and request.get("archived"))


def request_created_on_report_date(request: dict[str, Any], report_date: date) -> bool:
    return date_matches(request_list_value(request, "created_at", "createdAt"), report_date)


def date_matches(value: Any, expected: date) -> bool:
    parsed = parse_date(value)
    return bool(parsed and parsed == expected)


def categorize_request_type(value: Any) -> str:
    text = normalize_text(value).lower().replace("ё", "е")
    is_outbound = "отгруз" in text or "расход" in text
    if is_outbound and "брак" in text:
        return REQUEST_CATEGORY_DEFECT_SHIPMENT
    if "возврат" in text:
        return REQUEST_CATEGORY_RETURN
    if is_outbound:
        return REQUEST_CATEGORY_SHIPMENT
    if "прием" in text or "приемка" in text:
        return REQUEST_CATEGORY_RECEIVING
    return REQUEST_CATEGORY_OTHER


def category_sort_key(value: Any) -> int:
    return {category: index for index, category in enumerate(REQUEST_CATEGORIES, start=1)}.get(normalize_text(value), 9)


def fetch_daily_movements(
    client: Any,
    report_date: date,
    errors: list[str],
    coverage: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    limit = max(1, env_int("SKLADBOT_DAILY_REPORT_MOVEMENTS_LIMIT", 1000))
    if coverage is not None:
        coverage["movements_limit"] = limit
    result = []
    seen_movement_ids = set()
    for movement_type, direction in (("in", "Приход"), ("out", "Расход")):
        try:
            payload = read_style_post(client, "/warehouse/transactions", {
                "customer_id": getattr(client, "customer_id", None),
                "limit": limit,
                "type": movement_type,
                "from": report_date.isoformat(),
                "to": report_date.isoformat(),
            }, coverage)
        except Exception as exc:
            errors.append(f"Не удалось получить движения {direction}: {sanitize_skladbot_error(exc)}")
            continue
        list_items = extract_list_items(payload)
        if coverage is not None:
            coverage["movement_pages_fetched"] += 1
            mark_possible_truncation(
                coverage,
                "movements",
                len(list_items),
                limit,
                "movements_possible_truncation",
            )
        for item in list_items:
            if isinstance(item, dict):
                movement = normalize_movement(item, direction)
                if movement_on_report_date(movement, report_date):
                    movement_id = normalize_text(movement.get("source_id"))
                    if movement_id and movement_id in seen_movement_ids:
                        if coverage is not None:
                            coverage["duplicate_movement_ids"] += 1
                        continue
                    if movement_id:
                        seen_movement_ids.add(movement_id)
                    result.append(movement)
    result.sort(key=lambda item: (normalize_text(item.get("date")), normalize_text(item.get("request_number"))))
    return result


def movement_on_report_date(movement: dict[str, Any], report_date: date) -> bool:
    return date_matches(movement.get("date"), report_date)


def normalize_movement(item: dict[str, Any], direction: str) -> dict[str, Any]:
    product = first_nested_dict(item, "product", "nomenclature", "product_data", "productData")
    customer = first_nested_dict(item, "customer", "client")
    box = first_nested_dict(item, "box")
    cell = first_nested_dict(item, "cell", "place", "location")
    row = {
        "direction": direction,
        "date": first_text(item, "date", "created_at", "createdAt", "datetime", "created"),
        "request_number": first_text(item, "delivery_number", "request_number", "request", "document", "source"),
        "movement_type": first_text(item, "type", "movement_type", "operation"),
        "customer": nested_text(customer, "name", "title") or first_text(item, "customer", "client"),
        "product": nested_text(product, "name", "title") or first_text(item, "product", "name", "title"),
        "vendor_code": nested_text(product, "vendorCode", "vendor_code", "article", "sku") or first_text(item, "vendorCode", "vendor_code", "article", "sku"),
        "barcode": nested_text(product, "barcode") or first_text(item, "barcode"),
        "amount": first_int(item, "amount", "quantity", "count", "qty"),
        "box": nested_text(box, "name", "number", "title") or first_text(item, "box"),
        "cell": nested_text(cell, "name", "title", "code") or first_text(item, "cell", "place", "location"),
        "raw": item,
    }
    movement_id = first_text(item, "id", "uuid", "movement_id", "transaction_id")
    row["source_type"] = "movement"
    row["source_id"] = movement_id
    if movement_id:
        row["source_identity_key"] = f"movement:{movement_id}"
    else:
        row["source_identity_key"] = "|".join([
            "movement",
            normalize_text(row.get("request_number")),
            normalize_text(row.get("vendor_code") or row.get("barcode") or row.get("product")),
            normalize_text(row.get("amount")),
            normalize_text(row.get("date")),
            normalize_text(row.get("customer")),
            normalize_text(row.get("movement_type")),
        ])
    return row


def fetch_current_stock(
    client: Any,
    errors: list[str],
    coverage: dict[str, Any] | None = None,
) -> dict[str, Any]:
    products_stock = fetch_products_stock(client, errors, coverage)
    stock_limit = max(1, env_int("SKLADBOT_DAILY_REPORT_STOCK_LIMIT", 1000))
    if coverage is not None:
        coverage["stock_limit"] = stock_limit
    try:
        payload = read_style_post(client, "/report/stock", {
            "customer_id": getattr(client, "customer_id", None),
            "with_details": True,
            "limit": stock_limit,
        }, coverage)
    except Exception as exc:
        error = f"Не удалось получить остаток SkladBot: {sanitize_skladbot_error(exc)}"
        errors.append(error)
        if products_stock["rows"]:
            products_stock["error"] = error
            return products_stock
        return {"total": 0, "rows": [], "raw": {}, "error": error}
    report_stock_rows = normalize_stock_rows(payload)
    if coverage is not None:
        mark_possible_truncation(
            coverage,
            "stock",
            len(report_stock_rows),
            stock_limit,
            "stock_possible_truncation",
        )
    if products_stock["rows"]:
        products_stock["raw"] = {"products": products_stock.get("raw") or {}, "report_stock": payload}
        products_stock["report_total"] = stock_total(payload, report_stock_rows)
        return products_stock
    rows = report_stock_rows
    total = stock_total(payload, rows)
    return {"total": total, "rows": rows, "raw": payload, "error": ""}


def fetch_products_stock(
    client: Any,
    errors: list[str],
    coverage: dict[str, Any] | None = None,
) -> dict[str, Any]:
    limit = max(1, env_int("SKLADBOT_DAILY_REPORT_PRODUCTS_LIMIT", 1000))
    if coverage is not None:
        coverage["products_limit"] = limit
    try:
        payload = read_style_post(client, "/products", {
            "customer_id": getattr(client, "customer_id", None),
            "limit": limit,
        }, coverage)
    except Exception as exc:
        errors.append(f"Не удалось получить товары SkladBot: {sanitize_skladbot_error(exc)}")
        return {"total": 0, "rows": [], "raw": {}, "error": ""}
    rows = normalize_stock_rows(payload)
    if coverage is not None:
        mark_possible_truncation(
            coverage,
            "products",
            len(rows),
            limit,
            "products_possible_truncation",
        )
    total = sum(parse_int(row.get("stock")) for row in rows)
    return {"total": total, "rows": rows, "raw": payload, "error": ""}


def normalize_stock_rows(payload: Any) -> list[dict[str, Any]]:
    rows = []
    for item in find_stock_like_dicts(payload):
        row = normalize_stock_row(item)
        if any(row.get(key) for key in ("product", "vendor_code", "barcode")) or row.get("stock") or row.get("available"):
            rows.append(row)
    seen = set()
    result = []
    for row in rows:
        key = (
            row.get("customer"),
            row.get("product"),
            row.get("vendor_code"),
            row.get("barcode"),
            row.get("stock"),
            row.get("available"),
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    result.sort(key=lambda item: (normalize_text(item.get("product")), normalize_text(item.get("barcode"))))
    return result


def find_stock_like_dicts(value: Any) -> list[dict[str, Any]]:
    result = []
    if isinstance(value, dict):
        keys = {normalize_text(key).lower() for key in value.keys()}
        if keys.intersection({"stock", "balance", "available", "quantity", "amount", "product", "product_data", "productdata"}):
            result.append(value)
        for nested in value.values():
            result.extend(find_stock_like_dicts(nested))
    elif isinstance(value, list):
        for item in value:
            result.extend(find_stock_like_dicts(item))
    return result


def normalize_stock_row(item: dict[str, Any]) -> dict[str, Any]:
    product = first_nested_dict(item, "product", "product_data", "productData", "nomenclature")
    customer = first_nested_dict(item, "customer", "client")
    return {
        "customer": nested_text(customer, "name", "title") or first_text(item, "customer", "client"),
        "product": nested_text(product, "name", "title") or first_text(item, "product", "name", "title"),
        "vendor_code": nested_text(product, "vendorCode", "vendor_code", "article", "sku") or first_text(item, "vendorCode", "vendor_code", "article", "sku"),
        "barcode": nested_text(product, "barcode") or first_text(item, "barcode"),
        "stock": first_int(item, "stock", "balance", "quantity", "amount", "count", "total"),
        "regular_stock": first_int(item, "regular_stock", "ordinary_stock", "stock_regular", "normal_stock"),
        "nominal_stock": first_int(item, "nominal_stock", "nominale_stock", "stock_nominal", "nominale"),
        "available": first_int(item, "available", "available_stock", "free", "free_stock"),
        "raw": item,
    }


def stock_total(payload: Any, rows: list[dict[str, Any]]) -> int:
    root_total = first_int(payload, "total", "stock", "balance", "amount", "quantity") if isinstance(payload, dict) else 0
    if root_total:
        return root_total
    row_total = sum(parse_int(row.get("stock")) for row in rows)
    return row_total


def first_nested_dict(item: dict[str, Any], *keys: str) -> dict[str, Any]:
    for key in keys:
        value = item.get(key)
        if isinstance(value, dict):
            return value
    return {}


def nested_text(item: dict[str, Any], *keys: str) -> str:
    return first_text(item, *keys) if item else ""


def first_text(item: Any, *keys: str) -> str:
    if not isinstance(item, dict):
        return ""
    for key in keys:
        value = item.get(key)
        if isinstance(value, dict):
            text = nested_text(value, "name", "title", "number", "code")
        else:
            text = normalize_text(value)
        if text:
            return text
    return ""


def first_int(item: Any, *keys: str) -> int:
    if not isinstance(item, dict):
        return 0
    for key in keys:
        if key not in item:
            continue
        value = parse_int(item.get(key))
        if value:
            return value
    return 0


def summarize_daily_report(report: dict[str, Any]) -> dict[str, Any]:
    requests = report.get("requests") or []
    movements = report.get("movements") or []
    category_counts = {category: 0 for category in REQUEST_CATEGORIES}
    type_counts: dict[str, int] = {}
    request_blocks_by_category = {category: 0 for category in REQUEST_CATEGORIES}
    for request in requests:
        category = normalize_text(request.get("category")) or REQUEST_CATEGORY_OTHER
        category_counts[category] = category_counts.get(category, 0) + 1
        type_name = normalize_text(request.get("type")) or "Без типа"
        type_counts[type_name] = type_counts.get(type_name, 0) + 1
        request_blocks_by_category[category] = request_blocks_by_category.get(category, 0) + request_report_blocks(request)
    movement_in = [item for item in movements if item.get("direction") == "Приход"]
    movement_out = [item for item in movements if item.get("direction") == "Расход"]
    return {
        "requests_total": len(requests),
        "category_counts": category_counts,
        "type_counts": type_counts,
        "request_blocks_by_category": request_blocks_by_category,
        "movements_total": len(movements),
        "movement_in_rows": len(movement_in),
        "movement_out_rows": len(movement_out),
        "movement_in_amount": sum(parse_int(item.get("amount")) for item in movement_in),
        "movement_out_amount": sum(parse_int(item.get("amount")) for item in movement_out),
        "stock_total": parse_int((report.get("stock") or {}).get("total")),
        "stock_rows": len((report.get("stock") or {}).get("rows") or []),
        "errors": len(report.get("errors") or []),
    }


def request_blocks(request: dict[str, Any]) -> int:
    return sum(parse_int(product.get("amount")) for product in request.get("products") or [])


def request_report_blocks(request: dict[str, Any]) -> int:
    category = normalize_text(request.get("category")) or REQUEST_CATEGORY_OTHER
    return sum(report_product_blocks(product, category) for product in request.get("products") or [])


def report_product_blocks(product: dict[str, Any], category: str) -> int:
    if normalize_text(category) == REQUEST_CATEGORY_RECEIVING:
        accepted_amount = parse_int(product.get("accepted_amount"))
        if product.get("accepted_amount_present"):
            return accepted_amount
        if accepted_amount > 0:
            return accepted_amount
    return parse_int(product.get("amount"))


def product_key(name: Any, vendor_code: Any = "", barcode: Any = "") -> str:
    aliases = product_aliases(name, vendor_code, barcode)
    return aliases[0] if aliases else ""


def product_aliases(name: Any, vendor_code: Any = "", barcode: Any = "") -> list[str]:
    aliases = []
    product_name = normalize_text(name).lower()
    if product_name:
        aliases.append(f"name:{product_name}")
    vendor = normalize_text(vendor_code).lower()
    if vendor:
        aliases.append(f"vendor:{vendor}")
    product_barcode = normalize_text(barcode).lower()
    if product_barcode:
        aliases.append(f"barcode:{product_barcode}")
    return aliases


def product_label(name: Any, vendor_code: Any = "", barcode: Any = "") -> str:
    return (
        normalize_text(name)
        or normalize_text(vendor_code)
        or normalize_text(barcode)
        or "Товар не найден"
    )


def product_breakdown_for_summary(report: dict[str, Any]) -> list[dict[str, Any]]:
    products: dict[str, dict[str, Any]] = {}
    aliases_by_product: dict[str, str] = {}

    def ensure_product(name: Any, vendor_code: Any = "", barcode: Any = "") -> dict[str, Any]:
        aliases = product_aliases(name, vendor_code, barcode)
        key = next((
            aliases_by_product[alias]
            for alias in aliases
            if alias in aliases_by_product
        ), "")
        if not key:
            key = product_key(name, vendor_code, barcode)
        if not key:
            key = f"unknown:{len(products) + 1}"
        if key not in products:
            products[key] = {
                "key": key,
                "name": product_label(name, vendor_code, barcode),
                "ending_stock": 0,
                "inbound": 0,
                "outbound": 0,
                "defect_outbound": 0,
                "returns": 0,
            }
        elif normalize_text(name) and not normalize_text(products[key].get("name")):
            products[key]["name"] = product_label(name, vendor_code, barcode)
        for alias in aliases:
            aliases_by_product[alias] = key
        return products[key]

    for row in (report.get("stock") or {}).get("rows") or []:
        product = ensure_product(row.get("product"), row.get("vendor_code"), row.get("barcode"))
        product["ending_stock"] += parse_int(row.get("stock"))

    for request in report.get("requests") or []:
        category = normalize_text(request.get("category")) or REQUEST_CATEGORY_OTHER
        for request_product in request.get("products") or []:
            product = ensure_product(
                request_product.get("name"),
                request_product.get("vendor_code"),
                request_product.get("barcode"),
            )
            amount = report_product_blocks(request_product, category)
            if category == REQUEST_CATEGORY_RECEIVING:
                product["inbound"] += amount
            elif category == REQUEST_CATEGORY_SHIPMENT:
                product["outbound"] += amount
            elif category == REQUEST_CATEGORY_DEFECT_SHIPMENT:
                product["defect_outbound"] += amount
            elif category == REQUEST_CATEGORY_RETURN:
                product["returns"] += amount

    result = list(products.values())
    result.sort(key=lambda item: normalize_text(item.get("name")).lower())
    return result


def build_skladbot_daily_report_xlsx(report: dict[str, Any]) -> tuple[bytes, str]:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    write_summary_sheet(summary_sheet, report)
    write_requests_sheet(workbook.create_sheet("Заявки"), report.get("requests") or [])
    write_request_products_sheet(workbook.create_sheet("Товары заявок"), report.get("requests") or [])
    for sheet in workbook.worksheets:
        autosize_columns(sheet)
    apply_report_template_widths(workbook)
    buffer = BytesIO()
    force_workbook_text_literals(workbook)
    workbook.save(buffer)
    return buffer.getvalue(), daily_report_filename(report.get("report_date"))


def write_summary_sheet(sheet, report: dict[str, Any]) -> None:
    summary = report.get("summary") or {}
    category_counts = summary.get("category_counts") or {}
    blocks = summary.get("request_blocks_by_category") or {}
    sheet.append(["Показатель", "Блоков", "Заявок"])
    for category in (
        REQUEST_CATEGORY_SHIPMENT,
        REQUEST_CATEGORY_DEFECT_SHIPMENT,
        REQUEST_CATEGORY_RETURN,
        REQUEST_CATEGORY_RECEIVING,
    ):
        sheet.append([
            category,
            parse_int(blocks.get(category)),
            parse_int(category_counts.get(category)),
        ])
    sheet.append(["Актуальный остаток", parse_int(summary.get("stock_total")), None])
    apply_header_style(sheet)
    for cell in ("A6", "B6"):
        sheet[cell].font = Font(bold=True)
    apply_thin_border(sheet, "A2:C6")


def write_requests_sheet(sheet, requests: list[dict[str, Any]]) -> None:
    sheet.append(REQUEST_HEADERS)
    for request in requests:
        planned_blocks = request_blocks(request)
        actual_blocks = request_report_blocks(request)
        representative = request_representative(request)
        representative_zone = request_representative_zone(request)
        sheet.append([
            request.get("id") or "",
            request.get("number") or "",
            request_smartup_id(request),
            request.get("category") or "",
            request.get("type") or "",
            "Выполнена" if request.get("is_completed") else "Не выполнена",
            "Да" if request.get("archived") else "Нет",
            request.get("created_at") or "",
            request.get("updated_at") or "",
            request.get("unloading_date") or "",
            request.get("recipient") or "",
            representative,
            representative_zone,
            request.get("customer_name") or "",
            request.get("address") or "",
            request.get("comment") or "",
            planned_blocks,
            actual_blocks,
            actual_blocks - planned_blocks,
            len(request.get("products") or []),
            ", ".join(request.get("include_reasons") or []),
        ])
    apply_header_style(sheet)


def write_request_products_sheet(sheet, requests: list[dict[str, Any]]) -> None:
    sheet.append(REQUEST_PRODUCT_HEADERS)
    for request in requests:
        category = normalize_text(request.get("category")) or REQUEST_CATEGORY_OTHER
        representative = request_representative(request)
        representative_zone = request_representative_zone(request)
        for product in request.get("products") or []:
            planned_blocks = parse_int(product.get("amount"))
            actual_blocks = report_product_blocks(product, category)
            accepted_amount = parse_int(product.get("accepted_amount"))
            sheet.append([
                request.get("number") or "",
                request.get("id") or "",
                request_smartup_id(request),
                request.get("type") or "",
                request.get("unloading_date") or "",
                request.get("recipient") or "",
                representative,
                representative_zone,
                product.get("name") or "",
                product.get("vendor_code") or "",
                product.get("barcode") or "",
                planned_blocks,
                accepted_amount,
                actual_blocks,
                actual_blocks - planned_blocks,
            ])
    apply_header_style(sheet)

def request_representative(request: dict[str, Any]) -> str:
    explicit = normalize_text(request.get("representative"))
    if explicit:
        return display_representative_name(explicit)
    return display_representative_name(representative_from_comment(request.get("comment")))

def request_smartup_id(request: dict[str, Any]) -> str:
    return normalize_text(request.get("smartup_id"))


def canonical_daily_request_number(value: Any) -> str:
    return canonical_skladbot_request_number(value)


def canonical_daily_request_evidence_pair(request: dict[str, Any]) -> tuple[str, str]:
    return canonical_skladbot_request_evidence_link(
        request,
        allow_missing_raw=True,
        allow_single_raw_side=True,
    )


def enrich_smartup_ids_from_orders(db, report: dict[str, Any]) -> None:
    """Attach Smartup ids only when one durable ID/number pair has one owner."""
    requests = [request for request in report.get("requests") or [] if isinstance(request, dict)]
    report_pairs = set()
    for request in requests:
        request["smartup_id"] = ""
        request_id, request_number = canonical_daily_request_evidence_pair(request)
        if request_id and request_number:
            report_pairs.add((request_id, request_number))
    if not report_pairs:
        return

    report_ids = sorted({request_id for request_id, _request_number in report_pairs})
    report_numbers = sorted({request_number for _request_id, request_number in report_pairs})
    raw_payload = Order.raw_payload
    candidate_item = aliased(OrderItem)
    item_raw_payload = candidate_item.raw_payload
    order_link_values = (
        raw_payload["skladbot_request_id"].as_string(),
        raw_payload["skladbot_return_request_id"].as_string(),
    )
    order_number_values = (
        raw_payload["skladbot_request_number"].as_string(),
        raw_payload["skladbot_return_request_number"].as_string(),
    )
    item_link_values = (
        item_raw_payload["skladbot_request_id"].as_string(),
        item_raw_payload["skladbot_return_request_id"].as_string(),
    )
    item_number_values = (
        item_raw_payload["skladbot_request_number"].as_string(),
        item_raw_payload["skladbot_return_request_number"].as_string(),
    )
    order_matches = or_(
        *(func.trim(func.coalesce(value, "")).in_(report_ids) for value in order_link_values),
        *(func.trim(func.coalesce(value, "")).in_(report_numbers) for value in order_number_values),
    )
    item_matches = or_(
        *(func.trim(func.coalesce(value, "")).in_(report_ids) for value in item_link_values),
        *(func.trim(func.coalesce(value, "")).in_(report_numbers) for value in item_number_values),
    )
    owner_ids = select(Order.id).where(or_(
        order_matches,
        exists(
            select(1).where(
                candidate_item.order_id == Order.id,
                item_matches,
            )
        ),
    ))
    rows = db.execute(
        select(Order.id, Order.raw_payload, OrderItem.raw_payload.label("item_raw_payload"))
        .outerjoin(OrderItem, OrderItem.order_id == Order.id)
        .where(Order.id.in_(owner_ids))
    ).all()

    order_sources: dict[str, list[Any]] = {}
    owner_payloads: dict[str, list[dict[str, Any]]] = {}
    for order_id, order_raw, item_raw in rows:
        owner_id = str(order_id)
        normalized_order_raw = order_raw if isinstance(order_raw, dict) else {}
        normalized_item_raw = item_raw if isinstance(item_raw, dict) else {}
        payloads = owner_payloads.setdefault(owner_id, [])
        payloads.extend((normalized_order_raw, normalized_item_raw))
        sources = order_sources.setdefault(owner_id, [normalized_order_raw.get("source_order_id")])
        sources.append(normalized_item_raw.get("source_order_id"))

    pair_owners: dict[tuple[str, str], set[str]] = {}
    id_links: dict[str, set[tuple[str, str]]] = {}
    number_links: dict[str, set[tuple[str, str]]] = {}
    for owner_id, payloads in owner_payloads.items():
        for payload in payloads:
            for prefix in ("skladbot_request", "skladbot_return_request"):
                request_id = canonical_remote_request_id(payload.get(f"{prefix}_id"))
                request_number = canonical_daily_request_number(payload.get(f"{prefix}_number"))
                link = (request_id, request_number)
                if request_id:
                    id_links.setdefault(request_id, set()).add(link)
                if request_number:
                    number_links.setdefault(request_number, set()).add(link)
                if request_id and request_number:
                    pair_owners.setdefault(link, set()).add(owner_id)

    for request in requests:
        pair = canonical_daily_request_evidence_pair(request)
        owners = pair_owners.get(pair, set())
        if (
            not all(pair)
            or len(owners) != 1
            or id_links.get(pair[0], set()) != {pair}
            or number_links.get(pair[1], set()) != {pair}
        ):
            continue
        owner_id = next(iter(owners))
        request["smartup_id"] = format_internal_smartup_ids(order_sources.get(owner_id, []))


def representative_from_comment(comment: Any) -> str:
    lines = [normalize_text(line) for line in normalize_text(comment).splitlines()]
    lines = [line for line in lines if line]
    if len(lines) < 2 or not is_payment_comment_line(lines[0]):
        return ""
    candidate = lines[1]
    if ":" in candidate:
        return ""
    return candidate


def request_representative_zone(request: dict[str, Any]) -> str:
    explicit = normalize_text(
        request.get("representative_zone")
        or request.get("work_zone")
        or request.get("zone")
    )
    if explicit:
        return explicit
    return representative_zone_from_comment(request.get("comment"))


def representative_zone_from_comment(comment: Any) -> str:
    lines = [normalize_text(line) for line in normalize_text(comment).splitlines()]
    for line in lines:
        if ":" not in line:
            continue
        label, value = line.split(":", 1)
        if "зона" in label.lower().replace("ё", "е"):
            return normalize_text(value)
    return ""


def is_payment_comment_line(value: Any) -> bool:
    text = normalize_text(value).lower().replace("ё", "е")
    return "терминал" in text or "перечис" in text or "безнал" in text


def write_movements_sheet(sheet, movements: list[dict[str, Any]]) -> None:
    sheet.append(MOVEMENT_HEADERS)
    for item in movements:
        sheet.append([
            item.get("direction") or "",
            item.get("date") or "",
            item.get("request_number") or "",
            item.get("movement_type") or "",
            item.get("customer") or "",
            item.get("product") or "",
            item.get("vendor_code") or "",
            item.get("barcode") or "",
            item.get("amount") or 0,
            item.get("box") or "",
            item.get("cell") or "",
        ])
    apply_header_style(sheet)


def write_stock_sheet(sheet, report: dict[str, Any]) -> None:
    rows = (report.get("stock") or {}).get("rows") or []
    summary = report.get("summary") or {}
    sheet.append(STOCK_HEADERS)
    if rows:
        for row in rows:
            sheet.append([
                row.get("customer") or "",
                row.get("product") or "",
                row.get("vendor_code") or "",
                row.get("barcode") or "",
                parse_int(row.get("stock")),
                parse_int(row.get("regular_stock")),
                parse_int(row.get("nominal_stock")),
                parse_int(row.get("available")),
            ])
    else:
        sheet.append([
            "",
            "",
            "",
            "",
            parse_int(summary.get("stock_total")),
            0,
            0,
            0,
        ])
    apply_header_style(sheet)


def write_coverage_sheet(sheet, report: dict[str, Any]) -> None:
    coverage = report.get("coverage") or default_coverage(report.get("report_date"))
    sheet.append(["Поле", "Значение"])
    for field in COVERAGE_FIELDS:
        value = coverage.get(field)
        if isinstance(value, bool):
            value = bool(value)
        sheet.append([field, value])
    apply_header_style(sheet)


def write_excluded_requests_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    sheet.append(EXCLUDED_REQUEST_HEADERS)
    for row in rows:
        sheet.append([row.get(header) for header in EXCLUDED_REQUEST_HEADERS])
    apply_header_style(sheet)


def write_date_diagnostics_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    sheet.append(DATE_DIAGNOSTIC_HEADERS)
    for row in rows:
        sheet.append([row.get(header) for header in DATE_DIAGNOSTIC_HEADERS])
    apply_header_style(sheet)


def write_errors_sheet(sheet, errors: list[str], api_errors: list[dict[str, Any]] | None = None) -> None:
    sheet.append(["Ошибка"])
    for error in errors:
        sheet.append([normalize_text(error)])
    for error in api_errors or []:
        message = normalize_text(error.get("message"))
        source = normalize_text(error.get("source"))
        request_id = normalize_text(error.get("request_id"))
        prefix = source
        if request_id:
            prefix = f"{prefix} request_id={request_id}" if prefix else f"request_id={request_id}"
        sheet.append([f"{prefix}: {message}" if prefix else message])
    apply_header_style(sheet)


def format_date(value: Any) -> str:
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    parsed = parse_date(value)
    return parsed.strftime("%d.%m.%Y") if parsed else normalize_text(value)


def format_datetime(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M:%S")
    return normalize_text(value)


def apply_header_style(sheet, rows: tuple[int, ...] = (1,)) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for row_number in rows:
        if row_number > sheet.max_row:
            continue
        for cell in sheet[row_number]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"


def apply_thin_border(sheet, range_ref: str) -> None:
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in sheet[range_ref]:
        for cell in row:
            cell.border = border


def autosize_columns(sheet) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        width = min(60, max(10, max(len(normalize_text(cell.value)) for cell in column) + 2))
        sheet.column_dimensions[letter].width = width


def apply_report_template_widths(workbook: Workbook) -> None:
    widths_by_sheet = {
        "Сводка": {"A": 28, "B": 13, "C": 10},
        "Заявки": {"A": 10, "B": 13, "C": 11, "D": 20, "E": 11, "F": 10, "G": 15, "H": 17, "I": 15, "J": 45, "K": 24, "L": 33, "M": 60, "N": 13, "O": 12, "P": 12, "Q": 12, "R": 10, "S": 24},
        "Товары заявок": {"A": 13, "B": 11, "C": 20, "D": 15, "E": 45, "F": 24, "G": 36, "H": 17, "I": 15, "J": 12, "K": 13, "L": 12, "M": 12},
    }
    for sheet_name, widths in widths_by_sheet.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
