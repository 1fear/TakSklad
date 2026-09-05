import re
from datetime import datetime
from io import BytesIO
from re import sub

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .client_points_service import client_point_delivery_slot_map, delivery_slot_for_order, point_key
from .logistics_manual_stops_service import manual_stop_rows
from .logistics_zone_service import (
    ZONE_CITY,
    ZONE_REGION,
    ZONE_UNASSIGNED,
    classify_order,
    load_region_index,
    parse_coordinates,
)
from .models import Order, OrderItem
from .orders_service import ApiError, STATUS_RETURNED
from .reports_service import parse_report_date
from .spreadsheet_safety import force_workbook_text_literals
from .telegram_output_contract import logistics_report_filename


LOGISTICS_HEADERS = [
    "Тип заказа",
    "Внешний ID",
    "Описание",
    "Имя клиента",
    "Телефон",
    "Email",
    "Заметки",
    "Широта (забор)",
    "Долгота (забор)",
    "Адрес забора",
    "Окно времени С (забор)",
    "Окно времени ПО (забор)",
    "Окно перерыва С (забор)",
    "Окно перерыва ПО (забор)",
    "Детали адреса забора",
    "Время обслуживания забора",
    "Широта (доставка)",
    "Долгота (доставка)",
    "Адрес доставки",
    "Окно времени С (доставка)",
    "Окно времени ПО (доставка)",
    "Окно перерыва С (доставка)",
    "Окно перерыва ПО (доставка)",
    "Детали адреса доставки",
    "Время обслуживания доставки",
    "Приоритет заказа",
    "Навыки",
    "Тег заказа",
    "Название товара",
    "Айди товара",
    "Количество товара",
    "Вес (кг)",
    "Объем (m3)",
    "Короба",
    "Цена товара",
]

LOGISTICS_COORDINATE_PROBLEM_HEADERS = [
    "Клиент",
    "Адрес",
    "Внешний ID",
    "Причина",
    "Товары",
    "Тип оплаты",
    "Дата отгрузки",
    "Складская заявка",
]

PICKUP_ADDRESS = "Самовывоз со склада"
LOGISTICS_DATETIME_FORMAT = "yyyy-mm-dd hh:mm"
LOGISTICS_TEMPLATE_COLUMN_WIDTHS = {
    "A": 14,
    "B": 18,
    "C": 22,
    "D": 30,
    "E": 18,
    "F": 28,
    "G": 30,
    "H": 14,
    "I": 14,
    "J": 30,
    "K": 22,
    "L": 22,
    "Q": 14,
    "R": 14,
    "S": 30,
    "T": 22,
    "U": 22,
    "Z": 16,
    "AA": 24,
    "AB": 14,
    "AC": 22,
    "AD": 18,
    "AE": 18,
    "AF": 14,
    "AG": 14,
    "AH": 10,
    "AI": 14,
}


def list_logistics_dates(db: Session):
    orders = db.execute(
        select(Order)
        .where(Order.order_date.is_not(None))
        .order_by(Order.order_date.asc())
    ).scalars().all()
    dates = []
    for order in orders:
        if not order.order_date or not is_logistics_candidate_order(order):
            continue
        value = order.order_date.isoformat()
        if value not in dates:
            dates.append(value)
    return dates


def build_logistics_reports(db: Session, shipment_date: str):
    """Split candidate orders into city and region reports in a single pass."""
    report_date = parse_report_date(shipment_date)
    orders = db.execute(
        select(Order)
        .options(selectinload(Order.items))
        .where(Order.order_date == report_date)
        .order_by(Order.client.asc(), Order.created_at.asc())
    ).scalars().all()
    candidate_orders = [order for order in orders if is_logistics_candidate_order(order)]

    region_index = load_region_index(db)
    zone_orders = {ZONE_CITY: [], ZONE_REGION: []}
    unassigned_orders = []
    # Страховка: пустой справочник означал бы, что вся область выпадает из
    # обоих отчётов. Тогда возвращаемся к прежнему поведению, один городской
    # файл, а о самой пустоте сообщаем отдельным алертом
    region_directory_empty = len(region_index) == 0
    # Ручные точки живут только в логистике, заказа за ними нет. Обе проверки
    # на пустоту считают и их, иначе день, набранный одними ручными точками,
    # отдавал бы 404 вместо файла
    zone_manual_stops = {ZONE_CITY: [], ZONE_REGION: []}
    for stop in manual_stop_rows(db, report_date, region_index, region_directory_empty):
        zone_manual_stops.setdefault(stop["zone"], []).append(stop)
    manual_stops_count = sum(len(stops) for stops in zone_manual_stops.values())
    if not orders and not manual_stops_count:
        raise ApiError(404, f"No orders for shipment date {report_date.isoformat()}")
    if not candidate_orders and not manual_stops_count:
        raise ApiError(404, f"No logistics delivery orders for shipment date {report_date.isoformat()}")
    for order in candidate_orders:
        if region_directory_empty:
            zone_orders[ZONE_CITY].append(order)
            continue
        zone = classify_order(
            order.client,
            (order.raw_payload or {}).get("coordinates"),
            region_index,
        )
        if zone == ZONE_UNASSIGNED:
            unassigned_orders.append(order)
        else:
            zone_orders[zone].append(order)

    reports = {
        ZONE_CITY: None,
        ZONE_REGION: None,
        ZONE_UNASSIGNED: unassigned_orders,
        "region_directory_empty": region_directory_empty,
        # Заказы по зонам для приписки к отчёту: заказы, а не товарные строки
        "order_counts": {
            ZONE_CITY: len(zone_orders[ZONE_CITY]),
            ZONE_REGION: len(zone_orders[ZONE_REGION]),
        },
    }
    for zone in (ZONE_CITY, ZONE_REGION):
        if zone_orders[zone] or zone_manual_stops[zone]:
            reports[zone] = build_zone_report_xlsx(
                db, report_date, zone, zone_orders[zone], zone_manual_stops[zone]
            )
    return reports


def build_logistics_report_xlsx(db: Session, shipment_date: str, zone: str):
    if zone not in (ZONE_CITY, ZONE_REGION):
        raise ApiError(422, f"Unsupported logistics zone: {zone}")
    reports = build_logistics_reports(db, shipment_date)
    report = reports.get(zone)
    if report is None:
        report_date = parse_report_date(shipment_date)
        raise ApiError(
            404,
            f"No {zone} logistics delivery orders for shipment date {report_date.isoformat()}",
        )
    return report


def release_read_transaction(db: Session) -> bool:
    """Закрывает транзакцию чтения перед долгой сборкой книги.

    Приложение ставит своим соединениям idle_in_transaction_session_timeout,
    а сборка XLSX это чистый CPU без запросов, поэтому открытая транзакция
    висит простаивающей и соединение обрывается. 18.08.2026 так упал автоотчёт
    логистики: файлы ушли в Telegram, а `commit` факта отправки уже не прошёл,
    и повтор по таймауту отправил город и область второй раз.

    Незакоммиченные изменения вызывающего не трогаем: для такой сессии
    закрытие пропускается, поведение остаётся прежним.
    """
    if db.new or db.dirty or db.deleted:
        return False
    previous_expire_on_commit = db.expire_on_commit
    # Загруженные заказы нужны сборке уже после commit, истечение атрибутов
    # вернуло бы ленивые запросы построчно
    db.expire_on_commit = False
    try:
        db.commit()
    finally:
        db.expire_on_commit = previous_expire_on_commit
    return True


def build_zone_report_xlsx(db: Session, report_date, zone: str, zone_orders, zone_manual_stops=()):
    delivery_orders = [order for order in zone_orders if is_logistics_delivery_order(order)]
    coordinate_problem_orders = [order for order in zone_orders if not is_logistics_delivery_order(order)]
    delivery_slots = client_point_delivery_slot_map(db, delivery_orders)
    release_read_transaction(db)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(LOGISTICS_HEADERS)
    apply_orders_template_style(sheet)

    # Шаблон платформы «Orders via Excel» (35 колонок) ждёт строку на товарную
    # позицию: «Внешний ID» собирает строки в заказ, а различает их «Айди товара»,
    # сквозной уникальный номер по файлу. Без него платформа склеивала строки одного
    # внешнего ID и теряла товары, с ним 04.09.2026 прошли все 528 строк без ошибок.
    # «Количество товара» платформа ждёт равным 1, блоки живут в «Короба»
    line_id = 0

    def append_row(row):
        nonlocal line_id
        line_id += 1
        set_cell(row, 30, line_id)
        set_cell(row, 31, 1)
        sheet.append(row)
        apply_orders_row_style(sheet, sheet.max_row)

    for stop_orders in group_delivery_stops(delivery_orders):
        # Один и тот же клиент по тем же координатам это одна остановка, даже когда
        # заказы приехали разными слотами: все её строки несут общий склеенный номер
        lead_order = stop_orders[0]
        coordinates = normalize_coordinates((lead_order.raw_payload or {}).get("coordinates"))
        latitude, longitude = split_coordinates(coordinates)
        delivery_from, delivery_to = delivery_slot_for_order(lead_order, delivery_slots)
        external_id = stop_external_id(stop_orders)
        for order in stop_orders:
            for item in sorted(order.items, key=lambda value: (value.product, str(value.id))):
                row = [""] * len(LOGISTICS_HEADERS)
                set_cell(row, 1, "delivery")
                set_cell(row, 2, external_id)
                set_cell(row, 4, lead_order.client)
                set_cell(row, 7, lead_order.representative or "")
                set_cell(row, 17, latitude)
                set_cell(row, 18, longitude)
                set_cell(row, 19, lead_order.address)
                set_cell(row, 20, delivery_window_datetime(report_date, delivery_from))
                set_cell(row, 21, delivery_window_datetime(report_date, delivery_to))
                set_cell(row, 29, item.product)
                set_cell(row, 32, 0)
                set_cell(row, 33, 0)
                set_cell(row, 34, item_quantity_blocks(item))
                append_row(row)

    for stop in zone_manual_stops:
        append_row(manual_stop_report_row(report_date, stop))

    if coordinate_problem_orders:
        problem_sheet = workbook.create_sheet("Требуют координаты")
        problem_sheet.append(LOGISTICS_COORDINATE_PROBLEM_HEADERS)
        apply_header_style(problem_sheet)
        for order in coordinate_problem_orders:
            problem_sheet.append([
                order.client,
                order.address,
                logistics_external_id(order),
                logistics_coordinate_problem_reason(order),
                order_product_summary(order),
                order.payment_type,
                report_date.strftime("%d.%m.%Y"),
                (order.raw_payload or {}).get("skladbot_request_number") or "",
            ])
        autosize_columns(problem_sheet)
    buffer = BytesIO()
    force_workbook_text_literals(workbook)
    workbook.save(buffer)
    return buffer.getvalue(), logistics_report_filename(report_date, zone)


def manual_stop_report_row(report_date, stop):
    """Строка ручной точки для маршрутного листа

    Внешний ID и название товара остаются пустыми: заказа за такой точкой нет,
    подставлять туда выдуманный номер или выдуманный товар нельзя, файл уходит
    клиенту. Блоки пишутся как есть, включая ноль: точка с нулём это заезд без
    груза, и водитель обязан её увидеть
    """
    latitude, longitude = split_coordinates(normalize_coordinates(stop["coordinates"]))
    row = [""] * len(LOGISTICS_HEADERS)
    set_cell(row, 1, "delivery")
    set_cell(row, 4, stop["client"])
    set_cell(row, 7, stop["representative"] or "")
    set_cell(row, 17, latitude)
    set_cell(row, 18, longitude)
    set_cell(row, 19, stop["address"])
    set_cell(row, 20, delivery_window_datetime(report_date, stop["delivery_from"]))
    set_cell(row, 21, delivery_window_datetime(report_date, stop["delivery_to"]))
    set_cell(row, 32, 0)
    set_cell(row, 33, 0)
    set_cell(row, 34, int(stop["blocks"] or 0))
    return row


def group_delivery_stops(delivery_orders):
    """Собрать заказы одного клиента по одним координатам в одну остановку.

    Ключ намеренно включает клиента: по одному адресу может стоять несколько
    юрлиц, у них разные получатели и своё окно доставки, склеивать их нельзя.
    Остановки идут в порядке первого появления, а заказы внутри остановки
    сортируются по времени создания с добором по id: без этого добора два
    заказа с одинаковым created_at давали бы разный склеенный внешний ID при
    каждой пересборке отчёта.
    """
    stops = {}
    order_of_keys = []
    for order in delivery_orders:
        coordinates = normalize_coordinates((order.raw_payload or {}).get("coordinates"))
        key = (coordinates, point_key(order.client))
        if key not in stops:
            stops[key] = []
            order_of_keys.append(key)
        stops[key].append(order)
    return [sorted(stops[key], key=stop_order_sort_key) for key in order_of_keys]


def stop_order_sort_key(order):
    created_at = order.created_at
    return (created_at is None, created_at.isoformat() if created_at else "", str(order.id))


def stop_external_id(stop_orders):
    """Внешний ID остановки: номера всех её заказов через плюс, без повторов."""
    identifiers = []
    for order in stop_orders:
        identifier = logistics_external_id(order)
        if identifier and identifier not in identifiers:
            identifiers.append(identifier)
    return "+".join(identifiers)


def set_cell(row, one_based_index, value):
    row[one_based_index - 1] = value


def normalize_coordinates(value):
    point = parse_coordinates(value)
    if point is None:
        return ""
    latitude, longitude = point
    return f"{format_coordinate(latitude)},{format_coordinate(longitude)}"


def format_coordinate(value):
    return f"{value:.12f}".rstrip("0").rstrip(".")


def is_logistics_delivery_order(order):
    if not is_logistics_candidate_order(order):
        return False
    return bool(normalize_coordinates((order.raw_payload or {}).get("coordinates")))


def is_logistics_candidate_order(order):
    if is_returned_order(order):
        return False
    if is_skladbot_stock_shortage_blocked_order(order):
        return False
    if is_pickup_address(order.address):
        return False
    return True


def is_returned_order(order):
    raw_payload = order.raw_payload or {}
    return (
        str(order.status or "").strip().casefold() == STATUS_RETURNED
        or str(raw_payload.get("return_status") or "").strip().casefold() in {"returned", "return", "возврат"}
    )


def logistics_coordinate_problem_reason(order):
    raw_coordinates = str((order.raw_payload or {}).get("coordinates") or "").strip()
    if raw_coordinates:
        return "Невалидные координаты"
    return "Нет координат"


def order_product_summary(order):
    parts = []
    for item in sorted(order.items, key=lambda value: (value.product, str(value.id))):
        quantity = item_quantity_blocks(item)
        suffix = f" - {quantity} блоков" if quantity else ""
        parts.append(f"{item.product}{suffix}")
    return "; ".join(parts)


def is_skladbot_stock_shortage_blocked_order(order):
    raw_payload = order.raw_payload or {}
    skladbot_status = str(raw_payload.get("skladbot_status") or "").strip()
    if skladbot_status == "cancelled_stock_shortage":
        return True
    if skladbot_status == "create_failed" and "автоотмена пропущена" in str(raw_payload.get("skladbot_error") or ""):
        return True
    if skladbot_status == "create_failed" and "недостат" in str(raw_payload.get("skladbot_error") or "").casefold():
        return True
    return False


def is_pickup_address(value):
    text = normalize_lookup_text(value)
    return text == normalize_lookup_text(PICKUP_ADDRESS) or text.startswith("самовывоз")


def normalize_lookup_text(value):
    text = str(value or "").strip().casefold().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", "", text)


def split_coordinates(value):
    parts = [part.strip() for part in str(value or "").split(",")]
    if len(parts) < 2:
        return "", ""
    return parts[0], parts[1]


def delivery_window_datetime(report_date, value):
    text = str(value or "").strip()
    match = re.fullmatch(r"(\d{1,2}):(\d{2})", text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    if hour > 23 or minute > 59:
        return None
    return datetime(report_date.year, report_date.month, report_date.day, hour, minute)


def logistics_external_id(order, item=None):
    raw_payload = order.raw_payload or {}
    if raw_payload.get("skladbot_request_number"):
        return raw_payload.get("skladbot_request_number")

    def public_source_id(value):
        text = str(value or "").strip()
        if text.casefold().startswith("smartup:"):
            return ""
        return text

    value = public_source_id(raw_payload.get("source_order_id"))
    if value:
        return value
    if item is not None:
        item_payload = item.raw_payload or {}
        for key in ("source_order_id", "source_import_id"):
            value = public_source_id(item_payload.get(key))
            if value:
                return value
    for order_item in sorted(order.items, key=lambda value: (value.product, str(value.id))):
        item_payload = order_item.raw_payload or {}
        for key in ("source_order_id", "source_import_id"):
            value = public_source_id(item_payload.get(key))
            if value:
                return value
    return ""


def item_quantity_blocks(item):
    if item.quantity_blocks and item.quantity_blocks > 0:
        return item.quantity_blocks
    pieces = item.quantity_pieces or 0
    pieces_per_block = item.pieces_per_block or 10
    if pieces <= 0:
        return 0
    return (pieces + pieces_per_block - 1) // pieces_per_block


def apply_header_style(sheet, *, freeze_panes=True):
    fill = PatternFill("solid", fgColor="1E293B")
    bottom_border = Border(bottom=Side(style="thin", color="000000"))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bottom_border
    if freeze_panes:
        sheet.freeze_panes = "A2"


def apply_orders_template_style(sheet):
    apply_header_style(sheet, freeze_panes=False)
    sheet.row_dimensions[1].height = 17.55
    for column_letter, width in LOGISTICS_TEMPLATE_COLUMN_WIDTHS.items():
        sheet.column_dimensions[column_letter].width = width


def apply_orders_row_style(sheet, row_number):
    for column_letter in ("T", "U"):
        sheet[f"{column_letter}{row_number}"].number_format = LOGISTICS_DATETIME_FORMAT


def autosize_columns(sheet):
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def safe_filename(value):
    return sub(r"[^0-9A-Za-zА-Яа-я_.-]+", "_", str(value or "")).strip("_")
