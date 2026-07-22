#!/usr/bin/env python3
"""Fail-closed repair of the exact Telegram logistics order set.

The tool never imports a workbook and never sends Telegram messages.  It may
only download the original completed Telegram documents, parse them locally,
and update the fixed 62 PostgreSQL orders in one transaction.  Stdout is
counts/hashes only; exact before-images stay in PostgreSQL audit rows and in an
optional mode-0600 preimage file used by the production workflow.
"""

from __future__ import annotations

import argparse
import hashlib
import importlib
import json
import os
import re
import tempfile
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo


TARGET_DATE = date(2026, 7, 23)
TARGET_REFS = tuple(
    [f"WH-R-{value}" for value in range(208826, 208850)]
    + [f"WH-R-{value}" for value in range(209244, 209282)]
)
TARGET_REF_SET = frozenset(TARGET_REFS)
EXPECTED_TARGET_COUNT = 62
EXPECTED_TARGET_ITEM_COUNT = 230
EXPECTED_SOURCE_FILE_COUNT = 2
EXPECTED_SCAN_COUNT = 560
APPROVAL = "REPAIR-62-LOGISTICS-2026-07-23"
APPLY_APPROVAL = APPROVAL
ROLLBACK_APPROVAL = APPROVAL
REQUIRED_GOOGLE_DECOMMISSION_MIGRATION = "20260716_0019"
REPAIR_ACTION = "telegram_logistics_order_repaired"
BATCH_ACTION = "telegram_logistics_orders_repair"
ROLLBACK_ACTION = "telegram_logistics_order_repair_rolled_back"
ROLLBACK_BATCH_ACTION = "telegram_logistics_orders_repair_rollback"
SMARTUP_LOGISTICS_EVENT = "smartup_logistics_report"
SMARTUP_LOGISTICS_AUDIT = "smartup_auto_import_logistics_report"
NAMESPACE = uuid.UUID("92becc99-77ba-5683-8890-3378740ba431")
ALLOWED_ORDER_STATUSES = frozenset({"not_completed", "completed", "done", "closed"})
HEX_SHA_RE = re.compile(r"[0-9a-f]{64}")


class RepairBlocked(RuntimeError):
    """A sanitized, fail-closed repair blocker."""

    def __init__(self, code: str):
        self.code = str(code)
        super().__init__(self.code)


@dataclass(frozen=True)
class ParsedSource:
    sha256: str
    rows: tuple[dict, ...]


@dataclass
class PlanEntry:
    order: object
    items: tuple[object, ...]
    ref: str
    preimage: dict
    applied: dict
    immutable_sha256: str
    source_import_ids: tuple[str, ...]
    source_event_ids: tuple[uuid.UUID, ...]


@dataclass
class RepairPlan:
    summary: dict
    entries: tuple[PlanEntry, ...]
    parsed_sources: dict[uuid.UUID, ParsedSource]
    source_contexts: dict[uuid.UUID, dict]
    related_sha256: str
    client_points_sha256: str
    source_provenance_sha256: str
    no_send_sha256: str
    lock_ids: dict[str, tuple]
    preimage: dict = field(default_factory=dict)


def backend_module(name):
    try:
        return importlib.import_module(f"app.{name}")
    except ModuleNotFoundError as exc:
        if exc.name not in {"app", f"app.{name}"}:
            raise
        return importlib.import_module(f"backend.app.{name}")


def normalize(value) -> str:
    return str(value or "").strip()


def canonical_value(value):
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value.isoformat(timespec="microseconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dict):
        return {
            str(key): canonical_value(item)
            for key, item in sorted(value.items(), key=lambda pair: str(pair[0]))
        }
    if isinstance(value, (list, tuple)):
        return [canonical_value(item) for item in value]
    if isinstance(value, (set, frozenset)):
        items = [canonical_value(item) for item in value]
        return sorted(items, key=canonical_json)
    if isinstance(value, bytes):
        return {"bytes_sha256": hashlib.sha256(value).hexdigest(), "bytes_length": len(value)}
    return value


def canonical_json(value) -> str:
    return json.dumps(
        canonical_value(value),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def payload_sha256(value) -> str:
    return hashlib.sha256(canonical_json(value).encode("utf-8")).hexdigest()


def deterministic_uuid(kind: str, *values) -> uuid.UUID:
    return uuid.uuid5(NAMESPACE, "|".join([kind, *(str(value) for value in values)]))


def valid_sha(value) -> bool:
    return bool(HEX_SHA_RE.fullmatch(normalize(value).lower()))


def parse_uuid(value, blocker: str) -> uuid.UUID:
    try:
        return uuid.UUID(normalize(value))
    except (ValueError, TypeError, AttributeError):
        raise RepairBlocked(blocker) from None


def model_payload(instance, *, exclude=()) -> dict:
    excluded = set(exclude)
    return {
        column.name: canonical_value(getattr(instance, column.name))
        for column in instance.__table__.columns
        if column.name not in excluded
    }


def sorted_model_payload(rows, *, exclude=()) -> list[dict]:
    return [
        model_payload(row, exclude=exclude)
        for row in sorted(rows, key=lambda value: str(getattr(value, "id", "")))
    ]


def order_state(order) -> dict:
    raw_payload = dict(order.raw_payload or {})
    return {
        "order_date": order.order_date.isoformat() if order.order_date else "",
        "address": normalize(order.address),
        "coordinates_present": "coordinates" in raw_payload,
        "coordinates": normalize(raw_payload.get("coordinates")),
    }


def order_immutable_sha256(order) -> str:
    raw_payload = dict(order.raw_payload or {})
    raw_payload.pop("coordinates", None)
    return payload_sha256({
        "id": str(order.id),
        "source": order.source,
        "external_id": order.external_id,
        "import_order_key": order.import_order_key,
        "import_source_order_key": order.import_source_order_key,
        "payment_type": order.payment_type,
        "client": order.client,
        "representative": order.representative,
        "status": order.status,
        "created_at": order.created_at,
        "raw_payload_without_coordinates": raw_payload,
    })


def target_ref(order) -> str:
    return normalize((order.raw_payload or {}).get("skladbot_request_number"))


def runtime_guards(db) -> None:
    from sqlalchemy import text

    if normalize(os.environ.get("TAKSKLAD_ENV")).casefold() != "production":
        raise RepairBlocked("PRODUCTION_RUNTIME_REQUIRED")
    try:
        versions = [normalize(value) for value in db.execute(text("SELECT version_num FROM alembic_version")).scalars()]
    except Exception:
        raise RepairBlocked("MIGRATION_UNVERIFIED") from None
    if len(versions) != 1 or not re.fullmatch(r"\d{8}_\d{4}", versions[0]):
        raise RepairBlocked("MIGRATION_UNVERIFIED")
    if versions[0] < REQUIRED_GOOGLE_DECOMMISSION_MIGRATION:
        raise RepairBlocked("GOOGLE_DECOMMISSION_UNVERIFIED")

    timezone_name = normalize(os.environ.get("TAKSKLAD_TIMEZONE")) or "Asia/Tashkent"
    try:
        local_zone = ZoneInfo(timezone_name)
    except Exception:
        raise RepairBlocked("REPAIR_TIMEZONE_INVALID") from None
    due_text = (
        normalize(os.environ.get("SMARTUP_AUTO_IMPORT_LOGISTICS_DUE_TIME"))
        or normalize(os.environ.get("SMARTUP_AUTO_IMPORT_FINAL_TIME"))
        or "17:50"
    )
    match = re.fullmatch(r"(\d{1,2}):(\d{2})", due_text)
    if not match or int(match.group(1)) > 23 or int(match.group(2)) > 59:
        raise RepairBlocked("LOGISTICS_DUE_TIME_INVALID")
    deadline = datetime.combine(
        TARGET_DATE - timedelta(days=1),
        time(int(match.group(1)), int(match.group(2))),
        tzinfo=local_zone,
    )
    if datetime.now(local_zone) >= deadline:
        raise RepairBlocked("LOGISTICS_DUE_WINDOW_CLOSED")


def load_target_orders(db):
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    Order = backend_module("models").Order
    orders = db.execute(
        select(Order)
        .options(selectinload(Order.items))
        .where(Order.raw_payload["skladbot_request_number"].as_string().in_(TARGET_REFS))
        .order_by(Order.id)
    ).scalars().all()
    refs = [target_ref(order) for order in orders]
    if len(orders) != EXPECTED_TARGET_COUNT or len(set(refs)) != EXPECTED_TARGET_COUNT:
        raise RepairBlocked("TARGET_SCOPE_MISMATCH")
    if set(refs) != TARGET_REF_SET:
        raise RepairBlocked("TARGET_SCOPE_MISMATCH")

    request_ids = []
    for order in orders:
        raw_payload = order.raw_payload if isinstance(order.raw_payload, dict) else None
        if raw_payload is None:
            raise RepairBlocked("ORDER_RAW_PAYLOAD_INVALID")
        request_id = normalize(raw_payload.get("skladbot_request_id"))
        if not request_id:
            raise RepairBlocked("SKLADBOT_LINK_INCOMPLETE")
        request_ids.append(request_id)
        if normalize(order.status).casefold() not in ALLOWED_ORDER_STATUSES:
            raise RepairBlocked("ORDER_STATUS_BLOCKED")
        if normalize(raw_payload.get("return_status")).casefold() in {"returned", "return", "возврат"}:
            raise RepairBlocked("ORDER_RETURNED_BLOCKED")
        skladbot_status = normalize(raw_payload.get("skladbot_status")).casefold()
        skladbot_error = normalize(raw_payload.get("skladbot_error")).casefold()
        if skladbot_status == "cancelled_stock_shortage" or (
            skladbot_status == "create_failed" and "недостат" in skladbot_error
        ):
            raise RepairBlocked("ORDER_STOCK_SHORTAGE_BLOCKED")
        if order.order_date == TARGET_DATE:
            raise RepairBlocked("TARGET_ALREADY_REPAIRED")
        if not order.items:
            raise RepairBlocked("TARGET_ITEMS_MISSING")
    if len(set(request_ids)) != EXPECTED_TARGET_COUNT:
        raise RepairBlocked("SKLADBOT_LINK_NOT_UNIQUE")
    return orders


def load_source_contexts(db, orders):
    from sqlalchemy import select

    models = backend_module("models")
    ImportFile, ImportJob, PendingEvent = models.ImportFile, models.ImportJob, models.PendingEvent
    imports_service = backend_module("imports_service")
    items = [item for order in orders for item in order.items]
    import_id_by_item = {}
    import_ids = set()
    for item in items:
        raw_payload = item.raw_payload if isinstance(item.raw_payload, dict) else {}
        source_import_id = normalize(item.source_import_id)
        if (
            not source_import_id
            or normalize(raw_payload.get("source_import_id")) != source_import_id
            or normalize(item.source_import_key)
            != normalize(imports_service.source_import_lookup_key(source_import_id))
        ):
            raise RepairBlocked("SOURCE_IDENTITY_MISMATCH")
        import_id = parse_uuid(raw_payload.get("backend_import_id"), "SOURCE_IMPORT_JOB_MISSING")
        import_id_by_item[item.id] = import_id
        import_ids.add(import_id)

    import_jobs = db.execute(
        select(ImportJob).where(ImportJob.id.in_(sorted(import_ids, key=str))).order_by(ImportJob.id)
    ).scalars().all()
    if len(import_jobs) != len(import_ids):
        raise RepairBlocked("SOURCE_IMPORT_JOB_MISSING")
    jobs_by_id = {job.id: job for job in import_jobs}
    event_ids = set()
    event_id_by_job = {}
    expected_sha_by_event = {}
    for job in import_jobs:
        raw_payload = job.raw_payload if isinstance(job.raw_payload, dict) else {}
        if normalize(job.source).casefold() != "telegram" or normalize(job.status) not in {
            "completed", "completed_with_errors"
        }:
            raise RepairBlocked("SOURCE_IMPORT_JOB_INVALID")
        event_id = parse_uuid(raw_payload.get("telegram_event_id"), "SOURCE_EVENT_MISSING")
        sha256 = normalize(raw_payload.get("sha256")).lower()
        if not valid_sha(sha256):
            raise RepairBlocked("SOURCE_SHA_INVALID")
        if event_id in expected_sha_by_event and expected_sha_by_event[event_id] != sha256:
            raise RepairBlocked("SOURCE_EVENT_SHA_CONFLICT")
        event_ids.add(event_id)
        event_id_by_job[job.id] = event_id
        expected_sha_by_event[event_id] = sha256

    events = db.execute(
        select(PendingEvent).where(PendingEvent.id.in_(sorted(event_ids, key=str))).order_by(PendingEvent.id)
    ).scalars().all()
    if len(events) != len(event_ids):
        raise RepairBlocked("SOURCE_EVENT_MISSING")
    events_by_id = {event.id: event for event in events}
    expected_shas = set(expected_sha_by_event.values())
    import_files = db.execute(
        select(ImportFile).where(ImportFile.sha256.in_(sorted(expected_shas))).order_by(ImportFile.id)
    ).scalars().all()
    files_by_sha = {normalize(row.sha256).lower(): row for row in import_files}
    if set(files_by_sha) != expected_shas or len(import_files) != len(expected_shas):
        raise RepairBlocked("SOURCE_FILE_METADATA_MISSING")

    contexts = {}
    for event_id in sorted(event_ids, key=str):
        event = events_by_id[event_id]
        payload = event.payload if isinstance(event.payload, dict) else {}
        document = payload.get("document") if isinstance(payload.get("document"), dict) else {}
        file_id = normalize(document.get("file_id"))
        file_name = normalize(document.get("file_name") or payload.get("file_name"))
        if event.event_type != "telegram_excel_import" or event.status != "completed":
            raise RepairBlocked("SOURCE_EVENT_INVALID")
        if not file_id or not file_name.lower().endswith((".xlsx", ".xlsm")):
            raise RepairBlocked("SOURCE_DOCUMENT_UNAVAILABLE")
        expected_sha = expected_sha_by_event[event_id]
        contexts[event_id] = {
            "event": event,
            "file_id": file_id,
            "file_name": file_name,
            "shipment_date": normalize(payload.get("shipment_date")),
            "expected_sha256": expected_sha,
            "import_file": files_by_sha[expected_sha],
            "job_ids": tuple(sorted(
                (job.id for job in import_jobs if event_id_by_job[job.id] == event_id),
                key=str,
            )),
        }

    for item in items:
        import_id = import_id_by_item[item.id]
        item._repair_source_event_id = event_id_by_job[jobs_by_id[import_id].id]
    return contexts, import_jobs, import_files


def download_and_parse_sources(contexts) -> dict[uuid.UUID, ParsedSource]:
    """Ingress-only adapter: Telegram GET + local parser, with no import/send call."""

    token = normalize(os.environ.get("TELEGRAM_BOT_TOKEN"))
    if not token:
        raise RepairBlocked("TELEGRAM_DOWNLOAD_NOT_CONFIGURED")
    clients = backend_module("telegram_clients")
    excel_importer = backend_module("excel_importer")
    ensure_coordinates_header_alias(excel_importer)
    telegram_common = backend_module("telegram_common")
    timeout = int(os.environ.get("TELEGRAM_WORKER_FILE_TIMEOUT_SECONDS", "120") or "120")
    max_size = int(os.environ.get("TELEGRAM_WORKER_MAX_FILE_BYTES", str(20 * 1024 * 1024)) or 0)
    api = clients.TelegramApiClient(token, file_timeout=timeout)
    parsed_sources = {}
    with tempfile.TemporaryDirectory(prefix="taksklad-logistics-repair-") as temp_dir:
        for event_id, context in sorted(contexts.items(), key=lambda pair: str(pair[0])):
            path = Path(temp_dir) / f"{event_id}.xlsx"
            api.download_file(context["file_id"], path, max_size)
            actual_sha = hashlib.sha256(path.read_bytes()).hexdigest()
            if actual_sha != context["expected_sha256"]:
                raise RepairBlocked("SOURCE_SHA_MISMATCH")
            shipment_date = context["shipment_date"]
            payload = excel_importer.excel_file_to_import_payload(
                path,
                file_name=context["file_name"],
                source="telegram",
                shipment_date=shipment_date,
                force_shipment_date=bool(telegram_common.parse_date_from_text(shipment_date)),
            )
            if normalize(payload.get("sha256")).lower() != actual_sha:
                raise RepairBlocked("SOURCE_PARSER_SHA_MISMATCH")
            parsed_sources[event_id] = ParsedSource(
                sha256=actual_sha,
                rows=tuple(row for row in (payload.get("rows") or ()) if isinstance(row, dict)),
            )
    return parsed_sources


def ensure_coordinates_header_alias(excel_importer) -> None:
    """Apply the one required parser alias in memory without changing source bytes."""

    alias = "GPS-координаты клиента"
    mapping_names = [
        name
        for name in ("HEADER_ALIASES", "OPTIONAL_ALIASES")
        if hasattr(excel_importer, name)
    ]
    if not mapping_names:
        raise RepairBlocked("SOURCE_PARSER_ALIAS_STRUCTURE_INVALID")
    for mapping_name in mapping_names:
        aliases_by_field = getattr(excel_importer, mapping_name)
        coordinates_aliases = (
            aliases_by_field.get("coordinates")
            if isinstance(aliases_by_field, dict)
            else None
        )
        if (
            not isinstance(coordinates_aliases, list)
            or any(not isinstance(value, str) for value in coordinates_aliases)
        ):
            raise RepairBlocked("SOURCE_PARSER_ALIAS_STRUCTURE_INVALID")
        if alias not in coordinates_aliases:
            coordinates_aliases.append(alias)


def target_scan_counts(db, orders) -> tuple[int, int]:
    from sqlalchemy import or_, select

    models = backend_module("models")
    KizCode, KizMovement, ScanCode = models.KizCode, models.KizMovement, models.ScanCode
    items = [item for order in orders for item in order.items]
    try:
        scanned_blocks = sum(int(item.scanned_blocks or 0) for item in items)
    except (TypeError, ValueError):
        raise RepairBlocked("TARGET_SCAN_SCOPE_MISMATCH") from None
    item_ids = tuple(sorted((item.id for item in items), key=str))
    scan_rows = db.execute(
        select(ScanCode.id, ScanCode.code)
        .where(ScanCode.order_item_id.in_(item_ids))
        .order_by(ScanCode.id)
    ).all()
    scan_ids = tuple(row.id for row in scan_rows)
    normalized_codes = [normalize(row.code) for row in scan_rows]
    unique_scan_codes = set(normalized_codes)
    order_ids = tuple(sorted((order.id for order in orders), key=str))
    movement_kiz_ids = set(db.execute(
        select(KizMovement.kiz_id)
        .where(or_(
            KizMovement.order_id.in_(order_ids),
            KizMovement.order_item_id.in_(item_ids),
            KizMovement.scan_code_id.in_(scan_ids),
        ))
        .order_by(KizMovement.id)
    ).scalars().all())
    kiz_codes = db.execute(
        select(KizCode.code)
        .where(KizCode.id.in_(tuple(sorted(movement_kiz_ids, key=str))))
        .order_by(KizCode.id)
    ).scalars().all() if movement_kiz_ids else []
    unique_kiz_codes = {normalize(code) for code in kiz_codes}
    if (
        scanned_blocks != EXPECTED_SCAN_COUNT
        or len(scan_rows) != EXPECTED_SCAN_COUNT
        or "" in unique_scan_codes
        or len(unique_scan_codes) != EXPECTED_SCAN_COUNT
        or "" in unique_kiz_codes
        or len(unique_kiz_codes) != EXPECTED_SCAN_COUNT
        or unique_kiz_codes != unique_scan_codes
    ):
        raise RepairBlocked("TARGET_SCAN_SCOPE_MISMATCH")
    return len(scan_rows), len(unique_kiz_codes)


def parsed_rows_by_event(contexts, parsed_sources):
    imports_service = backend_module("imports_service")
    result = {}
    if set(parsed_sources) != set(contexts):
        raise RepairBlocked("SOURCE_DOCUMENT_SCOPE_MISMATCH")
    for event_id, context in contexts.items():
        parsed = parsed_sources[event_id]
        actual_sha = normalize(parsed.sha256).lower()
        if actual_sha != context["expected_sha256"]:
            raise RepairBlocked("SOURCE_SHA_MISMATCH")
        rows = {}
        for raw_row in parsed.rows:
            try:
                normalized = imports_service.normalize_import_row(raw_row)
            except Exception:
                raise RepairBlocked("SOURCE_ROW_INVALID") from None
            source_import_id = normalize(normalized.get("source_import_id"))
            if not source_import_id or source_import_id in rows:
                raise RepairBlocked("SOURCE_ROW_ID_NOT_UNIQUE")
            rows[source_import_id] = normalized
        result[event_id] = rows
    return result


def no_send_snapshot(db) -> tuple[str, dict, tuple]:
    from sqlalchemy import select

    models = backend_module("models")
    AuditLog, PendingEvent = models.AuditLog, models.PendingEvent
    events = db.execute(
        select(PendingEvent)
        .where(PendingEvent.event_type == SMARTUP_LOGISTICS_EVENT)
        .where(PendingEvent.payload["delivery_date"].as_string() == TARGET_DATE.isoformat())
        .order_by(PendingEvent.id)
    ).scalars().all()
    audits = db.execute(
        select(AuditLog)
        .where(AuditLog.action == SMARTUP_LOGISTICS_AUDIT)
        .where(AuditLog.entity_id == TARGET_DATE.isoformat())
        .order_by(AuditLog.id)
    ).scalars().all()
    payload = {
        "events": sorted_model_payload(events),
        "audits": sorted_model_payload(audits),
    }
    return payload_sha256(payload), {"events": len(events), "audits": len(audits)}, tuple(event.id for event in events)


def target_related_snapshot(db, orders) -> tuple[str, str, dict[str, tuple]]:
    from sqlalchemy import or_, select

    models = backend_module("models")
    ClientPoint = models.ClientPoint
    KizCode, KizMovement = models.KizCode, models.KizMovement
    OrderItem, PendingEvent, ScanCode = models.OrderItem, models.PendingEvent, models.ScanCode
    SmartupFulfillmentOrder = models.SmartupFulfillmentOrder
    client_points_service = backend_module("client_points_service")

    order_ids = tuple(sorted((order.id for order in orders), key=str))
    items = db.execute(
        select(OrderItem).where(OrderItem.order_id.in_(order_ids)).order_by(OrderItem.id)
    ).scalars().all()
    item_ids = tuple(sorted((item.id for item in items), key=str))
    scans = db.execute(
        select(ScanCode).where(ScanCode.order_item_id.in_(item_ids)).order_by(ScanCode.id)
    ).scalars().all() if item_ids else []
    movements = db.execute(
        select(KizMovement)
        .where(or_(KizMovement.order_id.in_(order_ids), KizMovement.order_item_id.in_(item_ids)))
        .order_by(KizMovement.id)
    ).scalars().all() if item_ids else []
    kiz_ids = tuple(sorted({movement.kiz_id for movement in movements}, key=str))
    kiz_codes = db.execute(
        select(KizCode).where(KizCode.id.in_(kiz_ids)).order_by(KizCode.id)
    ).scalars().all() if kiz_ids else []
    fulfillment = db.execute(
        select(SmartupFulfillmentOrder)
        .where(SmartupFulfillmentOrder.order_id.in_(order_ids))
        .order_by(SmartupFulfillmentOrder.id)
    ).scalars().all()
    explicit_event_ids = {
        parse_uuid((order.raw_payload or {}).get("skladbot_create_event_id"), "SKLADBOT_EVENT_ID_INVALID")
        for order in orders
        if normalize((order.raw_payload or {}).get("skladbot_create_event_id"))
    }
    explicit_event_ids.update(link.skladbot_event_id for link in fulfillment if link.skladbot_event_id)
    order_id_text = tuple(str(value) for value in order_ids)
    event_filter = (
        (PendingEvent.aggregate_type == "order") & PendingEvent.aggregate_id.in_(order_id_text)
    )
    if explicit_event_ids:
        event_filter = or_(event_filter, PendingEvent.id.in_(tuple(explicit_event_ids)))
    events = db.execute(select(PendingEvent).where(event_filter).order_by(PendingEvent.id)).scalars().all()
    related = {
        "items": sorted_model_payload(items),
        "scans": sorted_model_payload(scans),
        "kiz_codes": sorted_model_payload(kiz_codes),
        "kiz_movements": sorted_model_payload(movements),
        "pending_events": sorted_model_payload(events),
        "fulfillment_orders": sorted_model_payload(fulfillment),
    }

    client_keys = tuple(sorted({client_points_service.point_key(order.client) for order in orders}))
    client_points = db.execute(
        select(ClientPoint).where(ClientPoint.normalized_client.in_(client_keys)).order_by(ClientPoint.id)
    ).scalars().all() if client_keys else []
    lock_ids = {
        "orders": order_ids,
        "items": item_ids,
        "scans": tuple(scan.id for scan in scans),
        "kiz_codes": tuple(row.id for row in kiz_codes),
        "kiz_movements": tuple(row.id for row in movements),
        "pending_events": tuple(row.id for row in events),
        "fulfillment_orders": tuple(row.id for row in fulfillment),
        "client_points": tuple(row.id for row in client_points),
    }
    return payload_sha256(related), payload_sha256(sorted_model_payload(client_points)), lock_ids


def source_provenance_snapshot(contexts, import_jobs, import_files) -> tuple[str, dict[str, tuple]]:
    events = [context["event"] for context in contexts.values()]
    payload = {
        "events": sorted_model_payload(events),
        "imports": sorted_model_payload(import_jobs),
        "files": sorted_model_payload(import_files),
        "document_file_id_sha256": sorted(
            hashlib.sha256(context["file_id"].encode("utf-8")).hexdigest()
            for context in contexts.values()
        ),
    }
    return payload_sha256(payload), {
        "source_events": tuple(sorted((row.id for row in events), key=str)),
        "source_imports": tuple(sorted((row.id for row in import_jobs), key=str)),
        "source_files": tuple(sorted((row.id for row in import_files), key=str)),
    }


def existing_batch_audits(db, action=BATCH_ACTION):
    from sqlalchemy import select

    AuditLog = backend_module("models").AuditLog
    return db.execute(
        select(AuditLog).where(AuditLog.action == action).order_by(AuditLog.id)
    ).scalars().all()


def create_plan(db, *, parsed_sources=None, enforce_runtime_guards=True) -> RepairPlan:
    if len(TARGET_REFS) != EXPECTED_TARGET_COUNT or len(TARGET_REF_SET) != EXPECTED_TARGET_COUNT:
        raise RepairBlocked("COMPILED_TARGET_SCOPE_INVALID")
    if enforce_runtime_guards:
        runtime_guards(db)
    if existing_batch_audits(db, BATCH_ACTION) or existing_batch_audits(db, ROLLBACK_BATCH_ACTION):
        raise RepairBlocked("REPAIR_ALREADY_RECORDED")

    orders = load_target_orders(db)
    scan_count, unique_kiz_count = target_scan_counts(db, orders)
    contexts, import_jobs, import_files = load_source_contexts(db, orders)
    if (
        len(contexts) != EXPECTED_SOURCE_FILE_COUNT
        or len(import_jobs) != EXPECTED_SOURCE_FILE_COUNT
        or len(import_files) != EXPECTED_SOURCE_FILE_COUNT
    ):
        raise RepairBlocked("SOURCE_DOCUMENT_SCOPE_MISMATCH")
    parsed_sources = dict(parsed_sources or download_and_parse_sources(contexts))
    if len(parsed_sources) != EXPECTED_SOURCE_FILE_COUNT:
        raise RepairBlocked("SOURCE_DOCUMENT_SCOPE_MISMATCH")
    parsed_by_event = parsed_rows_by_event(contexts, parsed_sources)
    imports_service = backend_module("imports_service")
    logistics_service = backend_module("logistics_service")

    entries = []
    seen_source_ids = set()
    for order in sorted(orders, key=lambda value: target_ref(value)):
        proposed = set()
        source_ids = []
        source_events = set()
        for item in sorted(order.items, key=lambda value: str(value.id)):
            source_import_id = normalize(item.source_import_id)
            if source_import_id in seen_source_ids:
                raise RepairBlocked("SOURCE_IDENTITY_NOT_UNIQUE")
            seen_source_ids.add(source_import_id)
            event_id = getattr(item, "_repair_source_event_id", None)
            row = (parsed_by_event.get(event_id) or {}).get(source_import_id)
            if row is None:
                raise RepairBlocked("SOURCE_ROW_MISSING")
            if (
                normalize(row.get("product")) != normalize(item.product)
                or int(row.get("quantity_pieces") or 0) != int(item.quantity_pieces or 0)
                or int(row.get("quantity_blocks") or 0) != int(item.quantity_blocks or 0)
                or normalize(row.get("client")) != normalize(order.client)
                or normalize(row.get("payment_type")) != normalize(order.payment_type)
                or normalize(row.get("representative")) != normalize(order.representative)
            ):
                raise RepairBlocked("SOURCE_ROW_MISMATCH")
            row_number = normalize(row.get("skladbot_request_number"))
            row_request_id = normalize(row.get("skladbot_request_id"))
            raw_payload = order.raw_payload or {}
            if row_number and row_number != target_ref(order):
                raise RepairBlocked("SOURCE_SKLADBOT_LINK_MISMATCH")
            if row_request_id and row_request_id != normalize(raw_payload.get("skladbot_request_id")):
                raise RepairBlocked("SOURCE_SKLADBOT_LINK_MISMATCH")
            address = normalize(row.get("address"))
            coordinates = logistics_service.normalize_coordinates(row.get("coordinates"))
            if not imports_service.is_real_address(address) or logistics_service.is_pickup_address(address):
                raise RepairBlocked("SOURCE_ADDRESS_INVALID")
            if not coordinates or not logistics_service.normalize_coordinates(coordinates):
                raise RepairBlocked("SOURCE_COORDINATES_INVALID")
            proposed.add((address, coordinates))
            source_ids.append(source_import_id)
            source_events.add(event_id)
        if len(proposed) != 1:
            raise RepairBlocked("SOURCE_ORDER_LOCATION_AMBIGUOUS")
        address, coordinates = proposed.pop()
        entries.append(PlanEntry(
            order=order,
            items=tuple(sorted(order.items, key=lambda value: str(value.id))),
            ref=target_ref(order),
            preimage=order_state(order),
            applied={
                "order_date": TARGET_DATE.isoformat(),
                "address": address,
                "coordinates_present": True,
                "coordinates": coordinates,
            },
            immutable_sha256=order_immutable_sha256(order),
            source_import_ids=tuple(sorted(source_ids)),
            source_event_ids=tuple(sorted(source_events, key=str)),
        ))

    target_item_count = sum(len(row.items) for row in entries)
    if (
        len(entries) != EXPECTED_TARGET_COUNT
        or target_item_count != EXPECTED_TARGET_ITEM_COUNT
        or len(seen_source_ids) != target_item_count
    ):
        raise RepairBlocked("TARGET_ITEM_SCOPE_MISMATCH")
    related_sha, client_points_sha, lock_ids = target_related_snapshot(db, orders)
    pending_event_ids = tuple(lock_ids.get("pending_events") or ())
    if pending_event_ids:
        from sqlalchemy import select

        PendingEvent = backend_module("models").PendingEvent
        linked_events = db.execute(
            select(PendingEvent).where(PendingEvent.id.in_(pending_event_ids))
        ).scalars().all()
        if any(
            event.event_type == "skladbot_request_create"
            and normalize(event.status).casefold() not in {"completed", "cancelled", "dead"}
            for event in linked_events
        ):
            raise RepairBlocked("SKLADBOT_CREATE_EVENT_ACTIVE")
    source_sha, source_locks = source_provenance_snapshot(contexts, import_jobs, import_files)
    lock_ids.update(source_locks)
    no_send_sha, no_send_counts, no_send_event_ids = no_send_snapshot(db)
    lock_ids["no_send_events"] = no_send_event_ids
    if no_send_counts != {"events": 0, "audits": 0}:
        raise RepairBlocked("NO_SEND_STATE_NOT_EMPTY")

    plan_payload = {
        "schema_version": 1,
        "repair": "telegram_logistics_orders_2026_07_23_v1",
        "target_date": TARGET_DATE,
        "target_refs": TARGET_REFS,
        "entries": [{
            "order_id": str(entry.order.id),
            "ref": entry.ref,
            "preimage": entry.preimage,
            "applied": entry.applied,
            "immutable_sha256": entry.immutable_sha256,
            "source_import_ids": entry.source_import_ids,
            "source_event_ids": entry.source_event_ids,
        } for entry in entries],
        "source_documents": [{
            "event_id": str(event_id),
            "sha256": source.sha256,
            "row_ids_sha256": payload_sha256(sorted(
                normalize(backend_module("imports_service").normalize_import_row(row).get("source_import_id"))
                for row in source.rows
            )),
        } for event_id, source in sorted(parsed_sources.items(), key=lambda pair: str(pair[0]))],
        "related_sha256": related_sha,
        "client_points_sha256": client_points_sha,
        "source_provenance_sha256": source_sha,
        "no_send_sha256": no_send_sha,
    }
    plan_sha = payload_sha256(plan_payload)
    preimage = {
        "schema_version": 1,
        "plan_sha256": plan_sha,
        "target_date": TARGET_DATE.isoformat(),
        "target_item_count": EXPECTED_TARGET_ITEM_COUNT,
        "related_sha256": related_sha,
        "client_points_sha256": client_points_sha,
        "source_provenance_sha256": source_sha,
        "no_send_sha256": no_send_sha,
        "entries": [{
            "order_id": str(entry.order.id),
            "target_ref_sha256": hashlib.sha256(entry.ref.encode("utf-8")).hexdigest(),
            "preimage": entry.preimage,
            "applied": entry.applied,
            "immutable_sha256": entry.immutable_sha256,
        } for entry in entries],
    }
    preimage_sha = payload_sha256(preimage)
    summary = {
        "schema_version": 1,
        "mode": "plan_counts_only",
        "target_count": len(entries),
        "target_items": target_item_count,
        "source_files": len(parsed_sources),
        "scan_count": scan_count,
        "unique_kiz_count": unique_kiz_count,
        "safe_to_repair": True,
        "conflicts": 0,
        "plan_sha256": plan_sha,
        "preimage_sha256": preimage_sha,
        "mutations_expected": len(entries),
    }
    return RepairPlan(
        summary=summary,
        entries=tuple(entries),
        parsed_sources=parsed_sources,
        source_contexts=contexts,
        related_sha256=related_sha,
        client_points_sha256=client_points_sha,
        source_provenance_sha256=source_sha,
        no_send_sha256=no_send_sha,
        lock_ids={key: tuple(value) for key, value in lock_ids.items()},
        preimage=preimage,
    )


def lock_plan_rows(db, plan) -> None:
    from sqlalchemy import select

    models = backend_module("models")
    mapping = {
        "orders": models.Order,
        "items": models.OrderItem,
        "scans": models.ScanCode,
        "kiz_codes": models.KizCode,
        "kiz_movements": models.KizMovement,
        "pending_events": models.PendingEvent,
        "fulfillment_orders": models.SmartupFulfillmentOrder,
        "client_points": models.ClientPoint,
        "source_events": models.PendingEvent,
        "source_imports": models.ImportJob,
        "source_files": models.ImportFile,
        "no_send_events": models.PendingEvent,
    }
    for key, model in mapping.items():
        values = tuple(plan.lock_ids.get(key) or ())
        if values:
            db.execute(select(model.id).where(model.id.in_(values)).order_by(model.id).with_for_update()).all()


def write_preimage_file(plan, path) -> str:
    destination = Path(path)
    if not destination.is_absolute():
        raise RepairBlocked("PREIMAGE_PATH_NOT_ABSOLUTE")
    data = canonical_json(plan.preimage).encode("utf-8")
    digest = hashlib.sha256(data).hexdigest()
    if digest != plan.summary["preimage_sha256"]:
        raise RepairBlocked("PREIMAGE_HASH_INTERNAL_MISMATCH")
    fd = os.open(destination, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
    except Exception:
        try:
            destination.unlink()
        except OSError:
            pass
        raise
    if (destination.stat().st_mode & 0o777) != 0o600:
        raise RepairBlocked("PREIMAGE_MODE_INVALID")
    return digest


def apply_state(order, state) -> None:
    order.order_date = date.fromisoformat(state["order_date"]) if state.get("order_date") else None
    order.address = normalize(state.get("address"))
    raw_payload = dict(order.raw_payload or {})
    if state.get("coordinates_present"):
        raw_payload["coordinates"] = normalize(state.get("coordinates"))
    else:
        raw_payload.pop("coordinates", None)
    order.raw_payload = raw_payload
    order.updated_at = datetime.now(timezone.utc)


def apply_plan(db, plan, *, preimage_out="") -> dict:
    models = backend_module("models")
    AuditLog = models.AuditLog
    if existing_batch_audits(db, BATCH_ACTION):
        raise RepairBlocked("REPAIR_ALREADY_APPLIED")
    lock_plan_rows(db, plan)
    db.expire_all()
    locked = create_plan(db, parsed_sources=plan.parsed_sources, enforce_runtime_guards=False)
    if locked.summary["plan_sha256"] != plan.summary["plan_sha256"]:
        raise RepairBlocked("PLAN_CHANGED_UNDER_LOCK")
    if preimage_out:
        write_preimage_file(locked, preimage_out)

    now = datetime.now(timezone.utc)
    for entry in locked.entries:
        apply_state(entry.order, entry.applied)
        audit_id = deterministic_uuid("target-audit", entry.order.id, locked.summary["plan_sha256"])
        if db.get(AuditLog, audit_id) is not None:
            raise RepairBlocked("REPAIR_ALREADY_APPLIED")
        db.add(AuditLog(
            id=audit_id,
            action=REPAIR_ACTION,
            entity_type="order",
            entity_id=str(entry.order.id),
            payload={
                "plan_sha256": locked.summary["plan_sha256"],
                "preimage_sha256": locked.summary["preimage_sha256"],
                "order_id": str(entry.order.id),
                "target_ref_sha256": hashlib.sha256(entry.ref.encode("utf-8")).hexdigest(),
                "preimage": entry.preimage,
                "applied": entry.applied,
                "immutable_sha256": entry.immutable_sha256,
                "repaired_at": now.isoformat(),
                "internal_before_image": True,
            },
        ))
    batch_id = deterministic_uuid("batch-audit", locked.summary["plan_sha256"])
    if db.get(AuditLog, batch_id) is not None:
        raise RepairBlocked("REPAIR_ALREADY_APPLIED")
    db.add(AuditLog(
        id=batch_id,
        action=BATCH_ACTION,
        entity_type="order_batch",
        entity_id="telegram_logistics_orders_2026_07_23_v1",
        payload={
            "plan_sha256": locked.summary["plan_sha256"],
            "preimage_sha256": locked.summary["preimage_sha256"],
            "target_count": EXPECTED_TARGET_COUNT,
            "target_item_count": EXPECTED_TARGET_ITEM_COUNT,
            "target_date": TARGET_DATE.isoformat(),
            "related_sha256": locked.related_sha256,
            "client_points_sha256": locked.client_points_sha256,
            "source_provenance_sha256": locked.source_provenance_sha256,
            "no_send_sha256": locked.no_send_sha256,
            "source_event_ids": [str(value) for value in locked.lock_ids.get("source_events", ())],
            "source_import_ids": [str(value) for value in locked.lock_ids.get("source_imports", ())],
            "source_file_ids": [str(value) for value in locked.lock_ids.get("source_files", ())],
            "values_redacted": True,
            "repaired_at": now.isoformat(),
        },
    ))
    db.flush()
    orders = [entry.order for entry in locked.entries]
    if any(order_state(entry.order) != entry.applied for entry in locked.entries):
        raise RepairBlocked("POST_FLUSH_ORDER_INVARIANT_FAILED")
    if any(order_immutable_sha256(entry.order) != entry.immutable_sha256 for entry in locked.entries):
        raise RepairBlocked("POST_FLUSH_IMMUTABLE_INVARIANT_FAILED")
    related_sha, client_sha, _locks = target_related_snapshot(db, orders)
    no_send_sha, _counts, _events = no_send_snapshot(db)
    if (
        related_sha != locked.related_sha256
        or client_sha != locked.client_points_sha256
        or no_send_sha != locked.no_send_sha256
    ):
        raise RepairBlocked("POST_FLUSH_RELATED_INVARIANT_FAILED")
    db.commit()
    return {
        "schema_version": 1,
        "mode": "apply_counts_only",
        "target_count": EXPECTED_TARGET_COUNT,
        "safe_to_repair": True,
        "conflicts": 0,
        "plan_sha256": locked.summary["plan_sha256"],
        "preimage_sha256": locked.summary["preimage_sha256"],
        "mutations_applied": EXPECTED_TARGET_COUNT,
        "target_audits": EXPECTED_TARGET_COUNT,
        "batch_audits": 1,
    }


def repair_audits(db, plan_sha, *, for_update=False):
    from sqlalchemy import select

    AuditLog = backend_module("models").AuditLog
    target_statement = (
        select(AuditLog)
        .where(AuditLog.action == REPAIR_ACTION)
        .where(AuditLog.payload["plan_sha256"].as_string() == plan_sha)
        .order_by(AuditLog.id)
    )
    batch_statement = (
        select(AuditLog)
        .where(AuditLog.action == BATCH_ACTION)
        .where(AuditLog.payload["plan_sha256"].as_string() == plan_sha)
        .order_by(AuditLog.id)
    )
    if for_update:
        target_statement = target_statement.with_for_update()
        batch_statement = batch_statement.with_for_update()
    target_rows = db.execute(
        target_statement.execution_options(populate_existing=True)
    ).scalars().all()
    batch_rows = db.execute(
        batch_statement.execution_options(populate_existing=True)
    ).scalars().all()
    return target_rows, batch_rows


def source_provenance_from_batch(db, batch_payload):
    from sqlalchemy import select

    models = backend_module("models")
    event_ids = tuple(
        parse_uuid(value, "AUDIT_SOURCE_ID_INVALID")
        for value in batch_payload.get("source_event_ids") or ()
    )
    import_ids = tuple(
        parse_uuid(value, "AUDIT_SOURCE_ID_INVALID")
        for value in batch_payload.get("source_import_ids") or ()
    )
    file_ids = tuple(
        parse_uuid(value, "AUDIT_SOURCE_ID_INVALID")
        for value in batch_payload.get("source_file_ids") or ()
    )
    events = db.execute(select(models.PendingEvent).where(models.PendingEvent.id.in_(event_ids))).scalars().all()
    imports = db.execute(select(models.ImportJob).where(models.ImportJob.id.in_(import_ids))).scalars().all()
    files = db.execute(select(models.ImportFile).where(models.ImportFile.id.in_(file_ids))).scalars().all()
    if len(events) != len(event_ids) or len(imports) != len(import_ids) or len(files) != len(file_ids):
        return ""
    contexts = {}
    jobs_by_event = {}
    for job in imports:
        raw = job.raw_payload if isinstance(job.raw_payload, dict) else {}
        event_id = parse_uuid(raw.get("telegram_event_id"), "AUDIT_SOURCE_ID_INVALID")
        jobs_by_event.setdefault(event_id, []).append(job.id)
    files_by_sha = {normalize(row.sha256).lower(): row for row in files}
    for event in events:
        payload = event.payload if isinstance(event.payload, dict) else {}
        document = payload.get("document") if isinstance(payload.get("document"), dict) else {}
        matching_jobs = [
            job for job in imports
            if parse_uuid(
                (job.raw_payload or {}).get("telegram_event_id"),
                "AUDIT_SOURCE_ID_INVALID",
            ) == event.id
        ]
        if not matching_jobs:
            return ""
        sha = normalize((matching_jobs[0].raw_payload or {}).get("sha256")).lower()
        if sha not in files_by_sha:
            return ""
        contexts[event.id] = {
            "event": event,
            "file_id": normalize(document.get("file_id")),
            "import_file": files_by_sha[sha],
            "job_ids": tuple(sorted(jobs_by_event.get(event.id, ()), key=str)),
        }
    digest, _locks = source_provenance_snapshot(contexts, imports, files)
    return digest


def verify_report(db, orders, audits_by_order):
    from openpyxl import load_workbook

    logistics = backend_module("logistics_service")
    content, _filename = logistics.build_logistics_report_xlsx(db, TARGET_DATE.isoformat())
    workbook = load_workbook(BytesIO(content), data_only=False, read_only=True)
    try:
        if "Orders" not in workbook.sheetnames:
            return 0, EXPECTED_TARGET_COUNT, 1
        sheet = workbook["Orders"]
        rows_by_ref = {ref: [] for ref in TARGET_REFS}
        date_conflicts = 0
        value_conflicts = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ref = normalize(row[1] if len(row) > 1 else "")
            if ref not in rows_by_ref:
                continue
            rows_by_ref[ref].append(row)
            order = next(value for value in orders if target_ref(value) == ref)
            audit_payload = audits_by_order[str(order.id)].payload
            applied = audit_payload["applied"]
            coordinates = logistics.normalize_coordinates(applied["coordinates"])
            expected_latitude, expected_longitude = logistics.split_coordinates(coordinates)
            if (
                normalize(row[16]) != expected_latitude
                or normalize(row[17]) != expected_longitude
                or normalize(row[18]) != normalize(applied["address"])
            ):
                value_conflicts += 1
            for cell in (row[19], row[20]):
                if isinstance(cell, datetime) and cell.date() != TARGET_DATE:
                    date_conflicts += 1
        problem_rows = 0
        if "Требуют координаты" in workbook.sheetnames:
            problem = workbook["Требуют координаты"]
            for row in problem.iter_rows(min_row=2, values_only=True):
                if normalize(row[2] if len(row) > 2 else "") in TARGET_REF_SET or normalize(
                    row[7] if len(row) > 7 else ""
                ) in TARGET_REF_SET:
                    problem_rows += 1
        report_rows = sum(len(values) for values in rows_by_ref.values())
        for order in orders:
            ref = target_ref(order)
            rows = rows_by_ref[ref]
            expected_boxes = sum(logistics.item_quantity_blocks(item) for item in order.items)
            actual_boxes = sum(int(row[30] or 0) for row in rows)
            if len(rows) != len(order.items) or actual_boxes != expected_boxes:
                value_conflicts += 1
        return report_rows, problem_rows, value_conflicts + date_conflicts
    finally:
        workbook.close()


def verify_applied(db, expected_plan_sha, *, no_send=False) -> dict:
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    if not valid_sha(expected_plan_sha):
        raise RepairBlocked("EXPECTED_PLAN_SHA_INVALID")
    if not no_send:
        raise RepairBlocked("NO_SEND_ACK_REQUIRED")
    models = backend_module("models")
    target_audits, batch_audits = repair_audits(db, expected_plan_sha)
    conflicts = 0
    if len(target_audits) != EXPECTED_TARGET_COUNT or len(batch_audits) != 1:
        conflicts += 1
    batch = batch_audits[0].payload if len(batch_audits) == 1 else {}
    order_ids = []
    for audit in target_audits:
        try:
            order_ids.append(uuid.UUID(str(audit.entity_id)))
        except ValueError:
            conflicts += 1
    orders = db.execute(
        select(models.Order)
        .options(selectinload(models.Order.items))
        .where(models.Order.id.in_(order_ids))
        .order_by(models.Order.id)
    ).scalars().all() if order_ids else []
    if len(orders) != EXPECTED_TARGET_COUNT or {target_ref(order) for order in orders} != TARGET_REF_SET:
        conflicts += 1
    audits_by_order = {row.entity_id: row for row in target_audits}
    verified = 0
    for order in orders:
        audit = audits_by_order.get(str(order.id))
        payload = audit.payload if audit is not None and isinstance(audit.payload, dict) else {}
        if (
            order_state(order) != payload.get("applied")
            or order_immutable_sha256(order) != payload.get("immutable_sha256")
            or hashlib.sha256(target_ref(order).encode("utf-8")).hexdigest()
            != payload.get("target_ref_sha256")
        ):
            conflicts += 1
        else:
            verified += 1
    related_sha = client_sha = ""
    if len(orders) == EXPECTED_TARGET_COUNT:
        related_sha, client_sha, _locks = target_related_snapshot(db, orders)
    source_sha = source_provenance_from_batch(db, batch) if batch else ""
    no_send_sha, _counts, _events = no_send_snapshot(db)
    no_send_unchanged = bool(batch and no_send_sha == batch.get("no_send_sha256"))
    if (
        not batch
        or related_sha != batch.get("related_sha256")
        or client_sha != batch.get("client_points_sha256")
        or source_sha != batch.get("source_provenance_sha256")
        or not no_send_unchanged
    ):
        conflicts += 1
    report_rows = problem_rows = 0
    report_conflicts = 1
    if len(orders) == EXPECTED_TARGET_COUNT and len(target_audits) == EXPECTED_TARGET_COUNT:
        try:
            report_rows, problem_rows, report_conflicts = verify_report(db, orders, audits_by_order)
        except Exception:
            report_conflicts = 1
    conflicts += report_conflicts
    current_item_count = sum(len(order.items) for order in orders)
    if (
        current_item_count != EXPECTED_TARGET_ITEM_COUNT
        or report_rows != EXPECTED_TARGET_ITEM_COUNT
        or report_rows != current_item_count
        or problem_rows != 0
    ):
        conflicts += 1
    db.rollback()
    return {
        "schema_version": 1,
        "mode": "verify_counts_only",
        "target_count": EXPECTED_TARGET_COUNT,
        "safe_to_repair": conflicts == 0,
        "conflicts": conflicts,
        "plan_sha256": expected_plan_sha,
        "preimage_sha256": normalize(batch.get("preimage_sha256")),
        "verified_count": verified,
        "report_rows": report_rows,
        "problem_rows": problem_rows,
        "no_send_unchanged": no_send_unchanged,
    }


def validate_approved_preimage_payload(
    payload,
    expected_plan_sha,
    expected_preimage_sha,
) -> dict[str, dict]:
    top_level_fields = {
        "schema_version",
        "plan_sha256",
        "target_date",
        "target_item_count",
        "related_sha256",
        "client_points_sha256",
        "source_provenance_sha256",
        "no_send_sha256",
        "entries",
    }
    entry_fields = {
        "order_id",
        "target_ref_sha256",
        "preimage",
        "applied",
        "immutable_sha256",
    }
    state_fields = {"order_date", "address", "coordinates_present", "coordinates"}
    if not valid_sha(expected_plan_sha) or not valid_sha(expected_preimage_sha):
        raise RepairBlocked("ROLLBACK_HASH_INVALID")
    if not isinstance(payload, dict) or set(payload) != top_level_fields:
        raise RepairBlocked("PREIMAGE_SCOPE_MISMATCH")
    if (
        payload.get("schema_version") != 1
        or payload.get("plan_sha256") != expected_plan_sha
        or payload.get("target_date") != TARGET_DATE.isoformat()
        or payload.get("target_item_count") != EXPECTED_TARGET_ITEM_COUNT
        or payload_sha256(payload) != expected_preimage_sha
    ):
        raise RepairBlocked("PREIMAGE_SCOPE_MISMATCH")
    for field_name in (
        "related_sha256",
        "client_points_sha256",
        "source_provenance_sha256",
        "no_send_sha256",
    ):
        if not valid_sha(payload.get(field_name)):
            raise RepairBlocked("PREIMAGE_SCOPE_MISMATCH")
    entries = payload.get("entries")
    if not isinstance(entries, list) or len(entries) != EXPECTED_TARGET_COUNT:
        raise RepairBlocked("PREIMAGE_SCOPE_MISMATCH")
    entries_by_order = {}
    for entry in entries:
        if not isinstance(entry, dict) or set(entry) != entry_fields:
            raise RepairBlocked("PREIMAGE_SCOPE_MISMATCH")
        order_id = str(parse_uuid(entry.get("order_id"), "PREIMAGE_SCOPE_MISMATCH"))
        if order_id in entries_by_order:
            raise RepairBlocked("PREIMAGE_SCOPE_MISMATCH")
        if not valid_sha(entry.get("target_ref_sha256")) or not valid_sha(entry.get("immutable_sha256")):
            raise RepairBlocked("PREIMAGE_SCOPE_MISMATCH")
        preimage = entry.get("preimage")
        applied = entry.get("applied")
        if (
            not isinstance(preimage, dict)
            or not isinstance(applied, dict)
            or set(preimage) != state_fields
            or set(applied) != state_fields
            or applied.get("order_date") != TARGET_DATE.isoformat()
            or applied.get("coordinates_present") is not True
            or not isinstance(preimage.get("coordinates_present"), bool)
        ):
            raise RepairBlocked("PREIMAGE_SCOPE_MISMATCH")
        entries_by_order[order_id] = entry
    return entries_by_order


def validate_repair_audits_against_preimage(
    target_audits,
    batch_audit,
    entries_by_order,
    approved_preimage,
    expected_plan_sha,
    expected_preimage_sha,
) -> None:
    if len(target_audits) != EXPECTED_TARGET_COUNT or batch_audit is None:
        raise RepairBlocked("ROLLBACK_AUDIT_SCOPE_INVALID")
    batch = batch_audit.payload if isinstance(batch_audit.payload, dict) else {}
    if (
        batch_audit.entity_type != "order_batch"
        or batch.get("plan_sha256") != expected_plan_sha
        or batch.get("preimage_sha256") != expected_preimage_sha
        or batch.get("target_count") != EXPECTED_TARGET_COUNT
        or batch.get("target_item_count") != EXPECTED_TARGET_ITEM_COUNT
        or batch.get("target_date") != TARGET_DATE.isoformat()
        or batch.get("related_sha256") != approved_preimage.get("related_sha256")
        or batch.get("client_points_sha256") != approved_preimage.get("client_points_sha256")
        or batch.get("source_provenance_sha256") != approved_preimage.get("source_provenance_sha256")
        or batch.get("no_send_sha256") != approved_preimage.get("no_send_sha256")
    ):
        raise RepairBlocked("ROLLBACK_AUDIT_MISMATCH")
    if {audit.entity_id for audit in target_audits} != set(entries_by_order):
        raise RepairBlocked("ROLLBACK_AUDIT_MISMATCH")
    for audit in target_audits:
        payload = audit.payload if isinstance(audit.payload, dict) else {}
        approved = entries_by_order[audit.entity_id]
        if (
            audit.entity_type != "order"
            or payload.get("plan_sha256") != expected_plan_sha
            or payload.get("preimage_sha256") != expected_preimage_sha
            or payload.get("order_id") != approved.get("order_id")
            or audit.entity_id != approved.get("order_id")
            or payload.get("target_ref_sha256") != approved.get("target_ref_sha256")
            or payload.get("preimage") != approved.get("preimage")
            or payload.get("applied") != approved.get("applied")
            or payload.get("immutable_sha256") != approved.get("immutable_sha256")
        ):
            raise RepairBlocked("ROLLBACK_AUDIT_MISMATCH")


def rollback_applied(
    db,
    expected_plan_sha,
    *,
    approved_preimage,
    expected_preimage_sha,
) -> dict:
    from sqlalchemy import select

    entries_by_order = validate_approved_preimage_payload(
        approved_preimage,
        expected_plan_sha,
        expected_preimage_sha,
    )
    models = backend_module("models")
    AuditLog, Order = models.AuditLog, models.Order
    existing_rollbacks = [
        row for row in existing_batch_audits(db, ROLLBACK_BATCH_ACTION)
        if isinstance(row.payload, dict) and row.payload.get("plan_sha256") == expected_plan_sha
    ]
    if existing_rollbacks:
        rollback_payload = existing_rollbacks[0].payload or {}
        if (
            rollback_payload.get("plan_sha256") != expected_plan_sha
            or rollback_payload.get("preimage_sha256") != expected_preimage_sha
        ):
            raise RepairBlocked("ROLLBACK_AUDIT_MISMATCH")
        return {
            "schema_version": 1,
            "mode": "rollback_already_completed_counts_only",
            "target_count": EXPECTED_TARGET_COUNT,
            "safe_to_repair": True,
            "conflicts": 0,
            "plan_sha256": expected_plan_sha,
            "preimage_sha256": normalize(existing_rollbacks[0].payload.get("preimage_sha256")),
            "rollback_count": EXPECTED_TARGET_COUNT,
            "rollback_audits": EXPECTED_TARGET_COUNT,
        }
    db.expire_all()
    target_audits, batch_audits = repair_audits(
        db,
        expected_plan_sha,
        for_update=True,
    )
    if len(target_audits) != EXPECTED_TARGET_COUNT or len(batch_audits) != 1:
        raise RepairBlocked("ROLLBACK_AUDIT_SCOPE_INVALID")
    batch = batch_audits[0].payload
    validate_repair_audits_against_preimage(
        target_audits,
        batch_audits[0],
        entries_by_order,
        approved_preimage,
        expected_plan_sha,
        expected_preimage_sha,
    )
    order_ids = tuple(
        parse_uuid(value, "ROLLBACK_AUDIT_SCOPE_INVALID")
        for value in sorted(entries_by_order)
    )
    db.execute(select(Order.id).where(Order.id.in_(order_ids)).order_by(Order.id).with_for_update()).all()
    orders = db.execute(select(Order).where(Order.id.in_(order_ids)).order_by(Order.id)).scalars().all()
    guard_failed = len(orders) != EXPECTED_TARGET_COUNT
    for order in orders:
        approved = entries_by_order[str(order.id)]
        guard_failed = guard_failed or (
            order_state(order) != approved.get("applied")
            or order_immutable_sha256(order) != approved.get("immutable_sha256")
        )
    related_sha, client_sha, _locks = target_related_snapshot(db, orders) if orders else ("", "", {})
    no_send_sha, _counts, _events = no_send_snapshot(db)
    source_sha = source_provenance_from_batch(db, batch)
    guard_failed = guard_failed or any((
        related_sha != approved_preimage.get("related_sha256"),
        client_sha != approved_preimage.get("client_points_sha256"),
        no_send_sha != approved_preimage.get("no_send_sha256"),
        source_sha != approved_preimage.get("source_provenance_sha256"),
    ))
    if guard_failed:
        db.rollback()
        raise RepairBlocked("ROLLBACK_GUARD_FAILED")

    now = datetime.now(timezone.utc)
    for order in orders:
        approved = entries_by_order[str(order.id)]
        apply_state(order, approved["preimage"])
        db.add(AuditLog(
            id=deterministic_uuid("rollback-target-audit", order.id, expected_plan_sha),
            action=ROLLBACK_ACTION,
            entity_type="order",
            entity_id=str(order.id),
            payload={
                "plan_sha256": expected_plan_sha,
                "preimage_sha256": expected_preimage_sha,
                "target_ref_sha256": approved.get("target_ref_sha256"),
                "rolled_back_at": now.isoformat(),
                "values_redacted": True,
            },
        ))
    db.add(AuditLog(
        id=deterministic_uuid("rollback-batch-audit", expected_plan_sha),
        action=ROLLBACK_BATCH_ACTION,
        entity_type="order_batch",
        entity_id="telegram_logistics_orders_2026_07_23_v1",
        payload={
            "plan_sha256": expected_plan_sha,
            "preimage_sha256": expected_preimage_sha,
            "target_count": EXPECTED_TARGET_COUNT,
            "rolled_back_at": now.isoformat(),
            "values_redacted": True,
        },
    ))
    db.flush()
    for order in orders:
        approved = entries_by_order[str(order.id)]
        if (
            order_state(order) != approved.get("preimage")
            or order_immutable_sha256(order) != approved.get("immutable_sha256")
        ):
            raise RepairBlocked("ROLLBACK_POST_FLUSH_FAILED")
    post_related, post_client, _locks = target_related_snapshot(db, orders)
    post_no_send, _counts, _events = no_send_snapshot(db)
    if post_related != related_sha or post_client != client_sha or post_no_send != no_send_sha:
        raise RepairBlocked("ROLLBACK_POST_FLUSH_FAILED")
    db.commit()
    return {
        "schema_version": 1,
        "mode": "rollback_counts_only",
        "target_count": EXPECTED_TARGET_COUNT,
        "safe_to_repair": True,
        "conflicts": 0,
        "plan_sha256": expected_plan_sha,
        "preimage_sha256": expected_preimage_sha,
        "rollback_count": EXPECTED_TARGET_COUNT,
        "rollback_audits": EXPECTED_TARGET_COUNT,
    }


def validate_preimage_file(path, expected_sha, expected_plan_sha) -> dict:
    source = Path(path)
    if not source.is_absolute() or not source.is_file():
        raise RepairBlocked("PREIMAGE_FILE_INVALID")
    if (source.stat().st_mode & 0o777) != 0o600:
        raise RepairBlocked("PREIMAGE_MODE_INVALID")
    data = source.read_bytes()
    if hashlib.sha256(data).hexdigest() != expected_sha:
        raise RepairBlocked("PREIMAGE_SHA_MISMATCH")
    try:
        payload = json.loads(data.decode("utf-8"))
    except Exception:
        raise RepairBlocked("PREIMAGE_FILE_INVALID") from None
    validate_approved_preimage_payload(payload, expected_plan_sha, expected_sha)
    return payload


def parse_args(argv=None):
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--plan", action="store_true")
    mode.add_argument("--apply", action="store_true")
    mode.add_argument("--verify", action="store_true")
    mode.add_argument("--rollback", action="store_true")
    parser.add_argument("--approval", default="")
    parser.add_argument("--expected-plan-sha", default="")
    parser.add_argument("--preimage-out", default="")
    parser.add_argument("--preimage-file", default="")
    parser.add_argument("--expected-preimage-sha", default="")
    parser.add_argument("--no-send", action="store_true")
    return parser.parse_args(argv)


def run(argv=None):
    from sqlalchemy import text

    args = parse_args(argv)
    SessionLocal = backend_module("db").SessionLocal
    with SessionLocal() as db:
        if args.verify:
            result = verify_applied(db, args.expected_plan_sha, no_send=args.no_send)
            print(json.dumps(result, sort_keys=True))
            return 0 if result["safe_to_repair"] else 3
        if args.rollback:
            if args.approval != ROLLBACK_APPROVAL:
                raise RepairBlocked("ROLLBACK_APPROVAL_REJECTED")
            if not valid_sha(args.expected_plan_sha) or not valid_sha(args.expected_preimage_sha):
                raise RepairBlocked("ROLLBACK_HASH_INVALID")
            approved_preimage = validate_preimage_file(
                args.preimage_file,
                args.expected_preimage_sha,
                args.expected_plan_sha,
            )
            db.execute(text("SET LOCAL lock_timeout = '15s'"))
            db.execute(text("SET LOCAL statement_timeout = '120s'"))
            db.execute(text("SELECT pg_advisory_xact_lock(hashtextextended(:identity, 0))"), {
                "identity": "taksklad:telegram-logistics-orders-repair:2026-07-23:v1",
            })
            result = rollback_applied(
                db,
                args.expected_plan_sha,
                approved_preimage=approved_preimage,
                expected_preimage_sha=args.expected_preimage_sha,
            )
            if result["preimage_sha256"] != args.expected_preimage_sha:
                raise RepairBlocked("ROLLBACK_PREIMAGE_SHA_MISMATCH")
            print(json.dumps(result, sort_keys=True))
            return 0

        if args.plan:
            plan = create_plan(db)
            db.rollback()
            print(json.dumps(plan.summary, sort_keys=True))
            return 0
        if (
            args.approval != APPLY_APPROVAL
            or not valid_sha(args.expected_plan_sha)
            or not args.preimage_out
        ):
            db.rollback()
            raise RepairBlocked("APPLY_APPROVAL_REJECTED")
        db.execute(text("SET LOCAL lock_timeout = '15s'"))
        db.execute(text("SET LOCAL statement_timeout = '120s'"))
        db.execute(text("SELECT pg_advisory_xact_lock(hashtextextended(:identity, 0))"), {
            "identity": "taksklad:telegram-logistics-orders-repair:2026-07-23:v1",
        })
        plan = create_plan(db)
        if args.expected_plan_sha != plan.summary["plan_sha256"]:
            db.rollback()
            raise RepairBlocked("APPLY_PLAN_SHA_MISMATCH")
        result = apply_plan(db, plan, preimage_out=args.preimage_out)
        print(json.dumps(result, sort_keys=True))
        return 0


if __name__ == "__main__":
    try:
        raise SystemExit(run())
    except RepairBlocked as exc:
        print(json.dumps({
            "schema_version": 1,
            "mode": "blocked_counts_only",
            "safe_to_repair": False,
            "conflicts": 1,
            "blocker": exc.code,
        }, sort_keys=True))
        raise SystemExit(3)
    except SystemExit:
        raise
    except Exception as exc:
        print(json.dumps({
            "schema_version": 1,
            "mode": "failed_counts_only",
            "safe_to_repair": False,
            "conflicts": -1,
            "error_type": type(exc).__name__,
        }, sort_keys=True))
        raise SystemExit(6)
