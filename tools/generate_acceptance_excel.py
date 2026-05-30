#!/usr/bin/env python3
import argparse
import os
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_MARKER = "ACCEPTANCE TELEGRAM 20260531"
DEFAULT_SHIPMENT_DATE = "31.05.2026"
DEFAULT_OUTPUT = Path("outputs/taksklad_acceptance/TakSklad_Telegram_Acceptance_2026-05-31.xlsx")
FIXED_XLSX_TIMESTAMP = (2026, 5, 31, 0, 0, 0)

HEADERS = [
    "Дата отгрузки",
    "Тип оплаты",
    "Клиент",
    "Адрес",
    "Координаты",
    "Торговый представитель",
    "Товары",
    "Кол-во блок",
    "Кол-во ШТ",
    "Цена за блок",
    "Сумма позиции",
]


def acceptance_rows(marker=DEFAULT_MARKER, shipment_date=DEFAULT_SHIPMENT_DATE):
    return [
        [
            shipment_date,
            "Перечисление",
            marker,
            "Ташкент, acceptance address",
            "41.311081, 69.240562",
            "+998900000001",
            "Chapman Brown OP 20",
            2,
            20,
            240000,
            480000,
        ],
        [
            shipment_date,
            "Перечисление",
            marker,
            "Ташкент, acceptance address",
            "41.311081, 69.240562",
            "+998900000001",
            "Chapman Gold SSL 20",
            1,
            10,
            240000,
            240000,
        ],
    ]


def build_workbook(marker=DEFAULT_MARKER, shipment_date=DEFAULT_SHIPMENT_DATE):
    workbook = Workbook()
    workbook.properties.creator = "TakSklad"
    workbook.properties.lastModifiedBy = "TakSklad"
    workbook.properties.created = datetime(2026, 5, 31, 0, 0, 0)
    workbook.properties.modified = datetime(2026, 5, 31, 0, 0, 0)
    sheet = workbook.active
    sheet.title = "Заявки"
    sheet.append(HEADERS)
    for row in acceptance_rows(marker=marker, shipment_date=shipment_date):
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="F0E68C")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 14,
        "B": 16,
        "C": 32,
        "D": 30,
        "E": 22,
        "F": 22,
        "G": 24,
        "H": 14,
        "I": 12,
        "J": 14,
        "K": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row_number in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = 24
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    return workbook


def normalize_xlsx_archive(path):
    path = Path(path)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as source:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for name in source.namelist():
                target_info = zipfile.ZipInfo(name, date_time=FIXED_XLSX_TIMESTAMP)
                target_info.compress_type = zipfile.ZIP_DEFLATED
                target_info.create_system = 0
                target_info.external_attr = 0
                target.writestr(target_info, source.read(name))
    os.replace(temp_path, path)


def save_acceptance_excel(output_path=DEFAULT_OUTPUT, marker=DEFAULT_MARKER, shipment_date=DEFAULT_SHIPMENT_DATE):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(marker=marker, shipment_date=shipment_date)
    workbook.save(output_path)
    workbook.close()
    normalize_xlsx_archive(output_path)
    return output_path


def parse_args():
    parser = argparse.ArgumentParser(description="Generate TakSklad Telegram/Windows acceptance Excel file.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output .xlsx path.")
    parser.add_argument("--marker", default=DEFAULT_MARKER, help="Acceptance client marker.")
    parser.add_argument("--shipment-date", default=DEFAULT_SHIPMENT_DATE, help="Shipment date in DD.MM.YYYY format.")
    return parser.parse_args()


def main():
    args = parse_args()
    output_path = save_acceptance_excel(
        output_path=args.output,
        marker=args.marker,
        shipment_date=args.shipment_date,
    )
    print(output_path)


if __name__ == "__main__":
    main()
