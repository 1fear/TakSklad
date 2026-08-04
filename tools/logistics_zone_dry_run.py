"""Report how a shipment date would split between city and region, without sending."""

from __future__ import annotations

import argparse
from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from backend.app.logistics_service import build_logistics_reports  # noqa: E402


def count_rows(report) -> int:
    if report is None:
        return 0
    payload, _filename = report
    workbook = load_workbook(BytesIO(payload), read_only=True)
    count = workbook["Orders"].max_row - 1
    workbook.close()
    return max(0, count)


def summarize(db: Session, shipment_date: str) -> dict:
    reports = build_logistics_reports(db, shipment_date)
    unassigned = reports.get("unassigned") or []
    return {
        "shipment_date": shipment_date,
        "city_rows": count_rows(reports.get("city")),
        "region_rows": count_rows(reports.get("region")),
        "unassigned": len(unassigned),
        "unassigned_clients": [order.client for order in unassigned],
        "region_directory_empty": bool(reports.get("region_directory_empty")),
    }


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Dry-run the logistics zone split")
    parser.add_argument("dates", nargs="+", help="shipment dates in YYYY-MM-DD")
    args = parser.parse_args(argv)

    from backend.app.db import SessionLocal

    db = SessionLocal()
    try:
        for shipment_date in args.dates:
            try:
                summary = summarize(db, shipment_date)
            except Exception as exc:
                print(f"{shipment_date}: пропущено, {exc.__class__.__name__}")
                continue
            print(f"{shipment_date}:")
            if summary["region_directory_empty"]:
                print("  ВНИМАНИЕ: справочник областных точек пуст,")
                print("  сработала страховка, весь отчёт ушёл бы городским файлом")
            print(f"  город строк:        {summary['city_rows']}")
            print(f"  область строк:      {summary['region_rows']}")
            print(f"  вне зон заказов:    {summary['unassigned']}")
            for client in summary["unassigned_clients"]:
                print(f"    - {client}")
    finally:
        db.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
