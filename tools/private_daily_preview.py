#!/usr/bin/env python3
"""Send two read-only SkladBot daily previews to the exact personal admin route."""

from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from io import BytesIO
import json
from typing import Any

from openpyxl import load_workbook

from app.telegram_output_contract import daily_report_caption
from app.telegram_worker import TelegramWorker


EXPECTED_SHEETS = [
    "Сводка",
    "Заявки",
    "Товары заявок",
    "Коды маркировок",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--report-date", action="append", required=True)
    args = parser.parse_args()
    try:
        args.report_dates = [date.fromisoformat(value) for value in args.report_date]
    except ValueError as exc:
        raise SystemExit("PRIVATE_DAILY_PREVIEW_DATE_INVALID") from exc
    if len(args.report_dates) != 2 or len(set(args.report_dates)) != 2:
        raise SystemExit("PRIVATE_DAILY_PREVIEW_EXACTLY_TWO_DATES_REQUIRED")
    return args


def prepare_one(report_date: date) -> dict[str, Any]:
    worker = TelegramWorker()
    prepared = worker.prepare_skladbot_daily_report(
        report_date=report_date,
        scheduled=True,
        build_for_dry_run=True,
    )
    if prepared.get("blocker"):
        raise RuntimeError("PRIVATE_DAILY_PREVIEW_REPORT_BLOCKED")
    if int(prepared.get("requests_count") or 0) <= 0:
        raise RuntimeError("PRIVATE_DAILY_PREVIEW_NO_REQUESTS")
    content = prepared.get("content")
    message = str(prepared.get("message") or "").strip()
    filename = str(prepared.get("filename") or "").strip()
    if not isinstance(content, bytes) or not content or not message or not filename:
        raise RuntimeError("PRIVATE_DAILY_PREVIEW_PAYLOAD_INVALID")
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        if workbook.sheetnames != EXPECTED_SHEETS:
            raise RuntimeError("PRIVATE_DAILY_PREVIEW_SHEETS_INVALID")
    finally:
        workbook.close()
    return prepared


def main() -> int:
    args = parse_args()
    with ThreadPoolExecutor(max_workers=2) as executor:
        prepared = list(executor.map(prepare_one, args.report_dates))

    sender = TelegramWorker()
    chat_id = str(sender.automation_alert_chat_id or "").strip()
    if (
        not chat_id
        or not chat_id.isdigit()
        or int(chat_id) <= 0
        or chat_id not in sender.admin_chat_ids
        or chat_id not in sender.allowed_chat_ids
    ):
        raise RuntimeError("PRIVATE_DAILY_PREVIEW_ROUTE_INVALID")

    results = []
    for report_date, payload in sorted(zip(args.report_dates, prepared)):
        sender.send_message(chat_id, payload["message"])
        sender.send_document(
            chat_id,
            payload["content"],
            payload["filename"],
            caption=daily_report_caption(report_date),
        )
        results.append({
            "report_date": report_date.isoformat(),
            "requests_count": int(payload.get("requests_count") or 0),
            "order_kiz_count": int(payload.get("order_kiz_count") or 0),
            "xlsx_bytes": len(payload["content"]),
            "message_count": 1,
            "document_count": 1,
        })

    print(json.dumps({
        "status": "sent",
        "target": "personal_admin",
        "dates_count": 2,
        "database_writes": 0,
        "client_sends": 0,
        "results": results,
    }, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
