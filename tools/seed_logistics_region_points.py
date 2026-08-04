"""Load the region address programme into logistics_region_points.

The source workbook stays outside the repository: it holds counterparty names
and coordinates, and this repository is public.
"""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from backend.app.logistics_zone_service import normalize_client_key  # noqa: E402
from backend.app.models import LogisticsRegionPoint  # noqa: E402


def read_region_rows(path) -> list[dict]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    rows = []
    seen = set()
    for index, values in enumerate(sheet.iter_rows(values_only=True)):
        if index == 0:
            continue
        client_name = str(values[0] or "").strip()
        latitude = values[1] if len(values) > 1 else None
        longitude = values[2] if len(values) > 2 else None
        agent = str(values[3] or "").strip() if len(values) > 3 else ""
        if not client_name:
            continue
        if not isinstance(latitude, (int, float)) or not isinstance(longitude, (int, float)):
            continue
        latitude = round(float(latitude), 6)
        longitude = round(float(longitude), 6)
        key = (normalize_client_key(client_name), latitude, longitude)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            "client_name": client_name,
            "normalized_client": key[0],
            "latitude": latitude,
            "longitude": longitude,
            "agent": agent or None,
        })
    workbook.close()
    return rows


def existing_by_key(db: Session) -> dict:
    points = db.execute(select(LogisticsRegionPoint)).scalars().all()
    return {
        (point.normalized_client, round(float(point.latitude), 6), round(float(point.longitude), 6)): point
        for point in points
    }


def plan_changes(db: Session, rows: list[dict]) -> dict:
    current = existing_by_key(db)
    insert = 0
    update = 0
    unchanged = 0
    for row in rows:
        key = (row["normalized_client"], row["latitude"], row["longitude"])
        point = current.get(key)
        if point is None:
            insert += 1
        elif point.client_name != row["client_name"] or point.agent != row["agent"] or not point.is_active:
            update += 1
        else:
            unchanged += 1
    return {
        "existing": len(current),
        "insert": insert,
        "update": update,
        "unchanged": unchanged,
        "total_after": len(current) + insert,
    }


def apply_changes(db: Session, rows: list[dict]) -> dict:
    plan = plan_changes(db, rows)
    current = existing_by_key(db)
    for row in rows:
        key = (row["normalized_client"], row["latitude"], row["longitude"])
        point = current.get(key)
        if point is None:
            db.add(LogisticsRegionPoint(
                client_name=row["client_name"],
                normalized_client=row["normalized_client"],
                latitude=row["latitude"],
                longitude=row["longitude"],
                agent=row["agent"],
                is_active=True,
                raw_payload={"source": "address_programme"},
            ))
            continue
        point.client_name = row["client_name"]
        point.agent = row["agent"]
        point.is_active = True
        point.raw_payload = {**(point.raw_payload or {}), "source": "address_programme"}
    db.commit()
    return plan


def print_plan(plan: dict) -> None:
    print(f"было записей:      {plan['existing']}")
    print(f"будет добавлено:   {plan['insert']}")
    print(f"будет обновлено:   {plan['update']}")
    print(f"без изменений:     {plan['unchanged']}")
    print(f"станет записей:    {plan['total_after']}")


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Load region address programme")
    parser.add_argument("source", help="path to the address programme xlsx")
    parser.add_argument("--dry-run", action="store_true", help="only print the plan")
    args = parser.parse_args(argv)

    from backend.app.db import SessionLocal

    rows = read_region_rows(args.source)
    print(f"строк в файле после дедупликации: {len(rows)}")
    db = SessionLocal()
    try:
        plan = plan_changes(db, rows)
        print_plan(plan)
        if args.dry_run:
            print("режим --dry-run, запись не выполнялась")
            return 0
        apply_changes(db, rows)
        print("записано")
    finally:
        db.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
