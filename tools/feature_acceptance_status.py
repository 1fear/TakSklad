#!/usr/bin/env python3
import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_REGISTER_PATH = PROJECT_ROOT / "docs" / "taksklad-feature-user-stories.xlsx"

MANUAL_PASS_STATUSES = {"passed", "accepted", "done", "ok"}
RESOLVED_ERROR_STATUSES = {"fixed", "fixed_retested", "accepted", "done", "ok", "not_applicable"}
OPEN_MANUAL_STATUSES = {"pending", "failed", "blocked", "manual_required", "not_run"}
KNOWN_MANUAL_STATUSES = MANUAL_PASS_STATUSES | OPEN_MANUAL_STATUSES
OPEN_ERROR_STATUSES = {
    "needs_validation",
    "documented_fix_pending",
    "docs_fixed_env_rebuild_pending",
    "fix_needed",
    "open",
    "failed",
    "blocked",
}
KNOWN_ERROR_STATUSES = RESOLVED_ERROR_STATUSES | OPEN_ERROR_STATUSES
REQUIRED_COLUMNS = {
    "User Stories": {
        "Feature ID",
        "Feature",
        "Test Type",
        "Automated Result",
        "Manual Result",
        "Status",
        "Test Status",
    },
    "Test Loop": {
        "Feature ID",
        "Test Type",
        "Command/Manual Step",
        "Result",
        "Automated Result",
        "Manual Result",
        "Evidence",
    },
    "Errors": {
        "Error ID",
        "Feature ID",
        "Type",
        "Severity",
        "Description",
        "Status",
    },
    "Manual Acceptance": {
        "Feature ID",
        "Feature",
        "Dependency",
        "Manual Step",
        "Expected",
        "Status",
    },
}


def normalize(value):
    return str(value or "").strip()


def rows_by_header(workbook, sheet_name):
    sheet = workbook[sheet_name]
    headers = [normalize(cell.value) for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(normalize(value) for value in row):
            continue
        rows.append({headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))})
    return rows


def validate_headers(workbook, required_columns):
    problems = []
    for sheet_name, required in required_columns.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        headers = [normalize(cell.value) for cell in sheet[1]]
        blank_indexes = [str(index + 1) for index, header in enumerate(headers) if not header]
        if blank_indexes:
            problems.append(f"{sheet_name} has blank header cells: {', '.join(blank_indexes)}")
        duplicate_headers = sorted({header for header in headers if header and headers.count(header) > 1})
        if duplicate_headers:
            problems.append(f"{sheet_name} has duplicate headers: {', '.join(duplicate_headers)}")
        missing = sorted(required - set(headers))
        if missing:
            problems.append(f"{sheet_name} missing required columns: {', '.join(missing)}")
    return problems


def status_key(value):
    return normalize(value).casefold()


def manual_acceptance_summary(rows):
    counts = {}
    pending = []
    failed = []
    passed = []
    for row in rows:
        status = status_key(row.get("Status"))
        counts[status or "blank"] = counts.get(status or "blank", 0) + 1
        item = {
            "feature_id": normalize(row.get("Feature ID")),
            "feature": normalize(row.get("Feature")),
            "dependency": normalize(row.get("Dependency")),
            "status": status or "blank",
        }
        if status in MANUAL_PASS_STATUSES:
            passed.append(item)
        elif status in {"failed", "blocked"}:
            failed.append(item)
        else:
            pending.append(item)
    return {
        "total": len(rows),
        "counts": counts,
        "passed": len(passed),
        "pending": len(pending),
        "failed": len(failed),
        "pending_items": pending,
        "failed_items": failed,
    }


def error_summary(rows):
    counts = {}
    open_items = []
    fixed_items = []
    for row in rows:
        status = status_key(row.get("Status"))
        counts[status or "blank"] = counts.get(status or "blank", 0) + 1
        item = {
            "error_id": normalize(row.get("Error ID")),
            "feature_id": normalize(row.get("Feature ID")),
            "type": normalize(row.get("Type")),
            "severity": normalize(row.get("Severity")),
            "status": status or "blank",
            "description": normalize(row.get("Description")),
        }
        if status.startswith("fixed") or status in RESOLVED_ERROR_STATUSES:
            fixed_items.append(item)
        elif status in OPEN_ERROR_STATUSES or not status:
            open_items.append(item)
        else:
            open_items.append(item)
    return {
        "total": len(rows),
        "counts": counts,
        "fixed": len(fixed_items),
        "open": len(open_items),
        "open_items": open_items,
    }


def evaluate_register(path=DEFAULT_REGISTER_PATH):
    path = Path(path)
    if not path.exists():
        return {
            "status": "error",
            "path": str(path),
            "problems": [f"feature register not found: {path}"],
        }

    workbook = load_workbook(path, read_only=True, data_only=True)
    problems = []
    required_sheets = {"User Stories", "Test Loop", "Errors", "Manual Acceptance"}
    missing_sheets = sorted(required_sheets - set(workbook.sheetnames))
    if missing_sheets:
        problems.append(f"required sheet missing: {', '.join(missing_sheets)}")
        return {
            "status": "error",
            "path": str(path),
            "problems": problems,
        }
    problems.extend(validate_headers(workbook, REQUIRED_COLUMNS))

    stories = rows_by_header(workbook, "User Stories")
    test_loop = rows_by_header(workbook, "Test Loop")
    errors = rows_by_header(workbook, "Errors")
    manual = rows_by_header(workbook, "Manual Acceptance")

    story_ids = {normalize(row.get("Feature ID")) for row in stories if normalize(row.get("Feature ID"))}
    test_ids = {normalize(row.get("Feature ID")) for row in test_loop if normalize(row.get("Feature ID"))}
    manual_ids = {normalize(row.get("Feature ID")) for row in manual if normalize(row.get("Feature ID"))}
    expected_manual_ids = {
        normalize(row.get("Feature ID"))
        for row in stories
        if normalize(row.get("Feature ID"))
        and (
            status_key(row.get("Manual Result")) != "not_applicable"
            or status_key(row.get("Test Type")) in {"manual", "auto+manual"}
        )
    }
    if story_ids != test_ids:
        problems.append("User Stories and Test Loop feature IDs differ")
    if manual_ids != expected_manual_ids:
        missing_manual = sorted(expected_manual_ids - manual_ids)
        extra_manual = sorted(manual_ids - expected_manual_ids)
        if missing_manual:
            problems.append(f"Manual Acceptance missing required feature IDs: {', '.join(missing_manual)}")
        if extra_manual:
            problems.append(f"Manual Acceptance contains unexpected feature IDs: {', '.join(extra_manual)}")
    elif not manual_ids.issubset(story_ids):
        problems.append("Manual Acceptance contains unknown feature IDs")
    unknown_manual_statuses = sorted({
        status_key(row.get("Status"))
        for row in manual
        if status_key(row.get("Status")) and status_key(row.get("Status")) not in KNOWN_MANUAL_STATUSES
    })
    if unknown_manual_statuses:
        problems.append(f"Manual Acceptance has unknown statuses: {', '.join(unknown_manual_statuses)}")
    unknown_error_statuses = sorted({
        status_key(row.get("Status"))
        for row in errors
        if status_key(row.get("Status")) and status_key(row.get("Status")) not in KNOWN_ERROR_STATUSES
    })
    if unknown_error_statuses:
        problems.append(f"Errors has unknown statuses: {', '.join(unknown_error_statuses)}")

    manual_summary = manual_acceptance_summary(manual)
    errors_summary = error_summary(errors)
    automated_passed = sum(1 for row in stories if status_key(row.get("Automated Result")) == "passed")
    manual_pending = manual_summary["pending"] + manual_summary["failed"]
    open_errors = errors_summary["open"]
    status = "ok" if not problems else "error"

    return {
        "scope": "feature_register_status",
        "release_gate_note": "This is not production release GO/NO-GO; release canon is tools/release_go_no_go.py with outputs/taksklad_acceptance/ACCEPTANCE_RESULTS.md.",
        "status": status,
        "path": str(path),
        "problems": problems,
        "stories": {
            "total": len(stories),
            "automated_passed": automated_passed,
        },
        "test_loop": {
            "total": len(test_loop),
        },
        "manual_acceptance": manual_summary,
        "errors": errors_summary,
        "ready": {
            "manual_complete": manual_pending == 0,
            "no_open_errors": open_errors == 0,
            "all_feature_ids_consistent": not problems,
        },
    }


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Evaluate TakSklad feature user stories acceptance register. "
            "This is not the production release GO/NO-GO gate."
        )
    )
    parser.add_argument("--register", default=str(DEFAULT_REGISTER_PATH), help="Path to taksklad-feature-user-stories.xlsx")
    parser.add_argument("--require-manual-complete", action="store_true", help="Exit non-zero while any Manual Acceptance row is not passed/accepted.")
    parser.add_argument("--require-no-open-errors", action="store_true", help="Exit non-zero while Errors sheet has open validation/fix rows.")
    return parser.parse_args()


def main():
    args = parse_args()
    result = evaluate_register(args.register)
    print(json.dumps(result, ensure_ascii=False, sort_keys=True))
    if result["status"] != "ok":
        return 2
    if args.require_manual_complete and not result["ready"]["manual_complete"]:
        return 3
    if args.require_no_open_errors and not result["ready"]["no_open_errors"]:
        return 4
    return 0


if __name__ == "__main__":
    sys.exit(main())
