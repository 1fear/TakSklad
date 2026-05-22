import os
from datetime import datetime

from catalog import calculate_blocks, load_product_catalog, save_product_catalog
from config import (
    CHAPMAN_DATA_SHEET_NAME,
    ORDER_DATE_COLUMN,
    SHEET_NAME,
    SOURCE_OPTIONAL_ALIASES,
    SOURCE_REQUIRED_ALIASES,
    SPREADSHEET_ID,
    STATUS_COLUMN,
    STATUS_NOT_COMPLETED,
)
from geocoding import reverse_geocode_yandex
from orders import make_order_id
from sheets import (
    append_chapman_data_records,
    build_import_record_row,
    ensure_import_sheet_layout,
    get_existing_import_keys,
    get_google_client,
)
from storage import load_data_section, save_data_section
from utils import (
    clean_file_name,
    file_sha1,
    file_sha256,
    make_hash,
    normalize_lookup_text,
    normalize_text,
    parse_date_to_standard,
    parse_int_value,
)


def get_source_header_index(header):
    return {normalize_lookup_text(col): idx for idx, col in enumerate(header) if normalize_lookup_text(col)}


def find_source_column(header_idx, aliases):
    for alias in aliases:
        key = normalize_lookup_text(alias)
        if key in header_idx:
            return header_idx[key]
    return None


def get_source_cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    return normalize_text(value)


def parse_excel_order_files(file_paths, source_names=None):
    import openpyxl

    catalog = load_product_catalog()
    raw_rows = []
    errors = []
    warnings = []
    source_rows_count = 0
    geocoded_count = 0
    geocode_failed_count = 0
    geocode_cache = {}
    source_names = source_names or {}
    source_names_by_path = {
        os.path.abspath(path): clean_file_name(name, os.path.basename(path))
        for path, name in source_names.items()
    }

    for file_path in file_paths:
        file_name = source_names_by_path.get(os.path.abspath(file_path)) or clean_file_name(os.path.basename(file_path))
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as exc:
            errors.append(f"{file_name}: не удалось открыть файл ({exc})")
            continue

        sheet_name = "Заявки" if "Заявки" in workbook.sheetnames else workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            errors.append(f"{file_name}: лист пустой")
            continue

        header_idx = get_source_header_index(header)
        columns = {}
        missing = []
        for key, aliases in SOURCE_REQUIRED_ALIASES.items():
            idx = find_source_column(header_idx, aliases)
            if idx is None:
                missing.append(aliases[0])
            columns[key] = idx

        for key, aliases in SOURCE_OPTIONAL_ALIASES.items():
            columns[key] = find_source_column(header_idx, aliases)

        if missing:
            errors.append(f"{file_name}: нет обязательных колонок: {', '.join(missing)}")
            continue

        source_file_hash = file_sha1(file_path)
        source_file_sha256 = file_sha256(file_path)
        for row_number, row in enumerate(rows_iter, start=2):
            if not row or not any(normalize_text(cell) for cell in row if cell is not None):
                continue

            source_rows_count += 1
            client = get_source_cell(row, columns["client"])
            payment = get_source_cell(row, columns["payment"])
            product = get_source_cell(row, columns["product"])
            quantity = parse_int_value(get_source_cell(row, columns["quantity"]))

            if not client or not payment or not product or quantity <= 0:
                warnings.append(f"{file_name}, строка {row_number}: пропущена, не заполнены клиент/оплата/товар/количество")
                continue

            date_value = parse_date_to_standard(get_source_cell(row, columns.get("date"))) or datetime.now().strftime("%d.%m.%Y")
            source_address = get_source_cell(row, columns.get("address"))
            address = source_address
            coords = get_source_cell(row, columns.get("coords"))
            if not address and coords:
                geocoded_address, geocode_error = reverse_geocode_yandex(coords, cache=geocode_cache)
                if geocoded_address:
                    address = geocoded_address
                    geocoded_count += 1
                else:
                    geocode_failed_count += 1
                    address = f"Координаты: {coords}"
                    warnings.append(f"{file_name}, строка {row_number}: адрес по координатам не получен ({geocode_error})")
            if not address:
                warnings.append(f"{file_name}, строка {row_number}: адрес пустой")
                address = "Адрес не указан"

            representative = get_source_cell(row, columns.get("representative"))
            inn = get_source_cell(row, columns.get("inn"))
            lead_status = get_source_cell(row, columns.get("lead_status"))
            source_id = make_hash({
                "file_hash": source_file_hash,
                "sheet": sheet_name,
                "row": row_number,
            })

            raw_rows.append({
                "date": date_value,
                "payment": payment,
                "client": client,
                "address": address,
                "representative": representative,
                "inn": inn,
                "product": product,
                "quantity": quantity,
                "coords": coords,
                "source_address": source_address,
                "lead_status": lead_status,
                "source_id": source_id,
                "source_file": file_name,
                "source_file_sha256": source_file_sha256,
                "source_row": row_number,
            })

    grouped = {}
    for row in raw_rows:
        key = (
            row["date"],
            normalize_lookup_text(row["payment"]),
            normalize_lookup_text(row["client"]),
            normalize_lookup_text(row["address"]),
            normalize_lookup_text(row["representative"]),
            normalize_lookup_text(row["inn"]),
            normalize_lookup_text(row["product"]),
            normalize_lookup_text(row["coords"]),
            normalize_lookup_text(row["source_address"]),
            normalize_lookup_text(row["lead_status"]),
        )
        if key not in grouped:
            grouped[key] = row.copy()
            grouped[key]["source_ids"] = []
            grouped[key]["source_rows"] = []
            grouped[key]["source_files"] = set()
            grouped[key]["source_file_sha256"] = set()
        else:
            grouped[key]["quantity"] += row["quantity"]
        grouped[key]["source_ids"].append(row["source_id"])
        grouped[key]["source_rows"].append(str(row["source_row"]))
        grouped[key]["source_files"].add(row["source_file"])
        grouped[key]["source_file_sha256"].add(row["source_file_sha256"])

    records = []
    for item in grouped.values():
        blocks, pieces_per_block = calculate_blocks(item["quantity"], item["product"], catalog, warnings)
        record = {
            ORDER_DATE_COLUMN: item["date"],
            "Тип оплаты": item["payment"],
            "Клиент": item["client"],
            "Адрес": item["address"],
            "Торговый представитель": item["representative"],
            "Товары": item["product"],
            "Кол-во ШТ": item["quantity"],
            "Кол-во блок": blocks,
            "Отсканированные коды": "",
            STATUS_COLUMN: STATUS_NOT_COMPLETED,
            "ID импорта": make_hash(sorted(item["source_ids"])),
            "Источник файла": ", ".join(sorted(item["source_files"])),
            "Строка файла": ", ".join(item["source_rows"]),
            "Дата импорта": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
        }
        record["ID заказа"] = make_order_id(record)
        record["_pieces_per_block"] = pieces_per_block
        record["_source_file_sha256"] = sorted(item["source_file_sha256"])
        record["_chapman_data"] = {
            "delivery_date": item["date"],
            "inn": item["inn"],
            "coords": item["coords"],
            "address": item["source_address"],
            "lead_status": item["lead_status"],
        }
        records.append(record)

    save_product_catalog(catalog)

    return {
        "records": records,
        "errors": errors,
        "warnings": warnings,
        "source_rows_count": source_rows_count,
        "files_count": len(file_paths),
        "geocoded_count": geocoded_count,
        "geocode_failed_count": geocode_failed_count,
    }


def prepare_excel_import(file_paths, source_names=None):
    parsed = parse_excel_order_files(file_paths, source_names=source_names)
    client = get_google_client()
    spreadsheet = client.open_by_key(SPREADSHEET_ID)
    sheet = spreadsheet.worksheet(SHEET_NAME)
    ensure_import_sheet_layout(sheet)
    all_rows = sheet.get_all_values()

    existing_import_ids, existing_order_ids = get_existing_import_keys(all_rows)
    new_records = []
    duplicate_records = []

    for record in parsed["records"]:
        if record.get("ID импорта") in existing_import_ids or record.get("ID заказа") in existing_order_ids:
            duplicate_records.append(record)
        else:
            new_records.append(record)

    parsed["new_records"] = new_records
    parsed["duplicate_records"] = duplicate_records
    parsed["clients_count"] = len({record.get("Клиент") for record in new_records})
    parsed["products_count"] = len({record.get("Товары") for record in new_records})
    parsed["blocks_count"] = sum(parse_int_value(record.get("Кол-во блок")) for record in new_records)
    parsed["quantity_count"] = sum(parse_int_value(record.get("Кол-во ШТ")) for record in new_records)
    return parsed


def extract_record_file_hashes(records):
    hashes = set()
    for record in records:
        raw_hashes = record.get("_source_file_sha256", [])
        if isinstance(raw_hashes, str):
            raw_hashes = [raw_hashes]
        for file_hash in raw_hashes:
            normalized_hash = normalize_text(file_hash).lower()
            if normalized_hash:
                hashes.add(normalized_hash)
    return hashes


def find_successful_import_by_file_hash(file_hash):
    normalized_target = normalize_text(file_hash).lower()
    if not normalized_target:
        return None

    history = load_data_section("import_history", [])
    if not isinstance(history, list):
        return None

    for item in reversed(history):
        if not isinstance(item, dict) or parse_int_value(item.get("imported")) <= 0:
            continue
        entry_hashes = set()
        for key in ("source_file_hashes_sha256", "file_hashes_sha256", "file_sha256"):
            raw_hashes = item.get(key, [])
            if isinstance(raw_hashes, str):
                raw_hashes = [raw_hashes]
            for value in raw_hashes:
                normalized_hash = normalize_text(value).lower()
                if normalized_hash:
                    entry_hashes.add(normalized_hash)
        if normalized_target in entry_hashes:
            return item
    return None


def append_import_records(records):
    if not records:
        return {"imported": 0, "duplicates": 0}

    client = get_google_client()
    spreadsheet = client.open_by_key(SPREADSHEET_ID)
    sheet = spreadsheet.worksheet(SHEET_NAME)
    ensure_import_sheet_layout(sheet)
    all_rows = sheet.get_all_values()
    existing_import_ids, existing_order_ids = get_existing_import_keys(all_rows)

    rows_to_append = []
    appended_records = []
    duplicates = 0
    for record in records:
        if record.get("ID импорта") in existing_import_ids or record.get("ID заказа") in existing_order_ids:
            duplicates += 1
            continue
        rows_to_append.append(build_import_record_row(record))
        appended_records.append(record)
        existing_import_ids.add(record.get("ID импорта"))
        existing_order_ids.add(record.get("ID заказа"))

    if rows_to_append:
        chapman_data_sheet = spreadsheet.worksheet(CHAPMAN_DATA_SHEET_NAME)
        chapman_result = append_chapman_data_records(chapman_data_sheet, appended_records)
        start_row = len(all_rows) + 1
        end_row = start_row + len(rows_to_append) - 1
        sheet.batch_update([{
            "range": f"A{start_row}:AE{end_row}",
            "values": rows_to_append,
        }], value_input_option="USER_ENTERED")
    else:
        chapman_result = {"appended": 0, "start_row": None, "end_row": None}

    history = load_data_section("import_history", [])
    if not isinstance(history, list):
        history = []
    history.append({
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "imported": len(rows_to_append),
        "duplicates": duplicates,
        "sources": sorted({record.get("Источник файла", "") for record in records}),
        "source_file_hashes_sha256": sorted(extract_record_file_hashes(appended_records)),
    })
    save_data_section("import_history", history[-200:])

    return {
        "imported": len(rows_to_append),
        "duplicates": duplicates,
        "chapman_data_rows": chapman_result["appended"],
    }
