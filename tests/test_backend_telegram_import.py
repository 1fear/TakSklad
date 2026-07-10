import shutil
import tempfile
import unittest
from unittest import mock
from datetime import datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import httpx
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app import telegram_worker as telegram_worker_module
from backend.app import telegram_admin_processor as telegram_admin_processor_module
from backend.app import telegram_import_processor as telegram_import_processor_module
from backend.app import telegram_clients as telegram_clients_module
from backend.app import excel_importer
from backend.app.excel_importer import excel_file_to_import_payload
from backend.app.models import AuditLog, Base, Incident, PendingEvent
from backend.app.telegram_admin_processor import TelegramAdminProcessor
from backend.app.telegram_clients import BackendApiClient, TelegramApiClient
from backend.app.telegram_import_processor import TelegramImportProcessor
from backend.app.telegram_report_processor import TelegramReportProcessor
from backend.app.telegram_scheduled_report_processor import TelegramScheduledReportProcessor
from backend.app.telegram_worker import (
    TELEGRAM_BUTTON_IMPORTS,
    TELEGRAM_BUTTON_KIZ_BY_FILES,
    TELEGRAM_BUTTON_LOGISTICS_REPORT,
    TELEGRAM_BUTTON_MANUAL,
    TELEGRAM_BUTTON_SHIPMENT_DATE,
    TELEGRAM_BUTTON_STATUS,
    TELEGRAM_KIZ_FILE_PREFIX,
    TELEGRAM_KIZ_RANGE_CALLBACK_PREFIX,
    TelegramConfigurationError,
    TelegramWorker,
    display_date,
    summarize_active_orders_by_date,
    telegram_main_reply_keyboard,
    validate_telegram_worker_config,
)


def create_orders_workbook(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Заявки"
    sheet.append(["Документ от 30.05.2026"])
    sheet.append([
        "Клиент",
        "Тип оплаты",
        "Товары",
        "Кол-во ШТ",
        "Кол-во блок",
        "Адрес",
        "Торговый представитель",
    ])
    sheet.append(["Client One", "Терминал", "Product One", 20, 2, "Address One", "Rep One"])
    sheet.append(["Итого", "", "", 20, 2, "", ""])
    workbook.save(path)


def create_conflicting_date_workbook(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Конструктор отчетов"
    sheet.append([
        "Торговый представитель",
        "Клиент",
        "Координаты клиента",
        "ТМЦ",
        "Тип оплаты",
        "Дата доставки",
        "Количество заказа",
        "Сумма с переоценкой",
    ])
    sheet.append([
        "ТП1",
        "Client One",
        "41.320075,69.298547",
        "Chapman Brown OP 20",
        "Перечисление",
        "09.06.2026",
        20,
        480000,
    ])
    workbook.save(path)


class BackendTelegramImportTests(unittest.TestCase):
    def test_all_processors_execute_with_constructor_injected_ports(self):
        telegram_calls = []
        backend_calls = []

        class FakeTelegramClient:
            def send_message(self, chat_id, text, reply_markup=None):
                telegram_calls.append(("message", chat_id, text, reply_markup))
                return {"message_id": len(telegram_calls)}

            def send_document(self, chat_id, content, filename, caption=""):
                telegram_calls.append(("document", chat_id, content, filename, caption))
                return {"document_id": len(telegram_calls)}

            def download_file(self, file_id, destination_path, max_file_size=0):
                telegram_calls.append(("download", file_id, max_file_size))
                Path(destination_path).write_bytes(b"synthetic-xlsx")

        class FakeBackendClient:
            def get(self, path, params=None):
                backend_calls.append(("get", path, params))
                return [{"date": "2026-06-01", "planned_blocks": 2, "scanned_blocks": 1}]

            def get_bytes(self, path, params=None):
                backend_calls.append(("get_bytes", path, params))
                return b"file", {}

            def post(self, path, payload=None):
                backend_calls.append(("post", path, payload))
                return {"ok": True}

        telegram_client = FakeTelegramClient()
        backend_client = FakeBackendClient()
        report = TelegramReportProcessor(
            telegram_api_client=telegram_client, backend_api_client=backend_client,
        )
        admin = TelegramAdminProcessor(
            telegram_api_client=telegram_client,
            allowed_chat_ids={"allowed"},
            admin_chat_ids=set(),
        )
        importer = TelegramImportProcessor(
            telegram_api_client=telegram_client, backend_api_client=backend_client, max_file_size=64,
        )
        scheduled = TelegramScheduledReportProcessor(telegram_api_client=telegram_client)

        report.show_kiz_dates("chat")
        self.assertFalse(admin.ensure_admin_chat("allowed"))
        with tempfile.TemporaryDirectory() as temp_dir:
            destination = Path(temp_dir) / "orders.xlsx"
            importer.download_telegram_document({"file_id": "file-1"}, destination)
            self.assertEqual(destination.read_bytes(), b"synthetic-xlsx")
        self.assertIsNotNone(scheduled.safe_send_message("chat", "scheduled-ready"))

        self.assertEqual(backend_calls, [
            ("get", "/api/v1/reports/kiz/dates", None),
            ("get_bytes", "/api/v1/reports/kiz/date", {"shipment_date": "2026-06-01"}),
        ])
        self.assertEqual([call[0] for call in telegram_calls], ["document", "message", "download", "message"])
        self.assertEqual(telegram_calls[2], ("download", "file-1", 64))

    def test_explicit_external_clients_build_exact_requests_once(self):
        calls = []

        class FakeResponse:
            content = b"xlsx"
            headers = {"Content-Type": "application/octet-stream"}

            def raise_for_status(self):
                return None

            def json(self):
                return {"ok": True, "result": {"message_id": 7}}

        class FakeClient:
            def __init__(self, timeout=None):
                self.timeout = timeout

            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def post(self, url, **kwargs):
                calls.append(("post", self.timeout, url, kwargs))
                return FakeResponse()

            def get(self, url, **kwargs):
                calls.append(("get", self.timeout, url, kwargs))
                return FakeResponse()

        fake_http = SimpleNamespace(
            Client=FakeClient,
            HTTPError=httpx.HTTPError,
            HTTPStatusError=httpx.HTTPStatusError,
        )
        telegram_client = TelegramApiClient("bot-token", timeout=9, http_client_module=fake_http)
        backend_client = BackendApiClient(
            "http://backend:8000", token="api-token", timeout=11, import_timeout=33,
            http_client_module=fake_http,
        )

        self.assertEqual(telegram_client.send_message("123", "hello"), {"message_id": 7})
        self.assertEqual(backend_client.post("/api/v1/imports", {"rows": []}), {"ok": True, "result": {"message_id": 7}})
        self.assertEqual(telegram_client.poll_updates(4, 7), {"message_id": 7})
        self.assertEqual(len(calls), 3)
        telegram_correlation_id = calls[0][3]["headers"]["X-Correlation-ID"]
        backend_correlation_id = calls[1][3]["headers"]["X-Correlation-ID"]
        poll_correlation_id = calls[2][3]["headers"]["X-Correlation-ID"]
        self.assertEqual(telegram_correlation_id, backend_correlation_id)
        self.assertEqual(telegram_correlation_id, poll_correlation_id)
        self.assertEqual(len(telegram_correlation_id), 36)
        self.assertEqual(calls[0], (
            "post", 9, "https://api.telegram.org/botbot-token/sendMessage",
            {
                "json": {"chat_id": "123", "text": "hello"},
                "headers": {"X-Correlation-ID": telegram_correlation_id},
            },
        ))
        self.assertEqual(calls[1], (
            "post", 33, "http://backend:8000/api/v1/imports",
            {
                "json": {"rows": []},
                "headers": {
                    "X-Correlation-ID": telegram_correlation_id,
                    "Authorization": "Bearer api-token",
                },
            },
        ))
        self.assertEqual(calls[2], (
            "post", 12, "https://api.telegram.org/botbot-token/getUpdates",
            {
                "json": {"offset": 5, "timeout": 7, "allowed_updates": ["message", "callback_query"]},
                "headers": {"X-Correlation-ID": telegram_correlation_id},
            },
        ))

    def test_admin_processor_sends_notification_with_characterized_ordered_calls(self):
        processor = TelegramAdminProcessor()
        calls = []
        events = [{
            "id": "notification-1",
            "payload": {"chat_id": "allowed", "text": "Synthetic notification"},
            "lease_owner": "lease-1",
        }]

        processor.reset_stale_telegram_notification_events = lambda: calls.append(("reset",))

        def take_next():
            calls.append(("take",))
            return events.pop(0) if events else None

        processor.take_next_telegram_notification_event = take_next
        processor.is_allowed_chat = lambda chat_id: calls.append(("authorize", chat_id)) or chat_id == "allowed"
        processor.send_message = lambda chat_id, text: calls.append(("send_message", chat_id, text))
        processor.finish_telegram_notification_event = (
            lambda event_id, success, error="", failure_status="failed", lease_owner="": calls.append(
                ("finish", event_id, success, error, failure_status, lease_owner)
            )
        )

        self.assertEqual(processor.process_pending_telegram_notifications(), 1)
        self.assertEqual(calls, [
            ("reset",),
            ("take",),
            ("authorize", "allowed"),
            ("send_message", "allowed", "Synthetic notification"),
            ("finish", "notification-1", True, "", "failed", "lease-1"),
            ("take",),
        ])

    def test_report_processor_builds_kiz_menu_with_characterized_ordered_calls(self):
        processor = TelegramReportProcessor()
        calls = []
        dates = [
            {"date": "2026-06-02", "planned_blocks": 3, "scanned_blocks": 1, "remaining_blocks": 2},
            {"date": "2026-06-01", "planned_blocks": 2, "scanned_blocks": 2, "remaining_blocks": 0},
        ]

        processor.backend_get = lambda path: calls.append(("backend_get", path)) or dates
        processor.get_chat_state = lambda chat_id: calls.append(("get_chat_state", chat_id)) or {"kept": True}
        processor.save_chat_state = lambda chat_id, state: calls.append(("save_chat_state", chat_id, state))
        processor.safe_send_message = lambda chat_id, text, reply_markup=None: calls.append(
            ("safe_send_message", chat_id, text, reply_markup)
        )

        processor.show_kiz_dates("chat-1")

        self.assertEqual([call[0] for call in calls], [
            "backend_get",
            "get_chat_state",
            "save_chat_state",
            "safe_send_message",
        ])
        self.assertEqual(calls[0], ("backend_get", "/api/v1/reports/kiz/dates"))
        self.assertEqual(calls[2][2]["kept"], True)
        self.assertEqual([item["date"] for item in calls[2][2]["kiz_dates"]], ["2026-06-01", "2026-06-02"])
        self.assertIn("Выберите дату отгрузки", calls[3][2])
        self.assertEqual(len(calls[3][3]["inline_keyboard"]), 3)

    def test_import_processor_runs_queue_with_characterized_ordered_calls(self):
        processor = TelegramImportProcessor()
        calls = []
        events = [
            {"id": "event-1", "payload": {"chat_id": "admin", "document": {"file_name": "one.xlsx"}}},
            {"id": "event-2", "payload": {"chat_id": "denied", "document": {"file_name": "two.xlsx"}}},
        ]

        processor.reset_stale_telegram_import_events = lambda: calls.append(("reset",))

        def take_next():
            calls.append(("take",))
            return events.pop(0) if events else None

        processor.take_next_telegram_import_event = take_next
        processor.is_admin_chat = lambda chat_id: calls.append(("authorize", chat_id)) or chat_id == "admin"
        processor.import_telegram_document = lambda chat_id, document, shipment_date="", event_id=None: (
            calls.append(("import", chat_id, document["file_name"], shipment_date, event_id)) or (True, "")
        )
        processor.finish_telegram_import_event = lambda event_id, success, error="": calls.append(
            ("finish", event_id, success, error)
        )

        self.assertEqual(processor.process_queued_telegram_imports(), 2)
        self.assertEqual(calls, [
            ("reset",),
            ("take",),
            ("authorize", "admin"),
            ("import", "admin", "one.xlsx", "", "event-1"),
            ("finish", "event-1", True, ""),
            ("take",),
            ("authorize", "denied"),
            ("finish", "event-2", False, "telegram import chat is not authorized"),
            ("take",),
        ])

    def test_unsafe_telegram_filename_is_rejected_before_download_and_redacted(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []
        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append(text)
        worker.download_telegram_document = mock.Mock()
        worker.backend_post = mock.Mock()
        unsafe_name = "../secret-customer.xlsx"

        result = worker.import_telegram_document(
            "123",
            {"file_name": unsafe_name, "file_id": "synthetic-file"},
        )

        self.assertEqual(result, (False, "unsafe_filename"))
        worker.download_telegram_document.assert_not_called()
        worker.backend_post.assert_not_called()
        self.assertNotIn(unsafe_name, "\n".join(messages))
        self.assertNotIn("secret-customer", "\n".join(messages))

    def test_telegram_configuration_requires_allowlist_and_admin_subset(self):
        with self.assertRaises(TelegramConfigurationError) as missing_allowed:
            validate_telegram_worker_config("synthetic-token", set(), set())
        self.assertEqual(missing_allowed.exception.setting_names, ("TELEGRAM_ALLOWED_CHAT_IDS",))

        with self.assertRaises(TelegramConfigurationError) as invalid_admin:
            validate_telegram_worker_config("synthetic-token", {"123"}, {"999"})
        self.assertEqual(invalid_admin.exception.setting_names, ("TELEGRAM_ADMIN_CHAT_IDS",))

        self.assertTrue(validate_telegram_worker_config(
            "synthetic-token",
            {"123", "999"},
            {"999"},
            {"123"},
            {"999"},
        ))

    def test_telegram_constructor_rejects_invalid_config_before_database_access(self):
        with mock.patch.dict(
            telegram_worker_module.os.environ,
            {"TELEGRAM_BOT_TOKEN": "synthetic-token"},
            clear=True,
        ), mock.patch.object(
            TelegramAdminProcessor,
            "load_offset",
            side_effect=AssertionError("database must not be touched"),
        ):
            with self.assertRaises(TelegramConfigurationError):
                TelegramWorker()

    def test_allowed_non_admin_cannot_enqueue_excel_document(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123", "999"}
        worker.admin_chat_ids = {"999"}
        messages = []
        worker.send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text))

        result = worker.enqueue_telegram_document(
            "123",
            {"file_name": "orders.xlsx", "file_id": "synthetic-file"},
            update_id=1,
        )

        self.assertFalse(result)
        self.assertIn("только администратору", messages[0][1])

    def test_allowed_non_admin_cannot_confirm_waiting_import_date(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123", "999"}
        worker.admin_chat_ids = {"999"}
        messages = []
        worker.send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text))
        worker.take_waiting_telegram_import_for_date = lambda chat_id: self.fail(
            "unauthorized date confirmation reached database lookup"
        )

        result = worker.confirm_waiting_telegram_import_shipment_date("123", "10.07.2026")

        self.assertFalse(result)
        self.assertIn("только администратору", messages[0][1])

    def test_queued_import_rechecks_current_admin_membership(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123", "999"}
        worker.admin_chat_ids = {"999"}
        events = [{
            "id": "11111111-1111-1111-1111-111111111111",
            "payload": {"chat_id": "123", "document": {"file_name": "orders.xlsx"}},
        }]
        finished = []
        worker.reset_stale_telegram_import_events = lambda: 0
        worker.take_next_telegram_import_event = lambda: events.pop(0) if events else None
        worker.import_telegram_document = lambda *args, **kwargs: self.fail("unauthorized import executed")
        worker.finish_telegram_import_event = lambda *args: finished.append(args)

        processed = worker.process_queued_telegram_imports()

        self.assertEqual(processed, 1)
        self.assertEqual(finished[0][1], False)
        self.assertEqual(finished[0][2], "telegram import chat is not authorized")

    def test_excel_file_to_import_payload_converts_rows_for_backend_import(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "orders_30_05_2026.xlsx"
            create_orders_workbook(path)

            payload = excel_file_to_import_payload(path, file_name=path.name, source="telegram")

        self.assertEqual(payload["source"], "telegram")
        self.assertEqual(payload["filename"], "orders_30_05_2026.xlsx")
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["meta"]["source_rows_count"], 1)
        row = payload["rows"][0]
        self.assertEqual(row["Дата отгрузки"], "30.05.2026")
        self.assertEqual(row["Клиент"], "Client One")
        self.assertEqual(row["Тип оплаты"], "Терминал")
        self.assertEqual(row["Товары"], "Product One")
        self.assertEqual(row["Кол-во ШТ"], 20)
        self.assertEqual(row["Кол-во блок"], 2)
        self.assertEqual(row["Источник файла"], "orders_30_05_2026.xlsx")

    def test_excel_file_to_import_payload_reads_delivery_date_from_upper_header(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "Шаблон_отправки_заказов_на_склад_04_06_2026.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Конструктор отчетов"
            sheet.append(["", "", "", "", "", "", "", "ИТОГО", "", "ДАТА ДОСТАВКИ"])
            sheet.append([
                "Торговый представитель",
                "Клиент",
                "Координаты клиента",
                "",
                "",
                "ТМЦ",
                "Тип оплаты",
                "Количество заказа",
                "Сумма с переоценкой",
                "",
            ])
            sheet.append([
                "ТП1",
                "Client One",
                "41.320075",
                "69.298547",
                "41.320075,69.298547",
                "Chapman Brown OP 20",
                "Перечисление",
                20,
                480000,
                "2026-06-05",
            ])
            workbook.save(path)

            payload = excel_file_to_import_payload(path, file_name=path.name, source="telegram")

        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["Дата отгрузки"], "05.06.2026")
        self.assertEqual(payload["meta"]["shipment_date"], "05.06.2026")
        self.assertEqual(payload["meta"]["shipment_dates"], ["05.06.2026"])

    def test_excel_file_to_import_payload_rejects_telegram_date_conflict_with_excel_date(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "Шаблон_отправки_заказов_на_склад_09_06_2026.xlsx"
            create_conflicting_date_workbook(path)

            with self.assertRaisesRegex(ValueError, "не совпадает с датой в Excel"):
                excel_file_to_import_payload(
                    path,
                    file_name=path.name,
                    source="telegram",
                    shipment_date="08.06.2026",
                )

    def test_telegram_worker_imports_document_through_backend_api(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "orders.xlsx"
            create_orders_workbook(source_path)
            calls = {}
            messages = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.token = "telegram-token"
            worker.timeout = 20
            worker.file_timeout = 120
            worker.max_file_size = 20 * 1024 * 1024

            def fake_send(chat_id, text):
                messages.append((chat_id, text))

            def fake_download(document, destination_path):
                shutil.copyfile(source_path, destination_path)

            def fake_backend_post(path, payload):
                calls["path"] = path
                calls["payload"] = payload
                return {
                    "status": "completed",
                    "items_created": len(payload["rows"]),
                    "orders_created": 1,
                    "duplicate_rows": 0,
                    "invalid_rows": 0,
                    "errors": [],
                    "backend_address_updates": 0,
                    "google_sheets_status": "completed",
                    "google_sheets_imported": len(payload["rows"]),
                    "google_sheets_duplicates": 0,
                    "google_sheets_updated": 0,
                    "google_sheets_error": "",
                }

            worker.send_message = fake_send
            worker.download_telegram_document = fake_download
            worker.backend_post = fake_backend_post

            worker.import_telegram_document("123", {"file_name": "orders.xlsx", "file_id": "file-1"}, shipment_date="30.05.2026")

        self.assertEqual(calls["path"], "/api/v1/imports")
        self.assertEqual(calls["payload"]["source"], "telegram")
        self.assertEqual(calls["payload"]["filename"], "orders.xlsx")
        self.assertEqual(calls["payload"]["telegram_chat_id"], "123")
        self.assertEqual(len(calls["payload"]["rows"]), 1)
        self.assertEqual(calls["payload"]["rows"][0]["Дата отгрузки"], "30.05.2026")
        self.assertEqual(messages[0][0], "123")
        self.assertIn("Начинаю импорт", messages[0][1])
        self.assertIn("Excel импортирован через Telegram", messages[-1][1])
        self.assertIn("Блоков импортировано: 2", messages[-1][1])
        self.assertIn("Адреса в backend обновлены: 0", messages[-1][1])
        self.assertIn("Google Sheets: записано 1, повторы 0, адреса обновлены 0", messages[-1][1])

    def test_telegram_worker_sends_pending_notification_event(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_admin_processor_module.SessionLocal
        try:
            with SessionLocal() as db:
                event = PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_NOTIFICATION_EVENT_TYPE,
                    status="pending",
                    payload={"chat_id": "123", "text": "Заказ отменён из-за недостатка товара"},
                )
                db.add(event)
                db.commit()

            sent = []
            observed_leases = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.allowed_chat_ids = {"123"}
            worker.admin_chat_ids = {"123"}

            def fake_send(chat_id, text):
                with SessionLocal() as observer:
                    leased = observer.execute(select(PendingEvent)).scalar_one()
                    observed_leases.append((leased.status, bool(leased.lease_owner)))
                sent.append((chat_id, text))

            worker.send_message = fake_send
            telegram_admin_processor_module.SessionLocal = SessionLocal

            with mock.patch.dict("os.environ", {"TAKSKLAD_EVENT_LEASES_ENABLED": "1"}, clear=False):
                processed = worker.process_pending_telegram_notifications()

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()

            self.assertEqual(processed, 1)
            self.assertEqual(observed_leases, [("processing", True)])
            self.assertEqual(sent, [("123", "Заказ отменён из-за недостатка товара")])
            self.assertEqual(event.status, "completed")
            self.assertEqual(event.last_error, "")
            self.assertIsNone(event.lease_owner)
            self.assertIsNotNone(event.completed_at)
        finally:
            telegram_admin_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_blocks_invalid_notification_events_without_retry(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_admin_processor_module.SessionLocal
        try:
            with SessionLocal() as db:
                db.add_all([
                    PendingEvent(
                        event_type=telegram_worker_module.TELEGRAM_NOTIFICATION_EVENT_TYPE,
                        status="pending",
                        payload={"chat_id": "123", "text": ""},
                    ),
                    PendingEvent(
                        event_type=telegram_worker_module.TELEGRAM_NOTIFICATION_EVENT_TYPE,
                        status="pending",
                        payload={"text": "Нет адресата"},
                    ),
                ])
                db.commit()

            sent = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.allowed_chat_ids = set()
            worker.admin_chat_ids = set()
            worker.send_message = lambda chat_id, text: sent.append((chat_id, text))
            telegram_admin_processor_module.SessionLocal = SessionLocal

            processed = worker.process_pending_telegram_notifications()

            with SessionLocal() as db:
                events = db.execute(
                    select(PendingEvent).order_by(PendingEvent.created_at, PendingEvent.id)
                ).scalars().all()
                audits = db.execute(select(AuditLog)).scalars().all()

            self.assertEqual(processed, 2)
            self.assertEqual(sent, [])
            self.assertEqual({event.status for event in events}, {"blocked"})
            self.assertEqual([event.attempts for event in events], [1, 1])
            blocked_by_error = {event.last_error: event for event in events}
            self.assertIn("telegram notification text is empty", blocked_by_error)
            self.assertIn("telegram notification target chat is empty", blocked_by_error)
            self.assertEqual([audit.action for audit in audits], ["telegram_notification_blocked", "telegram_notification_blocked"])
            self.assertEqual({audit.entity_id for audit in audits}, {str(event.id) for event in events})
        finally:
            telegram_admin_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_send_message_does_not_force_keyboard_by_default(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        calls = []

        worker.telegram_api_client = SimpleNamespace(
            send_message=lambda chat_id, text, reply_markup=None: calls.append(
                (chat_id, text, reply_markup)
            ) or {"ok": True},
        )

        worker.send_message("123", "hello")

        self.assertEqual(calls, [("123", "hello", None)])

    def test_telegram_worker_start_message_uses_hideable_reply_keyboard(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        calls = []

        def fake_send(chat_id, text, reply_markup=None):
            calls.append((chat_id, text, reply_markup))

        worker.send_message = fake_send

        worker.handle_update({
            "update_id": 10,
            "message": {
                "chat": {"id": 123},
                "text": "/start",
            },
        })

        self.assertEqual(calls[0][0], "123")
        self.assertIn("TakSklad backend online", calls[0][1])
        self.assertEqual(calls[0][2], telegram_main_reply_keyboard())

    def test_telegram_main_reply_keyboard_contains_user_buttons(self):
        keyboard = telegram_main_reply_keyboard()
        rows = keyboard["keyboard"]
        self.assertEqual(
            rows,
            [
                [{"text": TELEGRAM_BUTTON_LOGISTICS_REPORT}, {"text": TELEGRAM_BUTTON_KIZ_BY_FILES}],
                [{"text": TELEGRAM_BUTTON_STATUS}, {"text": TELEGRAM_BUTTON_IMPORTS}],
                [{"text": TELEGRAM_BUTTON_SHIPMENT_DATE}, {"text": TELEGRAM_BUTTON_MANUAL}],
            ],
        )
        self.assertTrue(keyboard["resize_keyboard"])
        self.assertNotIn("is_persistent", keyboard)

    def test_telegram_date_buttons_are_user_friendly(self):
        worker = TelegramWorker.__new__(TelegramWorker)

        logistics_keyboard = worker.logistics_date_keyboard(["2026-05-29", "2026-05-30"])
        self.assertEqual(logistics_keyboard["inline_keyboard"][0][0]["text"], "29.05.2026")
        self.assertEqual(logistics_keyboard["inline_keyboard"][0][0]["callback_data"], "logistics:2026-05-29")
        self.assertEqual(logistics_keyboard["inline_keyboard"][1][0]["text"], "30.05.2026")
        self.assertEqual(logistics_keyboard["inline_keyboard"][1][0]["callback_data"], "logistics:2026-05-30")

        self.assertEqual(display_date("2026-05-29"), "29.05.2026")
        self.assertEqual(display_date("29.05.2026"), "29.05.2026")
        self.assertEqual(display_date("не дата"), "не дата")

    def test_telegram_worker_limits_logistics_dates_menu_to_recent_seven_dates(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []

        def fake_backend_get(path, params=None):
            self.assertEqual(path, "/api/v1/logistics/dates")
            return [f"2026-06-{day:02d}" for day in range(1, 11)]

        def fake_send(chat_id, text, reply_markup=None):
            messages.append((chat_id, text, reply_markup))

        worker.backend_get = fake_backend_get
        worker.safe_send_message = fake_send

        worker.show_logistics_dates("123")

        keyboard = messages[0][2]["inline_keyboard"]
        self.assertEqual(len(keyboard), 7)
        self.assertEqual(keyboard[0][0]["callback_data"], "logistics:2026-06-04")
        self.assertEqual(keyboard[-1][0]["callback_data"], "logistics:2026-06-10")
        self.assertNotIn("01.06.2026", str(keyboard))
        self.assertIn("10.06.2026", str(keyboard))

    def test_telegram_worker_shows_single_logistics_date_as_menu(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []
        reports = []

        worker.backend_get = lambda path, params=None: ["2026-06-29"]
        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
        worker.send_logistics_report = lambda chat_id, shipment_date: reports.append((chat_id, shipment_date))

        worker.show_logistics_dates("123")

        self.assertEqual(reports, [])
        keyboard = messages[0][2]["inline_keyboard"]
        self.assertEqual(keyboard, [[{"text": "29.06.2026", "callback_data": "logistics:2026-06-29"}]])

    def test_telegram_worker_refreshes_logistics_dates_on_each_open(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []
        calls = []
        responses = [
            ["2026-06-28", "2026-06-29"],
            ["2026-06-29", "2026-06-30"],
        ]

        def fake_backend_get(path, params=None):
            self.assertEqual(path, "/api/v1/logistics/dates")
            calls.append(path)
            return responses[len(calls) - 1]

        worker.backend_get = fake_backend_get
        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))

        worker.show_logistics_dates("123")
        worker.show_logistics_dates("123")

        self.assertEqual(calls, ["/api/v1/logistics/dates", "/api/v1/logistics/dates"])
        first_keyboard = messages[0][2]["inline_keyboard"]
        second_keyboard = messages[1][2]["inline_keyboard"]
        self.assertEqual(first_keyboard[-1][0]["callback_data"], "logistics:2026-06-29")
        self.assertEqual(second_keyboard[-1][0]["callback_data"], "logistics:2026-06-30")

    def test_telegram_worker_send_document_does_not_force_bottom_reply_keyboard(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.token = "telegram-token"
        worker.file_timeout = 120
        captured = {}
        original_client = telegram_clients_module.httpx.Client

        class FakeResponse:
            def raise_for_status(self):
                return None

            def json(self):
                return {"ok": True, "result": {"message_id": 1}}

        class FakeClient:
            def __init__(self, timeout=None, **_kwargs):
                self.timeout = timeout

            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def post(self, url, data=None, files=None, headers=None):
                captured["url"] = url
                captured["data"] = data
                captured["files"] = files
                captured["headers"] = headers
                return FakeResponse()

        try:
            telegram_clients_module.httpx.Client = FakeClient

            result = worker.send_document("123", b"excel", "orders.xlsx", caption="done")
        finally:
            telegram_clients_module.httpx.Client = original_client

        self.assertEqual(result, {"message_id": 1})
        self.assertEqual(captured["data"]["chat_id"], "123")
        self.assertEqual(captured["data"]["caption"], "done")
        self.assertNotIn("reply_markup", captured["data"])
        self.assertEqual(captured["files"]["document"][0], "orders.xlsx")
        self.assertEqual(len(captured["headers"]["X-Correlation-ID"]), 36)

    def test_telegram_worker_clears_public_command_menu_once(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        calls = []

        worker.telegram_api_client = SimpleNamespace(
            configure_menu=lambda: calls.extend([
                ("deleteMyCommands", {}),
                ("setChatMenuButton", {"menu_button": {"type": "default"}}),
            ]),
        )

        worker.ensure_bot_menu()
        worker.ensure_bot_menu()

        self.assertEqual(len(calls), 2)
        self.assertEqual(calls[0], ("deleteMyCommands", {}))
        self.assertEqual(calls[1], ("setChatMenuButton", {"menu_button": {"type": "default"}}))
        self.assertTrue(worker.bot_menu_ready)

    def test_telegram_worker_menu_command_shows_hideable_reply_keyboard(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        calls = []

        def fake_send(chat_id, text, reply_markup=None):
            calls.append((chat_id, text, reply_markup))

        worker.send_message = fake_send

        worker.handle_update({
            "update_id": 11,
            "message": {
                "chat": {"id": 123},
                "text": "/menu",
            },
        })

        self.assertEqual(len(calls), 1)
        self.assertEqual(calls[0][0], "123")
        self.assertIn("TakSklad backend online", calls[0][1])
        self.assertEqual(calls[0][2], telegram_main_reply_keyboard())

    def test_telegram_worker_unknown_text_opens_menu_instead_of_dead_button_error(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        calls = []

        def fake_send(chat_id, text, reply_markup=None):
            calls.append((chat_id, text, reply_markup))

        worker.send_message = fake_send
        worker.confirm_waiting_telegram_import_shipment_date = lambda chat_id, text: False

        worker.handle_update({
            "update_id": 13,
            "message": {
                "chat": {"id": 123},
                "text": "Старая неизвестная кнопка",
            },
        })

        self.assertIn("Команда не распознана", calls[0][1])
        self.assertEqual(calls[0][2], telegram_main_reply_keyboard())

    def test_telegram_worker_handles_bottom_logistics_button(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        calls = []

        def fake_show_logistics_dates(chat_id):
            calls.append(chat_id)

        worker.show_logistics_dates = fake_show_logistics_dates

        worker.handle_update({
            "update_id": 10,
            "message": {
                "chat": {"id": 123},
                "text": TELEGRAM_BUTTON_LOGISTICS_REPORT,
            },
        })

        self.assertEqual(calls, ["123"])

    def test_telegram_worker_handles_status_button(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        messages = []
        calls = []

        def fake_backend_get(path, params=None):
            calls.append(path)
            if path == "/api/v1/reports/day":
                return {
                    "report_date": "2026-05-31",
                    "totals": {
                        "orders": 8,
                        "completed_orders": 3,
                        "active_orders": 5,
                        "planned_blocks": 20,
                        "scanned_blocks": 7,
                        "scanned_today": 4,
                        "remaining_blocks": 13,
                        "scan_codes": 7,
                        "total_price": 4_800_000,
                    },
                }
            if path == "/api/v1/orders/active":
                return [
                    {
                        "order_date": "2026-06-02",
                        "skladbot_request_number": "WH-R-1",
                        "items": [
                            {"quantity_blocks": 2, "scanned_blocks": 1, "line_total": 480000},
                        ],
                    },
                    {
                        "order_date": "2026-06-03",
                        "skladbot_request_number": "",
                        "skladbot_request_id": "",
                        "items": [
                            {"quantity_blocks": 3, "scanned_blocks": 0, "line_total": 720000},
                        ],
                    },
                ]
            raise AssertionError(path)

        def fake_send(chat_id, text, reply_markup=None):
            messages.append((chat_id, text, reply_markup))

        worker.backend_get = fake_backend_get
        worker.safe_send_message = fake_send

        worker.handle_update({
            "update_id": 12,
            "message": {
                "chat": {"id": 123},
                "text": TELEGRAM_BUTTON_STATUS,
            },
        })

        self.assertEqual(messages[0][0], "123")
        self.assertEqual(calls, ["/api/v1/reports/day", "/api/v1/orders/active"])
        self.assertIn("Статус TakSklad за 31.05.2026", messages[0][1])
        self.assertIn("КИЗов сегодня: 4", messages[0][1])
        self.assertIn("02.06.2026: 1 заказов, 1/2 блоков", messages[0][1])
        self.assertIn("03.06.2026: 1 заказов, 0/3 блоков", messages[0][1])
        self.assertIn("Без номера SkladBot: 1", messages[0][1])

    def test_status_summary_groups_active_orders_by_shipment_date(self):
        summary = summarize_active_orders_by_date([
            {
                "order_date": "2026-06-02",
                "skladbot_request_number": "WH-R-1",
                "items": [
                    {"quantity_blocks": 2, "scanned_blocks": 1, "line_total": 480000},
                    {"quantity_blocks": 1, "scanned_blocks": 1, "line_total": 240000},
                ],
            },
            {
                "order_date": "2026-06-02",
                "skladbot_request_number": "",
                "skladbot_request_id": "",
                "items": [
                    {"quantity_blocks": 3, "scanned_blocks": 0, "line_total": 720000},
                ],
            },
        ])

        self.assertEqual(summary["2026-06-02"]["orders"], 2)
        self.assertEqual(summary["2026-06-02"]["items"], 3)
        self.assertEqual(summary["2026-06-02"]["planned_blocks"], 6)
        self.assertEqual(summary["2026-06-02"]["scanned_blocks"], 2)
        self.assertEqual(summary["2026-06-02"]["remaining_blocks"], 4)
        self.assertEqual(summary["2026-06-02"]["missing_skladbot"], 1)
        self.assertEqual(summary["2026-06-02"]["total_price"], 1_440_000)

    def test_telegram_worker_handles_inline_logistics_callback(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        answered = []
        calls = []

        def fake_answer(callback_query_id, text=""):
            answered.append((callback_query_id, text))

        def fake_send_report(chat_id, shipment_date):
            calls.append((chat_id, shipment_date))
            return True

        worker.answer_callback_query = fake_answer
        worker.send_logistics_report = fake_send_report

        worker.handle_update({
            "update_id": 20,
            "callback_query": {
                "id": "cb-1",
                "data": "logistics:2026-05-29",
                "message": {"chat": {"id": 123}},
            },
        })

        self.assertEqual(answered, [("cb-1", "")])
        self.assertEqual(calls, [("123", "2026-05-29")])

    def test_telegram_worker_handles_inline_kiz_file_callback(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        answered = []
        calls = []

        def fake_answer(callback_query_id, text=""):
            answered.append((callback_query_id, text))

        def fake_send_by_index(chat_id, text):
            calls.append((chat_id, text))
            return True

        worker.answer_callback_query = fake_answer
        worker.send_kiz_source_file_by_index = fake_send_by_index

        worker.handle_update({
            "update_id": 21,
            "callback_query": {
                "id": "cb-2",
                "data": "kiz_file:2",
                "message": {"chat": {"id": 123}},
            },
        })

        self.assertEqual(answered, [("cb-2", "")])
        self.assertEqual(calls, [("123", "2")])

    def test_telegram_worker_handles_inline_kiz_date_callback(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        answered = []
        calls = []

        def fake_answer(callback_query_id, text=""):
            answered.append((callback_query_id, text))

        def fake_send_date(chat_id, shipment_date):
            calls.append((chat_id, shipment_date))
            return True

        worker.answer_callback_query = fake_answer
        worker.send_kiz_date_report = fake_send_date

        worker.handle_update({
            "update_id": 22,
            "callback_query": {
                "id": "cb-3",
                "data": "kiz_date:2026-05-30",
                "message": {"chat": {"id": 123}},
            },
        })

        self.assertEqual(answered, [("cb-3", "")])
        self.assertEqual(calls, [("123", "2026-05-30")])

    def test_telegram_worker_handles_inline_kiz_range_callback(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        answered = []
        calls = []

        def fake_answer(callback_query_id, text=""):
            answered.append((callback_query_id, text))

        def fake_send_range(chat_id, date_from, date_to):
            calls.append((chat_id, date_from, date_to))
            return True

        worker.answer_callback_query = fake_answer
        worker.send_kiz_range_report = fake_send_range

        worker.handle_update({
            "update_id": 23,
            "callback_query": {
                "id": "cb-4",
                "data": f"{TELEGRAM_KIZ_RANGE_CALLBACK_PREFIX}2026-06-01:2026-06-10",
                "message": {"chat": {"id": 123}},
            },
        })

        self.assertEqual(answered, [("cb-4", "")])
        self.assertEqual(calls, [("123", "2026-06-01", "2026-06-10")])

    def test_telegram_worker_handles_main_menu_callbacks(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        answered = []
        calls = []

        worker.answer_callback_query = lambda callback_id, text="": answered.append((callback_id, text))
        worker.send_date_help = lambda chat_id: calls.append(("date", chat_id))
        worker.show_logistics_dates = lambda chat_id: calls.append(("logistics", chat_id))
        worker.show_kiz_export_menu = lambda chat_id: calls.append(("kiz", chat_id))
        worker.send_status_report = lambda chat_id: calls.append(("status", chat_id))
        worker.send_imports_report = lambda chat_id: calls.append(("imports", chat_id))
        worker.show_manual_menu = lambda chat_id: calls.append(("manual", chat_id))

        for index, data in enumerate([
            "menu:date",
            "menu:logistics",
            "menu:kiz",
            "menu:status",
            "menu:imports",
            "menu:manual",
        ], start=1):
            worker.handle_update({
                "update_id": 30 + index,
                "callback_query": {
                    "id": f"cb-menu-{index}",
                    "data": data,
                    "message": {"chat": {"id": 123}},
                },
            })

        self.assertEqual(
            answered,
            [
                ("cb-menu-1", ""),
                ("cb-menu-2", ""),
                ("cb-menu-3", ""),
                ("cb-menu-4", ""),
                ("cb-menu-5", ""),
                ("cb-menu-6", ""),
            ],
        )
        self.assertEqual(
            calls,
            [
                ("date", "123"),
                ("logistics", "123"),
                ("kiz", "123"),
                ("status", "123"),
                ("imports", "123"),
                ("manual", "123"),
            ],
        )

    def test_telegram_worker_manual_add_order_imports_through_backend(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        state = {}
        messages = []
        posts = []
        answered = []

        worker.answer_callback_query = lambda callback_id, text="": answered.append((callback_id, text))
        worker.get_chat_state = lambda chat_id: dict(state)
        worker.save_chat_state = lambda chat_id, payload: state.clear() or state.update(payload)
        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))

        def fake_backend_post(path, payload=None):
            posts.append((path, payload))
            return {
                "id": "import-1",
                "orders_created": 1,
                "items_created": 1,
                "skladbot_dry_run_status": "queued",
            }

        worker.backend_post = fake_backend_post

        worker.handle_update({
            "update_id": 101,
            "callback_query": {"id": "cb-1", "data": "manual:add", "message": {"chat": {"id": 123}}},
        })
        worker.handle_update({"update_id": 102, "message": {"chat": {"id": 123}, "text": "12.06.2026"}})
        worker.handle_update({
            "update_id": 103,
            "callback_query": {"id": "cb-2", "data": "manual:payment:terminal", "message": {"chat": {"id": 123}}},
        })
        worker.handle_update({"update_id": 104, "message": {"chat": {"id": 123}, "text": "ООО Ручной Клиент"}})
        worker.handle_update({"update_id": 105, "message": {"chat": {"id": 123}, "text": "41.311081, 69.240562"}})
        worker.handle_update({"update_id": 106, "message": {"chat": {"id": 123}, "text": "ТП1 Тест"}})
        worker.handle_update({
            "update_id": 107,
            "callback_query": {"id": "cb-3", "data": "manual:product:brown_op", "message": {"chat": {"id": 123}}},
        })
        worker.handle_update({"update_id": 108, "message": {"chat": {"id": 123}, "text": "3"}})
        worker.handle_update({
            "update_id": 109,
            "callback_query": {"id": "cb-4", "data": "manual:create", "message": {"chat": {"id": 123}}},
        })

        self.assertEqual([item[0] for item in answered], ["cb-1", "cb-2", "cb-3", "cb-4"])
        self.assertEqual(posts[0][0], "/api/v1/imports")
        payload = posts[0][1]
        self.assertEqual(payload["source"], "telegram_manual")
        self.assertEqual(payload["telegram_chat_id"], "123")
        self.assertEqual(len(payload["rows"]), 1)
        row = payload["rows"][0]
        self.assertEqual(row["Дата отгрузки"], "12.06.2026")
        self.assertEqual(row["Тип оплаты"], "Терминал")
        self.assertEqual(row["Клиент"], "ООО Ручной Клиент")
        self.assertEqual(row["Адрес"], "Адрес не указан")
        self.assertEqual(row["Координаты"], "41.311081, 69.240562")
        self.assertEqual(row["Торговый представитель"], "ТП1 Тест")
        self.assertEqual(row["Товары"], "Chapman Brown OP 20")
        self.assertEqual(row["Кол-во блок"], 3)
        self.assertEqual(row["Кол-во ШТ"], 30)
        self.assertEqual(state.get("manual_flow"), {})
        self.assertIn("Заказ создан", messages[-1][1])

    def test_telegram_worker_manual_controls_are_admin_only_when_configured(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"999"}
        answered = []
        messages = []

        worker.answer_callback_query = lambda callback_id, text="": answered.append((callback_id, text))
        worker.send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
        worker.start_manual_add_order = lambda chat_id: self.fail("manual add must not start for non-admin")
        worker.show_manual_delete_orders = lambda chat_id: self.fail("manual delete must not start for non-admin")

        worker.handle_update({
            "update_id": 110,
            "callback_query": {"id": "cb-add", "data": "manual:add", "message": {"chat": {"id": 123}}},
        })
        worker.handle_update({
            "update_id": 111,
            "callback_query": {"id": "cb-delete", "data": "manual:delete", "message": {"chat": {"id": 123}}},
        })

        self.assertEqual([item[0] for item in answered], ["cb-add", "cb-delete"])
        self.assertEqual(len(messages), 2)
        self.assertTrue(all("только администратору" in item[1] for item in messages))

    def test_telegram_worker_manual_delete_active_order_calls_safe_backend_endpoint(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        state = {}
        messages = []
        posts = []

        active_order = {
            "id": "11111111-1111-1111-1111-111111111111",
            "order_date": "2026-06-12",
            "client": "Delete Client",
            "payment_type": "Терминал",
            "address": "Delete Address",
            "skladbot_request_number": "WH-R-123",
            "items": [{
                "product": "Chapman Brown OP 20",
                "quantity_blocks": 2,
                "scanned_blocks": 0,
                "scan_codes": [],
            }],
        }

        worker.get_chat_state = lambda chat_id: dict(state)
        worker.save_chat_state = lambda chat_id, payload: state.clear() or state.update(payload)
        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
        worker.answer_callback_query = lambda callback_id, text="": None
        worker.backend_get = lambda path, params=None: [active_order]

        def fake_backend_post(path, payload=None):
            posts.append((path, payload))
            return {
                "order_id": active_order["id"],
                "deleted": True,
                "skladbot_request_number": "WH-R-123",
            }

        worker.backend_post = fake_backend_post

        worker.show_manual_delete_orders("123")
        worker.handle_update({
            "update_id": 201,
            "callback_query": {"id": "cb-1", "data": "manual:delete:1", "message": {"chat": {"id": 123}}},
        })
        worker.handle_update({
            "update_id": 202,
            "callback_query": {
                "id": "cb-2",
                "data": f"manual:delete_confirm:{active_order['id']}",
                "message": {"chat": {"id": 123}},
            },
        })

        self.assertEqual(posts[0][0], f"/api/v1/admin/orders/{active_order['id']}/delete-active")
        self.assertEqual(posts[0][1]["actor"], "telegram")
        self.assertEqual(posts[0][1]["source"], "telegram")
        self.assertEqual(posts[0][1]["idempotency_key"], f"telegram:manual_delete:123:{active_order['id']}")
        self.assertIn("Заказ удалён из TakSklad", messages[-1][1])
        self.assertIn("WH-R-123", messages[-1][1])
        self.assertIn("осталась", messages[-1][1])

    def test_telegram_worker_manual_delete_refuses_started_order_before_backend_call(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        messages = []
        posts = []
        state = {
            "manual_delete_orders": [{
                "id": "11111111-1111-1111-1111-111111111111",
                "client": "Started Client",
                "items": [{
                    "product": "Chapman Brown OP 20",
                    "quantity_blocks": 2,
                    "scanned_blocks": 1,
                    "scan_codes": ["0101"],
                }],
            }],
        }

        worker.get_chat_state = lambda chat_id: dict(state)
        worker.save_chat_state = lambda chat_id, payload: state.clear() or state.update(payload)
        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
        worker.answer_callback_query = lambda callback_id, text="": None
        worker.backend_post = lambda path, payload=None: posts.append((path, payload))

        worker.handle_update({
            "update_id": 203,
            "callback_query": {"id": "cb-1", "data": "manual:delete:1", "message": {"chat": {"id": 123}}},
        })
        worker.handle_update({
            "update_id": 204,
            "callback_query": {
                "id": "cb-2",
                "data": "manual:delete_confirm:11111111-1111-1111-1111-111111111111",
                "message": {"chat": {"id": 123}},
            },
        })

        self.assertEqual(posts, [])
        self.assertIn("уже начал обрабатывать", messages[-1][1])

    def test_telegram_worker_unknown_inline_callback_opens_fresh_menu(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        answered = []
        messages = []

        worker.answer_callback_query = lambda callback_id, text="": answered.append((callback_id, text))
        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))

        worker.handle_update({
            "update_id": 36,
            "callback_query": {
                "id": "cb-old",
                "data": "old:button",
                "message": {"chat": {"id": 123}},
            },
        })

        self.assertEqual(answered, [("cb-old", "")])
        self.assertIn("Кнопка устарела", messages[0][1])
        self.assertEqual(messages[0][2], telegram_main_reply_keyboard())

    def test_telegram_worker_reports_logistics_backend_error_to_user(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []

        def fake_backend_get_bytes(path, params=None):
            request = httpx.Request("GET", "http://backend-api/api/v1/logistics/report")
            response = httpx.Response(
                409,
                json={"detail": "Missing coordinates for logistics report: Client One"},
                request=request,
            )
            raise httpx.HTTPStatusError("backend rejected report", request=request, response=response)

        def fake_send_message(chat_id, text, reply_markup=None):
            messages.append((chat_id, text))

        worker.backend_get_bytes = fake_backend_get_bytes
        worker.safe_send_message = fake_send_message

        worker.send_logistics_report("123", "2026-05-29")

        self.assertEqual(messages[0][0], "123")
        self.assertIn("Не удалось выгрузить отчёт логистики за 29.05.2026", messages[0][1])
        self.assertIn("Missing coordinates for logistics report: Client One", messages[0][1])

    def test_telegram_worker_sends_logistics_report_bytes_from_backend(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        sent_documents = []
        report_content = b"xlsx-content"

        def fake_backend_get_bytes(path, params=None):
            self.assertEqual(path, "/api/v1/logistics/report")
            self.assertEqual(params, {"shipment_date": "2026-05-29"})
            return report_content, {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}

        def fake_send_document(chat_id, content, filename, caption=""):
            sent_documents.append((chat_id, content, filename, caption))
            return {"message_id": 1}

        worker.backend_get_bytes = fake_backend_get_bytes
        worker.safe_send_document = fake_send_document

        result = worker.send_logistics_report("123", "29.05.2026")

        self.assertTrue(result)
        self.assertEqual(sent_documents, [(
            "123",
            report_content,
            "TakSklad_логистика_29.05.2026.xlsx",
            "Отчёт логистики за 29.05.2026",
        )])

    def test_telegram_worker_saves_shipment_date_from_message(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        saved = []
        messages = []

        def fake_set_chat_shipment_date(chat_id, shipment_date):
            saved.append((chat_id, shipment_date))

        def fake_send(chat_id, text):
            messages.append((chat_id, text))

        worker.set_chat_shipment_date = fake_set_chat_shipment_date
        worker.send_message = fake_send
        worker.confirm_waiting_telegram_import_shipment_date = lambda chat_id, shipment_date: False

        worker.handle_update({
            "update_id": 11,
            "message": {
                "chat": {"id": 123},
                "text": "29.05.2026",
            },
        })

        self.assertEqual(saved, [("123", "29.05.2026")])
        self.assertIn("Дата сохранена: 29.05.2026", messages[0][1])
        self.assertIn("бот всё равно спросит дату", messages[0][1])

    def test_telegram_worker_enqueues_excel_document_from_message(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        calls = []

        def fake_enqueue(chat_id, document, update_id=None, shipment_date=""):
            calls.append((chat_id, document, update_id, shipment_date))
            return True

        worker.enqueue_telegram_document = fake_enqueue

        worker.handle_update({
            "update_id": 77,
            "message": {
                "chat": {"id": 123},
                "document": {"file_name": "orders.xlsx", "file_id": "file-1"},
            },
        })

        self.assertEqual(calls, [("123", {"file_name": "orders.xlsx", "file_id": "file-1"}, 77, "")])

    def test_telegram_worker_enqueues_document_waiting_for_manual_shipment_date(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_import_processor_module.SessionLocal
        try:
            messages = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.allowed_chat_ids = {"123"}
            worker.admin_chat_ids = {"123"}
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
            telegram_import_processor_module.SessionLocal = SessionLocal

            result = worker.enqueue_telegram_document(
                "123",
                {"file_name": "orders.xlsx", "file_id": "file-1"},
                update_id=77,
            )

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()
                payload = event.payload

            self.assertTrue(result)
            self.assertEqual(event.status, telegram_worker_module.TELEGRAM_EXCEL_IMPORT_WAITING_SHIPMENT_DATE_STATUS)
            self.assertEqual(payload["shipment_date"], "")
            self.assertEqual(payload["shipment_date_source"], "")
            self.assertIn("Укажите дату отгрузки", messages[0][1])
            self.assertIn("ДД.ММ.ГГГГ", messages[0][1])
        finally:
            telegram_import_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_manual_date_moves_waiting_import_to_pending(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_import_processor_module.SessionLocal
        try:
            with SessionLocal() as db:
                event = PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_EVENT_TYPE,
                    status=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_WAITING_SHIPMENT_DATE_STATUS,
                    payload={
                        "chat_id": "123",
                        "file_name": "orders.xlsx",
                        "document": {"file_name": "orders.xlsx", "file_id": "file-1"},
                        "shipment_date": "",
                    },
                )
                db.add(event)
                db.commit()

            messages = []
            processed = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.allowed_chat_ids = {"123"}
            worker.admin_chat_ids = {"123"}
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
            worker.process_queued_telegram_imports = lambda: processed.append(True)
            telegram_import_processor_module.SessionLocal = SessionLocal

            worker.handle_update({
                "update_id": 78,
                "message": {
                    "chat": {"id": 123},
                    "text": "09.06.2026",
                },
            })

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()
                payload = event.payload

            self.assertEqual(event.status, "pending")
            self.assertEqual(event.last_error, "")
            self.assertEqual(payload["shipment_date"], "09.06.2026")
            self.assertEqual(payload["shipment_date_source"], "telegram_manual_input")
            self.assertEqual(processed, [True])
            self.assertIn("Дата принята", messages[0][1])
        finally:
            telegram_import_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_new_excel_supersedes_old_waiting_import_for_chat(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_import_processor_module.SessionLocal
        try:
            now = datetime.now(timezone.utc)
            with SessionLocal() as db:
                old_event = PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_EVENT_TYPE,
                    status=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_WAITING_SHIPMENT_DATE_STATUS,
                    payload={
                        "chat_id": "123",
                        "file_name": "old.xlsx",
                        "document": {"file_name": "old.xlsx", "file_id": "old-file"},
                        "shipment_date": "",
                    },
                    created_at=now - timedelta(minutes=10),
                    updated_at=now - timedelta(minutes=10),
                )
                db.add(old_event)
                db.commit()
                old_event_id = old_event.id

            messages = []
            processed = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.allowed_chat_ids = {"123"}
            worker.admin_chat_ids = {"123"}
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
            worker.process_queued_telegram_imports = lambda: processed.append(True)
            telegram_import_processor_module.SessionLocal = SessionLocal

            worker.enqueue_telegram_document(
                "123",
                {"file_name": "new.xlsx", "file_id": "new-file"},
                update_id=80,
            )
            worker.confirm_waiting_telegram_import_shipment_date("123", "26.06.2026")

            with SessionLocal() as db:
                old_event = db.get(PendingEvent, old_event_id)
                events = db.execute(
                    select(PendingEvent).order_by(PendingEvent.created_at, PendingEvent.id)
                ).scalars().all()
                new_event = next(event for event in events if (event.payload or {}).get("file_name") == "new.xlsx")

            self.assertEqual(old_event.status, "cancelled")
            self.assertEqual(old_event.last_error, "superseded_by_new_telegram_excel_file")
            self.assertEqual(new_event.status, "pending")
            self.assertEqual(new_event.payload["shipment_date"], "26.06.2026")
            self.assertEqual(processed, [True])
            self.assertIn("Файл: new.xlsx", messages[-1][1])
            self.assertNotIn("Файл: old.xlsx", messages[-1][1])
        finally:
            telegram_import_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_repeats_prompt_for_invalid_manual_import_date(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_import_processor_module.SessionLocal
        try:
            with SessionLocal() as db:
                event = PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_EVENT_TYPE,
                    status=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_WAITING_SHIPMENT_DATE_STATUS,
                    payload={
                        "chat_id": "123",
                        "file_name": "orders.xlsx",
                        "document": {"file_name": "orders.xlsx", "file_id": "file-1"},
                        "shipment_date": "",
                    },
                )
                db.add(event)
                db.commit()

            messages = []
            processed = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.allowed_chat_ids = {"123"}
            worker.admin_chat_ids = {"123"}
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
            worker.process_queued_telegram_imports = lambda: processed.append(True)
            telegram_import_processor_module.SessionLocal = SessionLocal

            worker.handle_update({
                "update_id": 79,
                "message": {
                    "chat": {"id": 123},
                    "text": "завтра",
                },
            })

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()

            self.assertEqual(event.status, telegram_worker_module.TELEGRAM_EXCEL_IMPORT_WAITING_SHIPMENT_DATE_STATUS)
            self.assertEqual(processed, [])
            self.assertIn("Ожидаю дату отгрузки", messages[0][1])
            self.assertIn("ДД.ММ.ГГГГ", messages[0][1])
        finally:
            telegram_import_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_processes_multiple_queued_imports(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        events = [
            {"id": "event-1", "payload": {"chat_id": "123", "document": {"file_name": "a.xlsx"}}},
            {"id": "event-2", "payload": {"chat_id": "123", "document": {"file_name": "b.xlsx"}}},
        ]
        imported = []
        finished = []

        def fake_take_next():
            return events.pop(0) if events else None

        def fake_import(chat_id, document, shipment_date="", event_id=None):
            imported.append((chat_id, document["file_name"], shipment_date))
            return True, ""

        def fake_finish(event_id, success, error=""):
            finished.append((event_id, success, error))

        worker.take_next_telegram_import_event = fake_take_next
        worker.import_telegram_document = fake_import
        worker.finish_telegram_import_event = fake_finish
        worker.reset_stale_telegram_import_events = lambda: 0

        processed = worker.process_queued_telegram_imports()

        self.assertEqual(processed, 2)
        self.assertEqual(imported, [("123", "a.xlsx", ""), ("123", "b.xlsx", "")])
        self.assertEqual(finished, [("event-1", True, ""), ("event-2", True, "")])

    def test_telegram_worker_failed_backend_import_returns_failed_and_sends_actionable_redacted_message(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []
        backend_payloads = []
        event_id = "11111111-1111-1111-1111-111111111111"

        def fake_backend_post(path, payload):
            backend_payloads.append((path, payload))
            return {
                "status": "failed",
                "errors": ["Bearer secret-service-token failed for 0104006396053978217SECRETKIZVALUE"],
            }

        original_excel_to_payload = telegram_import_processor_module.excel_file_to_import_payload
        try:
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
            worker.download_telegram_document = lambda document, temp_path: Path(temp_path).write_bytes(b"xlsx")
            worker.backend_post = fake_backend_post
            telegram_import_processor_module.excel_file_to_import_payload = lambda *args, **kwargs: {
                "rows": [{"Клиент": "Broken"}],
                "meta": {"source_rows_count": 1, "shipment_date": "09.06.2026"},
            }

            result = worker.import_telegram_document(
                "123",
                {"file_name": "broken.xlsx", "file_id": "file-1"},
                shipment_date="09.06.2026",
                event_id=event_id,
            )
        finally:
            telegram_import_processor_module.excel_file_to_import_payload = original_excel_to_payload

        self.assertEqual(result[0], False)
        self.assertEqual(backend_payloads[0][0], "/api/v1/imports")
        self.assertEqual(backend_payloads[0][1]["telegram_event_id"], event_id)
        text = messages[1][1]
        self.assertIn("Файл: broken.xlsx", text)
        self.assertIn("Причина: Bearer ***", text)
        self.assertIn("Что сделать:", text)
        self.assertIn("Заказы и заявки SkladBot не созданы.", text)
        self.assertNotIn("secret-service-token", text)
        self.assertNotIn("0104006396053978217SECRETKIZVALUE", text)

    def test_telegram_worker_recovers_completed_import_after_backend_post_timeout(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.timeout = 20
        worker.import_timeout = 120
        messages = []
        event_id = "11111111-1111-1111-1111-111111111111"

        def fake_backend_post(path, payload):
            self.assertEqual(path, "/api/v1/imports")
            self.assertEqual(payload["telegram_event_id"], event_id)
            raise httpx.ReadTimeout("response timed out")

        def fake_backend_get(path, params=None):
            self.assertEqual(path, "/api/v1/imports")
            return [{
                "id": "import-1",
                "source": "telegram",
                "status": "completed",
                "rows_total": 1,
                "rows_imported": 1,
                "raw_payload": {
                    "filename": "orders.xlsx",
                    "telegram_event_id": event_id,
                    "orders_created": 1,
                    "items_created": 1,
                    "duplicate_rows": 0,
                    "invalid_rows": 0,
                    "backend_address_updates": 0,
                    "errors": [],
                    "google_sheets": {"status": "disabled"},
                },
            }]

        original_excel_to_payload = telegram_import_processor_module.excel_file_to_import_payload
        try:
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
            worker.download_telegram_document = lambda document, temp_path: Path(temp_path).write_bytes(b"xlsx")
            worker.backend_post = fake_backend_post
            worker.backend_get = fake_backend_get
            telegram_import_processor_module.excel_file_to_import_payload = lambda *args, **kwargs: {
                "rows": [{"Клиент": "Client", "Кол-во блок": 2}],
                "meta": {"source_rows_count": 1, "shipment_date": "09.06.2026"},
            }

            result = worker.import_telegram_document(
                "123",
                {"file_name": "orders.xlsx", "file_id": "file-1"},
                shipment_date="09.06.2026",
                event_id=event_id,
            )
        finally:
            telegram_import_processor_module.excel_file_to_import_payload = original_excel_to_payload

        self.assertEqual(result, (True, ""))
        text = messages[-1][1]
        self.assertIn("Excel импортирован через Telegram", text)
        self.assertIn("Позиции добавлены: 1", text)
        self.assertIn("Заказы добавлены: 1", text)
        self.assertIn("результат подтверждён через историю импортов", text)
        self.assertNotIn("Заказы и заявки SkladBot не созданы", text)

    def test_telegram_worker_import_timeout_without_readback_uses_unconfirmed_message(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.timeout = 20
        worker.import_timeout = 120
        messages = []

        def fake_backend_post(path, payload):
            raise httpx.ReadTimeout("response timed out")

        def fake_backend_get(path, params=None):
            return []

        original_excel_to_payload = telegram_import_processor_module.excel_file_to_import_payload
        try:
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
            worker.download_telegram_document = lambda document, temp_path: Path(temp_path).write_bytes(b"xlsx")
            worker.backend_post = fake_backend_post
            worker.backend_get = fake_backend_get
            telegram_import_processor_module.excel_file_to_import_payload = lambda *args, **kwargs: {
                "rows": [{"Клиент": "Client", "Кол-во блок": 2}],
                "meta": {"source_rows_count": 1, "shipment_date": "09.06.2026"},
            }

            result = worker.import_telegram_document(
                "123",
                {"file_name": "orders.xlsx", "file_id": "file-1"},
                shipment_date="09.06.2026",
                event_id="11111111-1111-1111-1111-111111111111",
            )
        finally:
            telegram_import_processor_module.excel_file_to_import_payload = original_excel_to_payload

        self.assertFalse(result[0])
        text = messages[-1][1]
        self.assertIn("Не удалось подтвердить импорт Excel-файла", text)
        self.assertIn("не отправляйте файл повторно", text)
        self.assertNotIn("Заказы и заявки SkladBot не созданы", text)

    def test_telegram_worker_failed_import_event_creates_one_linked_incident(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_import_processor_module.SessionLocal
        try:
            with SessionLocal() as db:
                event = PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_EVENT_TYPE,
                    status="processing",
                    attempts=1,
                    payload={
                        "chat_id": "123",
                        "file_name": "broken.xlsx",
                        "document": {"file_name": "broken.xlsx", "file_id": "file-1"},
                    },
                    last_error="",
                )
                db.add(event)
                db.commit()
                event_id = str(event.id)

            worker = TelegramWorker.__new__(TelegramWorker)
            telegram_import_processor_module.SessionLocal = SessionLocal

            worker.finish_telegram_import_event(event_id, False, "parser failed")
            worker.finish_telegram_import_event(event_id, False, "parser failed again")

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()
                incidents = db.execute(select(Incident)).scalars().all()
                audits = db.execute(
                    select(AuditLog).where(AuditLog.action == "telegram_import_incident_created")
                ).scalars().all()

            self.assertEqual(event.status, "failed")
            self.assertEqual(event.last_error, "parser failed again")
            self.assertEqual(len(incidents), 1)
            self.assertEqual(incidents[0].source, "telegram_import")
            self.assertEqual(str(incidents[0].pending_event_id), event_id)
            self.assertEqual(len(audits), 1)
        finally:
            telegram_import_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_old_failed_excel_import_does_not_block_next_pending_import(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_import_processor_module.SessionLocal
        try:
            now = datetime.now(timezone.utc)
            with SessionLocal() as db:
                failed_event = PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_EVENT_TYPE,
                    status="failed",
                    attempts=3,
                    payload={
                        "chat_id": "123",
                        "document": {"file_name": "broken.xlsx", "file_id": "old-file"},
                        "shipment_date": "09.06.2026",
                    },
                    last_error="old broken import",
                    created_at=now - timedelta(days=2),
                    updated_at=now - timedelta(days=2),
                )
                pending_event = PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_EVENT_TYPE,
                    status="pending",
                    attempts=0,
                    payload={
                        "chat_id": "123",
                        "document": {"file_name": "valid.xlsx", "file_id": "new-file"},
                        "shipment_date": "10.06.2026",
                    },
                    created_at=now - timedelta(minutes=5),
                    updated_at=now - timedelta(minutes=5),
                )
                db.add_all([failed_event, pending_event])
                db.commit()
                failed_id = failed_event.id
                pending_id = pending_event.id

            imported = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.allowed_chat_ids = {"123"}
            worker.admin_chat_ids = {"123"}
            worker.import_telegram_document = lambda chat_id, document, shipment_date="", event_id="": imported.append(
                (chat_id, document["file_name"], shipment_date, event_id)
            ) or (True, "")
            telegram_import_processor_module.SessionLocal = SessionLocal

            processed = worker.process_queued_telegram_imports()

            with SessionLocal() as db:
                failed_event = db.get(PendingEvent, failed_id)
                pending_event = db.get(PendingEvent, pending_id)

            self.assertEqual(processed, 1)
            self.assertEqual(imported[0][1:3], ("valid.xlsx", "10.06.2026"))
            self.assertEqual(failed_event.status, "failed")
            self.assertEqual(failed_event.last_error, "old broken import")
            self.assertEqual(pending_event.status, "completed")
            self.assertEqual(pending_event.attempts, 1)
        finally:
            telegram_import_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_resets_stale_processing_import_before_processing(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_import_processor_module.SessionLocal
        try:
            with SessionLocal() as db:
                db.add(PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_EVENT_TYPE,
                    status="processing",
                    attempts=2,
                    payload={
                        "chat_id": "123",
                        "file_name": "orders.xlsx",
                        "document": {"file_name": "orders.xlsx", "file_id": "file-1"},
                        "shipment_date": "09.06.2026",
                    },
                    last_error="worker died",
                    updated_at=datetime.now(timezone.utc) - timedelta(minutes=30),
                ))
                db.commit()

            import_calls = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.allowed_chat_ids = {"123"}
            worker.admin_chat_ids = {"123"}
            worker.import_telegram_document = (
                lambda chat_id, document, shipment_date="", event_id="":
                import_calls.append((chat_id, shipment_date, event_id)) or (True, "")
            )
            telegram_import_processor_module.SessionLocal = SessionLocal

            processed = worker.process_queued_telegram_imports()

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()

            self.assertEqual(processed, 1)
            self.assertEqual(event.status, "completed")
            self.assertEqual(event.attempts, 3)
            self.assertEqual(event.last_error, "")
            self.assertEqual((event.payload or {}).get("reset_reason"), "stale Telegram Excel import reset")
            self.assertEqual(import_calls[0][0], "123")
            self.assertEqual(import_calls[0][1], "09.06.2026")
        finally:
            telegram_import_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_resets_stale_processing_notification_before_processing(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_admin_processor_module.SessionLocal
        try:
            with SessionLocal() as db:
                db.add(PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_NOTIFICATION_EVENT_TYPE,
                    status="processing",
                    attempts=2,
                    payload={"chat_id": "123", "text": "hello"},
                    last_error="worker died",
                    updated_at=datetime.now(timezone.utc) - timedelta(minutes=30),
                ))
                db.commit()

            sent = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.admin_chat_ids = {"123"}
            worker.allowed_chat_ids = {"123"}
            worker.send_message = lambda chat_id, text: sent.append((chat_id, text))
            telegram_admin_processor_module.SessionLocal = SessionLocal

            processed = worker.process_pending_telegram_notifications()

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()

            self.assertEqual(processed, 1)
            self.assertEqual(sent, [("123", "hello")])
            self.assertEqual(event.status, "completed")
            self.assertEqual(event.attempts, 3)
            self.assertEqual(event.last_error, "")
            self.assertEqual((event.payload or {}).get("reset_reason"), "stale Telegram notification reset")
        finally:
            telegram_admin_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_manual_shipment_date_overrides_excel_date(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "orders.xlsx"
            create_conflicting_date_workbook(source_path)
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.token = "telegram-token"
            worker.timeout = 20
            worker.file_timeout = 120
            worker.max_file_size = 20 * 1024 * 1024
            messages = []
            calls = {}

            def fake_download(document, destination_path):
                shutil.copyfile(source_path, destination_path)

            def fake_backend_post(path, payload):
                calls["path"] = path
                calls["payload"] = payload
                return {
                    "status": "completed",
                    "items_created": len(payload["rows"]),
                    "orders_created": 1,
                    "duplicate_rows": 0,
                    "invalid_rows": 0,
                    "errors": [],
                    "backend_address_updates": 0,
                    "google_sheets_status": "disabled",
                }

            def fake_send(chat_id, text, reply_markup=None):
                messages.append((chat_id, text, reply_markup))

            worker.download_telegram_document = fake_download
            worker.backend_post = fake_backend_post
            worker.safe_send_message = fake_send

            result = worker.import_telegram_document(
                "123",
                {"file_name": "orders.xlsx", "file_id": "file-1"},
                shipment_date="08.06.2026",
                event_id="11111111-1111-1111-1111-111111111111",
            )

        self.assertEqual(result, (True, ""))
        self.assertEqual(calls["path"], "/api/v1/imports")
        self.assertEqual(calls["payload"]["telegram_chat_id"], "123")
        self.assertEqual(calls["payload"]["rows"][0]["Дата отгрузки"], "08.06.2026")
        self.assertIn("Дата отгрузки: 08.06.2026", messages[-1][1])

    def test_telegram_worker_does_not_finish_event_while_waiting_for_date_choice(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        events = [{
            "id": "11111111-1111-1111-1111-111111111111",
            "payload": {
                "chat_id": "123",
                "document": {"file_name": "orders.xlsx"},
                "shipment_date": "08.06.2026",
            },
        }]
        finished = []

        def fake_take_next():
            return events.pop(0) if events else None

        def fake_import(chat_id, document, shipment_date="", event_id=None):
            self.assertEqual(event_id, "11111111-1111-1111-1111-111111111111")
            return None, "date_choice_required"

        worker.take_next_telegram_import_event = fake_take_next
        worker.import_telegram_document = fake_import
        worker.finish_telegram_import_event = lambda *args: finished.append(args)
        worker.reset_stale_telegram_import_events = lambda: 0

        processed = worker.process_queued_telegram_imports()

        self.assertEqual(processed, 1)
        self.assertEqual(finished, [])

    def test_telegram_worker_handles_date_choice_callbacks(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        answered = []
        confirmed = []
        cancelled = []

        worker.answer_callback_query = lambda callback_id, text="": answered.append((callback_id, text))
        worker.confirm_telegram_import_excel_date = lambda chat_id, event_id: confirmed.append((chat_id, event_id))
        worker.cancel_telegram_import_date_choice = lambda chat_id, event_id: cancelled.append((chat_id, event_id))

        worker.handle_update({
            "update_id": 78,
            "callback_query": {
                "id": "cb-1",
                "data": "excel_date:use_excel:11111111-1111-1111-1111-111111111111",
                "message": {"chat": {"id": 123}},
            },
        })
        worker.handle_update({
            "update_id": 79,
            "callback_query": {
                "id": "cb-2",
                "data": "excel_date:cancel:22222222-2222-2222-2222-222222222222",
                "message": {"chat": {"id": 123}},
            },
        })

        self.assertEqual(answered, [("cb-1", ""), ("cb-2", "")])
        self.assertEqual(confirmed, [("123", "11111111-1111-1111-1111-111111111111")])
        self.assertEqual(cancelled, [("123", "22222222-2222-2222-2222-222222222222")])

    def test_telegram_worker_confirm_excel_date_requeues_waiting_event(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_import_processor_module.SessionLocal
        try:
            with SessionLocal() as db:
                event = PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_EVENT_TYPE,
                    status=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_WAITING_DATE_CHOICE_STATUS,
                    payload={
                        "chat_id": "123",
                        "file_name": "orders.xlsx",
                        "document": {"file_name": "orders.xlsx"},
                        "shipment_date": "08.06.2026",
                        "date_conflict": {"telegram_date": "08.06.2026", "excel_date": "09.06.2026"},
                    },
                )
                db.add(event)
                db.commit()
                event_id = str(event.id)

            messages = []
            processed = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.allowed_chat_ids = {"123"}
            worker.admin_chat_ids = {"123"}
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
            worker.process_queued_telegram_imports = lambda: processed.append(True)
            telegram_import_processor_module.SessionLocal = SessionLocal

            result = worker.confirm_telegram_import_excel_date("123", event_id)

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()
                payload = event.payload

            self.assertTrue(result)
            self.assertEqual(event.status, "pending")
            self.assertEqual(event.last_error, "")
            self.assertEqual(payload["shipment_date"], "")
            self.assertEqual(payload["date_choice_resolution"]["action"], "use_excel")
            self.assertEqual(processed, [True])
            self.assertIn("Импорт будет выполнен по дате из Excel", messages[0][1])
        finally:
            telegram_import_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_cancel_date_choice_does_not_requeue_event(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_import_processor_module.SessionLocal
        try:
            with SessionLocal() as db:
                event = PendingEvent(
                    event_type=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_EVENT_TYPE,
                    status=telegram_worker_module.TELEGRAM_EXCEL_IMPORT_WAITING_DATE_CHOICE_STATUS,
                    payload={
                        "chat_id": "123",
                        "file_name": "orders.xlsx",
                        "document": {"file_name": "orders.xlsx"},
                        "shipment_date": "08.06.2026",
                        "date_conflict": {"telegram_date": "08.06.2026", "excel_date": "09.06.2026"},
                    },
                )
                db.add(event)
                db.commit()
                event_id = str(event.id)

            messages = []
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.allowed_chat_ids = {"123"}
            worker.admin_chat_ids = {"123"}
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))
            telegram_import_processor_module.SessionLocal = SessionLocal

            result = worker.cancel_telegram_import_date_choice("123", event_id)

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()
                payload = event.payload

            self.assertTrue(result)
            self.assertEqual(event.status, "cancelled")
            self.assertEqual(event.last_error, "cancelled_by_user")
            self.assertEqual(payload["shipment_date"], "08.06.2026")
            self.assertEqual(payload["date_choice_resolution"]["action"], "cancel")
            self.assertIn("Импорт отменён", messages[0][1])
        finally:
            telegram_import_processor_module.SessionLocal = original_session_local
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_telegram_worker_shows_kiz_export_menu_modes(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []

        def fake_send(chat_id, text, reply_markup=None):
            messages.append((chat_id, text, reply_markup))

        worker.safe_send_message = fake_send

        worker.show_kiz_export_menu("123")

        self.assertIn("Как выгрузить КИЗы", messages[0][1])
        keyboard = messages[0][2]["inline_keyboard"]
        self.assertEqual(keyboard[0][0]["callback_data"], "kiz_mode:dates")
        self.assertEqual(keyboard[1][0]["callback_data"], "kiz_mode:files")

    def test_telegram_worker_shows_kiz_dates_as_display_dates(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []
        states = []

        def fake_backend_get(path, params=None):
            self.assertEqual(path, "/api/v1/reports/kiz/dates")
            return [
                {
                    "date": "2026-05-29",
                    "planned_blocks": 3,
                    "scanned_blocks": 3,
                },
                {
                    "date": "2026-05-30",
                    "planned_blocks": 2,
                    "scanned_blocks": 2,
                },
            ]

        def fake_get_state(chat_id):
            return {}

        def fake_save_state(chat_id, payload):
            states.append((chat_id, payload))

        def fake_send(chat_id, text, reply_markup=None):
            messages.append((chat_id, text, reply_markup))

        worker.backend_get = fake_backend_get
        worker.get_chat_state = fake_get_state
        worker.save_chat_state = fake_save_state
        worker.safe_send_message = fake_send

        worker.show_kiz_dates("123")

        self.assertIn("29.05.2026", messages[0][1])
        self.assertIn("30.05.2026", messages[0][1])
        self.assertEqual(
            messages[0][2]["inline_keyboard"][0][0]["callback_data"],
            f"{TELEGRAM_KIZ_RANGE_CALLBACK_PREFIX}2026-05-29:2026-05-30",
        )
        self.assertEqual(messages[0][2]["inline_keyboard"][1][0]["callback_data"], "kiz_date:2026-05-29")
        self.assertEqual(states[0][1]["kiz_dates"][0]["date"], "2026-05-29")

    def test_telegram_worker_shows_all_kiz_dates_and_range_export_button(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []
        states = []

        def fake_backend_get(path, params=None):
            self.assertEqual(path, "/api/v1/reports/kiz/dates")
            return [
                {
                    "date": f"2026-06-{day:02d}",
                    "planned_blocks": day,
                    "scanned_blocks": day,
                    "completed": True,
                }
                for day in range(1, 11)
            ]

        def fake_get_state(chat_id):
            return {}

        def fake_save_state(chat_id, payload):
            states.append((chat_id, payload))

        def fake_send(chat_id, text, reply_markup=None):
            messages.append((chat_id, text, reply_markup))

        worker.backend_get = fake_backend_get
        worker.get_chat_state = fake_get_state
        worker.save_chat_state = fake_save_state
        worker.safe_send_message = fake_send

        worker.show_kiz_dates("123")

        self.assertIn("Все даты - 01.06.2026-10.06.2026", messages[0][1])
        self.assertIn("01.06.2026", messages[0][1])
        self.assertIn("02.06.2026", messages[0][1])
        self.assertIn("03.06.2026", messages[0][1])
        self.assertIn("04.06.2026", messages[0][1])
        self.assertIn("10.06.2026", messages[0][1])
        keyboard = messages[0][2]["inline_keyboard"]
        self.assertEqual(len(keyboard), 11)
        self.assertEqual(
            keyboard[0][0]["callback_data"],
            f"{TELEGRAM_KIZ_RANGE_CALLBACK_PREFIX}2026-06-01:2026-06-10",
        )
        self.assertEqual(keyboard[1][0]["callback_data"], "kiz_date:2026-06-01")
        self.assertEqual(keyboard[-1][0]["callback_data"], "kiz_date:2026-06-10")
        self.assertEqual([item["date"] for item in states[0][1]["kiz_dates"]], [
            "2026-06-01",
            "2026-06-02",
            "2026-06-03",
            "2026-06-04",
            "2026-06-05",
            "2026-06-06",
            "2026-06-07",
            "2026-06-08",
            "2026-06-09",
            "2026-06-10",
        ])

    def test_telegram_worker_shows_kiz_source_files_with_progress(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []
        states = []

        def fake_backend_get(path, params=None):
            self.assertEqual(path, "/api/v1/reports/kiz/source-files")
            return [
                {
                    "source_key": "import:done",
                    "source_file": "done.xlsx",
                    "planned_blocks": 3,
                    "scanned_blocks": 3,
                    "completed": True,
                    "dates": ["2026-06-04"],
                },
                {
                    "source_key": "import:open",
                    "source_file": "open.xlsx",
                    "planned_blocks": 5,
                    "scanned_blocks": 2,
                    "completed": False,
                    "dates": ["2026-06-08"],
                    "uploaded_at": "2026-07-01T12:00:00+00:00",
                },
            ]

        def fake_get_state(chat_id):
            return {}

        def fake_save_state(chat_id, payload):
            states.append((chat_id, payload))

        def fake_send(chat_id, text, reply_markup=None):
            messages.append((chat_id, text, reply_markup))

        worker.backend_get = fake_backend_get
        worker.get_chat_state = fake_get_state
        worker.save_chat_state = fake_save_state
        worker.safe_send_message = fake_send

        worker.show_kiz_source_files("123")

        self.assertIn("done.xlsx", messages[0][1])
        self.assertIn("3/3 блоков", messages[0][1])
        self.assertIn("open.xlsx", messages[0][1])
        self.assertIn("2/5 блоков", messages[0][1])
        self.assertEqual(messages[0][2]["inline_keyboard"][0][0]["callback_data"], "kiz_file:2")
        self.assertEqual(len(messages[0][2]["inline_keyboard"]), 1)
        self.assertEqual(states[0][1]["kiz_files"][0]["source_key"], "import:open")
        self.assertFalse(states[0][1]["kiz_files"][0]["completed"])
        self.assertTrue(states[0][1]["kiz_files"][1]["completed"])

    def test_telegram_worker_limits_kiz_source_files_menu_to_recent_seven_files_by_upload_time(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []
        states = []

        def fake_backend_get(path, params=None):
            self.assertEqual(path, "/api/v1/reports/kiz/source-files")
            return [
                {
                    "source_key": f"import:file-{day:02d}",
                    "source_file": f"file-{day:02d}.xlsx",
                    "planned_blocks": 1,
                    "scanned_blocks": 1,
                    "completed": True,
                    "dates": [f"2026-06-{11 - day:02d}"],
                    "uploaded_at": f"2026-07-{day:02d}T12:00:00+00:00",
                }
                for day in range(1, 11)
            ]

        def fake_get_state(chat_id):
            return {}

        def fake_save_state(chat_id, payload):
            states.append((chat_id, payload))

        def fake_send(chat_id, text, reply_markup=None):
            messages.append((chat_id, text, reply_markup))

        worker.backend_get = fake_backend_get
        worker.get_chat_state = fake_get_state
        worker.save_chat_state = fake_save_state
        worker.safe_send_message = fake_send

        worker.show_kiz_source_files("123")

        self.assertNotIn("file-01.xlsx", messages[0][1])
        self.assertNotIn("file-02.xlsx", messages[0][1])
        self.assertNotIn("file-03.xlsx", messages[0][1])
        self.assertIn("file-04.xlsx", messages[0][1])
        self.assertIn("file-10.xlsx", messages[0][1])
        keyboard = messages[0][2]["inline_keyboard"]
        self.assertEqual(len(keyboard), 7)
        self.assertEqual(keyboard[0][0]["callback_data"], "kiz_file:1")
        self.assertEqual([item["source_file"] for item in states[0][1]["kiz_files"]], [
            "file-10.xlsx",
            "file-09.xlsx",
            "file-08.xlsx",
            "file-07.xlsx",
            "file-06.xlsx",
            "file-05.xlsx",
            "file-04.xlsx",
        ])

    def test_telegram_worker_keeps_telegram_upload_above_newer_auto_imports(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []
        states = []

        def fake_backend_get(path, params=None):
            self.assertEqual(path, "/api/v1/reports/kiz/source-files")
            return [
                {
                    "source_key": f"smartup:{day}",
                    "source_file": f"smartup-{day}.xlsx",
                    "planned_blocks": 1,
                    "scanned_blocks": 1,
                    "completed": True,
                    "uploaded_at": f"2026-07-{day:02d}T12:00:00+00:00",
                    "import_source": "smartup_auto",
                }
                for day in range(1, 11)
            ] + [{
                "source_key": "telegram:manual",
                "source_file": "manual.xlsx",
                "planned_blocks": 1,
                "scanned_blocks": 1,
                "completed": True,
                "uploaded_at": "2026-06-01T12:00:00+00:00",
                "import_source": "telegram",
            }]

        worker.backend_get = fake_backend_get
        worker.get_chat_state = lambda chat_id: {}
        worker.save_chat_state = lambda chat_id, payload: states.append((chat_id, payload))
        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))

        worker.show_kiz_source_files("123")

        self.assertEqual(len(messages[0][2]["inline_keyboard"]), 7)
        self.assertEqual([item["source_file"] for item in states[0][1]["kiz_files"]], [
            "manual.xlsx",
            "smartup-10.xlsx",
            "smartup-9.xlsx",
            "smartup-8.xlsx",
            "smartup-7.xlsx",
            "smartup-6.xlsx",
            "smartup-5.xlsx",
        ])

    def test_telegram_worker_downloads_kiz_source_file_by_import_key(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        params_seen = []
        sent = []

        def fake_backend_get_bytes(path, params=None):
            params_seen.append((path, params))
            return b"xlsx", {"X-TakSklad-Filename": "TakSklad_KIZ.xlsx"}

        def fake_send_document(chat_id, content, filename, caption=""):
            sent.append((chat_id, content, filename, caption))

        worker.backend_get_bytes = fake_backend_get_bytes
        worker.safe_send_document = fake_send_document

        result = worker.send_kiz_source_file_report("123", "same-name.xlsx", "import:first")

        self.assertTrue(result)
        self.assertEqual(params_seen[0], (
            "/api/v1/reports/kiz/source-file",
            {"source_file": "same-name.xlsx", "source_key": "import:first"},
        ))
        self.assertEqual(sent[0][2], "TakSklad_KIZ.xlsx")

    def test_telegram_worker_kiz_source_file_backend_failure_is_actionable(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []

        def fake_backend_get_bytes(path, params=None):
            raise httpx.ConnectError("backend down")

        worker.backend_get_bytes = fake_backend_get_bytes
        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))

        result = worker.send_kiz_source_file_report("123", "orders.xlsx", "import:first")

        self.assertFalse(result)
        self.assertIn("Не удалось выгрузить КИЗы по файлу orders.xlsx", messages[0][1])
        self.assertIn("backend временно недоступен", messages[0][1])

    def test_telegram_worker_redacts_backend_error_details_for_kiz_report(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        messages = []

        def fake_backend_get_bytes(path, params=None):
            request = httpx.Request("GET", "http://backend-api/api/v1/reports/kiz/source-file")
            response = httpx.Response(
                409,
                json={
                    "detail": "Bearer secret-service-token failed for 0104006396053978217SECRETKIZVALUE",
                },
                request=request,
            )
            raise httpx.HTTPStatusError("backend rejected report", request=request, response=response)

        worker.backend_get_bytes = fake_backend_get_bytes
        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text, reply_markup))

        result = worker.send_kiz_source_file_report("123", "orders.xlsx", "import:first")

        self.assertFalse(result)
        self.assertIn("Не удалось выгрузить КИЗы по файлу orders.xlsx", messages[0][1])
        self.assertNotIn("secret-service-token", messages[0][1])
        self.assertNotIn("0104006396053978217SECRETKIZVALUE", messages[0][1])
        self.assertIn("Bearer ***", messages[0][1])

    def test_telegram_worker_keeps_kiz_source_key_when_file_selected_by_index(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        calls = []

        def fake_get_state(chat_id):
            return {
                "kiz_files": [
                    {"index": 1, "source_file": "same-name.xlsx", "source_key": "import:first"},
                ],
            }

        def fake_send_report(chat_id, source_file, source_key=""):
            calls.append((chat_id, source_file, source_key))
            return True

        worker.get_chat_state = fake_get_state
        worker.send_kiz_source_file_report = fake_send_report

        result = worker.send_kiz_source_file_by_index("123", "1")

        self.assertTrue(result)
        self.assertEqual(calls, [("123", "same-name.xlsx", "import:first")])

    def test_telegram_worker_keeps_polling_after_single_update_error(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.token = "telegram-token"
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        worker.timeout = 20
        worker.poll_timeout = 1
        worker.offset = 0
        worker.bot_menu_ready = True
        saved_offsets = []
        errors = []
        enqueued = []

        updates = [
            {
                "update_id": 101,
                "message": {
                    "chat": {"id": 123},
                    "text": TELEGRAM_BUTTON_LOGISTICS_REPORT,
                },
            },
            {
                "update_id": 102,
                "message": {
                    "chat": {"id": 123},
                    "document": {"file_name": "orders.xlsx", "file_id": "file-1"},
                },
            },
        ]

        def fake_show_logistics_dates(chat_id):
            raise RuntimeError("backend temporary unavailable")

        def fake_error(chat_id, text, reply_markup=None):
            errors.append((chat_id, text))

        def fake_enqueue(chat_id, document, update_id=None, shipment_date=""):
            enqueued.append((chat_id, document, update_id, shipment_date))
            return True

        def fake_save_offset():
            saved_offsets.append(worker.offset)

        worker.telegram_api_client = SimpleNamespace(
            poll_updates=lambda offset, timeout: updates,
        )
        worker.show_logistics_dates = fake_show_logistics_dates
        worker.safe_send_message = fake_error
        worker.enqueue_telegram_document = fake_enqueue
        worker.save_offset = fake_save_offset
        worker.process_queued_telegram_imports = lambda: 0
        worker.process_pending_telegram_notifications = lambda: 0

        worker.poll_once()

        self.assertEqual(worker.offset, 102)
        self.assertEqual(saved_offsets, [102])
        self.assertEqual(errors[0][0], "123")
        self.assertIn("Не удалось выполнить действие Telegram", errors[0][1])
        self.assertEqual(enqueued, [(
            "123",
            {"file_name": "orders.xlsx", "file_id": "file-1"},
            102,
            "",
        )])

    def test_telegram_worker_runs_scheduled_jobs_after_getupdates_conflict(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.token = "telegram-token"
        worker.timeout = 20
        worker.poll_timeout = 1
        worker.offset = 0
        calls = []

        worker.ensure_bot_menu = lambda: None
        worker.telegram_api_client = SimpleNamespace(
            poll_updates=lambda offset, timeout: (_ for _ in ()).throw(
                RuntimeError("Telegram API request failed: getUpdates: HTTP 409 Conflict")
            ),
        )
        worker.process_queued_telegram_imports = lambda: calls.append("imports") or 0
        worker.process_pending_telegram_notifications = lambda: calls.append("notifications") or 0
        worker.send_due_skladbot_daily_reports = lambda: calls.append("daily") or 0

        worker.poll_once()

        self.assertEqual(calls, ["imports", "notifications", "daily"])
        self.assertEqual(worker.offset, 0)

    def test_telegram_worker_handles_hidden_logs_command(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123"}
        worker.admin_chat_ids = {"123"}
        sent = []

        def fake_backend_get_bytes(path, params=None):
            return b"diagnostics", {"X-TakSklad-Filename": "TakSklad_backend_diagnostics.txt"}

        def fake_send_document(chat_id, content, filename, caption=""):
            sent.append((chat_id, content, filename, caption))

        worker.backend_get_bytes = fake_backend_get_bytes
        worker.safe_send_document = fake_send_document

        worker.handle_update({
            "update_id": 88,
            "message": {
                "chat": {"id": 123},
                "text": "/logs",
            },
        })

        self.assertEqual(sent[0][0], "123")
        self.assertEqual(sent[0][1], b"diagnostics")
        self.assertEqual(sent[0][2], "TakSklad_backend_diagnostics.txt")
        self.assertIn("backend", sent[0][3])

    def test_telegram_worker_restricts_hidden_admin_commands_when_admin_chat_ids_set(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = {"123", "999"}
        worker.admin_chat_ids = {"999"}
        messages = []
        backend_calls = []

        def fake_send_message(chat_id, text, reply_markup=None):
            messages.append((chat_id, text))

        def fake_backend_get(path, params=None):
            backend_calls.append(path)
            return {"status": "ok", "version": "test"}

        worker.send_message = fake_send_message
        worker.backend_get = fake_backend_get

        worker.handle_update({
            "update_id": 89,
            "message": {"chat": {"id": 123}, "text": "/health"},
        })

        self.assertEqual(backend_calls, [])
        self.assertEqual(messages[0][0], "123")
        self.assertIn("только администратору", messages[0][1])

        worker.handle_update({
            "update_id": 90,
            "message": {"chat": {"id": 999}, "text": "/health"},
        })

        self.assertEqual(backend_calls, ["/health"])
        self.assertIn("Backend: ok / test", messages[-1][1])

    def test_excel_file_to_import_payload_reads_smartup_coordinates_and_prices(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "smartup.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Конструктор отчетов"
            for _ in range(4):
                sheet.append([])
            sheet.append([
                "Торговый представитель",
                "Клиент",
                "Координаты клиента",
                "ТМЦ",
                "Тип оплаты",
                "Статус",
                "Количество заказа",
                "Сумма с переоценкой",
            ])
            sheet.append(["Rep One", "Client One", "41.31, 69.27", "Chapman Brown OP 20", "Терминал", "", 200, 4_800_000])
            workbook.save(path)

            payload = excel_file_to_import_payload(path, file_name=path.name, source="telegram", shipment_date="29.05.2026")

        self.assertEqual(len(payload["rows"]), 1)
        row = payload["rows"][0]
        self.assertEqual(row["Дата отгрузки"], "29.05.2026")
        self.assertEqual(row["Координаты"], "41.31, 69.27")
        self.assertEqual(row["Кол-во ШТ"], 200)
        self.assertEqual(row["Кол-во блок"], 20)
        self.assertEqual(row["Сумма позиции"], 4_800_000)
        self.assertEqual(row["Цена за блок"], 240000)

    def test_excel_file_to_import_payload_removes_country_prefix_from_address(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "address.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Заявки"
            sheet.append([
                "Клиент",
                "Тип оплаты",
                "Товары",
                "Кол-во ШТ",
                "Адрес",
                "Торговый представитель",
            ])
            sheet.append([
                "Client One",
                "Терминал",
                "Chapman Brown OP 20",
                20,
                "Uzbekistan, Tashkent, Chilanzar 10",
                "Rep One",
            ])
            sheet.append([
                "Client Two",
                "Перечисление",
                "Chapman Red SSL 20",
                10,
                "O'zbekiston, Toshkent, Yunusobod 4",
                "Rep Two",
            ])
            workbook.save(path)

            payload = excel_file_to_import_payload(path, file_name=path.name, source="telegram", shipment_date="29.05.2026")

        self.assertEqual(payload["rows"][0]["Адрес"], "Tashkent, Chilanzar 10")
        self.assertEqual(payload["rows"][1]["Адрес"], "Toshkent, Yunusobod 4")

    def test_excel_file_to_import_payload_geocodes_address_when_coordinates_missing(self):
        original_geocoder = excel_importer.geocode_address_yandex
        calls = []
        try:
            def fake_geocoder(address, cache=None):
                calls.append(address)
                return "41.311081, 69.240562", ""

            excel_importer.geocode_address_yandex = fake_geocoder

            with tempfile.TemporaryDirectory() as tmp_dir:
                path = Path(tmp_dir) / "address_without_coordinates.xlsx"
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Заявки"
                sheet.append([
                    "Клиент",
                    "Тип оплаты",
                    "Товары",
                    "Кол-во ШТ",
                    "Адрес",
                    "Торговый представитель",
                ])
                sheet.append([
                    "Client One",
                    "Терминал",
                    "Chapman Brown OP 20",
                    20,
                    "Uzbekistan, Tashkent, Chilanzar 10",
                    "Rep One",
                ])
                workbook.save(path)

                payload = excel_importer.excel_file_to_import_payload(
                    path,
                    file_name=path.name,
                    source="telegram",
                    shipment_date="29.05.2026",
                )
        finally:
            excel_importer.geocode_address_yandex = original_geocoder

        self.assertEqual(calls, ["Tashkent, Chilanzar 10"])
        self.assertEqual(payload["rows"][0]["Координаты"], "41.311081, 69.240562")
        self.assertEqual(payload["meta"]["geocoded_count"], 1)
        self.assertEqual(payload["meta"]["geocode_failed_count"], 0)

    def test_excel_file_to_import_payload_marks_missing_address_as_pickup(self):
        original_geocoder = excel_importer.geocode_address_yandex
        calls = []
        try:
            def fake_geocoder(address, cache=None):
                calls.append(address)
                return "41.311081, 69.240562", ""

            excel_importer.geocode_address_yandex = fake_geocoder

            with tempfile.TemporaryDirectory() as tmp_dir:
                path = Path(tmp_dir) / "pickup_without_address.xlsx"
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Заявки"
                sheet.append([
                    "Клиент",
                    "Тип оплаты",
                    "Товары",
                    "Кол-во ШТ",
                    "Адрес",
                    "Торговый представитель",
                ])
                sheet.append([
                    "Pickup Client",
                    "Терминал",
                    "Chapman Brown OP 20",
                    20,
                    "",
                    "Rep One",
                ])
                workbook.save(path)

                payload = excel_importer.excel_file_to_import_payload(
                    path,
                    file_name=path.name,
                    source="telegram",
                    shipment_date="29.05.2026",
                )
        finally:
            excel_importer.geocode_address_yandex = original_geocoder

        self.assertEqual(calls, [])
        self.assertEqual(payload["rows"][0]["Адрес"], "Самовывоз со склада")
        self.assertEqual(payload["rows"][0]["Координаты"], "")
        self.assertEqual(payload["meta"]["geocoded_count"], 0)
        self.assertEqual(payload["meta"]["geocode_failed_count"], 0)

    def test_excel_file_to_import_payload_reverse_geocodes_address_when_only_coordinates_present(self):
        original_reverse_geocoder = excel_importer.reverse_geocode_yandex
        calls = []
        try:
            def fake_reverse_geocoder(coordinates, cache=None):
                calls.append(coordinates)
                return "Ташкент, Чиланзарский район, 10", ""

            excel_importer.reverse_geocode_yandex = fake_reverse_geocoder

            with tempfile.TemporaryDirectory() as tmp_dir:
                path = Path(tmp_dir) / "coordinates_without_address.xlsx"
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Заявки"
                sheet.append([
                    "Клиент",
                    "Тип оплаты",
                    "Товары",
                    "Кол-во ШТ",
                    "Адрес доставки",
                    "Координаты",
                    "Торговый представитель",
                ])
                sheet.append([
                    "Client One",
                    "Терминал",
                    "Chapman Brown OP 20",
                    20,
                    "",
                    "41.311081,69.240562,15",
                    "Rep One",
                ])
                workbook.save(path)

                payload = excel_importer.excel_file_to_import_payload(
                    path,
                    file_name=path.name,
                    source="telegram",
                    shipment_date="29.05.2026",
                )
        finally:
            excel_importer.reverse_geocode_yandex = original_reverse_geocoder

        self.assertEqual(calls, ["41.311081, 69.240562"])
        self.assertEqual(payload["rows"][0]["Адрес"], "Ташкент, Чиланзарский район, 10")
        self.assertEqual(payload["rows"][0]["Координаты"], "41.311081, 69.240562")
        self.assertEqual(payload["meta"]["geocoded_count"], 1)
        self.assertEqual(payload["meta"]["geocode_failed_count"], 0)

    def test_excel_file_to_import_payload_reverse_geocodes_when_address_is_not_found_placeholder(self):
        original_reverse_geocoder = excel_importer.reverse_geocode_yandex
        calls = []
        try:
            def fake_reverse_geocoder(coordinates, cache=None):
                calls.append(coordinates)
                return "Ташкент, Яккасарайский район", ""

            excel_importer.reverse_geocode_yandex = fake_reverse_geocoder

            with tempfile.TemporaryDirectory() as tmp_dir:
                path = Path(tmp_dir) / "coordinates_with_placeholder_address.xlsx"
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Заявки"
                sheet.append([
                    "Клиент",
                    "Тип оплаты",
                    "Товары",
                    "Кол-во ШТ",
                    "Адрес",
                    "Координаты",
                    "Торговый представитель",
                ])
                sheet.append([
                    "Client One",
                    "Терминал",
                    "Chapman Brown OP 20",
                    20,
                    "Адрес не найден",
                    "41.311081,69.240562",
                    "Rep One",
                ])
                workbook.save(path)

                payload = excel_importer.excel_file_to_import_payload(
                    path,
                    file_name=path.name,
                    source="telegram",
                    shipment_date="29.05.2026",
                )
        finally:
            excel_importer.reverse_geocode_yandex = original_reverse_geocoder

        self.assertEqual(calls, ["41.311081, 69.240562"])
        self.assertEqual(payload["rows"][0]["Адрес"], "Ташкент, Яккасарайский район")
        self.assertEqual(payload["rows"][0]["Координаты"], "41.311081, 69.240562")
        self.assertEqual(payload["meta"]["geocoded_count"], 1)
        self.assertEqual(payload["meta"]["geocode_failed_count"], 0)

    def test_excel_file_to_import_payload_reads_full_coordinates_from_repeated_smartup_columns(self):
        original_reverse_geocoder = excel_importer.reverse_geocode_yandex
        calls = []
        try:
            def fake_reverse_geocoder(coordinates, cache=None):
                calls.append(coordinates)
                return "Ташкент, Юнусабадский район", ""

            excel_importer.reverse_geocode_yandex = fake_reverse_geocoder

            with tempfile.TemporaryDirectory() as tmp_dir:
                path = Path(tmp_dir) / "smartup_repeated_coordinates.xlsx"
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Конструктор отчетов"
                sheet.append([
                    "Торговый представитель",
                    "Клиент",
                    "Координаты клиента",
                    "Координаты клиента",
                    "Координаты клиента",
                    "ТМЦ",
                    "Тип оплаты",
                    "Статус",
                    "Количество заказа",
                    "Сумма с переоценкой",
                    "Дата отгрузки",
                ])
                sheet.append([
                    "Rep One",
                    "Client One",
                    "41.325658539017745",
                    "69.23166364431383",
                    "41.325658539017745,69.23166364431383",
                    "Chapman Brown OP 20",
                    "Перечисление",
                    "В обработке",
                    150,
                    3_600_000,
                    "2026-06-03",
                ])
                workbook.save(path)

                payload = excel_importer.excel_file_to_import_payload(
                    path,
                    file_name=path.name,
                    source="telegram",
                    shipment_date="03.06.2026",
                )
        finally:
            excel_importer.reverse_geocode_yandex = original_reverse_geocoder

        self.assertEqual(calls, ["41.325658539017745, 69.23166364431383"])
        self.assertEqual(payload["rows"][0]["Адрес"], "Ташкент, Юнусабадский район")
        self.assertEqual(payload["rows"][0]["Координаты"], "41.325658539017745, 69.23166364431383")
        self.assertEqual(payload["meta"]["geocoded_count"], 1)
        self.assertEqual(payload["meta"]["geocode_failed_count"], 0)

    def test_excel_file_to_import_payload_combines_split_smartup_coordinates_without_full_column_header(self):
        original_reverse_geocoder = excel_importer.reverse_geocode_yandex
        calls = []
        try:
            def fake_reverse_geocoder(coordinates, cache=None):
                calls.append(coordinates)
                return "Ташкент, Мирзо-Улугбекский район", ""

            excel_importer.reverse_geocode_yandex = fake_reverse_geocoder

            with tempfile.TemporaryDirectory() as tmp_dir:
                path = Path(tmp_dir) / "smartup_split_coordinates.xlsx"
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Конструктор отчетов"
                sheet.append([
                    "Торговый представитель",
                    "Клиент",
                    "Координаты клиента",
                    "",
                    "",
                    "ТМЦ",
                    "Тип оплаты",
                    "Статус",
                    "Количество заказа",
                    "Сумма с переоценкой",
                    "Дата отгрузки",
                ])
                sheet.append([
                    "Rep One",
                    "Client One",
                    "41.363127",
                    "69.287982",
                    "",
                    "Chapman Brown OP 20",
                    "Перечисление",
                    "В обработке",
                    40,
                    960_000,
                    "2026-06-03",
                ])
                workbook.save(path)

                payload = excel_importer.excel_file_to_import_payload(
                    path,
                    file_name=path.name,
                    source="telegram",
                    shipment_date="03.06.2026",
                )
        finally:
            excel_importer.reverse_geocode_yandex = original_reverse_geocoder

        self.assertEqual(calls, ["41.363127, 69.287982"])
        self.assertEqual(payload["rows"][0]["Адрес"], "Ташкент, Мирзо-Улугбекский район")
        self.assertEqual(payload["rows"][0]["Координаты"], "41.363127, 69.287982")
        self.assertEqual(payload["meta"]["geocoded_count"], 1)
        self.assertEqual(payload["meta"]["geocode_failed_count"], 0)


if __name__ == "__main__":
    unittest.main()
