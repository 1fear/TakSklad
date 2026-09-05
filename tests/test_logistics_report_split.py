import unittest
from datetime import date, datetime, timezone
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.logistics_service import (
    build_logistics_report_xlsx,
    build_logistics_reports,
    release_read_transaction,
)
from backend.app.models import Base, LogisticsRegionPoint, Order, OrderItem
from backend.app.orders_service import ApiError


SHIPMENT_DATE = date(2030, 1, 2)


class LogisticsReportSplitTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        self.db = self.SessionLocal()
        self.db.add(LogisticsRegionPoint(
            client_name="Тест Клиент Область",
            normalized_client="тестклиентобласть",
            latitude=41.018778,
            longitude=70.083423,
        ))
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def add_order(
        self,
        client,
        coordinates,
        address="Тестовый адрес",
        request_number=None,
        representative=None,
        product="Тестовый товар",
        created_at=None,
    ):
        raw_payload = {"coordinates": coordinates}
        if request_number:
            raw_payload["skladbot_request_number"] = request_number
        order = Order(
            payment_type="Наличные",
            client=client,
            address=address,
            representative=representative,
            status="not_completed",
            order_date=SHIPMENT_DATE,
            raw_payload=raw_payload,
        )
        if created_at is not None:
            order.created_at = created_at
        order.items.append(OrderItem(product=product, quantity_blocks=3))
        self.db.add(order)
        self.db.commit()
        return order

    def sheet_rows(self, payload):
        workbook = load_workbook(BytesIO(payload))
        return [
            [cell for cell in row]
            for row in workbook["Orders"].iter_rows(min_row=2, values_only=True)
        ]

    def test_orders_split_between_city_and_region_files(self):
        self.add_order("Тест Клиент Область", "41.018778,70.083423")
        self.add_order("Тест Клиент Город", "41.3200,69.2400")
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())

        self.assertIsNotNone(reports["city"])
        self.assertIsNotNone(reports["region"])
        self.assertEqual(reports["unassigned"], [])

        city_payload, city_filename = reports["city"]
        region_payload, region_filename = reports["region"]
        self.assertEqual(city_filename, "TakSklad_логистика_город_02.01.2030.xlsx")
        self.assertEqual(region_filename, "TakSklad_логистика_область_02.01.2030.xlsx")

        city_clients = {row[3] for row in self.sheet_rows(city_payload)}
        region_clients = {row[3] for row in self.sheet_rows(region_payload)}
        self.assertEqual(city_clients, {"Тест Клиент Город"})
        self.assertEqual(region_clients, {"Тест Клиент Область"})

    def test_same_client_at_same_point_becomes_one_stop_with_joined_id(self):
        # Два слота одного дня: ранний заказ задаёт клиента, адрес и окно остановки
        self.add_order(
            "Тест Клиент Город",
            "41.3200,69.2400",
            address="Ташкент, улица Первая, 1",
            request_number="WH-R-1",
            representative="ТП Первый",
            product="Товар А",
            created_at=datetime(2030, 1, 1, 7, 0, tzinfo=timezone.utc),
        )
        self.add_order(
            "Тест Клиент Город",
            "41.3200,69.2400",
            address="Биринчи кучаси, 1-уй",
            request_number="WH-R-2",
            representative="ТП Второй",
            product="Товар Б",
            created_at=datetime(2030, 1, 1, 12, 50, tzinfo=timezone.utc),
        )
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        rows = self.sheet_rows(reports["city"][0])

        # Шаблон «Orders via Excel» ждёт строку на товар: обе позиции остановки
        # идут отдельными строками с общим склеенным внешним ID, а различает их
        # уникальный «Айди товара»
        self.assertEqual(len(rows), 2)
        self.assertEqual({row[1] for row in rows}, {"WH-R-1+WH-R-2"})
        self.assertEqual({row[18] for row in rows}, {"Ташкент, улица Первая, 1"})
        self.assertEqual({row[6] for row in rows}, {"ТП Первый"})
        self.assertEqual([row[28] for row in rows], ["Товар А", "Товар Б"])
        self.assertEqual([row[33] for row in rows], [3, 3])

    def test_multi_product_order_is_one_row_per_product_line(self):
        order = self.add_order(
            "Тест Клиент Город",
            "41.3200,69.2400",
            request_number="WH-R-1",
            product="Товар А",
        )
        order.items.append(OrderItem(product="Товар Б", quantity_blocks=2))
        self.db.commit()
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        rows = self.sheet_rows(reports["city"][0])

        self.assertEqual(len(rows), 2)
        self.assertEqual({row[1] for row in rows}, {"WH-R-1"})
        self.assertEqual([row[28] for row in rows], ["Товар А", "Товар Б"])
        self.assertEqual([row[33] for row in rows], [3, 2])

    def test_headers_match_orders_via_excel_template_exactly(self):
        # Перечень взят из шаблона платформы «Orders via Excel» (excel_orders_template),
        # а не из кода: 35 колонок, четыре новые вставлены в середину
        expected = [
            "Тип заказа", "Внешний ID", "Описание", "Имя клиента", "Телефон", "Email",
            "Заметки", "Широта (забор)", "Долгота (забор)", "Адрес забора",
            "Окно времени С (забор)", "Окно времени ПО (забор)",
            "Окно перерыва С (забор)", "Окно перерыва ПО (забор)",
            "Детали адреса забора", "Время обслуживания забора",
            "Широта (доставка)", "Долгота (доставка)", "Адрес доставки",
            "Окно времени С (доставка)", "Окно времени ПО (доставка)",
            "Окно перерыва С (доставка)", "Окно перерыва ПО (доставка)",
            "Детали адреса доставки", "Время обслуживания доставки",
            "Приоритет заказа", "Навыки", "Тег заказа", "Название товара",
            "Айди товара", "Количество товара", "Вес (кг)", "Объем (m3)",
            "Короба", "Цена товара",
        ]
        self.add_order("Тест Клиент Город", "41.3200,69.2400")
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        workbook = load_workbook(BytesIO(reports["city"][0]))
        header = [cell.value for cell in workbook["Orders"][1]]

        self.assertEqual(header, expected)

    def test_line_id_is_unique_sequence_and_quantity_is_one(self):
        # «Айди товара» сквозной 1..N по файлу, «Количество товара» всегда 1,
        # три новые колонки пустые: ровно так прошёл импорт на 327 строках
        order = self.add_order(
            "Тест Клиент Город",
            "41.3200,69.2400",
            request_number="WH-R-1",
            product="Товар А",
        )
        order.items.append(OrderItem(product="Товар Б", quantity_blocks=2))
        self.db.commit()
        self.add_order(
            "Тест Сосед Город",
            "41.3300,69.2500",
            request_number="WH-R-2",
        )
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        rows = self.sheet_rows(reports["city"][0])

        self.assertEqual(len(rows), 3)
        self.assertEqual([row[29] for row in rows], [1, 2, 3])
        self.assertEqual([row[30] for row in rows], [1, 1, 1])
        for row in rows:
            self.assertIn(row[25], ("", None))   # Приоритет заказа
            self.assertIn(row[27], ("", None))   # Тег заказа
            self.assertIn(row[34], ("", None))   # Цена товара
            self.assertEqual(row[31], 0)         # Вес (кг)
            self.assertEqual(row[32], 0)         # Объем (m3)

    def test_order_counts_follow_zone_split(self):
        # Счёт заказов по зонам считает заказы, а не товарные строки: два городских
        # заказа с тремя позициями и один областной дают 2 и 1
        order = self.add_order("Тест Клиент Город", "41.3200,69.2400", product="Товар А")
        order.items.append(OrderItem(product="Товар Б", quantity_blocks=2))
        self.db.commit()
        self.add_order("Тест Сосед Город", "41.3300,69.2500")
        self.add_order("Тест Клиент Область", "41.018778,70.083423")
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())

        self.assertEqual(reports["order_counts"], {"city": 2, "region": 1})

    def test_stop_external_id_is_stable_when_created_at_matches(self):
        same_moment = datetime(2030, 1, 1, 7, 0, tzinfo=timezone.utc)
        self.add_order(
            "Тест Клиент Город",
            "41.3200,69.2400",
            request_number="WH-R-1",
            created_at=same_moment,
        )
        self.add_order(
            "Тест Клиент Город",
            "41.3200,69.2400",
            request_number="WH-R-2",
            created_at=same_moment,
        )

        first = self.sheet_rows(build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())["city"][0])
        second = self.sheet_rows(build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())["city"][0])

        identifiers = {row[1] for row in first}
        self.assertEqual(len(identifiers), 1)
        self.assertEqual(identifiers, {row[1] for row in second})
        self.assertEqual(sorted(identifiers.pop().split("+")), ["WH-R-1", "WH-R-2"])

    def test_different_clients_at_same_point_stay_separate_stops(self):
        self.add_order(
            "Тест Клиент Город",
            "41.3200,69.2400",
            request_number="WH-R-1",
        )
        self.add_order(
            "Тест Сосед Город",
            "41.3200,69.2400",
            request_number="WH-R-2",
        )
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        rows = self.sheet_rows(reports["city"][0])

        self.assertEqual(len(rows), 2)
        self.assertEqual({row[1] for row in rows}, {"WH-R-1", "WH-R-2"})

    def test_same_client_at_different_points_stays_separate_stops(self):
        self.add_order(
            "Тест Клиент Город",
            "41.3200,69.2400",
            request_number="WH-R-1",
        )
        self.add_order(
            "Тест Клиент Город",
            "41.3300,69.2500",
            request_number="WH-R-2",
        )
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        rows = self.sheet_rows(reports["city"][0])

        self.assertEqual(len(rows), 2)
        self.assertEqual({row[1] for row in rows}, {"WH-R-1", "WH-R-2"})

    def test_single_order_keeps_plain_external_id(self):
        self.add_order(
            "Тест Клиент Город",
            "41.3200,69.2400",
            request_number="WH-R-1",
        )
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        rows = self.sheet_rows(reports["city"][0])

        self.assertEqual([row[1] for row in rows], ["WH-R-1"])

    def test_unknown_client_outside_city_goes_to_region_file(self):
        self.add_order("Тест Клиент Город", "41.3200,69.2400")
        self.add_order("Незнакомый Загород", "41.4700,69.5800")
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())

        self.assertIsNotNone(reports["region"])
        self.assertEqual(reports["unassigned"], [])

        city_clients = {row[3] for row in self.sheet_rows(reports["city"][0])}
        region_clients = {row[3] for row in self.sheet_rows(reports["region"][0])}
        self.assertEqual(city_clients, {"Тест Клиент Город"})
        self.assertEqual(region_clients, {"Незнакомый Загород"})

    def test_zone_without_orders_produces_no_file(self):
        self.add_order("Тест Клиент Город", "41.3200,69.2400")
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        self.assertIsNone(reports["region"])
        self.assertIsNotNone(reports["city"])

    def test_unknown_client_without_coordinates_lands_in_city_problem_sheet(self):
        self.add_order("Незнакомый Без Координат", "")
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        payload, _filename = reports["city"]
        workbook = load_workbook(BytesIO(payload))
        self.assertIn("Требуют координаты", workbook.sheetnames)
        problem_clients = [
            row[0] for row in workbook["Требуют координаты"].iter_rows(min_row=2, values_only=True)
        ]
        self.assertEqual(problem_clients, ["Незнакомый Без Координат"])

    def test_known_client_without_coordinates_lands_in_region_problem_sheet(self):
        self.add_order("Тест Клиент Область", "")
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        self.assertIsNone(reports["city"])
        payload, _filename = reports["region"]
        workbook = load_workbook(BytesIO(payload))
        problem_clients = [
            row[0] for row in workbook["Требуют координаты"].iter_rows(min_row=2, values_only=True)
        ]
        self.assertEqual(problem_clients, ["Тест Клиент Область"])

    def test_single_zone_builder_raises_404_for_empty_zone(self):
        self.add_order("Тест Клиент Город", "41.3200,69.2400")
        payload, filename = build_logistics_report_xlsx(self.db, SHIPMENT_DATE.isoformat(), "city")
        self.assertTrue(payload)
        self.assertEqual(filename, "TakSklad_логистика_город_02.01.2030.xlsx")
        with self.assertRaises(ApiError) as raised:
            build_logistics_report_xlsx(self.db, SHIPMENT_DATE.isoformat(), "region")
        self.assertEqual(raised.exception.status_code, 404)

    def test_empty_directory_keeps_every_order_in_city_report(self):
        # Справочник пуст: страховка не даёт заказам выпасть из обоих отчётов
        self.db.query(LogisticsRegionPoint).delete()
        self.db.commit()
        self.add_order("Тест Клиент Область", "41.018778,70.083423")
        self.add_order("Тест Клиент Город", "41.3200,69.2400")
        self.add_order("Незнакомый Загород", "41.4700,69.5800")
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())

        self.assertTrue(reports["region_directory_empty"])
        self.assertIsNone(reports["region"])
        self.assertEqual(reports["unassigned"], [])
        city_clients = {row[3] for row in self.sheet_rows(reports["city"][0])}
        self.assertEqual(
            city_clients,
            {"Тест Клиент Область", "Тест Клиент Город", "Незнакомый Загород"},
        )

    def test_filled_directory_does_not_raise_empty_flag(self):
        self.add_order("Тест Клиент Город", "41.3200,69.2400")
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())
        self.assertFalse(reports["region_directory_empty"])

    def test_unknown_zone_is_rejected_with_422(self):
        self.add_order("Тест Клиент Город", "41.3200,69.2400")
        with self.assertRaises(ApiError) as raised:
            build_logistics_report_xlsx(self.db, SHIPMENT_DATE.isoformat(), "moon")
        self.assertEqual(raised.exception.status_code, 422)

    def test_build_leaves_no_open_transaction(self):
        # Сборка книги это долгий CPU без запросов, открытая транзакция на это
        # время попадает под idle_in_transaction_session_timeout и рвёт соединение
        self.add_order("Тест Клиент Город", "41.3200,69.2400")
        self.add_order("Тест Клиент Область", "41.018778,70.083423")
        reports = build_logistics_reports(self.db, SHIPMENT_DATE.isoformat())

        self.assertFalse(self.db.in_transaction())
        city_clients = {row[3] for row in self.sheet_rows(reports["city"][0])}
        region_clients = {row[3] for row in self.sheet_rows(reports["region"][0])}
        self.assertEqual(city_clients, {"Тест Клиент Город"})
        self.assertEqual(region_clients, {"Тест Клиент Область"})

    def test_release_read_transaction_keeps_uncommitted_changes_of_caller(self):
        self.add_order("Тест Клиент Город", "41.3200,69.2400")
        pending = Order(
            payment_type="Наличные",
            client="Незакоммиченный клиент",
            address="Тестовый адрес",
            status="not_completed",
            order_date=SHIPMENT_DATE,
            raw_payload={},
        )
        self.db.add(pending)

        released = release_read_transaction(self.db)

        self.assertFalse(released)
        self.assertIn(pending, self.db.new)


if __name__ == "__main__":
    unittest.main()
