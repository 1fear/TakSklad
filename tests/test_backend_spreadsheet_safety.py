import unittest
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook

from backend.app.spreadsheet_safety import force_workbook_text_literals


class SpreadsheetOutputSafetyTests(unittest.TestCase):
    def test_formula_prefixes_reload_as_exact_text(self):
        values = ["=1+1", "+1+1", "-1+1", "@SUM(A1:A2)"]
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(values)

        force_workbook_text_literals(workbook)
        content = BytesIO()
        workbook.save(content)
        workbook.close()

        reloaded = load_workbook(BytesIO(content.getvalue()), data_only=False)
        try:
            cells = list(reloaded.active[1])
            self.assertEqual([cell.value for cell in cells], values)
            self.assertEqual([cell.data_type for cell in cells], ["s"] * len(values))
        finally:
            reloaded.close()

    def test_normal_business_values_keep_their_types_and_values(self):
        values = ["Клиент", 42, 12.5, date(2026, 7, 10), datetime(2026, 7, 10, 17, 50)]
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(values)

        force_workbook_text_literals(workbook)
        content = BytesIO()
        workbook.save(content)
        workbook.close()

        reloaded = load_workbook(BytesIO(content.getvalue()), data_only=False)
        try:
            cells = list(reloaded.active[1])
            self.assertEqual([cell.value for cell in cells[:3]], values[:3])
            self.assertEqual(cells[3].value.date(), values[3])
            self.assertEqual(cells[4].value, values[4])
            self.assertEqual([cell.data_type for cell in cells], ["s", "n", "n", "d", "d"])
        finally:
            reloaded.close()


if __name__ == "__main__":
    unittest.main()
