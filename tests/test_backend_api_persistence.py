import unittest
import uuid
from io import BytesIO
from datetime import date, datetime, timedelta, timezone
from unittest import mock

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, select, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.db import get_db
from backend.app.google_sheets_exporter import update_missing_sheet_addresses
from backend.app.main import (
    app,
    require_admin_write_permission,
    require_client_points_write_permission,
    require_service_token,
)
from backend.app.models import AuditLog, Base, ClientPoint, ImportFile, ImportJob, Incident, KizCode, KizMovement, Order, OrderItem, PendingEvent, ScanCode, User
from backend.app.skladbot_return_requests import SKLADBOT_RETURN_REQUEST_CREATE_EVENT_TYPE
from backend.app.settings import load_settings
from backend.app.web_auth import SESSION_COOKIE_NAME, hash_password


class BackendApiPersistenceTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)

        def override_get_db():
            db = self.SessionLocal()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[get_db] = override_get_db
        app.dependency_overrides[require_service_token] = lambda: None
        app.dependency_overrides[require_admin_write_permission] = lambda: None
        app.dependency_overrides[require_client_points_write_permission] = lambda: None
        self.client = TestClient(app)

    def tearDown(self):
        app.dependency_overrides.clear()
        Base.metadata.drop_all(self.engine)
        self.engine.dispose()

    def seed_order(
        self,
        *,
        status="not_completed",
        quantity_blocks=2,
        scanned_blocks=0,
        item_status="not_completed",
        product="Test Product",
    ):
        with self.SessionLocal() as db:
            order = Order(
                payment_type="cash",
                client="Test Client",
                address="Test Address",
                representative="Test Rep",
                order_date=date(2026, 5, 30),
                status=status,
                raw_payload={"source": "test"},
            )
            item = OrderItem(
                order=order,
                product=product,
                quantity_pieces=20,
                quantity_blocks=quantity_blocks,
                pieces_per_block=10,
                scanned_blocks=scanned_blocks,
                status=item_status,
                raw_payload={"source": "test"},
            )
            db.add_all([order, item])
            db.commit()
            return str(order.id), str(item.id)

    def confirmed_return_items(self, item_id, product="Test Product", blocks=2, pieces=20):
        return [{
            "item_id": item_id,
            "product": product,
            "sku": product,
            "quantity_blocks": blocks,
            "quantity_pieces": pieces,
        }]

    def test_active_orders_returns_uncompleted_orders_with_items(self):
        active_order_id, _ = self.seed_order()
        self.seed_order(status="completed", item_status="completed")

        response = self.client.get("/api/v1/orders/active")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0]["id"], active_order_id)
        self.assertEqual(payload[0]["status"], "not_completed")
        self.assertEqual(payload[0]["items"][0]["product"], "Test Product")

    def test_admin_table_returns_flat_rows_totals_and_recent_activity(self):
        with self.SessionLocal() as db:
            order = Order(
                payment_type="cash",
                client="Admin Client",
                address="Admin Address",
                representative="Admin Rep",
                order_date=date(2026, 6, 2),
                status="not_completed",
                raw_payload={
                    "coordinates": "41.2,69.2",
                    "skladbot_request_number": "SB-77",
                    "skladbot_request_id": "771",
                    "skladbot_status": "found",
                },
            )
            item = OrderItem(
                order=order,
                product="Admin Product",
                quantity_pieces=30,
                quantity_blocks=3,
                pieces_per_block=10,
                scanned_blocks=1,
                status="not_completed",
                raw_payload={
                    "source_file": "orders.xlsx",
                    "google_sheet_row_number": 12,
                    "google_sheet_synced_at": "2026-06-01T12:00:00+00:00",
                    "block_price": 240000,
                    "line_total": 720000,
                },
            )
            db.add(order)
            db.commit()
            item_id = str(item.id)
            order_id = str(order.id)
            db.add(PendingEvent(
                event_type="google_sheets_export",
                status="pending",
                payload={
                    "action": "google_sheets_scan_export",
                    "entity_type": "order_item",
                    "entity_id": item_id,
                },
            ))
            db.add(AuditLog(
                action="admin_test_activity",
                entity_type="order",
                entity_id=order_id,
                payload={"client": "Admin Client"},
            ))
            db.commit()

        response = self.client.get("/api/v1/admin/table")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["totals"]["orders"], 1)
        self.assertEqual(payload["totals"]["items"], 1)
        self.assertEqual(payload["totals"]["active_orders"], 1)
        self.assertEqual(payload["totals"]["planned_blocks"], 3)
        self.assertEqual(payload["totals"]["scanned_blocks"], 1)
        self.assertEqual(payload["totals"]["remaining_blocks"], 2)
        self.assertEqual(payload["totals"]["pending_google_exports"], 1)

        row = payload["rows"][0]
        self.assertEqual(row["order_id"], order_id)
        self.assertEqual(row["item_id"], item_id)
        self.assertEqual(row["status_bucket"], "active")
        self.assertEqual(row["client"], "Admin Client")
        self.assertEqual(row["product"], "Admin Product")
        self.assertEqual(row["source_file"], "orders.xlsx")
        self.assertEqual(row["google_sheet_status"], "pending")
        self.assertEqual(row["skladbot_request_number"], "SB-77")
        self.assertEqual(row["line_total"], 720000)
        self.assertEqual(payload["recent_activity"][0]["action"], "admin_test_activity")
        self.assertEqual(payload["recent_activity"][0]["payload"]["client"], "Admin Client")

    def test_admin_table_shows_return_link_and_skladbot_return_status(self):
        with self.SessionLocal() as db:
            order = Order(
                payment_type="terminal",
                client="Return Client",
                address="Return Address",
                representative="Return Rep",
                order_date=date(2026, 6, 10),
                status="returned",
                raw_payload={
                    "skladbot_request_number": "WH-R-OUT-1",
                    "skladbot_request_id": "190001",
                    "return_status": "returned",
                    "returned_at": "2026-06-10T12:00:00+00:00",
                    "return_reference": "WH-R-OUT-1",
                    "skladbot_return_request_number": "WH-R-RET-1",
                    "skladbot_return_request_id": "190777",
                    "skladbot_return_request_status": "created",
                },
            )
            item = OrderItem(
                order=order,
                product="Chapman Brown OP 20",
                quantity_pieces=20,
                quantity_blocks=2,
                pieces_per_block=10,
                scanned_blocks=2,
                status="completed",
                raw_payload={"source_file": "return-smoke.xlsx"},
            )
            db.add(order)
            db.flush()
            db.add_all([
                ScanCode(order_item_id=item.id, code="0104006396053978217RETURN001", source="desktop"),
                ScanCode(order_item_id=item.id, code="0104006396053978217RETURN002", source="desktop"),
            ])
            db.commit()
            order_id = str(order.id)
            item_id = str(item.id)

        response = self.client.get("/api/v1/admin/table")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["totals"]["returned_orders"], 1)
        row = payload["rows"][0]
        self.assertEqual(row["order_id"], order_id)
        self.assertEqual(row["item_id"], item_id)
        self.assertEqual(row["status_bucket"], "returned")
        self.assertEqual(row["scan_codes_count"], 2)
        self.assertEqual(row["return_reference"], "WH-R-OUT-1")
        self.assertEqual(row["return_status"], "returned")
        self.assertEqual(row["skladbot_request_number"], "WH-R-OUT-1")
        self.assertEqual(row["skladbot_return_request_number"], "WH-R-RET-1")
        self.assertEqual(row["skladbot_return_request_id"], "190777")
        self.assertEqual(row["skladbot_return_status"], "created")

    def test_admin_table_totals_are_not_limited_by_row_limit(self):
        first_order_id, _first_item_id = self.seed_order(quantity_blocks=3)
        second_order_id, _second_item_id = self.seed_order(quantity_blocks=4)

        response = self.client.get("/api/v1/admin/table?limit=1")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["totals"]["orders"], 2)
        self.assertEqual(payload["totals"]["items"], 2)
        self.assertEqual(payload["totals"]["planned_blocks"], 7)
        self.assertEqual({first_order_id, second_order_id}, {row["order_id"] for row in self.client.get("/api/v1/admin/table").json()["rows"]})

    def test_admin_table_supports_offset_pagination_metadata(self):
        created_at_base = datetime(2026, 6, 1, 8, 0, tzinfo=timezone.utc)
        with self.SessionLocal() as db:
            for index, client in enumerate(("Alpha", "Bravo", "Charlie")):
                order = Order(
                    payment_type="cash",
                    client=client,
                    address=f"{client} Address",
                    representative="Admin Rep",
                    order_date=date(2026, 6, 2),
                    status="not_completed",
                    created_at=created_at_base + timedelta(minutes=index),
                    raw_payload={"source": "pagination-test"},
                )
                item = OrderItem(
                    order=order,
                    product=f"{client} Product",
                    quantity_pieces=10,
                    quantity_blocks=index + 1,
                    pieces_per_block=10,
                    scanned_blocks=0,
                    status="not_completed",
                    created_at=created_at_base + timedelta(minutes=index),
                    raw_payload={"line_total": 1000 * (index + 1)},
                )
                db.add(item)
            db.commit()

        response = self.client.get("/api/v1/admin/table?limit=1&offset=1")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["limit"], 1)
        self.assertEqual(payload["offset"], 1)
        self.assertEqual(payload["row_count"], 1)
        self.assertEqual(payload["total_rows"], 3)
        self.assertTrue(payload["has_more"])
        self.assertEqual(payload["rows"][0]["client"], "Bravo")
        self.assertEqual(payload["totals"]["orders"], 3)
        self.assertEqual(payload["totals"]["items"], 3)
        self.assertEqual(payload["totals"]["planned_blocks"], 6)

        tail = self.client.get("/api/v1/admin/table?limit=2&offset=2")
        self.assertEqual(tail.status_code, 200)
        tail_payload = tail.json()
        self.assertEqual(tail_payload["offset"], 2)
        self.assertEqual(tail_payload["row_count"], 1)
        self.assertFalse(tail_payload["has_more"])
        self.assertEqual(tail_payload["rows"][0]["client"], "Charlie")

    def test_admin_client_points_lists_order_points_and_updates_timeslot(self):
        self.seed_order()

        initial = self.client.get("/api/v1/admin/client-points")

        self.assertEqual(initial.status_code, 200)
        points = initial.json()
        self.assertEqual(len(points), 1)
        self.assertEqual(points[0]["client_name"], "Test Client")
        self.assertEqual(points[0]["address"], "Test Address")
        self.assertEqual(points[0]["delivery_from"], "10:00")
        self.assertEqual(points[0]["delivery_to"], "18:00")
        self.assertFalse(points[0]["is_saved"])
        self.assertFalse(points[0]["has_custom_timeslot"])

        updated = self.client.post(
            "/api/v1/admin/client-points/timeslot",
            json={
                "client_name": "Test Client",
                "address": "Test Address",
                "delivery_from": "09:30",
                "delivery_to": "12:00",
                "actor": "web",
                "reason": "точка принимает до обеда",
            },
        )

        self.assertEqual(updated.status_code, 200)
        payload = updated.json()
        self.assertEqual(payload["delivery_from"], "09:30")
        self.assertEqual(payload["delivery_to"], "12:00")
        self.assertTrue(payload["is_saved"])
        self.assertTrue(payload["has_custom_timeslot"])
        custom = self.client.get("/api/v1/admin/client-points?custom_timeslot=true")
        self.assertEqual(custom.status_code, 200)
        self.assertEqual(len(custom.json()), 1)
        with self.SessionLocal() as db:
            point = db.execute(select(ClientPoint)).scalar_one()
            self.assertEqual(point.delivery_from, "09:30")
            self.assertEqual(point.delivery_to, "12:00")
            audit = db.execute(
                select(AuditLog).where(AuditLog.action == "client_point_timeslot_updated")
            ).scalar_one()
            self.assertEqual(audit.entity_id, str(point.id))

    def test_admin_client_points_default_response_is_not_capped(self):
        with self.SessionLocal() as db:
            db.add_all([
                ClientPoint(
                    client_name=f"Client {index:04d}",
                    address=f"Address {index:04d}",
                    normalized_client=f"client {index:04d}",
                    normalized_address=f"address {index:04d}",
                    delivery_from="10:00",
                    delivery_to="18:00",
                    raw_payload={},
                )
                for index in range(1001)
            ])
            db.commit()

        response = self.client.get("/api/v1/admin/client-points")
        limited_response = self.client.get("/api/v1/admin/client-points?limit=3")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 1001)
        self.assertEqual(limited_response.status_code, 200)
        self.assertEqual(len(limited_response.json()), 3)

    def test_admin_client_point_order_summary_groups_dates_and_products(self):
        with self.SessionLocal() as db:
            older_order = Order(
                payment_type="cash",
                client="History Client",
                address="History Address",
                representative="Rep",
                order_date=date(2026, 6, 20),
                status="not_completed",
                raw_payload={"source": "test"},
            )
            newer_order = Order(
                payment_type="cash",
                client="History Client",
                address="Changed Address",
                representative="Rep",
                order_date=date(2026, 6, 21),
                status="not_completed",
                raw_payload={"source": "test"},
            )
            other_client_order = Order(
                payment_type="cash",
                client="Other Client",
                address="Other Address",
                representative="Rep",
                order_date=date(2026, 6, 21),
                status="not_completed",
                raw_payload={"source": "test"},
            )
            db.add_all([
                older_order,
                newer_order,
                other_client_order,
                OrderItem(
                    order=older_order,
                    product="Chapman Green OP 20",
                    quantity_pieces=30,
                    quantity_blocks=3,
                    pieces_per_block=10,
                    status="not_completed",
                    raw_payload={"source": "test"},
                ),
                OrderItem(
                    order=older_order,
                    product="Chapman Brown SSL 100`20",
                    quantity_pieces=20,
                    quantity_blocks=2,
                    pieces_per_block=10,
                    status="not_completed",
                    raw_payload={"source": "test"},
                ),
                OrderItem(
                    order=newer_order,
                    product="Chapman Green OP 20",
                    quantity_pieces=50,
                    quantity_blocks=5,
                    pieces_per_block=10,
                    status="not_completed",
                    raw_payload={"source": "test"},
                ),
                OrderItem(
                    order=other_client_order,
                    product="Ignored Product",
                    quantity_pieces=90,
                    quantity_blocks=9,
                    pieces_per_block=10,
                    status="not_completed",
                    raw_payload={"source": "test"},
                ),
            ])
            db.commit()

        response = self.client.get(
            "/api/v1/admin/client-points/order-summary",
            params={"client_name": "history client"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["client_name"], "History Client")
        self.assertEqual(payload["totals"], {
            "orders_count": 2,
            "positions_count": 3,
            "quantity_blocks": 10,
            "quantity_pieces": 100,
        })
        self.assertEqual([row["shipment_date"] for row in payload["dates"]], ["2026-06-21", "2026-06-20"])
        self.assertEqual([row["payment_type"] for row in payload["dates"]], ["cash", "cash"])
        self.assertEqual(payload["dates"][0]["orders_count"], 1)
        self.assertEqual(payload["dates"][0]["positions_count"], 1)
        self.assertEqual(payload["dates"][0]["products"], [{
            "product": "Chapman Green OP 20",
            "positions_count": 1,
            "quantity_blocks": 5,
            "quantity_pieces": 50,
        }])
        self.assertEqual(payload["dates"][1]["orders_count"], 1)
        self.assertEqual(payload["dates"][1]["positions_count"], 2)
        self.assertEqual(
            {product["product"]: product["quantity_blocks"] for product in payload["dates"][1]["products"]},
            {
                "Chapman Brown SSL 100`20": 2,
                "Chapman Green OP 20": 3,
            },
        )

    def test_admin_client_points_rejects_invalid_timeslot_order(self):
        response = self.client.post(
            "/api/v1/admin/client-points/timeslot",
            json={
                "client_name": "Client",
                "address": "Address",
                "delivery_from": "18:00",
                "delivery_to": "10:00",
            },
        )

        self.assertEqual(response.status_code, 422)
        self.assertIn("delivery_from must be earlier", response.json()["detail"])

    def test_admin_client_points_use_client_identity_when_address_changes(self):
        first = self.client.post(
            "/api/v1/admin/client-points/timeslot",
            json={
                "client_name": "Point Client",
                "address": "Old Address",
                "delivery_from": "08:30",
                "delivery_to": "11:45",
                "actor": "web",
            },
        )
        self.assertEqual(first.status_code, 200)

        second = self.client.post(
            "/api/v1/admin/client-points/timeslot",
            json={
                "client_name": "Point Client",
                "address": "New Address",
                "delivery_from": "09:00",
                "delivery_to": "12:00",
                "actor": "web",
            },
        )

        self.assertEqual(second.status_code, 200)
        self.assertEqual(second.json()["id"], first.json()["id"])
        self.assertEqual(second.json()["address"], "New Address")
        with self.SessionLocal() as db:
            points = db.execute(select(ClientPoint)).scalars().all()
            self.assertEqual(len(points), 1)
            self.assertEqual(points[0].client_name, "Point Client")
            self.assertEqual(points[0].address, "New Address")
            self.assertEqual(points[0].delivery_from, "09:00")

    def test_web_auth_login_sets_cookie_and_check_accepts_session(self):
        auth_settings = load_settings({
            "TAKSKLAD_ENV": "local",
            "TAKSKLAD_API_TOKEN": "service-token",
            "TAKSKLAD_WEB_LOGIN": "998000000000",
            "TAKSKLAD_WEB_PASSWORD_HASH": hash_password("test-password", salt="test-salt", iterations=1000),
            "TAKSKLAD_WEB_SESSION_SECRET": "test-session-secret",
            "TAKSKLAD_WEB_COOKIE_SECURE": "false",
        })

        with mock.patch("backend.app.main.settings", auth_settings):
            session_before = self.client.get("/api/v1/auth/session")
            self.assertEqual(session_before.status_code, 200)
            self.assertFalse(session_before.json()["authenticated"])

            bad_login = self.client.post(
                "/api/v1/auth/login",
                json={"login": "998000000000", "password": "wrong-password"},
            )
            self.assertEqual(bad_login.status_code, 401)
            self.assertNotIn(SESSION_COOKIE_NAME, bad_login.cookies)

            login = self.client.post(
                "/api/v1/auth/login",
                json={"login": "998000000000", "password": "test-password"},
            )
            self.assertEqual(login.status_code, 200)
            self.assertTrue(login.json()["authenticated"])
            self.assertEqual(login.json()["login"], "998000000000")
            self.assertIn(SESSION_COOKIE_NAME, login.cookies)
            set_cookie = login.headers["set-cookie"]
            self.assertIn("HttpOnly", set_cookie)
            self.assertIn("SameSite=lax", set_cookie)

            check = self.client.get("/api/v1/auth/check")
            self.assertEqual(check.status_code, 204)

            logout = self.client.post("/api/v1/auth/logout")
            self.assertEqual(logout.status_code, 200)
            self.assertFalse(logout.json()["authenticated"])

            check_after_logout = self.client.get("/api/v1/auth/check")
            self.assertEqual(check_after_logout.status_code, 401)

    def test_web_auth_session_allows_admin_api_without_service_token(self):
        app.dependency_overrides.pop(require_service_token, None)
        self.seed_order()
        auth_settings = load_settings({
            "TAKSKLAD_ENV": "local",
            "TAKSKLAD_API_TOKEN": "service-token",
            "TAKSKLAD_WEB_LOGIN": "998000000000",
            "TAKSKLAD_WEB_PASSWORD_HASH": hash_password("test-password", salt="test-salt", iterations=1000),
            "TAKSKLAD_WEB_SESSION_SECRET": "test-session-secret",
            "TAKSKLAD_WEB_COOKIE_SECURE": "false",
        })

        with mock.patch("backend.app.main.settings", auth_settings):
            unauthorized = self.client.get("/api/v1/admin/table")
            self.assertEqual(unauthorized.status_code, 401)
            unauthorized_events = self.client.get("/api/v1/admin/events")
            self.assertEqual(unauthorized_events.status_code, 401)

            login = self.client.post(
                "/api/v1/auth/login",
                json={"login": "998000000000", "password": "test-password"},
            )
            self.assertEqual(login.status_code, 200)

            admin = self.client.get("/api/v1/admin/table")
            self.assertEqual(admin.status_code, 200)
            self.assertEqual(len(admin.json()["rows"]), 1)
            events = self.client.get("/api/v1/admin/events")
            self.assertEqual(events.status_code, 200)

    def test_web_auth_configured_without_service_token_still_requires_session(self):
        app.dependency_overrides.pop(require_service_token, None)
        self.seed_order()
        auth_settings = load_settings({
            "TAKSKLAD_ENV": "local",
            "TAKSKLAD_API_TOKEN": "",
            "TAKSKLAD_WEB_LOGIN": "998000000000",
            "TAKSKLAD_WEB_PASSWORD_HASH": hash_password("test-password", salt="test-salt", iterations=1000),
            "TAKSKLAD_WEB_SESSION_SECRET": "test-session-secret",
            "TAKSKLAD_WEB_COOKIE_SECURE": "false",
        })

        with mock.patch("backend.app.main.settings", auth_settings):
            unauthorized = self.client.get("/api/v1/admin/table")
            self.assertEqual(unauthorized.status_code, 401)

            login = self.client.post(
                "/api/v1/auth/login",
                json={"login": "998000000000", "password": "test-password"},
            )
            self.assertEqual(login.status_code, 200)

            admin = self.client.get("/api/v1/admin/table")
            self.assertEqual(admin.status_code, 200)
            self.assertEqual(len(admin.json()["rows"]), 1)

    def test_logistics_slots_user_can_write_only_client_point_timeslots(self):
        app.dependency_overrides.pop(require_service_token, None)
        app.dependency_overrides.pop(require_admin_write_permission, None)
        app.dependency_overrides.pop(require_client_points_write_permission, None)
        order_id, _item_id = self.seed_order()
        with self.SessionLocal() as db:
            db.add(User(
                username="998933456753",
                password_hash=hash_password("limited-password", salt="limited-salt", iterations=1000),
                role="logistics_slots",
                is_active=True,
            ))
            db.commit()
        auth_settings = load_settings({
            "TAKSKLAD_ENV": "local",
            "TAKSKLAD_API_TOKEN": "service-token",
            "TAKSKLAD_WEB_LOGIN": "998000000000",
            "TAKSKLAD_WEB_PASSWORD_HASH": hash_password("admin-password", salt="admin-salt", iterations=1000),
            "TAKSKLAD_WEB_SESSION_SECRET": "test-session-secret",
            "TAKSKLAD_WEB_COOKIE_SECURE": "false",
        })

        with mock.patch("backend.app.main.settings", auth_settings):
            login = self.client.post(
                "/api/v1/auth/login",
                json={"login": "998933456753", "password": "limited-password"},
            )
            self.assertEqual(login.status_code, 200)
            login_payload = login.json()
            self.assertTrue(login_payload["authenticated"])
            self.assertEqual(login_payload["login"], "998933456753")
            self.assertEqual(login_payload["role"], "logistics_slots")
            self.assertEqual(login_payload["permissions"], ["client_points:write"])

            table = self.client.get("/api/v1/admin/table")
            self.assertEqual(table.status_code, 200)

            timeslot = self.client.post(
                "/api/v1/admin/client-points/timeslot",
                json={
                    "client_name": "Test Client",
                    "address": "Test Address",
                    "delivery_from": "09:00",
                    "delivery_to": "12:00",
                    "actor": "web",
                    "reason": "limited user update",
                },
            )
            self.assertEqual(timeslot.status_code, 200)
            self.assertEqual(timeslot.json()["delivery_from"], "09:00")

            cancel = self.client.post(
                f"/api/v1/admin/orders/{order_id}/cancel",
                json={"reason": "should be forbidden", "actor": "web"},
            )
            self.assertEqual(cancel.status_code, 403)

            import_create = self.client.post("/api/v1/imports", json={"source": "excel", "rows": []})
            self.assertEqual(import_create.status_code, 403)

            reconciliation = self.client.get("/api/v1/reports/reconciliation/day?report_date=2026-06-10")
            self.assertEqual(reconciliation.status_code, 403)

    def test_api_allows_local_no_auth_only_when_no_auth_is_configured(self):
        app.dependency_overrides.pop(require_service_token, None)
        self.seed_order()
        auth_settings = load_settings({
            "TAKSKLAD_ENV": "local",
            "TAKSKLAD_API_TOKEN": "",
            "TAKSKLAD_WEB_LOGIN": "",
            "TAKSKLAD_WEB_PASSWORD_HASH": "",
        })

        with mock.patch("backend.app.main.settings", auth_settings):
            admin = self.client.get("/api/v1/admin/table")
            self.assertEqual(admin.status_code, 200)
            self.assertEqual(len(admin.json()["rows"]), 1)

    def test_reset_order_for_rescan_clears_scans_and_keeps_order_active(self):
        order_id, item_id = self.seed_order(quantity_blocks=2, scanned_blocks=2, item_status="completed")
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            item.scan_codes = [
                ScanCode(order_item_id=item.id, code="010000000001"),
                ScanCode(order_item_id=item.id, code="010000000002"),
            ]
            order = db.get(Order, uuid.UUID(order_id))
            order.status = "completed"
            db.commit()

        response = self.client.post(
            f"/api/v1/admin/orders/{order_id}/reset-rescan",
            json={"reason": "Wrong scan batch", "actor": "anton"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "not_completed")
        self.assertEqual(response.json()["items"][0]["scanned_blocks"], 0)
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 0)
            self.assertEqual(item.status, "not_completed")
            self.assertEqual(len(db.execute(select(ScanCode)).scalars().all()), 0)
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertIn("order_reset_for_rescan", actions)
            event = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalar_one()
            self.assertEqual(event.payload["action"], "google_sheets_restore_order_export")

    def test_reset_order_for_rescan_rejects_returned_order(self):
        order_id, _item_id = self.seed_order(status="returned", item_status="returned")

        response = self.client.post(
            f"/api/v1/admin/orders/{order_id}/reset-rescan",
            json={"reason": "Wrong order", "actor": "anton"},
        )

        self.assertEqual(response.status_code, 409)

    def test_restore_order_returns_cancelled_order_to_active(self):
        order_id, _item_id = self.seed_order(status="cancelled", item_status="cancelled")

        response = self.client.post(
            f"/api/v1/admin/orders/{order_id}/restore",
            json={"reason": "Wrong cancellation", "actor": "anton"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "not_completed")
        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        self.assertEqual(active.json()[0]["id"], order_id)
        with self.SessionLocal() as db:
            event = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalar_one()
            self.assertEqual(event.payload["action"], "google_sheets_restore_order_export")
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertIn("order_restored", actions)

    def test_resync_order_skladbot_keeps_cached_number_until_worker_updates_it(self):
        order_id, _item_id = self.seed_order()
        with self.SessionLocal() as db:
            order = db.get(Order, uuid.UUID(order_id))
            order.raw_payload = {
                **(order.raw_payload or {}),
                "skladbot_request_number": "WH-R-OLD",
                "skladbot_request_id": "OLD",
                "skladbot_status": "found",
            }
            db.commit()

        with mock.patch("backend.app.skladbot_worker.update_orders_from_skladbot", return_value={"status": "completed"}) as sync:
            response = self.client.post(
                f"/api/v1/admin/orders/{order_id}/resync-skladbot",
                json={"reason": "Retry match", "actor": "anton"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["skladbot_request_number"], "WH-R-OLD")
        sync.assert_called_once()
        with self.SessionLocal() as db:
            order = db.get(Order, uuid.UUID(order_id))
            self.assertEqual(order.raw_payload["skladbot_request_number"], "WH-R-OLD")
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertIn("order_skladbot_resync_requested", actions)

    def test_admin_state_changing_actions_require_reason(self):
        archive_id, _ = self.seed_order()
        cancel_id, _ = self.seed_order()
        delete_id, _ = self.seed_order()
        reset_id, _ = self.seed_order()
        restore_id, _ = self.seed_order(status="cancelled", item_status="cancelled")
        resync_google_id, _ = self.seed_order()
        resync_skladbot_id, _ = self.seed_order()
        bulk_id, _ = self.seed_order()

        checks = [
            ("post", f"/api/v1/admin/orders/{archive_id}/archive-without-kiz", {"actor": "anton"}),
            ("post", f"/api/v1/admin/orders/{cancel_id}/cancel", {"actor": "anton"}),
            ("post", f"/api/v1/admin/orders/{delete_id}/delete-active", {"actor": "anton"}),
            ("post", f"/api/v1/admin/orders/{reset_id}/reset-rescan", {"actor": "anton"}),
            ("post", f"/api/v1/admin/orders/{restore_id}/restore", {"actor": "anton"}),
            ("post", f"/api/v1/admin/orders/{resync_google_id}/resync-google", {"actor": "anton"}),
            ("post", f"/api/v1/admin/orders/{resync_skladbot_id}/resync-skladbot", {"actor": "anton"}),
            ("post", "/api/v1/admin/orders/bulk/complete-without-kiz", {"order_ids": [bulk_id], "actor": "anton"}),
        ]

        for _method, path, body in checks:
            with self.subTest(path=path):
                response = self.client.post(path, json=body)
                self.assertEqual(response.status_code, 422)
                self.assertEqual(response.json()["detail"], "Reason is required")

    def test_archive_without_kiz_moves_unscanned_order_out_of_active_without_marking_completed(self):
        order_id, item_id = self.seed_order(quantity_blocks=3)

        response = self.client.post(
            f"/api/v1/admin/orders/{order_id}/archive-without-kiz",
            json={"reason": "Emergency close", "actor": "anton"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "archived_no_kiz")
        self.assertEqual(payload["items"][0]["status"], "archived_no_kiz")

        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        self.assertEqual(active.json(), [])

        admin = self.client.get("/api/v1/admin/table")
        self.assertEqual(admin.status_code, 200)
        self.assertEqual(admin.json()["rows"][0]["status_bucket"], "archive_no_kiz")

        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 0)
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertIn("order_archived_without_kiz", actions)
            self.assertNotIn("order_completed", actions)
            audit = db.execute(
                select(AuditLog).where(AuditLog.action == "order_archived_without_kiz")
            ).scalar_one()
            self.assertEqual(audit.payload["action"], "order_archived_without_kiz")
            self.assertEqual(audit.payload["actor"], "anton")
            self.assertEqual(audit.payload["source"], "anton")
            self.assertEqual(audit.payload["reason"], "Emergency close")
            self.assertEqual(audit.payload["affected_order_ids"], [order_id])
            self.assertEqual(audit.payload["affected_item_ids"], [item_id])
            self.assertTrue(audit.payload["timestamp"])
            self.assertEqual(audit.payload["raw_context"]["order_status"], "archived_no_kiz")
            event = db.execute(select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")).scalar_one()
            self.assertEqual(event.status, "pending")
            self.assertEqual(event.payload["action"], "google_sheets_archive_no_kiz_export")

    def test_archive_without_kiz_rejects_order_with_scans(self):
        order_id, _item_id = self.seed_order(scanned_blocks=1)

        response = self.client.post(
            f"/api/v1/admin/orders/{order_id}/archive-without-kiz",
            json={"reason": "Emergency close", "actor": "anton"},
        )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["detail"]["message"], "Order already has scanned KIZ codes")

    def test_bulk_complete_without_kiz_marks_unscanned_orders_completed_and_queues_archive_export(self):
        first_order_id, _first_item_id = self.seed_order(quantity_blocks=3)
        second_order_id, _second_item_id = self.seed_order(quantity_blocks=1)

        response = self.client.post(
            "/api/v1/admin/orders/bulk/complete-without-kiz",
            json={
                "order_ids": [first_order_id, second_order_id],
                "reason": "Manual close without scans",
                "actor": "anton",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["completed"], 2)
        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        self.assertEqual(active.json(), [])

        admin = self.client.get("/api/v1/admin/table")
        self.assertEqual(admin.status_code, 200)
        self.assertEqual({row["status_bucket"] for row in admin.json()["rows"]}, {"archive"})
        with self.SessionLocal() as db:
            orders = db.execute(select(Order)).scalars().all()
            self.assertEqual({order.status for order in orders}, {"completed"})
            items = db.execute(select(OrderItem)).scalars().all()
            self.assertEqual({item.status for item in items}, {"completed"})
            self.assertEqual({item.scanned_blocks for item in items}, {0})
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertEqual(actions.count("order_completed_without_kiz"), 2)
            events = db.execute(select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")).scalars().all()
            self.assertEqual(len(events), 2)
            self.assertEqual({event.payload["action"] for event in events}, {"google_sheets_archive_export"})

    def test_bulk_complete_without_kiz_ignores_pending_skladbot_export(self):
        order_id, _item_id = self.seed_order(quantity_blocks=3)
        with self.SessionLocal() as db:
            db.add(PendingEvent(
                event_type="google_sheets_export",
                status="pending",
                payload={
                    "action": "google_sheets_skladbot_export",
                    "entity_id": "skladbot",
                    "order_ids": [order_id],
                },
            ))
            db.commit()

        response = self.client.post(
            "/api/v1/admin/orders/bulk/complete-without-kiz",
            json={
                "order_ids": [order_id],
                "reason": "Manual close without scans",
                "actor": "anton",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["completed"], 1)
        with self.SessionLocal() as db:
            order = db.get(Order, uuid.UUID(order_id))
            self.assertEqual(order.status, "completed")

    def test_bulk_complete_without_kiz_allows_partially_scanned_order_and_preserves_scans(self):
        clean_order_id, _clean_item_id = self.seed_order(quantity_blocks=3)
        scanned_order_id, scanned_item_id = self.seed_order(quantity_blocks=3, scanned_blocks=1)

        response = self.client.post(
            "/api/v1/admin/orders/bulk/complete-without-kiz",
            json={
                "order_ids": [clean_order_id, scanned_order_id],
                "reason": "Manual completed shipment",
                "actor": "anton",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["completed"], 2)
        with self.SessionLocal() as db:
            clean_order = db.get(Order, uuid.UUID(clean_order_id))
            scanned_order = db.get(Order, uuid.UUID(scanned_order_id))
            scanned_item = db.get(OrderItem, uuid.UUID(scanned_item_id))
            self.assertEqual(clean_order.status, "completed")
            self.assertEqual(scanned_order.status, "completed")
            self.assertEqual(scanned_item.status, "completed")
            self.assertEqual(scanned_item.scanned_blocks, 1)
            events = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalars().all()
            self.assertEqual(len(events), 2)
            self.assertEqual({event.payload["action"] for event in events}, {"google_sheets_archive_export"})

    def test_bulk_complete_without_kiz_allows_fully_scanned_order(self):
        order_id, _item_id = self.seed_order(quantity_blocks=2, scanned_blocks=2)

        response = self.client.post(
            "/api/v1/admin/orders/bulk/complete-without-kiz",
            json={
                "order_ids": [order_id],
                "reason": "Manual close completed order",
                "actor": "anton",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["completed"], 1)
        with self.SessionLocal() as db:
            order = db.get(Order, uuid.UUID(order_id))
            item = db.execute(select(OrderItem).where(OrderItem.order_id == uuid.UUID(order_id))).scalar_one()
            self.assertEqual(order.status, "completed")
            self.assertEqual(item.status, "completed")
            self.assertEqual(item.scanned_blocks, 2)

    def test_cancel_order_requires_no_scans_and_queues_google_export_when_google_is_down(self):
        order_id, _item_id = self.seed_order(quantity_blocks=4)

        response = self.client.post(
            f"/api/v1/admin/orders/{order_id}/cancel",
            json={"reason": "Wrong import", "actor": "anton"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "cancelled")
        with self.SessionLocal() as db:
            event = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalar_one()
            self.assertEqual(event.status, "pending")
            self.assertEqual(event.payload["action"], "google_sheets_cancel_export")
            self.assertEqual(event.payload["entity_id"], order_id)
            self.assertEqual(event.last_error, "")

    def test_delete_active_order_removes_unscanned_order_and_queues_google_delete(self):
        order_id, item_id = self.seed_order(quantity_blocks=4)

        response = self.client.post(
            f"/api/v1/admin/orders/{order_id}/delete-active",
            json={"reason": "Manual Telegram delete", "actor": "telegram"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["order_id"], order_id)
        self.assertEqual(payload["deleted"], True)
        self.assertEqual(payload["skladbot_request_number"], "")
        with self.SessionLocal() as db:
            self.assertIsNone(db.get(Order, uuid.UUID(order_id)))
            self.assertEqual(
                db.execute(select(OrderItem).where(OrderItem.id == uuid.UUID(item_id))).scalars().all(),
                [],
            )
            event = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalar_one()
            self.assertEqual(event.status, "pending")
            self.assertEqual(event.payload["action"], "google_sheets_delete_import_records_export")
            self.assertEqual(event.payload["entity_id"], order_id)
            self.assertEqual(len(event.payload["records"]), 1)
            self.assertEqual(event.payload["records"][0]["ID заказа"], order_id)
            actions = [item.action for item in db.execute(select(AuditLog)).scalars()]
            self.assertIn("order_deleted_from_active", actions)

    def test_delete_active_order_idempotency_prevents_duplicate_audit_and_export(self):
        order_id, _item_id = self.seed_order(quantity_blocks=4)
        body = {
            "reason": "Manual Telegram delete",
            "actor": "telegram",
            "source": "telegram",
            "idempotency_key": "delete-active-key",
        }

        first = self.client.post(f"/api/v1/admin/orders/{order_id}/delete-active", json=body)
        second = self.client.post(f"/api/v1/admin/orders/{order_id}/delete-active", json=body)

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(second.json()["message"], "Order delete already processed for this idempotency key")
        with self.SessionLocal() as db:
            self.assertEqual(
                len(db.execute(select(AuditLog).where(AuditLog.action == "order_deleted_from_active")).scalars().all()),
                1,
            )
            self.assertEqual(
                len(db.execute(select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")).scalars().all()),
                1,
            )

    def test_delete_active_order_rejects_order_with_scans(self):
        order_id, _item_id = self.seed_order(quantity_blocks=4, scanned_blocks=1)

        response = self.client.post(
            f"/api/v1/admin/orders/{order_id}/delete-active",
            json={"reason": "Manual Telegram delete", "actor": "telegram"},
        )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["detail"]["message"], "Order already has scanned KIZ codes")
        with self.SessionLocal() as db:
            self.assertIsNotNone(db.get(Order, uuid.UUID(order_id)))
            self.assertEqual(db.execute(select(PendingEvent)).scalars().all(), [])

    def test_resync_google_for_active_order_pushes_items_without_archiving_order(self):
        order_id, item_id = self.seed_order(quantity_blocks=2)

        response = self.client.post(
            f"/api/v1/admin/orders/{order_id}/resync-google",
            json={"reason": "Manual resync", "actor": "anton"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "not_completed")
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.status, "not_completed")
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertIn("order_google_resync_requested", actions)
            self.assertIn("google_sheets_scan_export", actions)
            event = db.execute(select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")).scalar_one()
            self.assertEqual(event.payload["action"], "google_sheets_scan_export")
            self.assertEqual(event.payload["entity_id"], item_id)

    def test_bulk_complete_without_kiz_idempotency_prevents_duplicate_audit_and_export(self):
        order_id, _item_id = self.seed_order(quantity_blocks=2)
        body = {
            "order_ids": [order_id],
            "reason": "Manual close",
            "actor": "anton",
            "idempotency_key": "bulk-complete-key",
        }

        first = self.client.post("/api/v1/admin/orders/bulk/complete-without-kiz", json=body)
        second = self.client.post("/api/v1/admin/orders/bulk/complete-without-kiz", json=body)

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        with self.SessionLocal() as db:
            self.assertEqual(
                len(db.execute(select(AuditLog).where(AuditLog.action == "order_completed_without_kiz")).scalars().all()),
                1,
            )
            self.assertEqual(
                len(db.execute(select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")).scalars().all()),
                1,
            )

    def test_sync_sources_runs_google_sheet_sync_then_skladbot_sync(self):
        with mock.patch("backend.app.main.sync_google_sheet_to_backend") as google_sync, mock.patch(
            "backend.app.main.update_orders_from_skladbot"
        ) as skladbot_sync:
            skladbot_sync.return_value = {
                "requests": 3,
                "updated": 2,
                "matched": 1,
                "not_found": 1,
                "multiple": 0,
            }

            response = self.client.post("/api/v1/sync/sources?wait_skladbot=1")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "completed")
        self.assertEqual(payload["google_sheets"]["status"], "skipped")
        self.assertEqual(payload["skladbot"]["status"], "completed")
        self.assertEqual(payload["skladbot"]["matched"], 1)
        google_sync.assert_not_called()
        skladbot_sync.assert_called_once()

    def test_sync_sources_can_skip_skladbot(self):
        with mock.patch("backend.app.main.sync_google_sheet_to_backend") as google_sync, mock.patch(
            "backend.app.main.update_orders_from_skladbot"
        ) as skladbot_sync:
            google_sync.return_value = {"rows": 1, "matched": 1, "orders_updated": 0, "items_updated": 0, "conflicts": 0}

            response = self.client.post("/api/v1/sync/sources?skladbot=0")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "completed")
        self.assertEqual(payload["skladbot"]["status"], "skipped")
        self.assertEqual(payload["google_sheets"]["status"], "skipped")
        google_sync.assert_not_called()
        skladbot_sync.assert_not_called()

    def test_sync_sources_starts_skladbot_in_background_by_default(self):
        with mock.patch("backend.app.main.sync_google_sheet_to_backend") as google_sync, mock.patch(
            "backend.app.main.start_skladbot_sync_background"
        ) as start_skladbot:
            google_sync.return_value = {"rows": 1, "matched": 1, "orders_updated": 0, "items_updated": 0, "conflicts": 0}
            start_skladbot.return_value = {"status": "started", "message": "background"}

            response = self.client.post("/api/v1/sync/sources")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "completed")
        self.assertEqual(payload["skladbot"]["status"], "started")
        self.assertEqual(payload["google_sheets"]["status"], "skipped")
        google_sync.assert_not_called()
        start_skladbot.assert_called_once()

    def test_scan_create_is_idempotent_for_same_item_and_rejects_cross_order_duplicate(self):
        order_id, item_id = self.seed_order(product="Chapman RED OP 20")
        _, other_item_id = self.seed_order(product="Chapman RED OP 20")
        with self.SessionLocal() as db:
            order = db.get(Order, uuid.UUID(order_id))
            order.client = "OOO Busy Client"
            order.raw_payload = {"skladbot_request_number": "WH-R-100500"}
            db.commit()

        response = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "  0104006396053947217ABCDEF  ", "workstation_id": "pc-1"},
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["code"], "0104006396053947217ABCDEF")
        self.assertEqual(payload["scanned_blocks"], 1)
        self.assertEqual(payload["item_status"], "not_completed")

        same_item_duplicate = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "0104006396053947217ABCDEF", "workstation_id": "pc-2"},
        )
        self.assertEqual(same_item_duplicate.status_code, 201)
        self.assertEqual(same_item_duplicate.json()["order_item_id"], item_id)
        self.assertEqual(same_item_duplicate.json()["scanned_blocks"], 1)

        other_item_duplicate = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": other_item_id, "code": "0104006396053947217ABCDEF", "workstation_id": "pc-2"},
        )
        self.assertEqual(other_item_duplicate.status_code, 409)
        detail = other_item_duplicate.json()["detail"]
        self.assertEqual(detail["message"], "Code already scanned in another order item")
        self.assertEqual(detail["existing_order"]["client"], "OOO Busy Client")
        self.assertEqual(detail["existing_order"]["order_date_display"], "30.05.2026")
        self.assertEqual(detail["existing_order"]["product"], "Chapman RED OP 20")
        self.assertEqual(detail["existing_order"]["skladbot_request_number"], "WH-R-100500")

        with self.SessionLocal() as db:
            self.assertEqual(len(db.execute(select(ScanCode)).scalars().all()), 1)
            movements = db.execute(
                select(KizMovement)
                .join(KizCode, KizMovement.kiz_id == KizCode.id)
                .where(KizCode.code == "0104006396053947217ABCDEF")
            ).scalars().all()
            self.assertEqual([movement.movement_type for movement in movements], ["outbound"])
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 1)
            other_item = db.get(OrderItem, uuid.UUID(other_item_id))
            self.assertEqual(other_item.scanned_blocks, 0)

    def test_scan_create_is_idempotent_for_same_completed_item(self):
        _, item_id = self.seed_order(quantity_blocks=1)

        first = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "010123456789", "workstation_id": "pc-1"},
        )
        self.assertEqual(first.status_code, 201)
        self.assertEqual(first.json()["item_status"], "completed")

        duplicate = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "010123456789", "workstation_id": "pc-2"},
        )
        self.assertEqual(duplicate.status_code, 201)
        self.assertEqual(duplicate.json()["order_item_id"], item_id)
        self.assertEqual(duplicate.json()["scanned_blocks"], 1)

        with self.SessionLocal() as db:
            self.assertEqual(len(db.execute(select(ScanCode)).scalars().all()), 1)
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 1)
            self.assertEqual(item.status, "completed")

    def test_scan_create_acknowledges_extra_scan_when_item_already_full(self):
        _, item_id = self.seed_order(quantity_blocks=2)

        first = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "010123456789", "workstation_id": "pc-1"},
        )
        self.assertEqual(first.status_code, 201)
        self.assertEqual(first.json()["item_status"], "not_completed")
        second = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "010123456780", "workstation_id": "pc-1"},
        )
        self.assertEqual(second.status_code, 201)
        self.assertEqual(second.json()["scanned_blocks"], 2)
        self.assertEqual(second.json()["item_status"], "completed")

        extra = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "010987654321", "workstation_id": "pc-1"},
        )

        self.assertEqual(extra.status_code, 201)
        self.assertEqual(extra.json()["order_item_id"], item_id)
        self.assertEqual(extra.json()["code"], "010123456780")
        self.assertEqual(extra.json()["scanned_blocks"], 2)
        self.assertEqual(extra.json()["item_status"], "completed")

        with self.SessionLocal() as db:
            scans = db.execute(select(ScanCode)).scalars().all()
            self.assertEqual(len(scans), 2)
            self.assertEqual({scan.code for scan in scans}, {"010123456789", "010123456780"})
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 2)
            self.assertEqual(item.status, "completed")

    def test_scan_create_counts_aggregate_box_as_fifty_blocks(self):
        _, item_id = self.seed_order(quantity_blocks=150, product="Chapman Gold SSL 100`20")

        response = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "010400639605401221UZ1112022525522513824013040046110ZIG1218229310000",
                "workstation_id": "pc-1",
            },
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["scanned_blocks"], 50)
        self.assertEqual(payload["item_status"], "not_completed")
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            scan = db.execute(select(ScanCode)).scalar_one()
            self.assertEqual(item.scanned_blocks, 50)
            self.assertEqual(scan.raw_payload["scan_type"], "aggregate_box")
            self.assertEqual(scan.raw_payload["block_quantity"], 50)
            audit = db.execute(select(AuditLog).where(AuditLog.action == "scan_code_created")).scalar_one()
            self.assertEqual(audit.payload["scan_type"], "aggregate_box")
            self.assertEqual(audit.payload["block_quantity"], 50)

    def test_scan_create_accepts_aggregate_box_when_next_ai_is_not_serial(self):
        _, item_id = self.seed_order(quantity_blocks=100, product="Chapman Brown SSL 100`20")

        response = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "010400639605407410BATCH21BOX",
                "workstation_id": "pc-1",
            },
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["scanned_blocks"], 50)
        self.assertEqual(payload["item_status"], "not_completed")
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            scan = db.execute(select(ScanCode)).scalar_one()
            self.assertEqual(item.scanned_blocks, 50)
            self.assertEqual(scan.raw_payload["scan_type"], "aggregate_box")
            self.assertEqual(scan.raw_payload["block_quantity"], 50)
            self.assertEqual(scan.raw_payload["product_key"], "brown:ssl")

    def test_scan_create_accepts_live_green_aggregate_box_gtin(self):
        _, item_id = self.seed_order(quantity_blocks=100, product="Chapman Green OP 20")

        first = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "010400639610445821UZ1112042611905354024013040030510ZIG1233389310000",
                "workstation_id": "pc-1",
            },
        )
        second = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "010400639610445821UZ1112042611909232924013040030510ZIG1233389310000",
                "workstation_id": "pc-1",
            },
        )

        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)
        self.assertEqual(first.json()["scanned_blocks"], 50)
        self.assertEqual(first.json()["item_status"], "not_completed")
        self.assertEqual(second.json()["scanned_blocks"], 100)
        self.assertEqual(second.json()["item_status"], "completed")
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            scans = db.execute(select(ScanCode).order_by(ScanCode.scanned_at, ScanCode.id)).scalars().all()
            self.assertEqual(item.scanned_blocks, 100)
            self.assertEqual(item.status, "completed")
            self.assertEqual(len(scans), 2)
            self.assertEqual({scan.raw_payload["scan_type"] for scan in scans}, {"aggregate_box"})
            self.assertEqual({scan.raw_payload["block_quantity"] for scan in scans}, {50})
            self.assertEqual({scan.raw_payload["product_key"] for scan in scans}, {"green:op"})

    def test_scan_create_accepts_live_brown_ssl_aggregate_box_gtin(self):
        _, item_id = self.seed_order(quantity_blocks=100, product="Chapman Brown SSL 100`20")

        first = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "010400639605407421UZ1112022612417151624013040046310ZIG1231569310000",
                "workstation_id": "pc-1",
            },
        )
        second = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "010400639605407421UZ1112022612416594224013040046310ZIG1231569310000",
                "workstation_id": "pc-1",
            },
        )

        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)
        self.assertEqual(first.json()["scanned_blocks"], 50)
        self.assertEqual(first.json()["item_status"], "not_completed")
        self.assertEqual(second.json()["scanned_blocks"], 100)
        self.assertEqual(second.json()["item_status"], "completed")
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            scans = db.execute(select(ScanCode).order_by(ScanCode.scanned_at, ScanCode.id)).scalars().all()
            self.assertEqual(item.scanned_blocks, 100)
            self.assertEqual(item.status, "completed")
            self.assertEqual(len(scans), 2)
            self.assertEqual({scan.raw_payload["scan_type"] for scan in scans}, {"aggregate_box"})
            self.assertEqual({scan.raw_payload["block_quantity"] for scan in scans}, {50})
            self.assertEqual({scan.raw_payload["product_key"] for scan in scans}, {"brown:ssl"})

    def test_scan_create_rejects_aggregate_box_when_remaining_blocks_are_less_than_fifty(self):
        _, item_id = self.seed_order(quantity_blocks=30, product="Chapman Gold SSL 100`20")

        response = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "010400639605401221UZ1112022525522513824013040046110ZIG1218229310000",
            },
        )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["detail"]["message"], "Aggregate box exceeds remaining order item blocks")
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 0)
            self.assertEqual(db.execute(select(ScanCode)).scalars().all(), [])

    def test_scan_create_rejects_aggregate_box_for_wrong_product(self):
        _, item_id = self.seed_order(quantity_blocks=150, product="Chapman RED OP 20")

        response = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "010400639605401221UZ1112022525522513824013040046110ZIG1218229310000",
            },
        )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["detail"]["message"], "Aggregate box product does not match order item")
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 0)
            self.assertEqual(db.execute(select(ScanCode)).scalars().all(), [])

    def test_scan_create_rejects_unit_kiz_for_wrong_chapman_product(self):
        _, item_id = self.seed_order(quantity_blocks=1, product="Chapman Gold SSL 100`20")

        response = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "0104006396053947217p-30o933ZXHZKjx",
            },
        )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["detail"]["message"], "Scan product does not match order item")
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 0)
            self.assertEqual(item.status, "not_completed")
            self.assertEqual(db.execute(select(ScanCode)).scalars().all(), [])
            self.assertEqual(db.execute(select(KizCode)).scalars().all(), [])
            self.assertEqual(db.execute(select(KizMovement)).scalars().all(), [])

    def test_scan_undo_subtracts_aggregate_box_block_quantity(self):
        _, item_id = self.seed_order(quantity_blocks=51, product="Chapman Gold SSL 100`20")
        aggregate = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "010400639605401221UZ1112022525522513824013040046110ZIG1218229310000",
            },
        )
        unit = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "010400639605400521UNIT"},
        )
        self.assertEqual(aggregate.status_code, 201)
        self.assertEqual(unit.status_code, 201)
        self.assertEqual(unit.json()["scanned_blocks"], 51)
        self.assertEqual(unit.json()["item_status"], "completed")

        undo_response = self.client.post(
            "/api/v1/scans/undo",
            json={
                "order_item_id": item_id,
                "code": "010400639605401221UZ1112022525522513824013040046110ZIG1218229310000",
                "actor": "desktop",
            },
        )

        self.assertEqual(undo_response.status_code, 200)
        self.assertEqual(undo_response.json()["scanned_blocks"], 1)
        self.assertEqual(undo_response.json()["item_status"], "not_completed")
        with self.SessionLocal() as db:
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 1)
            self.assertEqual([scan.code for scan in item.scan_codes], ["010400639605400521UNIT"])

    def test_scan_create_exports_scan_state_to_google_sheets_best_effort(self):
        _, item_id = self.seed_order()

        response = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "010000000001"},
        )

        self.assertEqual(response.status_code, 201)

        with self.SessionLocal() as db:
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertIn("google_sheets_scan_export", actions)
            event = db.execute(select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")).scalar_one()
            self.assertEqual(event.status, "pending")
            self.assertEqual(event.payload["action"], "google_sheets_scan_export")
            self.assertEqual(event.payload["entity_id"], item_id)

    def test_scan_undo_removes_backend_scan_and_queues_google_projection(self):
        _order_id, item_id = self.seed_order(quantity_blocks=1)
        create_response = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "010000000001", "workstation_id": "pc-1"},
        )
        self.assertEqual(create_response.status_code, 201)
        self.assertEqual(create_response.json()["item_status"], "completed")

        undo_response = self.client.post(
            "/api/v1/scans/undo",
            json={"order_item_id": item_id, "code": "010000000001", "workstation_id": "pc-1", "actor": "desktop"},
        )

        self.assertEqual(undo_response.status_code, 200)
        self.assertEqual(undo_response.json()["scanned_blocks"], 0)
        self.assertEqual(undo_response.json()["item_status"], "not_completed")
        with self.SessionLocal() as db:
            self.assertEqual(db.execute(select(ScanCode)).scalars().all(), [])
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 0)
            self.assertEqual(item.status, "not_completed")
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertIn("scan_code_deleted", actions)
            events = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalars().all()
            self.assertEqual([event.payload["action"] for event in events], ["google_sheets_scan_export"])

    def test_scan_undo_rejects_completed_order(self):
        order_id, item_id = self.seed_order(quantity_blocks=1)
        self.client.post("/api/v1/scans", json={"order_item_id": item_id, "code": "010000000001"})
        self.client.post(f"/api/v1/orders/{order_id}/complete")

        response = self.client.post(
            "/api/v1/scans/undo",
            json={"order_item_id": item_id, "code": "010000000001"},
        )

        self.assertEqual(response.status_code, 409)

    def test_scan_create_queues_google_sheets_export_when_google_is_down(self):
        _, item_id = self.seed_order()

        response = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "010000000901"},
        )

        self.assertEqual(response.status_code, 201)
        with self.SessionLocal() as db:
            event = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalar_one()
            self.assertEqual(event.status, "pending")
            self.assertEqual(event.payload["action"], "google_sheets_scan_export")
            self.assertEqual(event.payload["entity_id"], item_id)
            self.assertEqual(event.last_error, "")
            audit = db.execute(
                select(AuditLog).where(AuditLog.action == "google_sheets_scan_export")
            ).scalar_one()
            self.assertTrue(audit.payload["queued"])

    def test_sync_sources_flushes_pending_google_exports_before_google_refresh(self):
        _, item_id = self.seed_order()
        call_order = []
        with self.SessionLocal() as db:
            db.add(PendingEvent(
                event_type="google_sheets_export",
                status="pending",
                attempts=0,
                payload={
                    "action": "google_sheets_scan_export",
                    "entity_type": "order_item",
                    "entity_id": item_id,
                    "last_result": {"status": "error", "error": "Google timeout"},
                },
                last_error="Google timeout",
            ))
            db.commit()

        def fake_export(_items):
            call_order.append("pending_google_export")
            return {"status": "completed", "updated": 1}

        def fake_google_sync(_db):
            call_order.append("google_sheet_to_backend")
            return {"rows": 1, "matched": 1, "orders_updated": 0, "items_updated": 0, "conflicts": 0}

        with mock.patch(
            "backend.app.google_sheets_pending.sync_backend_order_items_to_google_sheets",
            side_effect=fake_export,
        ), mock.patch(
            "backend.app.main.sync_google_sheet_to_backend",
            side_effect=fake_google_sync,
        ):
            response = self.client.post("/api/v1/sync/sources?skladbot=0")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["google_sheets_pending"]["synced"], 1)
        self.assertEqual(payload["google_sheets"]["status"], "skipped")
        self.assertEqual(call_order, ["pending_google_export"])
        with self.SessionLocal() as db:
            event = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalar_one()
            self.assertEqual(event.status, "completed")
            self.assertEqual(event.last_error, "")

    def test_pending_google_export_malformed_event_does_not_block_newer_scan_export(self):
        _, item_id = self.seed_order()
        from backend.app.google_sheets_pending import process_pending_google_sheets_exports

        with self.SessionLocal() as db:
            invalid = PendingEvent(
                event_type="google_sheets_export",
                status="pending",
                attempts=0,
                payload={
                    "action": "google_sheets_scan_export",
                    "entity_type": "order_item",
                    "entity_id": "not-a-uuid",
                },
                created_at=datetime.now(timezone.utc) - timedelta(minutes=10),
            )
            valid = PendingEvent(
                event_type="google_sheets_export",
                status="pending",
                attempts=0,
                payload={
                    "action": "google_sheets_scan_export",
                    "entity_type": "order_item",
                    "entity_id": item_id,
                },
                created_at=datetime.now(timezone.utc) - timedelta(minutes=5),
            )
            db.add_all([invalid, valid])
            db.commit()
            invalid_id = invalid.id
            valid_id = valid.id

            with mock.patch(
                "backend.app.google_sheets_pending.sync_backend_order_items_to_google_sheets",
                return_value={"status": "completed", "updated": 1},
            ):
                result = process_pending_google_sheets_exports(db, limit=10)

            invalid = db.get(PendingEvent, invalid_id)
            valid = db.get(PendingEvent, valid_id)

        self.assertEqual(result["status"], "completed_with_errors")
        self.assertEqual(result["synced"], 1)
        self.assertEqual(result["failed"], 1)
        self.assertEqual(invalid.status, "failed")
        self.assertIn("invalid order item id", invalid.last_error)
        self.assertEqual(valid.status, "completed")
        self.assertEqual(valid.last_error, "")

    def test_complete_order_requires_required_blocks_and_closes_order(self):
        order_id, item_id = self.seed_order()

        too_early = self.client.post(f"/api/v1/orders/{order_id}/complete")

        self.assertEqual(too_early.status_code, 409)
        self.assertEqual(too_early.json()["detail"]["message"], "Order has incomplete required items")

        for code in ["010000000001", "010000000002"]:
            scan = self.client.post("/api/v1/scans", json={"order_item_id": item_id, "code": code})
            self.assertEqual(scan.status_code, 201)

        completed = self.client.post(f"/api/v1/orders/{order_id}/complete")

        self.assertEqual(completed.status_code, 200)
        self.assertEqual(completed.json()["status"], "completed")
        self.assertEqual(completed.json()["items"][0]["status"], "completed")

        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        self.assertEqual(active.json(), [])

        with self.SessionLocal() as db:
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertEqual(actions.count("scan_code_created"), 2)
            self.assertIn("order_completed", actions)

    def test_dashboard_day_summary_counts_loaded_items_not_shipment_or_scan_date(self):
        loaded_at = datetime(2026, 5, 30, 9, 0, tzinfo=timezone.utc)
        old_loaded_at = datetime(2026, 5, 29, 9, 0, tzinfo=timezone.utc)
        with self.SessionLocal() as db:
            import_job = ImportJob(
                source="excel",
                status="completed",
                rows_total=1,
                rows_imported=1,
                created_at=loaded_at,
                raw_payload={"source": "test"},
            )
            db.add(import_job)
            db.flush()
            active_order = Order(
                payment_type="cash",
                client="Loaded Active",
                address="Address",
                representative="Rep",
                order_date=date(2026, 5, 31),
                status="not_completed",
                created_at=loaded_at,
                raw_payload={"source": "test"},
            )
            active_item = OrderItem(
                order=active_order,
                product="Product A",
                quantity_pieces=50,
                quantity_blocks=5,
                scanned_blocks=2,
                status="not_completed",
                created_at=old_loaded_at,
                raw_payload={"backend_import_id": str(import_job.id), "line_total": 1200000},
            )
            completed_order = Order(
                payment_type="cash",
                client="Loaded Completed",
                address="Address",
                representative="Rep",
                order_date=date(2026, 5, 30),
                status="completed",
                created_at=loaded_at,
                raw_payload={"source": "test"},
            )
            completed_item = OrderItem(
                order=completed_order,
                product="Product B",
                quantity_pieces=30,
                quantity_blocks=3,
                scanned_blocks=3,
                status="completed",
                created_at=loaded_at,
                raw_payload={"line_total": 720000},
            )
            old_order = Order(
                payment_type="cash",
                client="Old Loaded",
                address="Address",
                representative="Rep",
                order_date=date(2026, 5, 30),
                status="not_completed",
                created_at=old_loaded_at,
                raw_payload={"source": "test"},
            )
            old_item = OrderItem(
                order=old_order,
                product="Product C",
                quantity_pieces=20,
                quantity_blocks=2,
                scanned_blocks=1,
                status="not_completed",
                created_at=old_loaded_at,
                raw_payload={"line_total": 480000},
            )
            db.add_all([active_order, active_item, completed_order, completed_item, old_order, old_item])
            db.flush()
            db.add(ScanCode(
                order_item_id=old_item.id,
                code="0104006396053978217OLDLOAD001",
                scanned_at=loaded_at,
                raw_payload={"scanned_at": loaded_at.isoformat()},
            ))
            db.commit()

        response = self.client.get("/api/v1/admin/dashboard/day-summary?report_date=2026-05-30")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_date"], "2026-05-30")
        self.assertEqual(payload["source"], "postgres_loaded_items")
        self.assertEqual(payload["totals"]["orders"], 2)
        self.assertEqual(payload["totals"]["completed_orders"], 1)
        self.assertEqual(payload["totals"]["active_orders"], 1)
        self.assertEqual(payload["totals"]["planned_blocks"], 8)
        self.assertEqual(payload["totals"]["scanned_blocks"], 5)
        self.assertEqual(payload["totals"]["scanned_today"], 0)
        self.assertEqual(payload["totals"]["remaining_blocks"], 3)
        self.assertEqual(payload["totals"]["total_price"], 1920000)

    def test_complete_order_exports_archive_to_google_sheets_best_effort(self):
        order_id, item_id = self.seed_order()
        for code in ["010000000001", "010000000002"]:
            scan = self.client.post("/api/v1/scans", json={"order_item_id": item_id, "code": code})
            self.assertEqual(scan.status_code, 201)

        completed = self.client.post(f"/api/v1/orders/{order_id}/complete")

        self.assertEqual(completed.status_code, 200)

        with self.SessionLocal() as db:
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertIn("google_sheets_archive_export", actions)
            events = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalars().all()
            self.assertIn("google_sheets_archive_export", [event.payload["action"] for event in events])

    def test_return_lookup_and_mark_returned_excludes_order_from_active_list(self):
        order_id, item_id = self.seed_order(status="completed", scanned_blocks=2, item_status="completed")
        with self.SessionLocal() as db:
            order = db.get(Order, uuid.UUID(order_id))
            order.raw_payload = {
                "skladbot_request_number": "WH-R-RETURN-100",
                "skladbot_request_id": "100500",
            }
            db.commit()

        lookup = self.client.get("/api/v1/returns/lookup", params={"lookup": " WH-R-RETURN-100 "})

        self.assertEqual(lookup.status_code, 200)
        self.assertEqual(lookup.json()["id"], order_id)
        self.assertEqual(lookup.json()["skladbot_request_number"], "WH-R-RETURN-100")

        returned = self.client.post(
            f"/api/v1/returns/{order_id}",
            json={
                "return_reference": "WH-R-RETURN-100",
                "returned_by": "test",
                "confirmed_items": self.confirmed_return_items(item_id),
            },
        )

        self.assertEqual(returned.status_code, 200)
        self.assertEqual(returned.json()["status"], "returned")
        self.assertEqual(returned.json()["return_status"], "returned")
        self.assertEqual(returned.json()["return_reference"], "WH-R-RETURN-100")
        self.assertEqual(returned.json()["skladbot_request_number"], "WH-R-RETURN-100")
        self.assertEqual(returned.json()["skladbot_return_status"], "queued")
        self.assertTrue(returned.json()["returned_at"])

        returns = self.client.get("/api/v1/returns")
        self.assertEqual(returns.status_code, 200)
        self.assertEqual(len(returns.json()), 1)
        self.assertEqual(returns.json()[0]["id"], order_id)
        self.assertEqual(returns.json()[0]["client"], "Test Client")
        self.assertEqual(returns.json()[0]["order_date"], "2026-05-30")
        self.assertEqual(returns.json()[0]["items"][0]["product"], "Test Product")
        self.assertEqual(returns.json()[0]["skladbot_request_number"], "WH-R-RETURN-100")
        self.assertEqual(returns.json()[0]["return_reference"], "WH-R-RETURN-100")
        self.assertEqual(returns.json()[0]["return_status"], "returned")
        self.assertEqual(returns.json()[0]["skladbot_return_status"], "queued")

        duplicate_return = self.client.post(
            f"/api/v1/returns/{order_id}",
            json={"return_reference": "WH-R-RETURN-100", "returned_by": "test"},
        )
        self.assertEqual(duplicate_return.status_code, 409)
        self.assertEqual(duplicate_return.json()["detail"], "Order is already returned")

        lookup_after_return = self.client.get("/api/v1/returns/lookup", params={"lookup": "100500"})
        self.assertEqual(lookup_after_return.status_code, 200)
        self.assertEqual(lookup_after_return.json()["status"], "returned")
        self.assertEqual(lookup_after_return.json()["return_reference"], "WH-R-RETURN-100")

        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        self.assertEqual(active.json(), [])

        with self.SessionLocal() as db:
            order = db.get(Order, uuid.UUID(order_id))
            self.assertEqual(order.raw_payload["return_status"], "returned")
            self.assertEqual(order.raw_payload["return_reference"], "WH-R-RETURN-100")
            self.assertEqual(order.raw_payload["skladbot_return_confirmed_items"][0]["item_id"], item_id)
            event = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == SKLADBOT_RETURN_REQUEST_CREATE_EVENT_TYPE)
            ).scalar_one()
            self.assertEqual(event.idempotency_key, f"skladbot:return_create:v1:order:{order_id}")
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertIn("order_returned", actions)
            self.assertIn("skladbot_return_request_create_queued", actions)

    def test_return_releases_kiz_for_new_outbound_scan_with_history(self):
        first_order_id, first_item_id = self.seed_order(quantity_blocks=1)
        code = "01040000000000000001"

        first_scan = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": first_item_id, "code": code, "workstation_id": "pc-1"},
        )
        self.assertEqual(first_scan.status_code, 201)
        self.assertEqual(first_scan.json()["item_status"], "completed")
        self.assertEqual(self.client.post(f"/api/v1/orders/{first_order_id}/complete").status_code, 200)

        returned = self.client.post(
            f"/api/v1/returns/{first_order_id}",
            json={
                "return_reference": "WH-R-RETURN-200",
                "returned_by": "warehouse-pc",
                "confirmed_items": self.confirmed_return_items(first_item_id, blocks=1, pieces=20),
            },
        )
        self.assertEqual(returned.status_code, 200)

        second_order_id, second_item_id = self.seed_order(quantity_blocks=1)
        second_scan = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": second_item_id, "code": code, "workstation_id": "pc-2"},
        )

        self.assertEqual(second_scan.status_code, 201)
        self.assertEqual(second_scan.json()["order_item_id"], second_item_id)
        self.assertEqual(second_scan.json()["item_status"], "completed")

        with self.SessionLocal() as db:
            scan_rows = db.execute(
                select(ScanCode).where(ScanCode.code == code).order_by(ScanCode.scanned_at, ScanCode.id)
            ).scalars().all()
            self.assertEqual(len(scan_rows), 2)
            self.assertEqual({str(scan.order_item_id) for scan in scan_rows}, {first_item_id, second_item_id})

            kiz_codes = db.execute(select(KizCode).where(KizCode.code == code)).scalars().all()
            self.assertEqual(len(kiz_codes), 1)
            movements = db.execute(
                select(KizMovement)
                .where(KizMovement.kiz_id == kiz_codes[0].id)
                .order_by(KizMovement.occurred_at, KizMovement.id)
            ).scalars().all()
            self.assertEqual([movement.movement_type for movement in movements], ["outbound", "return", "re_outbound"])
            self.assertEqual(str(movements[-1].order_item_id), second_item_id)

            first_order = db.get(Order, uuid.UUID(first_order_id))
            second_order = db.get(Order, uuid.UUID(second_order_id))
            self.assertEqual(first_order.status, "returned")
            self.assertEqual(second_order.status, "not_completed")

    def test_kiz_availability_reports_returned_code_as_reusable(self):
        first_order_id, first_item_id = self.seed_order(quantity_blocks=1)
        _second_order_id, second_item_id = self.seed_order(quantity_blocks=1)
        code = "01040000000000000021"

        first_scan = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": first_item_id, "code": code, "workstation_id": "pc-1"},
        )
        self.assertEqual(first_scan.status_code, 201)

        busy = self.client.get(
            "/api/v1/kiz/availability",
            params={"code": code, "order_item_id": second_item_id},
        )
        self.assertEqual(busy.status_code, 200)
        self.assertFalse(busy.json()["available"])
        self.assertEqual(busy.json()["latest_movement_type"], "outbound")

        self.assertEqual(self.client.post(f"/api/v1/orders/{first_order_id}/complete").status_code, 200)
        returned = self.client.post(
            f"/api/v1/returns/{first_order_id}",
            json={
                "return_reference": "WH-R-RETURN-210",
                "returned_by": "warehouse-pc",
                "confirmed_items": self.confirmed_return_items(first_item_id, blocks=1, pieces=20),
            },
        )
        self.assertEqual(returned.status_code, 200)

        reusable = self.client.get(
            "/api/v1/kiz/availability",
            params={"code": code, "order_item_id": second_item_id},
        )

        self.assertEqual(reusable.status_code, 200)
        self.assertTrue(reusable.json()["available"])
        self.assertEqual(reusable.json()["latest_movement_type"], "return")
        self.assertEqual(reusable.json()["existing_order_item_id"], first_item_id)

    def test_scan_flushes_scan_code_before_kiz_movement(self):
        _order_id, item_id = self.seed_order(quantity_blocks=1)
        code = "01040000000000000003"

        from backend.app import orders_service

        original_record_kiz_movement = orders_service.record_kiz_movement

        def assert_scan_code_is_persisted(db, **kwargs):
            scan_code_id = kwargs["scan_code_id"]
            persisted_scan = db.execute(
                select(ScanCode).where(ScanCode.id == scan_code_id)
            ).scalar_one_or_none()
            if persisted_scan is None:
                raise AssertionError("scan_code must be flushed before kiz_movement")
            return original_record_kiz_movement(db, **kwargs)

        with mock.patch.object(orders_service, "record_kiz_movement", side_effect=assert_scan_code_is_persisted):
            response = self.client.post(
                "/api/v1/scans",
                json={"order_item_id": item_id, "code": code, "workstation_id": "pc-1"},
            )

        self.assertEqual(response.status_code, 201)

        with self.SessionLocal() as db:
            scan = db.execute(select(ScanCode).where(ScanCode.code == code)).scalar_one()
            movement = db.execute(
                select(KizMovement).join(KizCode, KizMovement.kiz_id == KizCode.id).where(KizCode.code == code)
            ).scalar_one()
            self.assertEqual(movement.scan_code_id, scan.id)

    def test_failed_return_does_not_release_kiz_for_new_order(self):
        first_order_id, first_item_id = self.seed_order(quantity_blocks=1)
        code = "01040000000000000002"
        self.assertEqual(
            self.client.post("/api/v1/scans", json={"order_item_id": first_item_id, "code": code}).status_code,
            201,
        )
        self.assertEqual(self.client.post(f"/api/v1/orders/{first_order_id}/complete").status_code, 200)

        failed_return = self.client.post(
            f"/api/v1/returns/{first_order_id}",
            json={
                "return_reference": "WH-R-RETURN-201",
                "returned_by": "warehouse-pc",
                "confirmed_items": self.confirmed_return_items(first_item_id, blocks=2, pieces=20),
            },
        )
        self.assertEqual(failed_return.status_code, 422)

        _second_order_id, second_item_id = self.seed_order(quantity_blocks=1)
        second_scan = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": second_item_id, "code": code},
        )
        self.assertEqual(second_scan.status_code, 409)

        with self.SessionLocal() as db:
            movements = db.execute(
                select(KizMovement).join(KizCode, KizMovement.kiz_id == KizCode.id).where(KizCode.code == code)
            ).scalars().all()
            self.assertEqual([movement.movement_type for movement in movements], ["outbound"])

    def test_mark_return_exports_archive_and_returns_to_google_sheets_best_effort(self):
        order_id, item_id = self.seed_order(status="completed", scanned_blocks=2, item_status="completed")
        with self.SessionLocal() as db:
            order = db.get(Order, uuid.UUID(order_id))
            order.raw_payload = {
                "skladbot_request_number": "WR-RETURN-101",
                "skladbot_request_id": "100501",
            }
            db.commit()

        returned = self.client.post(
            f"/api/v1/returns/{order_id}",
            json={
                "return_reference": "WR-RETURN-101",
                "returned_by": "test",
                "confirmed_items": self.confirmed_return_items(item_id),
            },
        )

        self.assertEqual(returned.status_code, 200)

        with self.SessionLocal() as db:
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertIn("google_sheets_archive_export", actions)
            self.assertIn("google_sheets_return_export", actions)
            event_actions = [
                event.payload["action"]
                for event in db.execute(
                    select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
                ).scalars().all()
            ]
            self.assertIn("google_sheets_archive_export", event_actions)
            self.assertIn("google_sheets_return_export", event_actions)

    def test_mark_return_rejects_mismatched_confirmed_items_without_side_effects(self):
        order_id, item_id = self.seed_order(status="completed", scanned_blocks=2, item_status="completed")

        returned = self.client.post(
            f"/api/v1/returns/{order_id}",
            json={
                "return_reference": "WR-RETURN-102",
                "returned_by": "test",
                "confirmed_items": self.confirmed_return_items(item_id, blocks=1),
            },
        )

        self.assertEqual(returned.status_code, 422)
        with self.SessionLocal() as db:
            order = db.get(Order, uuid.UUID(order_id))
            self.assertEqual(order.status, "completed")
            self.assertEqual(db.execute(select(PendingEvent)).scalars().all(), [])

    def test_import_creates_grouped_order_items_and_history(self):
        rows = [
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "cash",
                "Клиент": "Import Client",
                "Адрес": "Import Address",
                "Торговый представитель": "Import Rep",
                "Товары": "Product One",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "source-order-1",
                "ID импорта": "import-row-1",
            },
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "cash",
                "Клиент": "Import Client",
                "Адрес": "Import Address",
                "Торговый представитель": "Import Rep",
                "Товары": "Product Two",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "ID заказа": "source-order-2",
                "ID импорта": "import-row-2",
            },
        ]

        response = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["status"], "completed")
        self.assertEqual(payload["orders_created"], 1)
        self.assertEqual(payload["items_created"], 2)
        self.assertEqual(payload["duplicate_rows"], 0)
        self.assertEqual(payload["google_sheets_status"], "queued")
        self.assertEqual(payload["google_sheets_imported"], 0)

        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        active_payload = active.json()
        self.assertEqual(len(active_payload), 1)
        self.assertEqual(active_payload[0]["client"], "Import Client")
        self.assertEqual(len(active_payload[0]["items"]), 2)
        self.assertEqual(active_payload[0]["items"][0]["scan_codes"], [])

        history = self.client.get("/api/v1/imports")
        self.assertEqual(history.status_code, 200)
        self.assertEqual(len(history.json()), 1)
        self.assertEqual(history.json()[0]["rows_imported"], 2)
        self.assertEqual(history.json()[0]["raw_payload"]["google_sheets"]["status"], "queued")
        with self.SessionLocal() as db:
            event = db.execute(select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")).scalar_one()
            sheet_records = event.payload["records"]
            self.assertEqual(event.status, "pending")
            self.assertEqual(event.payload["action"], "google_sheets_import_export")
            self.assertEqual(len(sheet_records), 2)
            self.assertEqual(sheet_records[0]["Дата отгрузки"], "30.05.2026")
            self.assertEqual(sheet_records[0]["Клиент"], "Import Client")
            self.assertEqual(sheet_records[0]["ID импорта"], "import-row-1")

    def test_import_skips_duplicate_rows_inside_same_payload(self):
        row = {
            "Дата отгрузки": "30.05.2026",
            "Тип оплаты": "cash",
            "Клиент": "Payload Duplicate Client",
            "Адрес": "Payload Duplicate Address",
            "Товары": "Product One",
            "Кол-во ШТ": "20",
            "Кол-во блок": "2",
            "ID заказа": "payload-duplicate-order",
            "ID импорта": "payload-duplicate-import",
        }

        response = self.client.post(
            "/api/v1/imports",
            json={"source": "telegram", "filename": "orders.xlsx", "rows": [row, dict(row)]},
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["status"], "completed")
        self.assertEqual(payload["orders_created"], 1)
        self.assertEqual(payload["items_created"], 1)
        self.assertEqual(payload["duplicate_rows"], 1)
        with self.SessionLocal() as db:
            self.assertEqual(len(db.execute(select(Order)).scalars().all()), 1)
            self.assertEqual(len(db.execute(select(OrderItem)).scalars().all()), 1)
            event = db.execute(select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")).scalar_one()
            self.assertEqual(len(event.payload["records"]), 1)
            self.assertEqual(event.payload["records"][0]["ID импорта"], "payload-duplicate-import")

    def test_import_keeps_backend_data_when_google_sheets_export_fails(self):
        rows = [
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "cash",
                "Клиент": "Google Fail Client",
                "Адрес": "Import Address",
                "Товары": "Product One",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "google-fail-order",
                "ID импорта": "google-fail-import",
            },
        ]

        response = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["status"], "completed")
        self.assertEqual(payload["items_created"], 1)
        self.assertEqual(payload["google_sheets_status"], "queued")
        self.assertEqual(payload["google_sheets_error"], "")

        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        self.assertEqual(active.json()[0]["client"], "Google Fail Client")
        with self.SessionLocal() as db:
            event = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalar_one()
            self.assertEqual(event.status, "pending")
            self.assertEqual(event.payload["action"], "google_sheets_import_export")
            self.assertEqual(len(event.payload["records"]), 1)

    def test_import_reports_google_queue_failure_without_rolling_back_backend_data(self):
        rows = [
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "cash",
                "Клиент": "Google Queue Failure Client",
                "Адрес": "Import Address",
                "Товары": "Product One",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "google-queue-failure-order",
                "ID импорта": "google-queue-failure-import",
            },
        ]

        with mock.patch(
            "backend.app.imports_service.queue_google_sheets_export",
            side_effect=RuntimeError("Google queue storage failed"),
        ):
            response = self.client.post(
                "/api/v1/imports",
                json={"source": "excel", "filename": "orders.xlsx", "rows": rows},
            )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["status"], "completed")
        self.assertEqual(payload["items_created"], 1)
        self.assertEqual(payload["google_sheets_status"], "error")
        self.assertIn("Google queue storage failed", payload["google_sheets_error"])

        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        self.assertEqual(active.json()[0]["client"], "Google Queue Failure Client")
        with self.SessionLocal() as db:
            google_events = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalars().all()
            self.assertEqual(google_events, [])
            import_job = db.execute(select(ImportJob)).scalar_one()
            self.assertEqual(import_job.rows_imported, 1)
            self.assertEqual(import_job.raw_payload["google_sheets"]["status"], "error")
            incident = db.execute(
                select(Incident).where(Incident.source == "google_sheets_import_export")
            ).scalar_one()
            self.assertEqual(incident.import_id, import_job.id)
            self.assertIn("Google queue storage failed", incident.message)
            audit = db.execute(
                select(AuditLog).where(AuditLog.action == "google_sheets_import_export_failed")
            ).scalar_one()
            self.assertEqual(audit.entity_id, str(import_job.id))

    def test_duplicate_backend_import_still_can_backfill_google_sheets(self):
        row = {
            "Дата отгрузки": "30.05.2026",
            "Тип оплаты": "cash",
            "Клиент": "Backfill Client",
            "Адрес": "Import Address",
            "Товары": "Product One",
            "Кол-во ШТ": "20",
            "Кол-во блок": "2",
            "ID заказа": "backfill-order",
            "ID импорта": "backfill-import",
        }

        first = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": [row]})
        self.assertEqual(first.status_code, 201)
        self.assertEqual(first.json()["items_created"], 1)

        second = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": [row]})

        self.assertEqual(second.status_code, 201)
        payload = second.json()
        self.assertEqual(payload["items_created"], 0)
        self.assertEqual(payload["duplicate_rows"], 1)
        self.assertEqual(payload["google_sheets_status"], "queued")
        with self.SessionLocal() as db:
            events = db.execute(
                select(PendingEvent).where(PendingEvent.event_type == "google_sheets_export")
            ).scalars().all()
            self.assertEqual(len(events), 2)
            self.assertEqual(events[-1].payload["records"][0]["ID импорта"], "backfill-import")

    def test_import_after_return_creates_new_active_order_instead_of_duplicate(self):
        row = {
            "Дата отгрузки": "30.05.2026",
            "Тип оплаты": "cash",
            "Клиент": "Returned Reissue Client",
            "Адрес": "Returned Reissue Address",
            "Координаты": "41.31,69.27",
            "Товары": "Product One",
            "Кол-во ШТ": "20",
            "Кол-во блок": "2",
            "ID заказа": "returned-reissue-order",
            "ID импорта": "returned-reissue-import",
        }

        first = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": [row]})
        self.assertEqual(first.status_code, 201)
        self.assertEqual(first.json()["items_created"], 1)
        with self.SessionLocal() as db:
            order = db.execute(select(Order)).scalar_one()
            order.status = "returned"
            order.raw_payload = {
                **(order.raw_payload or {}),
                "return_status": "returned",
                "return_reference": "WH-R-RETURN",
            }
            db.commit()

        second = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": [row]})

        self.assertEqual(second.status_code, 201)
        payload = second.json()
        self.assertEqual(payload["orders_created"], 1)
        self.assertEqual(payload["items_created"], 1)
        self.assertEqual(payload["duplicate_rows"], 0)
        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        self.assertEqual(len(active.json()), 1)
        self.assertEqual(active.json()[0]["client"], "Returned Reissue Client")
        with self.SessionLocal() as db:
            orders = db.execute(select(Order)).scalars().all()
            items = db.execute(select(OrderItem)).scalars().all()
            self.assertEqual(len(orders), 2)
            self.assertEqual(len(items), 2)
            self.assertEqual([order.status for order in orders].count("returned"), 1)
            self.assertEqual([order.status for order in orders].count("not_completed"), 1)

    def test_import_preview_reports_duplicates_invalid_rows_and_does_not_write(self):
        existing_row = {
            "Дата отгрузки": "30.05.2026",
            "Тип оплаты": "cash",
            "Клиент": "Preview Client",
            "Адрес": "Preview Address",
            "Товары": "Product One",
            "Кол-во ШТ": "20",
            "Кол-во блок": "2",
            "ID заказа": "preview-order",
            "ID импорта": "preview-import",
        }
        created = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": [existing_row]})
        self.assertEqual(created.status_code, 201)
        with self.SessionLocal() as db:
            pending_events_before_preview = len(db.execute(select(PendingEvent)).scalars().all())
        new_row = {
            "Дата отгрузки": "30.05.2026",
            "Тип оплаты": "cash",
            "Клиент": "Preview Client",
            "Адрес": "Preview Address",
            "Товары": "Product Two",
            "Кол-во ШТ": "10",
            "Кол-во блок": "1",
            "ID заказа": "preview-order-2",
            "ID импорта": "preview-import-2",
        }

        preview = self.client.post(
            "/api/v1/imports/preview",
            json={"source": "excel", "filename": "orders.xlsx", "rows": [existing_row, new_row, {"Клиент": "Broken"}]},
        )

        self.assertEqual(preview.status_code, 200)
        payload = preview.json()
        self.assertEqual(payload["rows_total"], 3)
        self.assertEqual(payload["rows_importable"], 1)
        self.assertEqual(payload["items_new"], 1)
        self.assertEqual(payload["duplicate_rows"], 1)
        self.assertEqual(payload["invalid_rows"], 1)
        self.assertEqual(payload["duplicate_row_numbers"], [1])
        self.assertEqual(payload["invalid_row_numbers"], [3])
        self.assertIn("row 3:", payload["errors"][0])
        with self.SessionLocal() as db:
            self.assertEqual(len(db.execute(select(Order)).scalars().all()), 1)
            self.assertEqual(len(db.execute(select(OrderItem)).scalars().all()), 1)
            self.assertEqual(len(db.execute(select(ImportJob)).scalars().all()), 1)
            self.assertEqual(len(db.execute(select(PendingEvent)).scalars().all()), pending_events_before_preview)

    def test_retrying_same_import_payload_does_not_duplicate_backend_records(self):
        row = {
            "Дата отгрузки": "30.05.2026",
            "Тип оплаты": "cash",
            "Клиент": "Retry Client",
            "Адрес": "Retry Address",
            "Товары": "Product One",
            "Кол-во ШТ": "20",
            "Кол-во блок": "2",
            "ID заказа": "retry-order",
            "ID импорта": "retry-import-row",
        }
        payload = {
            "source": "telegram",
            "filename": "retry.xlsx",
            "sha256": "a" * 64,
            "rows": [row],
        }

        first = self.client.post("/api/v1/imports", json=payload)
        second = self.client.post("/api/v1/imports", json=payload)

        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)
        self.assertEqual(first.json()["orders_created"], 1)
        self.assertEqual(first.json()["items_created"], 1)
        self.assertEqual(second.json()["orders_created"], 0)
        self.assertEqual(second.json()["items_created"], 0)
        self.assertEqual(second.json()["duplicate_rows"], 1)
        with self.SessionLocal() as db:
            self.assertEqual(len(db.execute(select(Order)).scalars().all()), 1)
            self.assertEqual(len(db.execute(select(OrderItem)).scalars().all()), 1)
            self.assertEqual(len(db.execute(select(ImportFile)).scalars().all()), 1)
            self.assertEqual(len(db.execute(select(ScanCode)).scalars().all()), 0)

    def test_failed_import_creates_linked_incident_and_resolve_removes_readiness_blocker(self):
        with self.SessionLocal() as db:
            db.execute(text("CREATE TABLE alembic_version (version_num VARCHAR(32) NOT NULL)"))
            db.execute(text("INSERT INTO alembic_version (version_num) VALUES ('20260623_0004')"))
            event = PendingEvent(
                event_type="telegram_excel_import",
                status="processing",
                attempts=1,
                payload={"document": {"file_name": "broken.xlsx", "file_id": "file-1"}},
            )
            db.add(event)
            db.commit()
            event_id = str(event.id)

        import_response = self.client.post(
            "/api/v1/imports",
            json={
                "source": "telegram",
                "filename": "broken.xlsx",
                "sha256": "b" * 64,
                "telegram_event_id": event_id,
                "rows": [{"Тип оплаты": "cash", "Клиент": "", "Товары": "", "Кол-во блок": "0"}],
            },
        )

        self.assertEqual(import_response.status_code, 201)
        self.assertEqual(import_response.json()["status"], "failed")
        import_id = import_response.json()["id"]
        with self.SessionLocal() as db:
            incidents = db.execute(select(Incident)).scalars().all()
            self.assertEqual(len(incidents), 1)
            self.assertEqual(incidents[0].source, "excel_import")
            self.assertEqual(str(incidents[0].import_id), import_id)
            self.assertEqual(str(incidents[0].pending_event_id), event_id)
            incident_id = str(incidents[0].id)

        before = self.client.get("/ready").json()
        self.assertEqual(before["status"], "degraded")
        self.assertEqual(before["imports"]["recent_errors"][0]["filename"], "broken.xlsx")

        status_response = self.client.post(
            f"/api/v1/admin/incidents/{incident_id}/status",
            json={"status": "resolved", "actor": "anton", "source": "web", "reason": "Malformed file reviewed"},
        )
        self.assertEqual(status_response.status_code, 200)

        after = self.client.get("/ready").json()
        self.assertEqual(after["status"], "ok")
        self.assertEqual(after["imports"]["recent_errors"], [])
        with self.SessionLocal() as db:
            import_job = db.get(ImportJob, uuid.UUID(import_id))
            self.assertEqual(import_job.status, "failed")
            audits = db.execute(
                select(AuditLog).where(AuditLog.action == "incident_status_changed")
            ).scalars().all()
            self.assertEqual(len(audits), 1)

    def test_changed_address_with_same_source_import_id_does_not_create_backend_duplicate(self):
        first_row = {
            "Дата отгрузки": "30.05.2026",
            "Тип оплаты": "cash",
            "Клиент": "Stable Import Client",
            "Адрес": "Адрес не указан",
            "Товары": "Product One",
            "Кол-во ШТ": "20",
            "Кол-во блок": "2",
            "ID заказа": "stable-order",
            "ID импорта": "stable-import",
        }
        second_row = {
            **first_row,
            "Адрес": "Ташкент, Чиланзарский район, 10",
        }

        first = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": [first_row]})
        self.assertEqual(first.status_code, 201)
        self.assertEqual(first.json()["items_created"], 1)

        second = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": [second_row]})

        self.assertEqual(second.status_code, 201)
        self.assertEqual(second.json()["items_created"], 0)
        self.assertEqual(second.json()["duplicate_rows"], 1)
        self.assertEqual(second.json()["backend_address_updates"], 1)
        self.assertEqual(second.json()["google_sheets_status"], "queued")
        with self.SessionLocal() as db:
            orders = db.execute(select(Order)).scalars().all()
            self.assertEqual(len(orders), 1)
            self.assertEqual(len(db.execute(select(OrderItem)).scalars().all()), 1)
            self.assertEqual(orders[0].address, "Ташкент, Чиланзарский район, 10")
            self.assertEqual(orders[0].raw_payload["address_backfill_source"], "import")

    def test_google_sheets_export_updates_missing_address_for_existing_import_id(self):
        class FakeSheet:
            def __init__(self):
                self.updates = []

            def batch_update(self, updates, value_input_option=None):
                self.updates.extend(updates)
                self.value_input_option = value_input_option

        rows = [
            ["Дата отгрузки", "Тип оплаты", "Клиент", "Адрес"] + [""] * 22 + ["ID заказа", "ID импорта"],
            ["30.05.2026", "cash", "Backfill Client", "Адрес не указан"] + [""] * 22 + ["order-1", "import-1"],
        ]
        records = [{
            "ID заказа": "order-1",
            "ID импорта": "import-1",
            "Адрес": "Ташкент, Чиланзарский район, 10",
        }]
        sheet = FakeSheet()

        updated = update_missing_sheet_addresses(sheet, rows, records)

        self.assertEqual(updated, 1)
        self.assertEqual(sheet.updates, [{"range": "D2", "values": [["Ташкент, Чиланзарский район, 10"]]}])
        self.assertEqual(sheet.value_input_option, "USER_ENTERED")

    def test_google_sheets_export_updates_not_found_address_for_existing_import_id(self):
        class FakeSheet:
            def __init__(self):
                self.updates = []

            def batch_update(self, updates, value_input_option=None):
                self.updates.extend(updates)
                self.value_input_option = value_input_option

        rows = [
            ["Дата отгрузки", "Тип оплаты", "Клиент", "Адрес"] + [""] * 22 + ["ID заказа", "ID импорта"],
            ["30.05.2026", "cash", "Backfill Client", "Адрес не найден"] + [""] * 22 + ["order-1", "import-1"],
        ]
        records = [{
            "ID заказа": "order-1",
            "ID импорта": "import-1",
            "Адрес": "Ташкент, Чиланзарский район, 10",
        }]
        sheet = FakeSheet()

        updated = update_missing_sheet_addresses(sheet, rows, records)

        self.assertEqual(updated, 1)
        self.assertEqual(sheet.updates, [{"range": "D2", "values": [["Ташкент, Чиланзарский район, 10"]]}])

    def test_google_sheets_export_updates_coordinate_address_by_business_key_when_ids_changed(self):
        class FakeSheet:
            def __init__(self):
                self.updates = []

            def batch_update(self, updates, value_input_option=None):
                self.updates.extend(updates)
                self.value_input_option = value_input_option

        rows = [
            [
                "Дата отгрузки",
                "Тип оплаты",
                "Клиент",
                "Адрес",
                "Торговый представитель",
                "Товары",
                "Кол-во ШТ",
                "Кол-во блок",
            ] + [""] * 18 + ["ID заказа", "ID импорта"],
            [
                "04.06.2026",
                "Терминал",
                "Backfill Client",
                "Координаты: 41.325341, 69.233731",
                "ТП2",
                "Chapman Brown OP 20",
                "500",
                "50",
            ] + [""] * 18 + ["old-order", "old-import"],
        ]
        records = [{
            "Дата отгрузки": "04.06.2026",
            "Тип оплаты": "Терминал",
            "Клиент": "Backfill Client",
            "Адрес": "Ташкент, улица Сакичмон, 1/18",
            "Торговый представитель": "ТП2",
            "Товары": "Chapman Brown OP 20",
            "Кол-во ШТ": 500,
            "Кол-во блок": 50,
            "ID заказа": "new-order",
            "ID импорта": "new-import",
        }]
        sheet = FakeSheet()

        updated = update_missing_sheet_addresses(sheet, rows, records)

        self.assertEqual(updated, 1)
        self.assertEqual(sheet.updates, [{"range": "D2", "values": [["Ташкент, улица Сакичмон, 1/18"]]}])
        self.assertEqual(rows[1][3], "Ташкент, улица Сакичмон, 1/18")

    def test_google_sheets_export_does_not_update_ambiguous_business_key_address(self):
        class FakeSheet:
            def __init__(self):
                self.updates = []

            def batch_update(self, updates, value_input_option=None):
                self.updates.extend(updates)

        rows = [
            [
                "Дата отгрузки",
                "Тип оплаты",
                "Клиент",
                "Адрес",
                "Торговый представитель",
                "Товары",
                "Кол-во ШТ",
                "Кол-во блок",
            ] + [""] * 18 + ["ID заказа", "ID импорта"],
            [
                "04.06.2026",
                "Терминал",
                "Backfill Client",
                "Координаты: 41.325341, 69.233731",
                "ТП2",
                "Chapman Brown OP 20",
                "500",
                "50",
            ] + [""] * 18 + ["old-order-1", "old-import-1"],
            [
                "04.06.2026",
                "Терминал",
                "Backfill Client",
                "Координаты: 41.325341, 69.233731",
                "ТП2",
                "Chapman Brown OP 20",
                "500",
                "50",
            ] + [""] * 18 + ["old-order-2", "old-import-2"],
        ]
        records = [{
            "Дата отгрузки": "04.06.2026",
            "Тип оплаты": "Терминал",
            "Клиент": "Backfill Client",
            "Адрес": "Ташкент, улица Сакичмон, 1/18",
            "Торговый представитель": "ТП2",
            "Товары": "Chapman Brown OP 20",
            "Кол-во ШТ": 500,
            "Кол-во блок": 50,
            "ID заказа": "new-order",
            "ID импорта": "new-import",
        }]
        sheet = FakeSheet()

        updated = update_missing_sheet_addresses(sheet, rows, records)

        self.assertEqual(updated, 0)
        self.assertEqual(sheet.updates, [])

    def test_import_stores_coordinates_blocks_and_prices(self):
        rows = [
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "Терминал",
                "Клиент": "Price Client",
                "Адрес": "Tashkent Address",
                "Координаты": "41.31, 69.27",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "200",
                "Сумма позиции": "4800000",
                "ID заказа": "price-source-order",
            },
        ]

        response = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})

        self.assertEqual(response.status_code, 201)
        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        order = active.json()[0]
        self.assertEqual(order["coordinates"], "41.31, 69.27")
        self.assertEqual(order["items"][0]["quantity_pieces"], 200)
        self.assertEqual(order["items"][0]["quantity_blocks"], 20)
        self.assertEqual(order["items"][0]["block_price"], 240000)
        self.assertEqual(order["items"][0]["line_total"], 4_800_000)

    def test_import_marks_missing_address_as_pickup(self):
        rows = [
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "Терминал",
                "Клиент": "Pickup Import Client",
                "Адрес": "",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "ID заказа": "pickup-import-order",
            },
        ]

        response = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})

        self.assertEqual(response.status_code, 201)
        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        self.assertEqual(active.json()[0]["address"], "Самовывоз со склада")

    def test_import_skips_duplicate_items_and_reports_invalid_rows(self):
        valid_row = {
            "Дата отгрузки": "2026-05-30",
            "Тип оплаты": "cash",
            "Клиент": "Duplicate Client",
            "Адрес": "Duplicate Address",
            "Товары": "Duplicate Product",
            "Кол-во ШТ": 20,
            "Кол-во блок": 2,
            "ID заказа": "duplicate-source-order",
        }
        first = self.client.post("/api/v1/imports", json={"source": "excel", "rows": [valid_row]})
        second = self.client.post("/api/v1/imports", json={"source": "excel", "rows": [valid_row, {"Клиент": "Broken"}]})

        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)
        payload = second.json()
        self.assertEqual(payload["items_created"], 0)
        self.assertEqual(payload["duplicate_rows"], 1)
        self.assertEqual(payload["invalid_rows"], 1)
        self.assertEqual(payload["status"], "failed")

        with self.SessionLocal() as db:
            self.assertEqual(len(db.execute(select(Order)).scalars().all()), 1)
            self.assertEqual(len(db.execute(select(OrderItem)).scalars().all()), 1)

    def test_day_report_summarizes_orders_scans_and_payment_groups(self):
        rows = [
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "Терминал",
                "Клиент": "Report Client",
                "Адрес": "Report Address",
                "Торговый представитель": "Report Rep",
                "Товары": "Report Product One",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "Номер заявки SkladBot": "WR-100",
                "ID заказа": "report-source-order-1",
            },
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "Терминал",
                "Клиент": "Report Client",
                "Адрес": "Report Address",
                "Торговый представитель": "Report Rep",
                "Товары": "Report Product Two",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "Номер заявки SkladBot": "WR-100",
                "ID заказа": "report-source-order-2",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        active = self.client.get("/api/v1/orders/active").json()
        order_id = active[0]["id"]
        item_ids = {item["product"]: item["id"] for item in active[0]["items"]}

        scans = [
            ("Report Product One", "010000000101"),
            ("Report Product One", "010000000102"),
            ("Report Product Two", "010000000201"),
        ]
        for product, code in scans:
            response = self.client.post(
                "/api/v1/scans",
                json={
                    "order_item_id": item_ids[product],
                    "code": code,
                    "scanned_at": "2026-05-30T12:00:00+00:00",
                },
            )
            self.assertEqual(response.status_code, 201)

        active_after_scans = self.client.get("/api/v1/orders/active")
        self.assertEqual(active_after_scans.status_code, 200)
        active_item = active_after_scans.json()[0]["items"][0]
        self.assertTrue(active_item["scan_codes"])

        completed = self.client.post(f"/api/v1/orders/{order_id}/complete")
        self.assertEqual(completed.status_code, 200)

        report = self.client.get("/api/v1/reports/day?report_date=2026-05-30")

        self.assertEqual(report.status_code, 200)
        payload = report.json()
        self.assertEqual(payload["report_date"], "2026-05-30")
        self.assertEqual(payload["source"], "postgres")
        self.assertEqual(payload["totals"]["orders"], 1)
        self.assertEqual(payload["totals"]["completed_orders"], 1)
        self.assertEqual(payload["totals"]["active_orders"], 0)
        self.assertEqual(payload["totals"]["items"], 2)
        self.assertEqual(payload["totals"]["completed_items"], 2)
        self.assertEqual(payload["totals"]["planned_blocks"], 3)
        self.assertEqual(payload["totals"]["scanned_blocks"], 3)
        self.assertEqual(payload["totals"]["scanned_today"], 3)
        self.assertEqual(payload["totals"]["remaining_blocks"], 0)
        self.assertEqual(payload["totals"]["scan_codes"], 3)
        self.assertEqual(payload["payment_groups"][0]["payment_group"], "terminal")
        self.assertEqual(payload["payment_groups"][0]["orders"], 1)
        self.assertEqual(payload["orders"][0]["skladbot_request_number"], "WR-100")

    def test_day_report_counts_scan_by_business_timezone(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-31",
                "Тип оплаты": "Терминал",
                "Клиент": "Timezone Client",
                "Адрес": "Tashkent Address",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "ID заказа": "timezone-source-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        active = self.client.get("/api/v1/orders/active").json()
        item_id = active[0]["items"][0]["id"]
        scan = self.client.post(
            "/api/v1/scans",
            json={
                "order_item_id": item_id,
                "code": "0104006396053978217TIMEZONE001",
                "scanned_at": "2026-05-31T20:30:00+00:00",
            },
        )
        self.assertEqual(scan.status_code, 201)

        with mock.patch.dict("os.environ", {"TAKSKLAD_TIMEZONE": "Asia/Tashkent"}):
            report = self.client.get("/api/v1/reports/day?report_date=2026-06-01")

        self.assertEqual(report.status_code, 200)
        payload = report.json()
        self.assertEqual(payload["report_date"], "2026-06-01")
        self.assertEqual(payload["totals"]["orders"], 1)
        self.assertEqual(payload["totals"]["scanned_today"], 1)
        self.assertEqual(payload["orders"][0]["client"], "Timezone Client")

    def test_logistics_report_uses_shipment_date_coordinates_and_prices(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Logistics Client",
                "Адрес": "Tashkent Address",
                "Координаты": "41.31, 69.27",
                "Торговый представитель": "Rep One",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "200",
                "Кол-во блок": "20",
                "Сумма позиции": "4800000",
                "ID заказа": "logistics-source-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        dates = self.client.get("/api/v1/logistics/dates")
        self.assertEqual(dates.status_code, 200)
        self.assertEqual(dates.json(), ["2026-05-30"])

        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-30")
        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        sheet = workbook["Заявки"]

        self.assertEqual(sheet["C2"].value, "Logistics Client")
        self.assertEqual(sheet["G2"].value, "41.31,69.27")
        self.assertEqual(sheet["J2"].value, "30.05.2026")
        self.assertEqual(sheet["K2"].value, "10:00")
        self.assertEqual(sheet["L2"].value, "18:00")
        self.assertEqual(sheet["R2"].value, "Chapman Brown OP 20")
        self.assertEqual(sheet["S2"].value, 20)
        self.assertEqual(sheet["V2"].value, 240000)
        self.assertEqual(sheet["W2"].value, 4_800_000)
        self.assertEqual(sheet["AE2"].value, "41.31,69.27")
        self.assertEqual(sheet["AF2"].value, "41.31")
        self.assertEqual(sheet["AG2"].value, "69.27")
        workbook.close()

    def test_logistics_report_uses_saved_client_point_timeslot(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Timeslot Legal Entity",
                "Адрес": "Timeslot Address",
                "Координаты": "41.31, 69.27",
                "Торговый представитель": "Rep One",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "timeslot-source-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)
        updated = self.client.post(
            "/api/v1/admin/client-points/timeslot",
            json={
                "client_name": "Timeslot Legal Entity",
                "address": "Timeslot Address",
                "delivery_from": "08:30",
                "delivery_to": "11:45",
                "actor": "web",
                "reason": "точка принимает утром",
            },
        )
        self.assertEqual(updated.status_code, 200)

        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-30")

        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        sheet = workbook["Заявки"]
        self.assertEqual(sheet["C2"].value, "Timeslot Legal Entity")
        self.assertEqual(sheet["K2"].value, "08:30")
        self.assertEqual(sheet["L2"].value, "11:45")
        workbook.close()

    def test_import_updates_client_point_address_and_keeps_timeslot_by_client(self):
        first_rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Timeslot Legal Entity",
                "Адрес": "Old Timeslot Address",
                "Координаты": "41.31, 69.27",
                "Торговый представитель": "Rep One",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "timeslot-source-order-old",
            },
        ]
        first_import = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "old.xlsx", "rows": first_rows})
        self.assertEqual(first_import.status_code, 201)
        updated = self.client.post(
            "/api/v1/admin/client-points/timeslot",
            json={
                "client_name": "Timeslot Legal Entity",
                "address": "Old Timeslot Address",
                "delivery_from": "08:30",
                "delivery_to": "11:45",
                "actor": "web",
                "reason": "точка принимает утром",
            },
        )
        self.assertEqual(updated.status_code, 200)

        second_rows = [
            {
                "Дата отгрузки": "2026-05-31",
                "Тип оплаты": "Терминал",
                "Клиент": "Timeslot Legal Entity",
                "Адрес": "New Timeslot Address",
                "Координаты": "41.32, 69.28",
                "Торговый представитель": "Rep Two",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "timeslot-source-order-new",
            },
        ]
        second_import = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "new.xlsx", "rows": second_rows})
        self.assertEqual(second_import.status_code, 201)

        with self.SessionLocal() as db:
            points = db.execute(select(ClientPoint)).scalars().all()
            self.assertEqual(len(points), 1)
            self.assertEqual(points[0].client_name, "Timeslot Legal Entity")
            self.assertEqual(points[0].address, "New Timeslot Address")
            self.assertEqual(points[0].coordinates, "41.32, 69.28")
            self.assertEqual(points[0].representative, "Rep Two")
            self.assertEqual(points[0].delivery_from, "08:30")
            self.assertEqual(points[0].delivery_to, "11:45")

        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-31")

        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        sheet = workbook["Заявки"]
        self.assertEqual(sheet["C2"].value, "Timeslot Legal Entity")
        self.assertEqual(sheet["F2"].value, "New Timeslot Address")
        self.assertEqual(sheet["K2"].value, "08:30")
        self.assertEqual(sheet["L2"].value, "11:45")
        workbook.close()

    def test_logistics_report_keeps_unrouteable_orders_on_separate_sheet(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Route Client",
                "Адрес": "Tashkent Address",
                "Координаты": "41.31, 69.27",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "route-order",
            },
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "No Coordinates Client",
                "Адрес": "Tashkent Address Without Coordinates",
                "Товары": "Chapman RED OP 20",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "ID заказа": "no-coordinates-order",
            },
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Pickup Client",
                "Адрес": "Самовывоз со склада",
                "Координаты": "41.32, 69.28",
                "Товары": "Chapman Gold SSL",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "ID заказа": "pickup-order",
            },
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Pickup Variant Client",
                "Адрес": "Самовывоз: склад",
                "Координаты": "41.33, 69.29",
                "Товары": "Chapman RED OP 20",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "ID заказа": "pickup-variant-order",
            },
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Invalid Coordinates Client",
                "Адрес": "Invalid Coordinates Address",
                "Координаты": "999, 999",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "ID заказа": "invalid-coordinates-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        dates = self.client.get("/api/v1/logistics/dates")
        self.assertEqual(dates.status_code, 200)
        self.assertEqual(dates.json(), ["2026-05-30"])

        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-30")
        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        sheet = workbook["Заявки"]
        problems = workbook["Требуют координаты"]

        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["C2"].value, "Route Client")
        self.assertEqual(sheet["F2"].value, "Tashkent Address")
        self.assertEqual(sheet["R2"].value, "Chapman Brown OP 20")
        self.assertEqual(problems.max_row, 3)
        problem_rows = {
            row[0]: row
            for row in problems.iter_rows(min_row=2, values_only=True)
        }
        self.assertEqual(problem_rows["No Coordinates Client"][1], "Tashkent Address Without Coordinates")
        self.assertEqual(problem_rows["No Coordinates Client"][3], "Нет координат")
        self.assertEqual(problem_rows["Invalid Coordinates Client"][1], "Invalid Coordinates Address")
        self.assertEqual(problem_rows["Invalid Coordinates Client"][3], "Невалидные координаты")
        workbook.close()

    def test_logistics_report_returns_unrouteable_sheet_when_date_has_no_routeable_orders(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Pickup Client",
                "Адрес": "Самовывоз со склада",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "pickup-only-order",
            },
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "No Coordinates Client",
                "Адрес": "Tashkent Address Without Coordinates",
                "Товары": "Chapman RED OP 20",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "ID заказа": "no-coordinates-only-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        dates = self.client.get("/api/v1/logistics/dates")
        self.assertEqual(dates.status_code, 200)
        self.assertEqual(dates.json(), ["2026-05-30"])

        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-30")

        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        sheet = workbook["Заявки"]
        problems = workbook["Требуют координаты"]

        self.assertEqual(sheet.max_row, 1)
        self.assertEqual(problems.max_row, 2)
        self.assertEqual(problems["A2"].value, "No Coordinates Client")
        workbook.close()

    def test_logistics_report_404_when_date_has_only_pickup_orders(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Pickup Client",
                "Адрес": "Самовывоз со склада",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "pickup-only-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        dates = self.client.get("/api/v1/logistics/dates")
        self.assertEqual(dates.status_code, 200)
        self.assertEqual(dates.json(), [])

        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-30")

        self.assertEqual(report.status_code, 404)
        self.assertIn("No logistics delivery orders", report.json()["detail"])

    def test_logistics_report_excludes_returned_orders(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Returned Logistics Client",
                "Адрес": "Returned Logistics Address",
                "Координаты": "41.31, 69.27",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "returned-logistics-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)
        with self.SessionLocal() as db:
            order = db.execute(select(Order)).scalar_one()
            order.status = "returned"
            order.raw_payload = {
                **(order.raw_payload or {}),
                "return_status": "returned",
                "return_reference": "WH-R-RETURN",
            }
            db.commit()

        dates = self.client.get("/api/v1/logistics/dates")
        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-30")

        self.assertEqual(dates.status_code, 200)
        self.assertEqual(dates.json(), [])
        self.assertEqual(report.status_code, 404)
        self.assertIn("No logistics delivery orders", report.json()["detail"])

    def test_logistics_report_normalizes_three_part_coordinates(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Coordinates Client",
                "Адрес": "Tashkent Address",
                "Координаты": "41.214609,69.223027,15",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "ID заказа": "coordinates-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-30")
        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        sheet = workbook["Заявки"]

        self.assertEqual(sheet["AE2"].value, "41.214609,69.223027")
        self.assertEqual(sheet["AF2"].value, "41.214609")
        self.assertEqual(sheet["AG2"].value, "69.223027")
        workbook.close()

    def test_kiz_reports_show_source_file_progress_and_allow_partial_date_export(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "KIZ Client",
                "Адрес": "KIZ Address",
                "Координаты": "41.31, 69.27",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "Сумма позиции": "480000",
                "Источник файла": "source-a.xlsx",
                "ID заказа": "kiz-source-order",
            },
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Partial Date Client",
                "Адрес": "Partial Address",
                "Товары": "Chapman RED OP 20",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "Источник файла": "source-c.xlsx",
                "ID заказа": "partial-date-order",
            },
            {
                "Дата отгрузки": "2026-05-31",
                "Тип оплаты": "Перечисление",
                "Клиент": "Open Client",
                "Адрес": "Open Address",
                "Товары": "Chapman Gold SSL 20",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "Источник файла": "source-b.xlsx",
                "ID заказа": "open-source-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        active = self.client.get("/api/v1/orders/active").json()
        kiz_order = next(order for order in active if order["client"] == "KIZ Client")
        item_id = kiz_order["items"][0]["id"]
        for code in ("0104006396053978217SOURCEA001", "0104006396053978217SOURCEA002"):
            response = self.client.post("/api/v1/scans", json={"order_item_id": item_id, "code": code})
            self.assertEqual(response.status_code, 201)

        source_files = self.client.get("/api/v1/reports/kiz/source-files")
        self.assertEqual(source_files.status_code, 200)
        source_payload = source_files.json()
        by_file = {item["source_file"]: item for item in source_payload}
        self.assertEqual(set(by_file), {"source-a.xlsx", "source-b.xlsx", "source-c.xlsx"})
        self.assertTrue(by_file["source-a.xlsx"]["source_key"].startswith("import:"))
        self.assertTrue(by_file["source-a.xlsx"]["completed"])
        self.assertEqual(by_file["source-a.xlsx"]["scanned_blocks"], 2)
        self.assertEqual(by_file["source-a.xlsx"]["planned_blocks"], 2)
        self.assertFalse(by_file["source-b.xlsx"]["completed"])
        self.assertEqual(by_file["source-b.xlsx"]["scanned_blocks"], 0)
        self.assertEqual(by_file["source-b.xlsx"]["planned_blocks"], 1)
        self.assertFalse(by_file["source-c.xlsx"]["completed"])

        report = self.client.get(
            "/api/v1/reports/kiz/source-file",
            params={"source_file": "source-a.xlsx", "source_key": by_file["source-a.xlsx"]["source_key"]},
        )
        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        summary = workbook["Сводка"]
        self.assertEqual(summary["C2"].value, "KIZ Client")
        self.assertEqual(summary["G2"].value, 2)
        self.assertEqual(summary["H2"].value, 2)
        self.assertEqual(summary["I2"].value, 480000)
        sheet = workbook["Терминал"]
        self.assertEqual(sheet["C2"].value, "KIZ Client")
        self.assertEqual(sheet["G2"].value, "Chapman Brown OP 20")
        self.assertEqual(sheet["H2"].value, 1)
        self.assertEqual(sheet["I2"].value, "0104006396053978217SOURCEA001")
        self.assertEqual(sheet["H3"].value, 1)
        self.assertEqual(sheet["I3"].value, "0104006396053978217SOURCEA002")
        self.assertEqual(sheet["K2"].value, "source-a.xlsx")
        workbook.close()

        dates = self.client.get("/api/v1/reports/kiz/dates")
        self.assertEqual(dates.status_code, 200)
        self.assertEqual(dates.json()[0]["date"], "2026-05-30")
        self.assertFalse(dates.json()[0]["completed"])
        self.assertEqual(dates.json()[0]["planned_blocks"], 3)
        self.assertEqual(dates.json()[0]["scanned_blocks"], 2)

        date_report = self.client.get("/api/v1/reports/kiz/date", params={"shipment_date": "2026-05-30"})
        self.assertEqual(date_report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(date_report.content), data_only=True)
        summary = workbook["Сводка"]
        self.assertEqual(summary["C2"].value, "KIZ Client")
        self.assertEqual(summary["G2"].value, 2)
        self.assertEqual(summary["H2"].value, 2)
        workbook.close()

    def test_kiz_source_file_report_separates_same_filename_by_import(self):
        first_import = self.client.post(
            "/api/v1/imports",
            json={
                "source": "excel",
                "filename": "same-name.xlsx",
                "rows": [
                    {
                        "Дата отгрузки": "2026-05-30",
                        "Тип оплаты": "Терминал",
                        "Клиент": "Done Client",
                        "Адрес": "Done Address",
                        "Товары": "Chapman Brown OP 20",
                        "Кол-во ШТ": "10",
                        "Кол-во блок": "1",
                        "Источник файла": "same-name.xlsx",
                        "ID заказа": "same-done-order",
                    },
                ],
            },
        )
        self.assertEqual(first_import.status_code, 201)

        active = self.client.get("/api/v1/orders/active").json()
        done_order = next(order for order in active if order["client"] == "Done Client")
        response = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": done_order["items"][0]["id"], "code": "0104006396053978217SAMENAME001"},
        )
        self.assertEqual(response.status_code, 201)

        second_import = self.client.post(
            "/api/v1/imports",
            json={
                "source": "excel",
                "filename": "same-name.xlsx",
                "rows": [
                    {
                        "Дата отгрузки": "2026-05-30",
                        "Тип оплаты": "Перечисление",
                        "Клиент": "Open Client",
                        "Адрес": "Open Address",
                        "Товары": "Chapman Gold SSL 20",
                        "Кол-во ШТ": "10",
                        "Кол-во блок": "1",
                        "Источник файла": "same-name.xlsx",
                        "ID заказа": "same-open-order",
                    },
                ],
            },
        )
        self.assertEqual(second_import.status_code, 201)

        source_files = self.client.get("/api/v1/reports/kiz/source-files")
        self.assertEqual(source_files.status_code, 200)
        same_name_files = [item for item in source_files.json() if item["source_file"] == "same-name.xlsx"]
        self.assertEqual(len(same_name_files), 2)
        completed_files = [item for item in same_name_files if item["completed"]]
        self.assertEqual(len(completed_files), 1)
        self.assertTrue(completed_files[0]["source_key"].startswith("import:"))

        report = self.client.get(
            "/api/v1/reports/kiz/source-file",
            params={"source_file": "same-name.xlsx", "source_key": completed_files[0]["source_key"]},
        )
        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        summary = workbook["Сводка"]
        self.assertEqual(summary["C2"].value, "Done Client")
        self.assertIsNone(summary["C3"].value)
        workbook.close()

    def test_day_report_rejects_invalid_report_date(self):
        response = self.client.get("/api/v1/reports/day?report_date=not-a-date")

        self.assertEqual(response.status_code, 422)
        self.assertIn("Invalid report_date", response.json()["detail"])

    def test_reconciliation_report_endpoint_is_db_first_and_does_not_alert(self):
        with mock.patch("backend.app.main.run_daily_reconciliation") as reconcile:
            reconcile.return_value = {
                "source": "postgres",
                "status": "ok",
                "report_date": "2026-06-10",
                "alerts": [],
            }

            response = self.client.get("/api/v1/reports/reconciliation/day?report_date=2026-06-10")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["source"], "postgres")
        self.assertEqual(response.json()["alerts"], [])
        reconcile.assert_called_once()
        self.assertEqual(reconcile.call_args.kwargs["report_date"], "2026-06-10")
        self.assertEqual(reconcile.call_args.kwargs["alert_chat_ids"], [])

    def test_reconciliation_report_endpoint_records_google_down_as_mirror_issue(self):
        with self.SessionLocal() as db:
            order = Order(
                payment_type="terminal",
                client="Mirror Client",
                address="Mirror Address",
                representative="Mirror Rep",
                order_date=date(2026, 6, 10),
                status="not_completed",
                raw_payload={
                    "skladbot_request_number": "WH-R-MIRROR",
                    "skladbot_request_id": "1001",
                    "skladbot_status": "found",
                },
            )
            db.add(OrderItem(
                order=order,
                product="Chapman RED OP 20",
                quantity_pieces=10,
                quantity_blocks=1,
                pieces_per_block=10,
                scanned_blocks=0,
                status="not_completed",
                raw_payload={"source_import_id": "mirror-import", "source_order_id": "mirror-order"},
            ))
            db.commit()

        with mock.patch("backend.app.reconciliation_service.load_google_sheet_records", side_effect=RuntimeError("Google timeout")):
            response = self.client.get("/api/v1/reports/reconciliation/day?report_date=2026-06-10")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["source"], "postgres")
        self.assertEqual(payload["status"], "mirror_issue")
        self.assertEqual(payload["google"]["status"], "error")
        self.assertEqual(payload["skladbot"]["missing_request_orders"], 0)
        self.assertEqual(payload["skladbot"]["problem_status_orders"], 0)
        self.assertEqual(payload["alerts"], [])

        with self.SessionLocal() as db:
            incidents = db.execute(select(Incident).where(Incident.source == "daily_reconciliation")).scalars().all()
            notifications = db.execute(select(PendingEvent).where(PendingEvent.event_type == "telegram_notification")).scalars().all()

        self.assertEqual(len(incidents), 1)
        self.assertEqual(incidents[0].severity, "warning")
        self.assertEqual(incidents[0].external_ref, "reconciliation:2026-06-10:google_mirror_unavailable")
        self.assertEqual(notifications, [])

    def test_diagnostics_logs_include_failed_events_import_errors_and_redact_secrets(self):
        with self.SessionLocal() as db:
            db.add(PendingEvent(
                event_type="telegram_excel_import",
                status="failed",
                attempts=2,
                payload={},
                last_error="Bearer secret-service-token failed",
            ))
            db.add(PendingEvent(
                event_type="google_sheets_export",
                idempotency_key="google:test:key",
                status="pending",
                attempts=1,
                payload={"next_attempt_at": "2026-06-16T12:01:00+00:00"},
                last_error="APIError: [429]: quota exceeded",
            ))
            db.add(PendingEvent(
                event_type="telegram_notification",
                status="processing",
                attempts=3,
                payload={},
                last_error="old processing",
                updated_at=datetime.now(timezone.utc) - timedelta(minutes=30),
            ))
            db.add(ImportJob(
                source="telegram",
                status="failed",
                rows_total=2,
                rows_imported=0,
                raw_payload={
                    "filename": "broken.xlsx",
                    "errors": ["row 1: missing client"],
                    "invalid_rows": 1,
                    "duplicate_rows": 0,
                },
            ))
            db.add(AuditLog(
                action="skladbot_worker_sync",
                entity_type="skladbot",
                entity_id="worker",
                payload={"matched": 0, "not_found": 1, "token": "should-hide"},
            ))
            db.add(AuditLog(
                action="scan_code_created",
                entity_type="scan_code",
                entity_id="scan",
                payload={"code": "010-secret-code"},
            ))
            db.commit()

        response = self.client.get("/api/v1/diagnostics/logs")

        self.assertEqual(response.status_code, 200)
        text = response.content.decode("utf-8")
        self.assertIn("Failed/Pending Events", text)
        self.assertIn("telegram_excel_import", text)
        self.assertIn("google_sheets_export", text)
        self.assertIn("status=pending", text)
        self.assertIn("idempotency_key=google:test:key", text)
        self.assertIn("next_attempt_at=2026-06-16T12:01:00+00:00", text)
        self.assertIn("Stale Processing Events", text)
        self.assertIn("broken.xlsx", text)
        self.assertIn("row 1: missing client", text)
        self.assertIn("skladbot_worker_sync", text)
        self.assertIn("Bearer ***", text)
        self.assertIn('"token": "***"', text)
        self.assertNotIn("secret-service-token", text)
        self.assertNotIn("010-secret-code", text)

    def test_admin_events_exposes_queue_diagnostics(self):
        with self.SessionLocal() as db:
            db.add(PendingEvent(
                event_type="google_sheets_export",
                idempotency_key="google:event:1",
                status="pending",
                attempts=2,
                payload={"next_attempt_at": "2026-06-16T12:01:00+00:00"},
                last_error="quota",
            ))
            db.add(PendingEvent(
                event_type="telegram_notification",
                status="processing",
                attempts=1,
                payload={},
                updated_at=datetime.now(timezone.utc) - timedelta(minutes=30),
                last_error="Bearer secret-token failed for 0104006396053978217ABCDE12345678901234567890",
            ))
            db.commit()

        response = self.client.get("/api/v1/admin/events")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["active"], 2)
        self.assertEqual(payload["summary"]["by_type"]["google_sheets_export"]["pending"], 1)
        self.assertEqual(payload["summary"]["by_type"]["telegram_notification"]["processing"], 1)
        self.assertEqual(len(payload["stale_processing"]), 1)
        recent = {event["idempotency_key"]: event for event in payload["recent_events"] if event.get("idempotency_key")}
        self.assertEqual(recent["google:event:1"]["attempts"], 2)
        self.assertEqual(recent["google:event:1"]["next_attempt_at"], "2026-06-16T12:01:00+00:00")
        dumped = str(payload)
        self.assertIn("Bearer ***", dumped)
        self.assertNotIn("secret-token", dumped)
        self.assertNotIn("0104006396053978217ABCDE12345678901234567890", dumped)

    def test_admin_operations_summarizes_attention_without_raw_payload_or_telegram_spam(self):
        with self.SessionLocal() as db:
            db.add(PendingEvent(
                event_type="google_sheets_export",
                idempotency_key="google:event:ops",
                status="failed",
                attempts=2,
                payload={
                    "action": "google_sheets_scan_export",
                    "entity_type": "order_item",
                    "entity_id": str(uuid.uuid4()),
                    "authorization": "secret-token",
                    "next_attempt_at": "2026-06-16T12:01:00+00:00",
                },
                last_error="APIError: [429] token=secret 0104006396053978217SECRETKIZVALUE",
            ))
            db.add(PendingEvent(
                event_type="telegram_notification",
                status="failed",
                attempts=1,
                payload={"chat_id": "123456789", "bot_token": "telegram-secret"},
                last_error="Bearer telegram-secret failed",
            ))
            stale_time = datetime.now(timezone.utc) - timedelta(minutes=20)
            db.add(PendingEvent(
                event_type="telegram_excel_import",
                status="processing",
                attempts=1,
                payload={"filename": "secret-client.xlsx", "chat_id": "987654321"},
                last_error="processing timeout",
                created_at=stale_time,
                updated_at=stale_time,
            ))
            db.add(Incident(
                source="telegram_import",
                severity="critical",
                status="open",
                title="Bearer secret-token incident",
                message="token=secret 0104006396053978217SECRETKIZVALUE",
            ))
            db.add(ImportJob(
                source="telegram",
                status="failed",
                rows_total=1,
                rows_imported=0,
                raw_payload={"filename": "broken.xlsx", "token": "secret-token"},
            ))
            db.commit()
            notifications_before = len(db.execute(select(PendingEvent).where(PendingEvent.event_type == "telegram_notification")).scalars().all())

        response = self.client.get("/api/v1/admin/operations")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "requires_attention")
        self.assertEqual(payload["summary"]["mirror"], 1)
        self.assertGreaterEqual(payload["summary"]["hot_path"], 3)
        categories = {item["category"] for item in payload["items"]}
        self.assertIn("google_mirror", categories)
        self.assertIn("telegram", categories)
        self.assertIn("incident", categories)
        self.assertIn("import", categories)
        shadow = payload["shadow_diagnostics"]
        self.assertEqual(shadow["backend_active_orders_source"], "postgres_backend")
        self.assertEqual(shadow["google_mirror_status"], "degraded")
        self.assertEqual(shadow["google_mirror_failed_exports"], 1)
        self.assertEqual(shadow["hot_path_stale_processing"], 1)
        self.assertEqual(shadow["telegram_worker_state"], "requires_attention")
        self.assertGreaterEqual(shadow["telegram_pending_events"], 2)
        self.assertIn("TakSklad: требуется внимание", payload["telegram_summary"])
        self.assertIn("retry", str(payload["items"]))
        dumped = str(payload)
        self.assertIn("error=present", dumped)
        self.assertNotIn("secret-token", dumped)
        self.assertNotIn("telegram-secret", dumped)
        self.assertNotIn("123456789", dumped)
        self.assertNotIn("987654321", dumped)
        self.assertNotIn("secret-client.xlsx", dumped)
        self.assertNotIn("broken.xlsx", dumped)
        self.assertNotIn("0104006396053978217SECRETKIZVALUE", dumped)
        with self.SessionLocal() as db:
            notifications_after = len(db.execute(select(PendingEvent).where(PendingEvent.event_type == "telegram_notification")).scalars().all())
        self.assertEqual(notifications_after, notifications_before)

    def test_admin_events_default_response_is_not_capped(self):
        with self.SessionLocal() as db:
            db.add_all([
                PendingEvent(
                    event_type="google_sheets_export",
                    idempotency_key=f"google:event:{index}",
                    status="completed",
                    attempts=1,
                    payload={},
                )
                for index in range(101)
            ])
            db.commit()

        response = self.client.get("/api/v1/admin/events")
        limited_response = self.client.get("/api/v1/admin/events?limit=3")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()["recent_events"]), 101)
        self.assertEqual(limited_response.status_code, 200)
        self.assertEqual(len(limited_response.json()["recent_events"]), 3)

    def test_admin_event_detail_retry_redacts_payload_and_writes_audit(self):
        order_id, item_id = self.seed_order()
        with self.SessionLocal() as db:
            import_job = ImportJob(
                source="telegram",
                status="failed",
                rows_total=1,
                rows_imported=0,
                raw_payload={"filename": "failed.xlsx"},
            )
            db.add(import_job)
            db.commit()
            retryable = PendingEvent(
                event_type="telegram_excel_import",
                idempotency_key="telegram:import:failed",
                status="failed",
                attempts=3,
                payload={
                    "order_id": order_id,
                    "import_id": str(import_job.id),
                    "entity_type": "order_item",
                    "entity_id": item_id,
                    "document": {"file_name": "failed.xlsx", "file_id": "telegram-file-1"},
                    "authorization": "secret-token",
                    "nested": {"bot_token": "telegram-secret"},
                    "next_attempt_at": "2026-06-16T12:01:00+00:00",
                },
                last_error="Bearer secret-token failed for 0104006396053978217ABCDE12345678901234567890",
            )
            terminal = PendingEvent(
                event_type="telegram_excel_import",
                status="completed",
                attempts=1,
                payload={},
            )
            state_event = PendingEvent(
                event_type="telegram_worker_state",
                status="failed",
                attempts=1,
                payload={},
            )
            db.add_all([retryable, terminal, state_event])
            db.commit()
            import_id = str(import_job.id)
            retryable_id = str(retryable.id)
            terminal_id = str(terminal.id)
            state_event_id = str(state_event.id)

        list_response = self.client.get("/api/v1/admin/events")
        self.assertEqual(list_response.status_code, 200)
        listed = {
            event["id"]: event
            for event in list_response.json()["recent_events"]
        }
        self.assertTrue(listed[retryable_id]["retryable"])
        self.assertEqual(listed[retryable_id]["linked_order_id"], order_id)
        self.assertEqual(listed[retryable_id]["linked_import_id"], import_id)
        self.assertEqual(listed[retryable_id]["raw_payload"]["authorization"], "***")
        self.assertEqual(listed[retryable_id]["raw_payload"]["nested"]["bot_token"], "***")
        dumped_list = str(list_response.json())
        self.assertIn("Bearer ***", dumped_list)
        self.assertNotIn("secret-token", dumped_list)
        self.assertNotIn("telegram-secret", dumped_list)
        self.assertNotIn("0104006396053978217ABCDE12345678901234567890", dumped_list)

        detail_response = self.client.get(f"/api/v1/admin/events/{retryable_id}")
        self.assertEqual(detail_response.status_code, 200)
        detail = detail_response.json()
        self.assertEqual(detail["id"], retryable_id)
        self.assertTrue(detail["retryable"])
        self.assertEqual(detail["raw_payload"]["authorization"], "***")

        missing_reason = self.client.post(
            f"/api/v1/admin/events/{retryable_id}/retry",
            json={"actor": "anton"},
        )
        self.assertEqual(missing_reason.status_code, 422)
        self.assertEqual(missing_reason.json()["detail"], "reason is required")

        retry_response = self.client.post(
            f"/api/v1/admin/events/{retryable_id}/retry",
            json={
                "reason": "Manual retry after operator review",
                "actor": "anton",
                "source": "web",
                "idempotency_key": "retry-1",
            },
        )
        self.assertEqual(retry_response.status_code, 200)
        retry_payload = retry_response.json()
        self.assertEqual(retry_payload["status"], "pending")
        self.assertEqual(retry_payload["last_error"], "")
        self.assertEqual(retry_payload["next_attempt_at"], "")
        self.assertTrue(retry_payload["retryable"])
        self.assertEqual(retry_payload["raw_payload"]["authorization"], "***")
        self.assertEqual(retry_payload["raw_payload"]["manual_retry_actor"], "anton")

        completed_retry = self.client.post(
            f"/api/v1/admin/events/{terminal_id}/retry",
            json={"reason": "Should not retry terminal", "actor": "anton"},
        )
        self.assertEqual(completed_retry.status_code, 409)
        state_retry = self.client.post(
            f"/api/v1/admin/events/{state_event_id}/retry",
            json={"reason": "Should not retry state event", "actor": "anton"},
        )
        self.assertEqual(state_retry.status_code, 409)

        with self.SessionLocal() as db:
            event = db.get(PendingEvent, uuid.UUID(retryable_id))
            self.assertEqual(event.status, "pending")
            self.assertEqual(event.last_error, "")
            self.assertNotIn("next_attempt_at", event.payload)
            audit = db.execute(
                select(AuditLog).where(AuditLog.action == "pending_event_retry_requested")
            ).scalar_one()
            self.assertEqual(audit.entity_type, "pending_event")
            self.assertEqual(audit.entity_id, retryable_id)
            self.assertEqual(audit.payload["old_status"], "failed")
            self.assertEqual(audit.payload["new_status"], "pending")
            self.assertEqual(audit.payload["actor"], "anton")
            self.assertEqual(audit.payload["source"], "web")
            self.assertEqual(audit.payload["reason"], "Manual retry after operator review")

    def test_admin_event_retry_rejects_telegram_import_when_original_file_is_unavailable(self):
        with self.SessionLocal() as db:
            event = PendingEvent(
                event_type="telegram_excel_import",
                status="failed",
                attempts=2,
                payload={"document": {"file_name": "missing.xlsx"}},
                last_error="download failed",
            )
            db.add(event)
            db.commit()
            event_id = str(event.id)

        response = self.client.post(
            f"/api/v1/admin/events/{event_id}/retry",
            json={"reason": "Try again after operator review", "actor": "anton"},
        )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["detail"]["message"], "Original Telegram file is unavailable for retry")
        with self.SessionLocal() as db:
            event = db.get(PendingEvent, uuid.UUID(event_id))
            self.assertEqual(event.status, "failed")
            self.assertEqual(event.last_error, "download failed")

    def test_admin_event_retry_accepts_failed_skladbot_create_event(self):
        order_id, _item_id = self.seed_order()
        with self.SessionLocal() as db:
            event = PendingEvent(
                event_type="skladbot_request_create",
                status="failed",
                attempts=1,
                payload={
                    "order_id": order_id,
                    "create_status": "create_failed",
                    "next_attempt_at": "2026-06-16T12:01:00+00:00",
                },
                last_error="Недостаточно товара на складе",
            )
            db.add(event)
            db.commit()
            event_id = str(event.id)

        response = self.client.post(
            f"/api/v1/admin/events/{event_id}/retry",
            json={"reason": "Stock replenished, retry SkladBot create", "actor": "anton", "source": "web"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "pending")
        self.assertEqual(payload["last_error"], "")
        self.assertEqual(payload["linked_order_id"], order_id)
        self.assertEqual(payload["raw_payload"]["manual_retry_actor"], "anton")
        with self.SessionLocal() as db:
            event = db.get(PendingEvent, uuid.UUID(event_id))
            self.assertEqual(event.status, "pending")
            self.assertNotIn("next_attempt_at", event.payload)

    def test_admin_incidents_link_filter_redact_and_change_status_with_audit(self):
        order_id, item_id = self.seed_order()
        with self.SessionLocal() as db:
            import_job = ImportJob(
                source="telegram",
                status="failed",
                rows_total=1,
                rows_imported=0,
                raw_payload={"filename": "broken.xlsx"},
            )
            pending_event = PendingEvent(
                event_type="telegram_excel_import",
                status="failed",
                attempts=2,
                payload={"file": "broken.xlsx"},
                last_error="Bearer secret-token failed",
            )
            scan_code = ScanCode(
                order_item_id=uuid.UUID(item_id),
                code="0104006396053978217SECRETKIZVALUE",
                raw_payload={},
            )
            db.add_all([import_job, pending_event, scan_code])
            db.commit()
            import_id = str(import_job.id)
            event_id = str(pending_event.id)
            scan_id = str(scan_code.id)

        create_response = self.client.post(
            "/api/v1/admin/incidents",
            json={
                "source": "telegram_import",
                "severity": "critical",
                "status": "open",
                "title": "Import failed",
                "message": "Bearer secret-token failed for 0104006396053978217SECRETKIZVALUE",
                "entity_type": "pending_event",
                "entity_id": event_id,
                "pending_event_id": event_id,
                "order_id": order_id,
                "order_item_id": item_id,
                "import_id": import_id,
                "scan_code_id": scan_id,
                "external_ref": "WH-R-INCIDENT",
                "raw_payload": {
                    "authorization": "super-secret",
                    "nested": ["0104006396053978217SECRETKIZVALUE"],
                },
            },
        )

        self.assertEqual(create_response.status_code, 201)
        incident = create_response.json()
        incident_id = incident["id"]
        self.assertEqual(incident["status"], "open")
        self.assertEqual(incident["severity"], "critical")
        self.assertEqual(incident["source"], "telegram_import")
        self.assertEqual(incident["entity_type"], "pending_event")
        self.assertEqual(incident["entity_id"], event_id)
        self.assertEqual(incident["pending_event_id"], event_id)
        self.assertEqual(incident["order_id"], order_id)
        self.assertEqual(incident["order_item_id"], item_id)
        self.assertEqual(incident["import_id"], import_id)
        self.assertEqual(incident["scan_code_id"], scan_id)
        self.assertEqual(incident["external_ref"], "WH-R-INCIDENT")
        dumped = str(incident)
        self.assertIn("Bearer ***", dumped)
        self.assertIn("'authorization': '***'", dumped)
        self.assertNotIn("secret-token", dumped)
        self.assertNotIn("super-secret", dumped)
        self.assertNotIn("0104006396053978217SECRETKIZVALUE", dumped)

        list_response = self.client.get(
            "/api/v1/admin/incidents",
            params={
                "status": "open",
                "severity": "critical",
                "source": "telegram_import",
                "entity_type": "pending_event",
                "date_from": "2000-01-01",
                "date_to": "2999-01-01",
            },
        )

        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(len(list_payload["items"]), 1)
        self.assertEqual(list_payload["items"][0]["id"], incident_id)
        self.assertEqual(list_payload["summary"]["by_status"]["open"], 1)
        self.assertEqual(list_payload["summary"]["by_severity"]["critical"], 1)

        with self.SessionLocal() as db:
            db.add_all([
                Incident(
                    source="bulk",
                    severity="warning",
                    status="open",
                    title=f"Bulk incident {index}",
                    raw_payload={},
                )
                for index in range(101)
            ])
            db.commit()

        uncapped_response = self.client.get("/api/v1/admin/incidents")
        limited_response = self.client.get("/api/v1/admin/incidents?limit=3")
        self.assertEqual(uncapped_response.status_code, 200)
        self.assertEqual(len(uncapped_response.json()["items"]), 102)
        self.assertEqual(limited_response.status_code, 200)
        self.assertEqual(len(limited_response.json()["items"]), 3)

        detail_response = self.client.get(f"/api/v1/admin/incidents/{incident_id}")
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(detail_response.json()["id"], incident_id)

        status_response = self.client.post(
            f"/api/v1/admin/incidents/{incident_id}/status",
            json={
                "status": "resolved",
                "actor": "anton",
                "source": "web",
                "reason": "Checked and fixed import",
            },
        )

        self.assertEqual(status_response.status_code, 200)
        self.assertEqual(status_response.json()["status"], "resolved")
        self.assertTrue(status_response.json()["resolved_at"])
        with self.SessionLocal() as db:
            audit = db.execute(
                select(AuditLog).where(AuditLog.action == "incident_status_changed")
            ).scalar_one()
            self.assertEqual(audit.entity_type, "incident")
            self.assertEqual(audit.entity_id, incident_id)
            self.assertEqual(audit.payload["old_status"], "open")
            self.assertEqual(audit.payload["new_status"], "resolved")
            self.assertEqual(audit.payload["actor"], "anton")
            self.assertEqual(audit.payload["source"], "web")
            self.assertEqual(audit.payload["reason"], "Checked and fixed import")

    def test_admin_incident_accepts_every_status_severity_and_external_ref(self):
        created_ids = []
        for index, incident_status in enumerate(["open", "in_progress", "manual_review", "resolved", "ignored", "cancelled"]):
            response = self.client.post(
                "/api/v1/admin/incidents",
                json={
                    "source": "manual",
                    "severity": ["info", "warning", "critical"][index % 3],
                    "status": incident_status,
                    "title": f"External incident {index}",
                    "entity_type": "external",
                    "external_ref": f"EXT-{index}",
                },
            )
            self.assertEqual(response.status_code, 201)
            created_ids.append(response.json()["id"])

        self.assertEqual(len(set(created_ids)), 6)
        with self.SessionLocal() as db:
            self.assertEqual(len(db.execute(select(Incident)).scalars().all()), 6)

    def test_health_is_lightweight_and_readiness_reports_sanitized_db_queue_status(self):
        with self.SessionLocal() as db:
            db.execute(text("CREATE TABLE alembic_version (version_num VARCHAR(32) NOT NULL)"))
            db.execute(text("INSERT INTO alembic_version (version_num) VALUES ('20260616_0001')"))
            db.add(PendingEvent(
                event_type="google_sheets_export",
                idempotency_key="google:event:pending",
                status="pending",
                attempts=1,
                payload={"next_attempt_at": "2026-06-16T12:01:00+00:00"},
                created_at=datetime.now(timezone.utc) - timedelta(minutes=20),
                last_error="APIError: [429]: quota exceeded",
            ))
            db.add(PendingEvent(
                event_type="telegram_notification",
                status="failed",
                attempts=2,
                payload={},
                last_error="Bearer secret-token failed for 0104006396053978217ABCDE12345678901234567890",
            ))
            db.add(PendingEvent(
                event_type="telegram_chat_state:123456789",
                status="pending",
                attempts=0,
                payload={"chat_id": "123456789"},
            ))
            db.add(ImportJob(
                source="telegram",
                status="failed",
                rows_total=1,
                rows_imported=0,
                raw_payload={
                    "filename": "broken.xlsx",
                    "errors": ["authorization token=super-secret failed"],
                },
            ))
            db.commit()

        health = self.client.get("/health")
        readiness = self.client.get("/ready")
        api_readiness = self.client.get("/api/v1/readiness")

        self.assertEqual(health.status_code, 200)
        self.assertEqual(health.json()["status"], "ok")
        self.assertNotIn("database", health.json())
        self.assertNotIn("queue", health.json())

        self.assertEqual(readiness.status_code, 200)
        payload = readiness.json()
        self.assertEqual(payload["status"], "degraded")
        self.assertEqual(payload["database"]["status"], "ok")
        self.assertEqual(payload["migrations"]["status"], "ok")
        self.assertEqual(payload["migrations"]["current_revision"], "20260616_0001")
        self.assertEqual(payload["queue"]["summary"]["by_type"]["google_sheets_export"]["pending"], 1)
        self.assertEqual(payload["queue"]["summary"]["by_type"]["telegram_chat_state:*"]["pending"], 1)
        self.assertEqual(payload["google_mirror"]["status"], "degraded")
        self.assertEqual(payload["google_mirror"]["role"], "mirror_export")
        self.assertEqual(payload["google_mirror"]["summary"]["pending"], 1)
        self.assertGreaterEqual(payload["queue"]["oldest_pending_age_seconds"], 60)
        self.assertEqual(payload["queue"]["stale_processing_count"], 0)
        dumped = str(payload)
        self.assertIn("Bearer ***", dumped)
        self.assertIn("authorization ***", dumped)
        self.assertNotIn("secret-token", dumped)
        self.assertNotIn("super-secret", dumped)
        self.assertNotIn("0104006396053978217ABCDE", dumped)
        self.assertNotIn("telegram_chat_state:123456789", dumped)
        self.assertNotIn("'chat_id': '123456789'", dumped)
        self.assertEqual(api_readiness.status_code, 200)

    def test_readiness_keeps_hot_path_ok_when_only_google_mirror_is_degraded(self):
        next_attempt_at = datetime.now(timezone.utc) + timedelta(minutes=5)
        with self.SessionLocal() as db:
            db.execute(text("CREATE TABLE alembic_version (version_num VARCHAR(32) NOT NULL)"))
            db.execute(text("INSERT INTO alembic_version (version_num) VALUES ('20260616_0001')"))
            db.add(PendingEvent(
                event_type="google_sheets_export",
                idempotency_key="google:event:rate-limited",
                status="failed",
                attempts=3,
                payload={
                    "action": "google_sheets_scan_export",
                    "entity_type": "order_item",
                    "entity_id": str(uuid.uuid4()),
                    "next_attempt_at": next_attempt_at.isoformat(),
                    "last_result": {
                        "status": "rate_limited",
                        "error": "APIError: [429] token=secret 0104006396053978217SECRETKIZVALUE",
                    },
                },
                created_at=datetime.now(timezone.utc) - timedelta(minutes=10),
                last_error="APIError: [429] token=secret 0104006396053978217SECRETKIZVALUE",
            ))
            db.commit()

        response = self.client.get("/ready")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["queue"]["hot_path_last_errors"], [])
        self.assertEqual(payload["queue"]["hot_path_stale_processing_count"], 0)
        self.assertEqual(payload["google_mirror"]["status"], "degraded")
        self.assertEqual(payload["google_mirror"]["role"], "mirror_export")
        self.assertEqual(payload["google_mirror"]["event_type"], "google_sheets_export")
        self.assertEqual(payload["google_mirror"]["summary"]["failed"], 1)
        self.assertTrue(payload["google_mirror"]["paused"])
        self.assertEqual(payload["google_mirror"]["last_errors"][0]["event_type"], "google_sheets_export")
        dumped = str(payload)
        self.assertIn("token=***", dumped)
        self.assertNotIn("secret", dumped)
        self.assertNotIn("0104006396053978217SECRETKIZVALUE", dumped)

    def test_readiness_accepts_user_password_hash_schema_head_revision(self):
        with self.SessionLocal() as db:
            db.execute(text("CREATE TABLE alembic_version (version_num VARCHAR(32) NOT NULL)"))
            db.execute(text("INSERT INTO alembic_version (version_num) VALUES ('20260623_0004')"))
            db.commit()

        response = self.client.get("/ready")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["migrations"]["status"], "ok")
        self.assertEqual(payload["migrations"]["expected_baseline"], "20260616_0001")
        self.assertEqual(payload["migrations"]["expected_head"], "20260623_0004")
        self.assertEqual(payload["migrations"]["current_revision"], "20260623_0004")

    def test_readiness_degrades_when_migration_state_is_missing_or_wrong(self):
        response = self.client.get("/ready")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["database"]["status"], "ok")
        self.assertEqual(payload["status"], "degraded")
        self.assertEqual(payload["migrations"]["status"], "not_configured")

        with self.SessionLocal() as db:
            db.execute(text("CREATE TABLE alembic_version (version_num VARCHAR(32) NOT NULL)"))
            db.execute(text("INSERT INTO alembic_version (version_num) VALUES ('old_revision')"))
            db.commit()

        response = self.client.get("/ready")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "degraded")
        self.assertEqual(payload["migrations"]["status"], "revision_mismatch")
        self.assertEqual(payload["migrations"]["current_revision"], "old_revision")


if __name__ == "__main__":
    unittest.main()
