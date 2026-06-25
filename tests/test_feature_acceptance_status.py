import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tools.feature_acceptance_status import DEFAULT_REGISTER_PATH, evaluate_register


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = PROJECT_ROOT / "tools" / "feature_acceptance_status.py"


class FeatureAcceptanceStatusTests(unittest.TestCase):
    def write_temp_register(self, tmp_dir):
        temp_register = Path(tmp_dir) / "feature-register.xlsx"
        temp_register.write_bytes(DEFAULT_REGISTER_PATH.read_bytes())
        return temp_register

    def test_current_register_reports_manual_acceptance_pending(self):
        result = evaluate_register(DEFAULT_REGISTER_PATH)

        self.assertEqual(result["scope"], "feature_register_status")
        self.assertIn("not production release GO/NO-GO", result["release_gate_note"])
        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["stories"]["total"], 47)
        self.assertEqual(result["stories"]["automated_passed"], 46)
        self.assertEqual(result["test_loop"]["total"], 47)
        self.assertEqual(result["manual_acceptance"]["total"], 45)
        self.assertFalse(result["ready"]["manual_complete"])
        self.assertGreater(result["manual_acceptance"]["pending"], 0)

    def test_cli_require_manual_complete_fails_until_manual_rows_are_accepted(self):
        proc = subprocess.run(
            [
                sys.executable,
                str(SCRIPT_PATH),
                "--register",
                str(DEFAULT_REGISTER_PATH),
                "--require-manual-complete",
            ],
            cwd=PROJECT_ROOT,
            text=True,
            capture_output=True,
        )

        self.assertEqual(proc.returncode, 3)
        payload = json.loads(proc.stdout)
        self.assertFalse(payload["ready"]["manual_complete"])

    def test_cli_require_no_open_errors_fails_until_errors_are_closed(self):
        proc = subprocess.run(
            [
                sys.executable,
                str(SCRIPT_PATH),
                "--register",
                str(DEFAULT_REGISTER_PATH),
                "--require-no-open-errors",
            ],
            cwd=PROJECT_ROOT,
            text=True,
            capture_output=True,
        )

        self.assertEqual(proc.returncode, 4)
        payload = json.loads(proc.stdout)
        self.assertFalse(payload["ready"]["no_open_errors"])

    def test_cli_require_no_open_errors_passes_when_errors_are_closed_in_temp_copy(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_register = Path(tmp_dir) / "feature-register.xlsx"
            temp_register.write_bytes(DEFAULT_REGISTER_PATH.read_bytes())
            workbook = load_workbook(temp_register)
            sheet = workbook["Errors"]
            headers = [cell.value for cell in sheet[1]]
            status_col = headers.index("Status") + 1
            retest_col = headers.index("Retest Evidence") + 1
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, status_col).value = "fixed_retested"
                sheet.cell(row, retest_col).value = "Closed in test copy"
            workbook.save(temp_register)

            proc = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    "--register",
                    str(temp_register),
                    "--require-no-open-errors",
                ],
                cwd=PROJECT_ROOT,
                text=True,
                capture_output=True,
            )

        self.assertEqual(proc.returncode, 0, proc.stdout + proc.stderr)
        payload = json.loads(proc.stdout)
        self.assertTrue(payload["ready"]["no_open_errors"])

    def test_cli_passes_when_manual_rows_are_accepted_in_temp_copy(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_register = self.write_temp_register(tmp_dir)
            workbook = load_workbook(temp_register)
            sheet = workbook["Manual Acceptance"]
            headers = [cell.value for cell in sheet[1]]
            status_col = headers.index("Status") + 1
            observed_col = headers.index("Observed") + 1
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, status_col).value = "accepted"
                sheet.cell(row, observed_col).value = "Accepted in test copy"
            workbook.save(temp_register)

            proc = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    "--register",
                    str(temp_register),
                    "--require-manual-complete",
                ],
                cwd=PROJECT_ROOT,
                text=True,
                capture_output=True,
            )

        self.assertEqual(proc.returncode, 0, proc.stdout + proc.stderr)
        payload = json.loads(proc.stdout)
        self.assertTrue(payload["ready"]["manual_complete"])

    def test_not_applicable_does_not_close_manual_acceptance(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_register = self.write_temp_register(tmp_dir)
            workbook = load_workbook(temp_register)
            sheet = workbook["Manual Acceptance"]
            headers = [cell.value for cell in sheet[1]]
            status_col = headers.index("Status") + 1
            observed_col = headers.index("Observed") + 1
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, status_col).value = "not_applicable"
                sheet.cell(row, observed_col).value = "Should not bypass manual gate"
            workbook.save(temp_register)

            result = evaluate_register(temp_register)

        self.assertFalse(result["ready"]["manual_complete"])
        self.assertEqual(result["manual_acceptance"]["passed"], 0)
        self.assertEqual(result["manual_acceptance"]["pending"], result["manual_acceptance"]["total"])

    def test_missing_manual_acceptance_row_is_register_error(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_register = self.write_temp_register(tmp_dir)
            workbook = load_workbook(temp_register)
            sheet = workbook["Manual Acceptance"]
            sheet.delete_rows(2)
            workbook.save(temp_register)

            result = evaluate_register(temp_register)

        self.assertEqual(result["status"], "error")
        self.assertFalse(result["ready"]["all_feature_ids_consistent"])
        self.assertTrue(
            any("Manual Acceptance missing required feature IDs" in problem for problem in result["problems"]),
            result["problems"],
        )

    def test_empty_manual_acceptance_sheet_is_register_error(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_register = self.write_temp_register(tmp_dir)
            workbook = load_workbook(temp_register)
            sheet = workbook["Manual Acceptance"]
            if sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row - 1)
            workbook.save(temp_register)

            result = evaluate_register(temp_register)

        self.assertEqual(result["status"], "error")
        self.assertEqual(result["manual_acceptance"]["total"], 0)
        self.assertTrue(
            any("Manual Acceptance missing required feature IDs" in problem for problem in result["problems"]),
            result["problems"],
        )

    def test_missing_required_column_is_register_error(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_register = self.write_temp_register(tmp_dir)
            workbook = load_workbook(temp_register)
            sheet = workbook["Manual Acceptance"]
            headers = [cell.value for cell in sheet[1]]
            sheet.delete_cols(headers.index("Status") + 1)
            workbook.save(temp_register)

            result = evaluate_register(temp_register)

        self.assertEqual(result["status"], "error")
        self.assertTrue(
            any("Manual Acceptance missing required columns: Status" in problem for problem in result["problems"]),
            result["problems"],
        )

    def test_unknown_error_status_is_open_and_register_error(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_register = self.write_temp_register(tmp_dir)
            workbook = load_workbook(temp_register)
            sheet = workbook["Errors"]
            headers = [cell.value for cell in sheet[1]]
            status_col = headers.index("Status") + 1
            sheet.cell(2, status_col).value = "needs_validaton"
            workbook.save(temp_register)

            result = evaluate_register(temp_register)

        self.assertEqual(result["status"], "error")
        self.assertFalse(result["ready"]["no_open_errors"])
        self.assertTrue(
            any("Errors has unknown statuses: needs_validaton" in problem for problem in result["problems"]),
            result["problems"],
        )


if __name__ == "__main__":
    unittest.main()
