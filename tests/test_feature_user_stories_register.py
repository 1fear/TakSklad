import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
REGISTER_PATH = ROOT_DIR / "docs" / "taksklad-feature-user-stories.xlsx"


class FeatureUserStoriesRegisterTests(unittest.TestCase):
    def setUp(self):
        self.workbook = load_workbook(REGISTER_PATH, read_only=True, data_only=True)

    def sheet_rows(self, sheet_name):
        sheet = self.workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        return headers, list(sheet.iter_rows(min_row=2, values_only=True))

    def test_register_has_required_sheets_and_columns(self):
        self.assertTrue(REGISTER_PATH.exists())
        self.assertEqual(
            set(self.workbook.sheetnames),
            {"Summary", "User Stories", "Test Loop", "Errors", "Sources", "Manual Acceptance"},
        )

        required_user_story_columns = {
            "Feature ID",
            "Area",
            "Feature",
            "Actor",
            "User Story",
            "Expected Behaviour",
            "Evidence Files",
            "Current Test Evidence",
            "Status",
            "Test Status",
            "Automated Result",
            "Manual Result",
            "Live/Hardware Dependency",
            "Last Auto Evidence",
        }
        headers, rows = self.sheet_rows("User Stories")
        self.assertGreaterEqual(len(rows), 40)
        self.assertTrue(required_user_story_columns.issubset(set(headers)))

    def test_feature_ids_are_unique_and_cross_sheets_match(self):
        story_headers, story_rows = self.sheet_rows("User Stories")
        story_idx = {header: index for index, header in enumerate(story_headers)}
        story_ids = [row[story_idx["Feature ID"]] for row in story_rows]

        self.assertEqual(len(story_ids), len(set(story_ids)))
        self.assertTrue(all(story_ids))

        test_headers, test_rows = self.sheet_rows("Test Loop")
        test_idx = {header: index for index, header in enumerate(test_headers)}
        test_ids = [row[test_idx["Feature ID"]] for row in test_rows]
        self.assertEqual(set(test_ids), set(story_ids))

        manual_headers, manual_rows = self.sheet_rows("Manual Acceptance")
        manual_idx = {header: index for index, header in enumerate(manual_headers)}
        manual_ids = {row[manual_idx["Feature ID"]] for row in manual_rows}
        self.assertTrue(manual_ids.issubset(set(story_ids)))

    def test_statuses_are_consistent_with_test_type(self):
        headers, rows = self.sheet_rows("User Stories")
        idx = {header: index for index, header in enumerate(headers)}
        allowed_statuses = {
            "documented_from_code",
            "covered_auto_manual_pending",
            "covered_auto",
            "manual_required",
            "gap_found",
            "tested_ok",
            "fix_needed",
            "fixed_retest_pending",
            "fixed_tested_ok",
        }
        allowed_test_statuses = {
            "pending_test_loop",
            "not_run_current_loop",
            "passed",
            "failed",
            "manual_required",
            "blocked",
            "not_applicable",
        }

        for row in rows:
            feature_id = row[idx["Feature ID"]]
            test_type = row[idx["Test Type"]]
            status = row[idx["Status"]]
            test_status = row[idx["Test Status"]]
            automated_result = row[idx["Automated Result"]]
            manual_result = row[idx["Manual Result"]]

            self.assertIn(status, allowed_statuses, feature_id)
            self.assertIn(test_status, allowed_test_statuses, feature_id)
            if test_type == "auto":
                self.assertEqual(automated_result, "passed", feature_id)
                self.assertEqual(manual_result, "not_applicable", feature_id)
            elif test_type == "auto+manual":
                self.assertEqual(automated_result, "passed", feature_id)
                self.assertEqual(manual_result, "pending", feature_id)
            elif test_type == "manual":
                self.assertEqual(automated_result, "not_applicable", feature_id)
                self.assertEqual(manual_result, "pending", feature_id)
            else:
                self.fail(f"Unknown Test Type for {feature_id}: {test_type}")

    def test_test_loop_uses_unittest_and_existing_evidence_files(self):
        story_headers, story_rows = self.sheet_rows("User Stories")
        story_idx = {header: index for index, header in enumerate(story_headers)}
        missing_paths = []
        for row in story_rows:
            feature_id = row[story_idx["Feature ID"]]
            evidence = row[story_idx["Evidence Files"]] or ""
            for part in evidence.split(";"):
                path_text = part.strip()
                if not path_text or path_text.startswith("/"):
                    continue
                candidate = ROOT_DIR / path_text
                if not candidate.exists():
                    missing_paths.append(f"{feature_id}: {path_text}")
        self.assertEqual(missing_paths, [])

        test_headers, test_rows = self.sheet_rows("Test Loop")
        test_idx = {header: index for index, header in enumerate(test_headers)}
        for row in test_rows:
            feature_id = row[test_idx["Feature ID"]]
            command = row[test_idx["Command/Manual Step"]] or ""
            self.assertNotIn("pytest", command, feature_id)
            if command.startswith("python -m unittest"):
                modules = command.split()[3:]
                self.assertTrue(modules, feature_id)
                for module in modules:
                    if module.startswith("tests."):
                        path = ROOT_DIR / (module.split(".")[0] + "/" + module.split(".")[1] + ".py")
                        self.assertTrue(path.exists(), f"{feature_id}: {module}")


if __name__ == "__main__":
    unittest.main()
