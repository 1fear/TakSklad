import importlib.util
import importlib
import unittest
from unittest import mock
from io import BytesIO
from urllib.parse import quote

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.db import get_db
from backend.app.main import app, require_admin_write_permission, require_service_token
from backend.app.models import AuditLog, Base, Order, OrderItem, PendingEvent


REMOVED_GOOGLE_MODULES = (
    "backend.app.google_sheets_exporter",
    "backend.app.google_sheets_pending",
    "backend.app.google_sheets_sync_worker",
    "backend.app.google_backend_sync_diagnostic",
)


class DbOnlyRuntimeTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.SessionLocal = sessionmaker(bind=self.engine)

        def override_get_db():
            with self.SessionLocal() as db:
                yield db

        app.dependency_overrides[get_db] = override_get_db
        app.dependency_overrides[require_service_token] = lambda: None
        app.dependency_overrides[require_admin_write_permission] = lambda: None
        self.client = TestClient(app)

    def tearDown(self):
        app.dependency_overrides.clear()
        Base.metadata.drop_all(self.engine)
        self.engine.dispose()

    def workbook_bytes(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Заявки"
        sheet.append(["Клиент", "Тип оплаты", "ТМЦ", "Количество заказа", "Адрес", "Дата заказа"])
        sheet.append(["Клиент 1", "Терминал", "Chapman Brown OP 20", 20, "Самовывоз", "16.07.2026"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def test_google_runtime_modules_and_routes_are_removed(self):
        for module_name in REMOVED_GOOGLE_MODULES:
            self.assertIsNone(importlib.util.find_spec(module_name), module_name)
        paths = {route.path for route in app.routes}
        self.assertNotIn("/api/v1/admin/google/pending/retry", paths)
        self.assertNotIn("/api/v1/admin/orders/{order_id}/resync-google", paths)

    def test_raw_excel_preview_and_commit_use_postgres_only(self):
        content = self.workbook_bytes()
        headers = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "X-TakSklad-Filename": quote("Заказы 16.07.2026.xlsx"),
        }
        preview = self.client.post("/api/v1/imports/excel/preview", content=content, headers=headers)
        self.assertEqual(preview.status_code, 200, preview.text)
        self.assertEqual(preview.json()["preview"]["rows_importable"], 1)
        self.assertEqual(preview.json()["filename"], "Заказы 16.07.2026.xlsx")

        committed = self.client.post("/api/v1/imports/excel", content=content, headers=headers)
        self.assertEqual(committed.status_code, 201, committed.text)
        self.assertEqual(committed.json()["result"]["items_created"], 1)
        self.assertNotIn("google_sheets_status", committed.json()["result"])
        with self.SessionLocal() as db:
            self.assertEqual(len(db.execute(select(Order)).scalars().all()), 1)
            event_types = {event.event_type for event in db.execute(select(PendingEvent)).scalars().all()}
            self.assertNotIn("google_sheets_export", event_types)

    def test_admin_export_is_filtered_xlsx_from_database(self):
        with self.SessionLocal() as db:
            order = Order(source="test", payment_type="Терминал", client="Клиент", address="Самовывоз", status="not_completed")
            order.items.append(OrderItem(product="Chapman", quantity_pieces=10, quantity_blocks=1, scanned_blocks=0))
            db.add(order)
            db.commit()
        response = self.client.get("/api/v1/admin/orders/export.xlsx?status_bucket=active")
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.headers["X-TakSklad-Row-Count"], "1")
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        try:
            rows = list(workbook["Заказы"].iter_rows(values_only=True))
        finally:
            workbook.close()
        self.assertEqual(rows[1][2], "Клиент")
        self.assertEqual(rows[1][6], "Chapman")

    def test_migration_cancels_legacy_events_with_per_event_audit(self):
        migration = importlib.import_module(
            "backend.migrations.versions.20260716_0019_google_runtime_decommission"
        )
        with self.SessionLocal() as db:
            event = PendingEvent(
                event_type="google_sheets_export",
                status="failed",
                payload={"action": "legacy"},
                last_error="provider unavailable",
            )
            db.add(event)
            db.commit()
            event_id = event.id
        with (
            self.engine.begin() as connection,
            mock.patch.object(migration.op, "get_bind", return_value=connection),
            mock.patch.object(migration.op, "execute"),
            mock.patch.object(migration.op, "alter_column"),
        ):
            migration.upgrade()
        with self.SessionLocal() as db:
            event = db.get(PendingEvent, event_id)
            self.assertEqual(event.status, "cancelled")
            self.assertEqual(event.payload["google_runtime_decommission"]["previous_status"], "failed")
            audit = db.execute(
                select(AuditLog).where(AuditLog.action == "google_runtime_event_cancelled")
            ).scalar_one()
            self.assertEqual(audit.entity_id, str(event_id))

        with self.assertRaisesRegex(RuntimeError, "forward-only"):
            migration.downgrade()


if __name__ == "__main__":
    unittest.main()
