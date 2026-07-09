import os
import unittest
import uuid
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from types import SimpleNamespace
from zoneinfo import ZoneInfo

import openpyxl
from sqlalchemy import create_engine, select, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app import telegram_worker as telegram_worker_module
from backend.app.health_service import build_readiness_report
from backend.app.models import Base, PendingEvent
from backend.app.skladbot_daily_report import (
    DEFAULT_DAILY_REPORT_MAX_PAGES,
    MOVEMENT_HEADERS,
    REQUEST_HEADERS,
    REQUEST_PRODUCT_HEADERS,
    SkladBotReadOnlyClient,
    STOCK_HEADERS,
    build_skladbot_daily_report_message,
    build_skladbot_daily_report_xlsx,
    categorize_request_type,
    collect_skladbot_daily_report,
    product_breakdown_for_summary,
    read_style_post,
    request_representative,
    request_representative_zone,
    summarize_daily_report,
)
from backend.app.telegram_worker import (
    SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE,
    TelegramWorker,
)


def collect_report_without_delay(client, report_date):
    original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
    try:
        os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
        return collect_skladbot_daily_report(
            report_date=report_date,
            client=client,
        )
    finally:
        if original_delay is None:
            os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
        else:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay


def collect_report_with_env(client, report_date, env):
    originals = {key: os.environ.get(key) for key in env}
    try:
        for key, value in env.items():
            os.environ[key] = value
        return collect_skladbot_daily_report(
            report_date=report_date,
            client=client,
        )
    finally:
        for key, value in originals.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value


def worksheet_rows_by_header(sheet):
    headers = [cell.value for cell in sheet[1]]
    return [
        {header: sheet.cell(row=row, column=index + 1).value for index, header in enumerate(headers)}
        for row in range(2, sheet.max_row + 1)
    ]


class FakeSkladBotDailyReportClient:
    configured = True
    customer_id = 6211
    limit = 500

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {
                "data": {
                    "types": [
                        {"id": 3389, "name": "Отгрузка 3PL"},
                        {"id": 3403, "name": "Возврат 3PL"},
                        {"id": 3391, "name": "Приемка с услугами"},
                    ]
                }
            }
        if path == "/requests":
            request_type = params.get("type_id")
            if request_type == 3389:
                return {"data": [
                    {
                        "id": 101,
                        "delivery_number": "WH-R-101",
                        "type": "Отгрузка 3PL",
                        "created_at": "2026-06-08T10:00:00+05:00",
                    },
                    {
                        "id": 999,
                        "delivery_number": "WH-R-999",
                        "type": "Отгрузка 3PL",
                        "created_at": "2026-06-01T10:00:00+05:00",
                    },
                ]}
            if request_type == 3403:
                return {"data": [
                    {
                        "id": 202,
                        "delivery_number": "WH-R-202",
                        "type": "Возврат 3PL",
                        "created_at": "2026-06-08T10:00:00+05:00",
                        "updated_at": "2026-06-08T14:00:00+05:00",
                    },
                ]}
            if request_type == 3391:
                return {"data": [
                    {
                        "id": 303,
                        "delivery_number": "WH-R-303",
                        "type": "Приемка с услугами",
                        "created_at": "2026-06-08T09:00:00+05:00",
                    },
                ]}
            return {"data": []}
        raise AssertionError(path)

    def get_request_detail(self, request_id):
        details = {
            101: self.request_detail(
                101,
                "WH-R-101",
                "Отгрузка 3PL",
                "2026-06-08T10:00:00+05:00",
                "2026-06-08T11:00:00+05:00",
                "2026-06-08",
                "XASAN XUSAN SAVDO SERVIS XK",
                "Ташкент, Карасу",
                "Терминал\nТП1 Суюнбеков Умид\nРаб зона: Юнусабад\nSmartup ID: smartup:259704266\nРабочий номер: +998 91 111 11 11\nЛичный номер: +998 90 222 22 22",
                [{"name": "Chapman Brown OP 20", "vendorCode": "130400353", "barcode": "4006396053978", "amount": 4}],
            ),
            999: {
                **self.request_detail(
                    999,
                    "WH-R-999",
                    "Отгрузка 3PL",
                    "2026-06-01T10:00:00+05:00",
                    "2026-06-01T11:00:00+05:00",
                    "2026-06-01",
                    "OLD REQUEST",
                    "Ташкент",
                    "Терминал",
                    [{"name": "Chapman Brown OP 20", "vendorCode": "130400353", "barcode": "4006396053978", "amount": 1}],
                ),
                "completedAt": "2026-06-01T11:00:00+05:00",
                "archivedAt": "2026-06-01T12:00:00+05:00",
            },
            202: self.request_detail(
                202,
                "WH-R-202",
                "Возврат 3PL",
                "2026-06-08T10:00:00+05:00",
                "2026-06-08T14:00:00+05:00",
                "2026-06-08",
                "THE BIG RICH RAKAT",
                "Ташкент, возврат",
                "Возврат",
                [{"name": "Chapman RED OP 20", "vendorCode": "130400237", "barcode": "4006396053947", "amount": 2}],
            ),
            303: self.request_detail(
                303,
                "WH-R-303",
                "Приемка с услугами",
                "2026-06-08T09:00:00+05:00",
                "2026-06-08T09:30:00+05:00",
                "2026-06-08",
                "BASTION IMPORT",
                "Склад",
                "Приемка",
                [{"name": "Chapman Gold SSL", "vendorCode": "4006396054012", "barcode": "4006396054005", "amount": 1, "acceptedAmount": 500}],
            ),
        }
        return details[request_id]

    def post(self, path, payload=None):
        payload = payload or {}
        if path == "/products":
            return {
                "total": 3,
                "data": {
                    "2010857": [{
                        "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                        "name": "Chapman Brown OP 20",
                        "vendor_code": "130400353",
                        "barcode": "4006396053978",
                        "amount": 42,
                    }],
                    "2010858": [{
                        "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                        "name": "Chapman Gold SSL",
                        "vendor_code": "4006396054012",
                        "barcode": "4006396054005",
                        "amount": 500,
                    }],
                    "2010859": [{
                        "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                        "name": "Chapman RED OP 20",
                        "vendor_code": "130400237",
                        "barcode": "4006396053947",
                        "amount": 2,
                    }],
                },
            }
        if path == "/warehouse/transactions":
            if payload.get("type") == "in":
                return {"data": [
                    {
                        "date": "2026-06-08 09:30:00",
                        "delivery_number": "WH-R-303",
                        "type": "in",
                        "product": {"name": "Chapman Gold SSL", "vendorCode": "4006396054012", "barcode": "4006396054005"},
                        "amount": 500,
                        "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                        "box": {"number": "BOX-1"},
                    }
                ]}
            return {"data": [
                {
                    "date": "2026-06-08 11:00:00",
                    "delivery_number": "WH-R-101",
                    "type": "out",
                    "product": {"name": "Chapman Brown OP 20", "vendorCode": "130400353", "barcode": "4006396053978"},
                    "amount": 4,
                    "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                }
            ]}
        if path == "/report/stock":
            return {
                "total": 544,
                "items": [
                    {
                        "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                        "product": {"name": "Chapman Brown OP 20", "vendorCode": "130400353", "barcode": "4006396053978"},
                        "stock": 42,
                    },
                    {
                        "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                        "product": {"name": "Chapman Gold SSL", "vendorCode": "4006396054012", "barcode": "4006396054005"},
                        "stock": 500,
                    },
                    {
                        "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                        "product": {"name": "Chapman RED OP 20", "vendorCode": "130400237", "barcode": "4006396053947"},
                        "stock": 2,
                    },
                ],
            }
        raise AssertionError(path)

    @staticmethod
    def request_detail(request_id, number, request_type, created_at, updated_at, unloading_date, company, address, comment, products):
        return {
            "id": request_id,
            "delivery_number": number,
            "type": request_type,
            "isCompleted": True,
            "archived": True,
            "createdAt": created_at,
            "updatedAt": updated_at,
            "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
            "comment": comment,
            "fields": [
                {"field": "company_name", "value": company},
                {"field": "address", "value": address},
                {"field": "unloading_date", "value": unloading_date},
                {"field": "comment", "value": comment},
            ],
            "products": products,
        }


class TransientRateLimitDailyReportClient(FakeSkladBotDailyReportClient):
    def __init__(self):
        self.detail_calls = {}

    def get_request_detail(self, request_id):
        self.detail_calls[request_id] = self.detail_calls.get(request_id, 0) + 1
        if request_id == 101 and self.detail_calls[request_id] == 1:
            raise RuntimeError("429 Too Many Requests")
        return super().get_request_detail(request_id)


class TransientReadStylePostDailyReportClient(FakeSkladBotDailyReportClient):
    def __init__(self):
        self.post_attempts = {}

    def post(self, path, payload=None):
        payload = payload or {}
        key = (path, payload.get("type") or "")
        self.post_attempts[key] = self.post_attempts.get(key, 0) + 1
        if path == "/warehouse/transactions" and payload.get("type") == "in" and self.post_attempts[key] == 1:
            raise RuntimeError("SkladBot API HTTP 500")
        return super().post(path, payload)


class ForbiddenWriteDailyReportClient(FakeSkladBotDailyReportClient):
    write_attempts = 0

    def create_request(self, payload):
        self.write_attempts += 1
        raise AssertionError("daily report must not create SkladBot requests")

    def post(self, path, payload=None):
        if path == "/requests":
            self.write_attempts += 1
            raise AssertionError("daily report must not POST write requests")
        return super().post(path, payload)


class PreviousDayReceivingCompletedTodayClient(FakeSkladBotDailyReportClient):
    request_id = 404

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3391, "name": "Приемка с услугами"}]}}
        if path == "/requests" and params.get("type_id") == 3391:
            return {"data": [{
                "id": self.request_id,
                "delivery_number": "WH-R-404",
                "type": "Приемка с услугами",
                "created_at": "2026-06-19T18:00:00+05:00",
                "archived": True,
                "isCompleted": True,
            }]}
        return {"data": []}

    def get_request_detail(self, request_id):
        if request_id != self.request_id:
            raise AssertionError(request_id)
        return self.request_detail(
            self.request_id,
            "WH-R-404",
            "Приемка с услугами",
            "2026-06-19T18:00:00+05:00",
            "",
            "2026-06-19",
            "BASTION IMPORT",
            "Склад",
            "Приемка",
            [{"name": "Chapman Green OP 20", "vendorCode": "CHPMGreenOP20UZ", "barcode": "4006396104441", "amount": 1, "acceptedAmount": 10000}],
        )


class StaleRequestDetailShouldNotBeCalledClient(PreviousDayReceivingCompletedTodayClient):
    def get_request_detail(self, request_id):
        raise AssertionError("stale request should not be detailed")


class CompletedAfterCutoffReceivingClient(PreviousDayReceivingCompletedTodayClient):
    request_id = 405

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3391, "name": "Приемка с услугами"}]}}
        if path == "/requests" and params.get("type_id") == 3391:
            return {"data": [{
                "id": self.request_id,
                "delivery_number": "WH-R-405",
                "type": "Приемка с услугами",
                "created_at": "2026-06-20T18:00:00+05:00",
                "archived": True,
                "isCompleted": True,
            }]}
        return {"data": []}

    def get_request_detail(self, request_id):
        if request_id != self.request_id:
            raise AssertionError(request_id)
        detail = self.request_detail(
            self.request_id,
            "WH-R-405",
            "Приемка с услугами",
            "2026-06-20T18:00:00+05:00",
            "",
            "2026-06-20",
            "BASTION IMPORT",
            "Склад",
            "Приемка",
            [{"name": "Chapman Green OP 20", "vendorCode": "CHPMGreenOP20UZ", "barcode": "4006396104441", "amount": 1, "acceptedAmount": 500}],
        )
        detail["completedAt"] = "2026-06-20T23:10:00+05:00"
        return detail


class OldListDateReceivingCompletedTodayClient(PreviousDayReceivingCompletedTodayClient):
    request_id = 406

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3391, "name": "Приемка с услугами"}]}}
        if path == "/requests" and params.get("type_id") == 3391:
            return {"data": [{
                "id": self.request_id,
                "delivery_number": "WH-R-406",
                "type": "Приемка с услугами",
                "created_at": "2026-06-10T18:00:00+05:00",
                "archived": True,
                "isCompleted": True,
            }]}
        return {"data": []}

    def get_request_detail(self, request_id):
        if request_id != self.request_id:
            raise AssertionError(request_id)
        detail = self.request_detail(
            self.request_id,
            "WH-R-406",
            "Приемка с услугами",
            "2026-06-10T18:00:00+05:00",
            "",
            "2026-06-10",
            "BASTION IMPORT",
            "Склад",
            "Приемка",
            [{"name": "Chapman Green OP 20", "vendorCode": "CHPMGreenOP20UZ", "barcode": "4006396104441", "amount": 1, "acceptedAmount": 700}],
        )
        detail["completedAt"] = "2026-06-20T12:10:00+05:00"
        return detail


class ArchivedAfterCutoffReceivingClient(CompletedAfterCutoffReceivingClient):
    request_id = 407

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3391, "name": "Приемка с услугами"}]}}
        if path == "/requests" and params.get("type_id") == 3391:
            return {"data": [{
                "id": self.request_id,
                "delivery_number": "WH-R-407",
                "type": "Приемка с услугами",
                "created_at": "2026-06-20T18:00:00+05:00",
                "archived": True,
                "isCompleted": True,
            }]}
        return {"data": []}

    def get_request_detail(self, request_id):
        if request_id != self.request_id:
            raise AssertionError(request_id)
        detail = self.request_detail(
            self.request_id,
            "WH-R-407",
            "Приемка с услугами",
            "2026-06-20T18:00:00+05:00",
            "",
            "2026-06-20",
            "BASTION IMPORT",
            "Склад",
            "Приемка",
            [{"name": "Chapman Green OP 20", "vendorCode": "CHPMGreenOP20UZ", "barcode": "4006396104441", "amount": 1, "acceptedAmount": 900}],
        )
        detail["completedAt"] = "2026-06-20T20:10:00+05:00"
        detail["archivedAt"] = "2026-06-20T23:10:00+05:00"
        return detail


class OldCreatedUnloadingTodayClient(PreviousDayReceivingCompletedTodayClient):
    request_id = 408

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests" and params.get("type_id") == 3389:
            return {"data": [{
                "id": self.request_id,
                "delivery_number": "WH-R-408",
                "type": "Отгрузка 3PL",
                "created_at": "2026-06-19T18:00:00+05:00",
                "archived": True,
                "isCompleted": True,
            }]}
        return {"data": []}

    def get_request_detail(self, request_id):
        if request_id != self.request_id:
            raise AssertionError(request_id)
        detail = self.request_detail(
            self.request_id,
            "WH-R-408",
            "Отгрузка 3PL",
            "2026-06-19T18:00:00+05:00",
            "",
            "2026-06-20",
            "OLD CREATED UNLOADING TODAY",
            "Ташкент",
            "Терминал",
            [{"name": "Chapman RED OP 20", "vendorCode": "CHPMRedOP20UZ", "barcode": "4006396053947", "amount": 11}],
        )
        return detail


class MovementTodayRequestClient(PreviousDayReceivingCompletedTodayClient):
    request_id = 409

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests" and params.get("type_id") == 3389:
            return {"data": [{
                "id": self.request_id,
                "delivery_number": "WH-R-409",
                "type": "Отгрузка 3PL",
                "created_at": "2026-06-19T18:00:00+05:00",
                "archived": False,
                "isCompleted": False,
            }]}
        return {"data": []}

    def get_request_detail(self, request_id):
        if request_id != self.request_id:
            raise AssertionError(request_id)
        detail = self.request_detail(
            self.request_id,
            "WH-R-409",
            "Отгрузка 3PL",
            "2026-06-19T18:00:00+05:00",
            "",
            "2026-06-19",
            "MOVEMENT TODAY",
            "Ташкент",
            "Терминал",
            [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": 5}],
        )
        detail["isCompleted"] = False
        detail["archived"] = False
        return detail

    def post(self, path, payload=None):
        payload = payload or {}
        if path == "/warehouse/transactions":
            if payload.get("type") == "out":
                return {"data": [{
                    "date": "2026-06-20 12:00:00",
                    "delivery_number": "WH-R-409",
                    "type": "out",
                    "product": {"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978"},
                    "amount": 5,
                    "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                }]}
            return {"data": []}
        return super().post(path, payload)


class CompletedArchivedMovementTodayRequestClient(MovementTodayRequestClient):
    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests" and params.get("type_id") == 3389:
            return {"data": [{
                "id": self.request_id,
                "delivery_number": "WH-R-409",
                "type": "Отгрузка 3PL",
                "created_at": "2026-06-19T18:00:00+05:00",
                "archived": True,
                "isCompleted": True,
            }]}
        return {"data": []}

    def get_request_detail(self, request_id):
        detail = super().get_request_detail(request_id)
        detail["isCompleted"] = True
        detail["archived"] = True
        return detail


class MixedDateMovementsClient(MovementTodayRequestClient):
    def post(self, path, payload=None):
        payload = payload or {}
        if path == "/warehouse/transactions":
            if payload.get("type") == "out":
                return {"data": [
                    {
                        "date": "2026-06-19 23:59:00",
                        "delivery_number": "WH-R-OLD",
                        "type": "out",
                        "product": {"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978"},
                        "amount": 99,
                        "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                    },
                    {
                        "date": "2026-06-20 12:00:00",
                        "delivery_number": "WH-R-409",
                        "type": "out",
                        "product": {"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978"},
                        "amount": 5,
                        "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                    },
                    {
                        "date": "",
                        "delivery_number": "WH-R-NO-DATE",
                        "type": "out",
                        "product": {"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978"},
                        "amount": 7,
                        "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                    },
                ]}
            return {"data": []}
        return super().post(path, payload)


class MovementTodayRequestAfterStaleListItemsClient(MovementTodayRequestClient):
    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests" and params.get("type_id") == 3389:
            return {"data": [
                {
                    "id": 999,
                    "delivery_number": "WH-R-999",
                    "type": "Отгрузка 3PL",
                    "created_at": "2026-06-01T18:00:00+05:00",
                    "archived": True,
                    "isCompleted": True,
                },
                {
                    "id": self.request_id,
                    "delivery_number": "WH-R-409",
                    "type": "Отгрузка 3PL",
                    "created_at": "2026-06-20T18:00:00+05:00",
                    "archived": True,
                    "isCompleted": True,
                },
            ]}
        return {"data": []}

    def get_request_detail(self, request_id):
        if request_id == 999:
            raise AssertionError("stale request should not consume today's request detail limit")
        if request_id != self.request_id:
            raise AssertionError(request_id)
        return self.request_detail(
            self.request_id,
            "WH-R-409",
            "Отгрузка 3PL",
            "2026-06-20T18:00:00+05:00",
            "",
            "2026-06-20",
            "MOVEMENT TODAY",
            "Ташкент",
            "Терминал",
            [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": 5}],
        )


class PaginatedStrictCoverageClient(FakeSkladBotDailyReportClient):
    limit = 2

    def __init__(self):
        self.request_calls = []
        self.detail_calls = []

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests" and params.get("type_id") == 3389:
            self.request_calls.append(dict(params))
            if "offset" in params:
                raise AssertionError("daily report must not use offset pagination")
            page = int(params.get("page") or 1)
            pages = {
                1: [
                    self.list_item(401, "WH-R-401", "2026-06-18T10:00:00+05:00", True, True),
                    self.list_item(402, "WH-R-402", "2026-06-20T10:00:00+05:00", True, True),
                ],
                2: [
                    self.list_item(401, "WH-R-401", "2026-06-18T10:00:00+05:00", True, True),
                    self.list_item(403, "WH-R-403", "2026-06-19T10:00:00+05:00", False, False),
                ],
                3: [
                    self.list_item(404, "WH-R-404", "2026-06-19T10:00:00+05:00", True, False),
                    self.list_item(405, "WH-R-405", "2026-06-19T10:00:00+05:00", False, True),
                ],
                4: [
                    self.list_item(406, "WH-R-406", "2026-06-19T10:00:00+05:00", True, True),
                    self.list_item(407, "WH-R-407", "2026-06-18T10:00:00+05:00", True, True),
                ],
                5: [],
            }
            return {"data": pages.get(page, [])}
        return {"data": []}

    @staticmethod
    def list_item(request_id, number, created_at, completed, archived):
        return {
            "id": request_id,
            "delivery_number": number,
            "type": "Отгрузка 3PL",
            "created_at": created_at,
            "isCompleted": completed,
            "archived": archived,
        }

    def get_request_detail(self, request_id):
        self.detail_calls.append(request_id)
        details = {
            401: self.request_detail(
                401,
                "WH-R-401",
                "Отгрузка 3PL",
                "2026-06-18T10:00:00+05:00",
                "2026-06-20T12:00:00+05:00",
                "2026-06-20",
                "UNLOADING TODAY",
                "Ташкент",
                "Терминал",
                [{"name": "Chapman RED OP 20", "vendorCode": "CHPMRedOP20UZ", "barcode": "4006396053947", "amount": 11}],
            ),
            402: self.request_detail(
                402,
                "WH-R-402",
                "Отгрузка 3PL",
                "2026-06-20T10:00:00+05:00",
                "2026-06-20T12:00:00+05:00",
                "2026-06-22",
                "FUTURE UNLOADING",
                "Ташкент",
                "Терминал",
                [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": 3}],
            ),
            403: {
                **self.request_detail(
                    403,
                    "WH-R-403",
                    "Отгрузка 3PL",
                    "2026-06-19T10:00:00+05:00",
                    "2026-06-20T12:00:00+05:00",
                    "2026-06-20",
                    "NEITHER STATUS",
                    "Ташкент",
                    "Терминал",
                    [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": 4}],
                ),
                "isCompleted": False,
                "archived": False,
            },
            404: {
                **self.request_detail(
                    404,
                    "WH-R-404",
                    "Отгрузка 3PL",
                    "2026-06-19T10:00:00+05:00",
                    "2026-06-20T12:00:00+05:00",
                    "2026-06-20",
                    "COMPLETED ONLY",
                    "Ташкент",
                    "Терминал",
                    [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": 5}],
                ),
                "isCompleted": True,
                "archived": False,
            },
            405: {
                **self.request_detail(
                    405,
                    "WH-R-405",
                    "Отгрузка 3PL",
                    "2026-06-19T10:00:00+05:00",
                    "2026-06-20T12:00:00+05:00",
                    "2026-06-20",
                    "ARCHIVED ONLY",
                    "Ташкент",
                    "Терминал",
                    [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": 6}],
                ),
                "isCompleted": False,
                "archived": True,
            },
            406: self.request_detail(
                406,
                "WH-R-406",
                "Отгрузка 3PL",
                "2026-06-19T10:00:00+05:00",
                "2026-06-19T12:00:00+05:00",
                "2026-06-19",
                "STALE",
                "Ташкент",
                "Терминал",
                [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": 7}],
            ),
            407: self.request_detail(
                407,
                "WH-R-407",
                "Отгрузка 3PL",
                "2026-06-18T10:00:00+05:00",
                "2026-06-19T12:00:00+05:00",
                "2026-06-19",
                "MOVEMENT DATE",
                "Ташкент",
                "Терминал",
                [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": 8}],
            ),
        }
        return details[request_id]

    def post(self, path, payload=None):
        payload = payload or {}
        if path == "/warehouse/transactions":
            if payload.get("type") == "out":
                return {"data": [{
                    "id": "MOV-407",
                    "date": "2026-06-20 12:00:00",
                    "delivery_number": "WH-R-407",
                    "type": "out",
                    "product": {"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978"},
                    "amount": 8,
                    "customer": {"name": "ООО Bastion Import Chapman MCHJ"},
                }]}
            return {"data": []}
        return super().post(path, payload)


class MaxPagesDailyReportClient(PaginatedStrictCoverageClient):
    limit = 1

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests" and params.get("type_id") == 3389:
            self.request_calls.append(dict(params))
            return {"data": [self.list_item(401, "WH-R-401", "2026-06-18T10:00:00+05:00", True, True)]}
        return super().get(path, params)


class MultiTypeMaxPagesDailyReportClient(PaginatedStrictCoverageClient):
    limit = 1

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [
                {"id": 3389, "name": "Отгрузка 3PL"},
                {"id": 3403, "name": "Возврат 3PL"},
            ]}}
        if path == "/requests" and params.get("type_id") in {3389, 3403}:
            self.request_calls.append(dict(params))
            request_id = 401 if params.get("type_id") == 3389 else 501
            number = "WH-R-401" if request_id == 401 else "WH-R-501"
            return {"data": [self.list_item(request_id, number, "2026-06-20T10:00:00+05:00", True, True)]}
        return {"data": []}

    def get_request_detail(self, request_id):
        self.detail_calls.append(request_id)
        number = "WH-R-401" if request_id == 401 else "WH-R-501"
        request_type = "Отгрузка 3PL" if request_id == 401 else "Возврат 3PL"
        return self.request_detail(
            request_id,
            number,
            request_type,
            "2026-06-20T10:00:00+05:00",
            "2026-06-20T12:00:00+05:00",
            "2026-06-20",
            "TODAY",
            "Ташкент",
            "Терминал",
            [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": 1}],
        )


class Type3389DeepPaginationClient(PaginatedStrictCoverageClient):
    limit = 100

    def __init__(self, full_pages=30):
        super().__init__()
        self.full_pages = full_pages

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests" and params.get("type_id") == 3389:
            self.request_calls.append(dict(params))
            page = int(params.get("page") or 1)
            if page > self.full_pages:
                return {"data": []}
            rows = []
            for index in range(self.limit):
                request_id = page * 1000 + index
                row = self.list_item(
                    request_id,
                    f"WH-R-{request_id}",
                    "2026-07-01T10:00:00+05:00",
                    True,
                    True,
                )
                if page == 1 and index == 0:
                    row["created_at"] = "2026-07-07T10:00:00+05:00"
                    row["unloading_date"] = "2026-07-07"
                rows.append(row)
            return {"data": rows}
        return {"data": []}

    def get_request_detail(self, request_id):
        self.detail_calls.append(request_id)
        return self.request_detail(
            request_id,
            f"WH-R-{request_id}",
            "Отгрузка 3PL",
            "2026-07-07T10:00:00+05:00",
            "2026-07-07T12:00:00+05:00",
            "2026-07-07",
            "TODAY",
            "Ташкент",
            "Терминал",
            [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": 1}],
        )


class RepeatedPageDailyReportClient(PaginatedStrictCoverageClient):
    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests" and params.get("type_id") == 3389:
            self.request_calls.append(dict(params))
            return {"data": [
                self.list_item(401, "WH-R-401", "2026-06-18T10:00:00+05:00", True, True),
                self.list_item(402, "WH-R-402", "2026-06-20T10:00:00+05:00", True, True),
            ]}
        return super().get(path, params)


class ListFailureDailyReportClient(FakeSkladBotDailyReportClient):
    def get(self, path, params=None):
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests":
            raise RuntimeError("SkladBot list unavailable")
        return super().get(path, params)


class DetailFailureDailyReportClient(PaginatedStrictCoverageClient):
    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests" and params.get("type_id") == 3389:
            self.request_calls.append(dict(params))
            return {"data": [self.list_item(501, "WH-R-501", "2026-06-20T10:00:00+05:00", True, True)]}
        return {"data": []}

    def get_request_detail(self, request_id):
        self.detail_calls.append(request_id)
        raise RuntimeError("detail 500")


JULY7_MISSING_TRANSFER_BATCH = [
    (204498, '"VERY VITAL PREMIUM" MChJ (Теген)', 5, "ТП8 Мурод"),
    (204499, '"CENTRAL TOBAKKO" MCHJ', 58, "ОПТ"),
    (204500, '"JAVOHIR FAYZ XAMKOR" MChJ', 5, "ТП1 Суюнбеков Умид"),
    (204501, '"ALO ESHITAMAN" MCHJ', 8, "ТП7 Султонов Азимжон"),
    (204502, "TEQUILA LIGHTS", 5, "ТП5 Авазов Азиз Бегжонович"),
    (204503, '"NUMBER ONE DRINK" MCHJ', 9, "ТП6 Хасанов Мираббос"),
    (204504, "TASHEVA LAYLO KAMOLOVNA", 3, "ТП2 Кобилов Достон Рустам угли"),
    (204505, '"MAYAN BUSINESS" MCHJ (1 филиал)', 2, "ТП4 Елчиев Сардор"),
]


class July7MissingTransferBatchClient(FakeSkladBotDailyReportClient):
    def __init__(self, unloading_date="2026-07-08", status_matrix=None):
        self.unloading_date = unloading_date
        self.status_matrix = status_matrix or {}
        self.detail_calls = []

    def get(self, path, params=None):
        params = params or {}
        if path == "/requests/filter/fields":
            return {"data": {"types": [{"id": 3389, "name": "Отгрузка 3PL"}]}}
        if path == "/requests" and params.get("type_id") == 3389:
            page = int(params.get("page") or 1)
            if page > 1:
                return {"data": []}
            rows = []
            for request_id, _client, _blocks, _representative in JULY7_MISSING_TRANSFER_BATCH:
                completed, archived = self.status_matrix.get(request_id, (True, True))
                rows.append({
                    "id": request_id,
                    "delivery_number": f"WH-R-{request_id}",
                    "type": "Отгрузка 3PL",
                    "created_at": "2026-07-07T15:30:00+05:00",
                    "unloading_date": self.unloading_date,
                    "isCompleted": completed,
                    "archived": archived,
                })
            return {"data": rows}
        return {"data": []}

    def get_request_detail(self, request_id):
        self.detail_calls.append(request_id)
        row = next(item for item in JULY7_MISSING_TRANSFER_BATCH if item[0] == request_id)
        _request_id, client, blocks, representative = row
        completed, archived = self.status_matrix.get(request_id, (True, True))
        detail = self.request_detail(
            request_id,
            f"WH-R-{request_id}",
            "Отгрузка 3PL",
            "2026-07-07T15:30:00+05:00",
            "2026-07-07T16:10:00+05:00",
            self.unloading_date,
            client,
            "Ташкент",
            f"Перечисление\n{representative}",
            [{"name": "Chapman Brown OP 20", "vendorCode": "CHPMBrownOP20UZ", "barcode": "4006396053978", "amount": blocks}],
        )
        detail["isCompleted"] = completed
        detail["archived"] = archived
        detail["completedAt"] = "2026-07-07T16:00:00+05:00" if completed else ""
        detail["archivedAt"] = "2026-07-07T16:05:00+05:00" if archived else ""
        return detail

    def post(self, path, payload=None):
        if path == "/warehouse/transactions":
            return {"data": []}
        return super().post(path, payload)


class SkladBotDailyReportTests(unittest.TestCase):
    def test_parses_new_representative_comment_with_phones_without_zone(self):
        request = {
            "comment": (
                "Терминал\n"
                "ТП6 Хасанов Мираббос\n"
                "Рабочий номер: +998 77 000 00 00\n"
                "Личный номер: +998 93 000 00 00"
            ),
        }

        self.assertEqual(request_representative(request), "ТП6 Хасанов Мираббос")
        self.assertEqual(request_representative_zone(request), "")

    def test_collects_requests_movements_stock_and_builds_xlsx(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 8),
                client=FakeSkladBotDailyReportClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        summary = report["summary"]
        self.assertEqual(summary["requests_total"], 3)
        self.assertEqual(summary["category_counts"]["Отгрузка"], 1)
        self.assertEqual(summary["category_counts"]["Отгрузка в браке"], 0)
        self.assertEqual(summary["category_counts"]["Возврат"], 1)
        self.assertEqual(summary["category_counts"]["Приемка"], 1)
        self.assertEqual(summary["request_blocks_by_category"]["Приемка"], 500)
        self.assertEqual(summary["movement_in_amount"], 500)
        self.assertEqual(summary["movement_out_amount"], 4)
        self.assertEqual(summary["stock_total"], 544)
        self.assertEqual(report["errors"], [])

        content, filename = build_skladbot_daily_report_xlsx(report)
        self.assertEqual(filename, "TakSklad_SkladBot_daily_08.06.2026.xlsx")
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
        self.assertEqual(workbook.sheetnames, [
            "Сводка",
            "Заявки",
            "Товары заявок",
            "Остатки",
        ])

        summary_sheet = workbook["Сводка"]
        self.assertEqual(summary_sheet["A1"].value, "Показатель")
        self.assertEqual(summary_sheet["B1"].value, "Блоков")
        self.assertEqual(summary_sheet["C1"].value, "Заявок")
        self.assertEqual(summary_sheet["A2"].value, "Отгрузка")
        self.assertEqual(summary_sheet["B2"].value, 4)
        self.assertEqual(summary_sheet["C2"].value, 1)
        self.assertEqual(summary_sheet["A3"].value, "Отгрузка в браке")
        self.assertEqual(summary_sheet["B3"].value, 0)
        self.assertEqual(summary_sheet["C3"].value, 0)
        self.assertEqual(summary_sheet["A4"].value, "Возврат")
        self.assertEqual(summary_sheet["B4"].value, 2)
        self.assertEqual(summary_sheet["C4"].value, 1)
        self.assertEqual(summary_sheet["A5"].value, "Приемка")
        self.assertEqual(summary_sheet["B5"].value, 500)
        self.assertEqual(summary_sheet["C5"].value, 1)
        self.assertEqual(summary_sheet["A6"].value, "Актуальный остаток")
        self.assertEqual(summary_sheet["B6"].value, 544)
        self.assertIsNone(summary_sheet["C6"].value)
        self.assertEqual(summary_sheet.freeze_panes, "A2")
        self.assertEqual(summary_sheet["A2"].border.left.style, "thin")
        self.assertEqual(summary_sheet["C6"].border.right.style, "thin")
        self.assertEqual(summary_sheet.column_dimensions["A"].width, 28)
        self.assertEqual(summary_sheet.column_dimensions["B"].width, 13)

        requests_sheet = workbook["Заявки"]
        self.assertEqual([cell.value for cell in requests_sheet[1]], REQUEST_HEADERS)
        self.assertEqual(requests_sheet.max_row, 4)
        request_rows = worksheet_rows_by_header(requests_sheet)
        request_by_number = {row["Номер"]: row for row in request_rows}
        self.assertEqual(request_by_number["WH-R-101"]["Юрлицо/точка"], "XASAN XUSAN SAVDO SERVIS XK")
        self.assertEqual(request_by_number["WH-R-101"]["ID заявки Smartup"], "smartup:259704266")
        self.assertEqual(request_by_number["WH-R-101"]["Торговый представитель"], "ТП1 Суюнбеков Умид")
        self.assertEqual(request_by_number["WH-R-101"]["Раб зона"], "Юнусабад")
        self.assertEqual(request_by_number["WH-R-101"]["Блоков план"], 4)
        self.assertEqual(request_by_number["WH-R-101"]["Блоков факт"], 4)
        self.assertEqual(request_by_number["WH-R-101"]["Отклонение"], 0)
        self.assertEqual(request_by_number["WH-R-303"]["Торговый представитель"], None)
        self.assertEqual(request_by_number["WH-R-303"]["Блоков план"], 1)
        self.assertEqual(request_by_number["WH-R-303"]["Блоков факт"], 500)
        self.assertEqual(request_by_number["WH-R-303"]["Отклонение"], 499)

        products_sheet = workbook["Товары заявок"]
        self.assertEqual([cell.value for cell in products_sheet[1]], REQUEST_PRODUCT_HEADERS)
        product_rows = worksheet_rows_by_header(products_sheet)
        self.assertIn("Chapman Gold SSL", [row["Товар"] for row in product_rows])
        brown_product_row = next(row for row in product_rows if row["Товар"] == "Chapman Brown OP 20")
        self.assertEqual(brown_product_row["ID заявки Smartup"], "smartup:259704266")
        self.assertEqual(brown_product_row["Торговый представитель"], "ТП1 Суюнбеков Умид")
        self.assertEqual(brown_product_row["Раб зона"], "Юнусабад")
        gold_product_row = next(row for row in product_rows if row["Товар"] == "Chapman Gold SSL")
        self.assertEqual(gold_product_row["Блоков план"], 1)
        self.assertEqual(gold_product_row["Принято факт"], 500)
        self.assertEqual(gold_product_row["Блоков факт"], 500)
        self.assertEqual(gold_product_row["Отклонение"], 499)

        stock_sheet = workbook["Остатки"]
        self.assertEqual([cell.value for cell in stock_sheet[1]], STOCK_HEADERS)
        self.assertEqual(stock_sheet.max_row, 4)
        stock_rows = worksheet_rows_by_header(stock_sheet)
        self.assertEqual(sum(row["Остаток"] for row in stock_rows), 544)
        self.assertIn("Chapman Brown OP 20", [row["Товар"] for row in stock_rows])
        self.assertIn("Chapman Gold SSL", [row["Товар"] for row in stock_rows])
        self.assertIn("Chapman RED OP 20", [row["Товар"] for row in stock_rows])

        message = build_skladbot_daily_report_message(report)
        self.assertEqual(message, "\n".join([
            "SkladBot daily за 2026-06-08",
            "Отгрузка: 1 заявок, 4 блоков",
            "Отгрузка в браке: 0 заявок, 0 блоков",
            "Возврат: 1 заявок, 2 блоков",
            "Приемка: 1 заявок, 500 блоков",
            "Актуальный остаток: 544",
        ]))
        for hidden_line in (
            "Срез:",
            "Статус покрытия:",
            "В операционных итогах:",
            "В диагностике/исключено:",
            "Ошибки API:",
            "Прочее:",
            "Движения:",
        ):
            self.assertNotIn(hidden_line, message)

    def test_daily_report_never_calls_skladbot_create_request(self):
        client = ForbiddenWriteDailyReportClient()
        report = collect_report_without_delay(client, date(2026, 6, 8))

        self.assertEqual(report["coverage"]["coverage_status"], "complete")
        self.assertEqual(client.write_attempts, 0)

    def test_daily_report_client_blocks_write_methods(self):
        client = ForbiddenWriteDailyReportClient()
        read_only = SkladBotReadOnlyClient(client)

        with self.assertRaises(AttributeError):
            read_only.create_request({"unsafe": True})
        with self.assertRaises(RuntimeError):
            read_only.post("/requests", {"unsafe": True})
        self.assertEqual(client.write_attempts, 0)

    def test_write_capable_workers_not_part_of_daily_report_flow(self):
        client = ForbiddenWriteDailyReportClient()
        collect_report_without_delay(client, date(2026, 6, 8))

        self.assertEqual(client.write_attempts, 0)

    def test_read_style_post_retries_429_5xx_like_get(self):
        client = TransientReadStylePostDailyReportClient()
        report = collect_report_with_env(
            client,
            date(2026, 6, 8),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_READ_POST_RETRY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_READ_POST_RETRIES": "1",
            },
        )

        self.assertEqual(report["errors"], [])
        self.assertGreaterEqual(report["coverage"]["read_style_post_retry_count"], 1)
        self.assertEqual(report["coverage"]["read_style_post_error_count"], 0)
        self.assertEqual(client.post_attempts[("/warehouse/transactions", "in")], 2)

    def test_write_post_does_not_use_read_retry_policy(self):
        client = TransientReadStylePostDailyReportClient()
        coverage = {"read_style_post_retry_count": 0, "read_style_post_error_count": 0}

        with self.assertRaises(RuntimeError):
            read_style_post(SkladBotReadOnlyClient(client), "/requests", {"unsafe": True}, coverage)

        self.assertEqual(coverage["read_style_post_retry_count"], 0)
        self.assertEqual(coverage["read_style_post_error_count"], 0)
        self.assertEqual(client.post_attempts, {})

    def test_movements_limit_reached_marks_partial_or_paginates(self):
        report = collect_report_with_env(
            FakeSkladBotDailyReportClient(),
            date(2026, 6, 8),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_MOVEMENTS_LIMIT": "1",
                "SKLADBOT_DAILY_REPORT_PRODUCTS_LIMIT": "1000",
                "SKLADBOT_DAILY_REPORT_STOCK_LIMIT": "1000",
            },
        )

        self.assertEqual(report["coverage"]["coverage_status"], "partial")
        self.assertTrue(report["coverage"]["movements_truncation_possible"])
        self.assertIn("movements_possible_truncation", report["coverage"]["warnings"])

    def test_products_limit_reached_marks_partial_or_paginates(self):
        report = collect_report_with_env(
            FakeSkladBotDailyReportClient(),
            date(2026, 6, 8),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_MOVEMENTS_LIMIT": "1000",
                "SKLADBOT_DAILY_REPORT_PRODUCTS_LIMIT": "3",
                "SKLADBOT_DAILY_REPORT_STOCK_LIMIT": "1000",
            },
        )

        self.assertEqual(report["coverage"]["coverage_status"], "partial")
        self.assertTrue(report["coverage"]["products_truncation_possible"])
        self.assertIn("products_possible_truncation", report["coverage"]["warnings"])

    def test_stock_report_limit_or_truncation_marks_partial(self):
        report = collect_report_with_env(
            FakeSkladBotDailyReportClient(),
            date(2026, 6, 8),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_MOVEMENTS_LIMIT": "1000",
                "SKLADBOT_DAILY_REPORT_PRODUCTS_LIMIT": "1000",
                "SKLADBOT_DAILY_REPORT_STOCK_LIMIT": "3",
            },
        )

        self.assertEqual(report["coverage"]["coverage_status"], "partial")
        self.assertTrue(report["coverage"]["stock_truncation_possible"])
        self.assertIn("stock_possible_truncation", report["coverage"]["warnings"])

    def test_movements_products_stock_counters_stay_in_internal_coverage(self):
        report = collect_report_with_env(
            FakeSkladBotDailyReportClient(),
            date(2026, 6, 8),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_MOVEMENTS_LIMIT": "1",
                "SKLADBOT_DAILY_REPORT_PRODUCTS_LIMIT": "3",
                "SKLADBOT_DAILY_REPORT_STOCK_LIMIT": "3",
            },
        )
        coverage_values = report["coverage"]

        self.assertEqual(coverage_values["coverage_status"], "partial")
        self.assertEqual(coverage_values["movements_rows_returned"], 2)
        self.assertEqual(coverage_values["products_rows_returned"], 3)
        self.assertEqual(coverage_values["stock_rows_returned"], 3)
        self.assertEqual(coverage_values["movements_truncation_possible"], True)
        self.assertEqual(coverage_values["products_truncation_possible"], True)
        self.assertEqual(coverage_values["stock_truncation_possible"], True)

    def test_movement_product_stock_truncation_blocks_scheduled_send(self):
        report = collect_report_with_env(
            FakeSkladBotDailyReportClient(),
            date(2026, 6, 8),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_MOVEMENTS_LIMIT": "1",
                "SKLADBOT_DAILY_REPORT_PRODUCTS_LIMIT": "3",
                "SKLADBOT_DAILY_REPORT_STOCK_LIMIT": "3",
            },
        )

        blocker = telegram_worker_module.scheduled_skladbot_daily_report_blocker(report)

        self.assertIn("coverage_status=partial", blocker)

    def test_daily_report_crawls_pages_and_records_coverage_diagnostics(self):
        client = PaginatedStrictCoverageClient()
        report = collect_report_with_env(
            client,
            date(2026, 6, 20),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_REQUESTS_LIMIT": "2",
                "SKLADBOT_DAILY_REPORT_MAX_PAGES": "10",
            },
        )

        self.assertGreaterEqual(len(client.request_calls), 5)
        self.assertTrue(all("page" in call for call in client.request_calls))
        self.assertTrue(all("offset" not in call for call in client.request_calls))
        self.assertEqual(sorted(set(client.detail_calls)), [401, 402, 403, 404, 405, 406, 407])

        coverage = report["coverage"]
        self.assertEqual(coverage["coverage_status"], "partial")
        self.assertEqual(coverage["pages_fetched"], 5)
        self.assertEqual(coverage["page_limit"], 2)
        self.assertEqual(coverage["list_rows_total"], 8)
        self.assertEqual(coverage["unique_request_ids"], 7)
        self.assertEqual(coverage["duplicate_request_ids"], 1)
        self.assertEqual(coverage["detail_attempted"], 7)
        self.assertEqual(coverage["detail_success"], 7)
        self.assertEqual(coverage["detail_errors"], 0)
        self.assertEqual(coverage["included_operational_requests"], 3)
        self.assertEqual(coverage["excluded_diagnostic_requests"], 4)
        self.assertEqual(coverage["out_of_scope_requests"], 1)
        self.assertEqual(coverage["completed_only_count"], 1)
        self.assertEqual(coverage["archived_only_count"], 1)
        self.assertEqual(coverage["neither_count"], 1)
        self.assertEqual(coverage["api_error_count"], 0)
        self.assertIn("status_not_completed_archived", coverage["warnings"])

        self.assertEqual(report["summary"]["requests_total"], 3)
        self.assertEqual(report["summary"]["request_blocks_by_category"]["Отгрузка"], 22)
        request_by_id = {item["id"]: item for item in report["requests"]}
        self.assertEqual(request_by_id[401]["date_field_used"], "unloading_date")
        self.assertEqual(request_by_id[401]["inclusion_reason"], "Дата выгрузки")
        self.assertEqual(request_by_id[402]["date_field_used"], "created_at")
        self.assertEqual(request_by_id[402]["inclusion_reason"], "Дата создания")
        self.assertEqual(request_by_id[407]["date_field_used"], "movement_date")
        self.assertEqual(request_by_id[407]["inclusion_reason"], "Движение склада")
        self.assertEqual(report["summary"]["movement_out_amount"], 8)

        excluded_by_id = {item["request_id"]: item for item in report["excluded_requests"]}
        self.assertEqual(excluded_by_id[403]["exclusion_reason"], "status_not_completed_archived")
        self.assertEqual(excluded_by_id[404]["diagnostic_reason"], "completed_only")
        self.assertEqual(excluded_by_id[405]["diagnostic_reason"], "archived_only")
        self.assertEqual(excluded_by_id[406]["exclusion_reason"], "out_of_scope")

    def test_july7_transfer_batch_with_next_day_unloading_is_in_regular_requests(self):
        report = collect_report_without_delay(
            July7MissingTransferBatchClient(unloading_date="2026-07-08"),
            date(2026, 7, 7),
        )

        self.assertEqual(report["summary"]["requests_total"], 8)
        self.assertEqual(report["summary"]["request_blocks_by_category"]["Отгрузка"], 95)
        self.assertEqual(report["coverage"]["coverage_status"], "complete")
        self.assertEqual(report["coverage"]["included_operational_requests"], 8)
        self.assertEqual(report["coverage"]["excluded_diagnostic_requests"], 0)

        request_by_number = {item["number"]: item for item in report["requests"]}
        self.assertEqual(sorted(request_by_number), [f"WH-R-{item[0]}" for item in JULY7_MISSING_TRANSFER_BATCH])
        self.assertEqual(request_by_number["WH-R-204499"]["recipient"], '"CENTRAL TOBAKKO" MCHJ')
        self.assertEqual(request_by_number["WH-R-204499"]["date_field_used"], "created_at")
        self.assertEqual(request_by_number["WH-R-204499"]["inclusion_reason"], "Дата создания")
        self.assertEqual(sum(sum(product["amount"] for product in request["products"]) for request in request_by_number.values()), 95)

        blocker = telegram_worker_module.scheduled_skladbot_daily_report_blocker(report)
        self.assertEqual(blocker, "")

    def test_july7_transfer_batch_next_day_unloading_is_written_to_regular_xlsx_and_message(self):
        report = collect_report_without_delay(
            July7MissingTransferBatchClient(unloading_date="2026-07-08"),
            date(2026, 7, 7),
        )
        content, _filename = build_skladbot_daily_report_xlsx(report)
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)

        self.assertNotIn("Будущие отгрузки", workbook.sheetnames)
        request_rows = worksheet_rows_by_header(workbook["Заявки"])
        self.assertEqual(len(request_rows), 8)
        row_by_number = {row["Номер"]: row for row in request_rows}
        self.assertEqual(row_by_number["WH-R-204499"]["Юрлицо/точка"], '"CENTRAL TOBAKKO" MCHJ')
        self.assertEqual(row_by_number["WH-R-204499"]["Дата выгрузки"], "2026-07-08")
        self.assertEqual(row_by_number["WH-R-204499"]["Причина включения"], "Дата создания")
        self.assertEqual(row_by_number["WH-R-204499"]["Блоков план"], 58)
        self.assertEqual(sum(row["Блоков план"] for row in request_rows), 95)

        product_rows = worksheet_rows_by_header(workbook["Товары заявок"])
        self.assertEqual(len(product_rows), 8)
        self.assertEqual(sum(row["Блоков план"] for row in product_rows), 95)

        self.assertEqual(workbook.sheetnames, ["Сводка", "Заявки", "Товары заявок", "Остатки"])
        self.assertEqual(report["coverage"]["included_operational_requests"], 8)
        self.assertEqual(report["coverage"]["excluded_diagnostic_requests"], 0)

        message = build_skladbot_daily_report_message(report)
        self.assertIn("Отгрузка: 8 заявок, 95 блоков", message)
        self.assertNotIn("будущую дату выгрузки", message)

    def test_july7_transfer_batch_includes_when_unloading_date_is_report_date(self):
        report = collect_report_without_delay(
            July7MissingTransferBatchClient(unloading_date="2026-07-07"),
            date(2026, 7, 7),
        )

        self.assertEqual(report["summary"]["requests_total"], 8)
        self.assertEqual(report["summary"]["request_blocks_by_category"]["Отгрузка"], 95)
        self.assertEqual(report["coverage"]["included_operational_requests"], 8)
        self.assertEqual(report["coverage"]["excluded_diagnostic_requests"], 0)
        self.assertEqual(report["coverage"]["coverage_status"], "complete")
        self.assertEqual(sorted(item["number"] for item in report["requests"]), [f"WH-R-{item[0]}" for item in JULY7_MISSING_TRANSFER_BATCH])

    def test_july7_transfer_batch_requires_completed_and_archived_status(self):
        status_matrix = {
            204498: (True, True),
            204499: (True, False),
            204500: (False, True),
            204501: (False, False),
        }
        report = collect_report_without_delay(
            July7MissingTransferBatchClient(unloading_date="2026-07-08", status_matrix=status_matrix),
            date(2026, 7, 7),
        )

        included_numbers = {item["number"] for item in report["requests"]}
        self.assertIn("WH-R-204498", included_numbers)
        self.assertNotIn("WH-R-204499", included_numbers)
        self.assertNotIn("WH-R-204500", included_numbers)
        self.assertNotIn("WH-R-204501", included_numbers)
        excluded_by_id = {item["request_id"]: item for item in report["excluded_requests"]}
        self.assertEqual(excluded_by_id[204499]["diagnostic_reason"], "completed_only")
        self.assertEqual(excluded_by_id[204500]["diagnostic_reason"], "archived_only")
        self.assertEqual(excluded_by_id[204501]["diagnostic_reason"], "neither")
        self.assertEqual(report["coverage"]["coverage_status"], "partial")

    def test_unknown_cancelled_problem_status_goes_to_diagnostics_or_partial(self):
        report = collect_report_with_env(
            PaginatedStrictCoverageClient(),
            date(2026, 6, 20),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_REQUESTS_LIMIT": "2",
                "SKLADBOT_DAILY_REPORT_MAX_PAGES": "10",
            },
        )

        self.assertEqual(report["coverage"]["coverage_status"], "partial")
        self.assertIn("status_not_completed_archived", report["coverage"]["warnings"])
        diagnostics = {row["request_id"]: row for row in report["excluded_requests"]}
        self.assertEqual(diagnostics[403]["diagnostic_reason"], "neither")
        self.assertEqual(diagnostics[404]["diagnostic_reason"], "completed_only")
        self.assertEqual(diagnostics[405]["diagnostic_reason"], "archived_only")

    def test_status_matrix_documented_and_enforced(self):
        report = collect_report_with_env(
            PaginatedStrictCoverageClient(),
            date(2026, 6, 20),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_REQUESTS_LIMIT": "2",
                "SKLADBOT_DAILY_REPORT_MAX_PAGES": "10",
            },
        )
        excluded_reasons = {row["request_id"]: row["diagnostic_reason"] for row in report["excluded_requests"]}

        self.assertEqual(excluded_reasons[403], "neither")
        self.assertEqual(excluded_reasons[404], "completed_only")
        self.assertEqual(excluded_reasons[405], "archived_only")
        self.assertEqual(report["coverage"]["coverage_status"], "partial")

    def test_daily_report_marks_max_page_truncation_partial(self):
        report = collect_report_with_env(
            MaxPagesDailyReportClient(),
            date(2026, 6, 20),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_REQUESTS_LIMIT": "1",
                "SKLADBOT_DAILY_REPORT_MAX_PAGES": "1",
            },
        )

        coverage = report["coverage"]
        self.assertEqual(coverage["coverage_status"], "partial")
        self.assertTrue(coverage["max_pages_reached"])
        self.assertIn("max_pages", coverage["warnings"])

    def test_request_type_3389_natural_stop_page_31_with_max_pages_60_complete(self):
        original_max_pages = os.environ.pop("SKLADBOT_DAILY_REPORT_MAX_PAGES", None)
        try:
            client = Type3389DeepPaginationClient(full_pages=30)
            report = collect_report_with_env(
                client,
                date(2026, 7, 7),
                {
                    "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                    "SKLADBOT_DAILY_REPORT_REQUESTS_LIMIT": "100",
                    "SKLADBOT_DAILY_REPORT_OUT_OF_SCOPE_DETAIL_SAMPLE_LIMIT": "0",
                },
            )
        finally:
            if original_max_pages is not None:
                os.environ["SKLADBOT_DAILY_REPORT_MAX_PAGES"] = original_max_pages

        coverage = report["coverage"]
        self.assertEqual(coverage["max_pages"], DEFAULT_DAILY_REPORT_MAX_PAGES)
        self.assertEqual(coverage["list_pages_fetched"], 31)
        self.assertFalse(coverage["max_pages_reached"])
        self.assertEqual(coverage["coverage_status"], "complete")
        self.assertEqual(report["summary"]["requests_total"], 1)
        self.assertEqual(client.detail_calls, [1000])

    def test_request_type_3389_max_pages_20_marks_partial_and_blocks_send(self):
        report = collect_report_with_env(
            Type3389DeepPaginationClient(full_pages=30),
            date(2026, 7, 7),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_REQUESTS_LIMIT": "100",
                "SKLADBOT_DAILY_REPORT_MAX_PAGES": "20",
                "SKLADBOT_DAILY_REPORT_OUT_OF_SCOPE_DETAIL_SAMPLE_LIMIT": "0",
            },
        )

        coverage = report["coverage"]
        self.assertEqual(coverage["max_pages"], 20)
        self.assertEqual(coverage["list_pages_fetched"], 20)
        self.assertTrue(coverage["max_pages_reached"])
        self.assertEqual(coverage["coverage_status"], "partial")
        self.assertIn("coverage_status=partial", telegram_worker_module.scheduled_skladbot_daily_report_blocker(report))

    def test_max_pages_60_still_blocks_if_page_60_full(self):
        report = collect_report_with_env(
            Type3389DeepPaginationClient(full_pages=60),
            date(2026, 7, 7),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_REQUESTS_LIMIT": "100",
                "SKLADBOT_DAILY_REPORT_MAX_PAGES": "60",
                "SKLADBOT_DAILY_REPORT_OUT_OF_SCOPE_DETAIL_SAMPLE_LIMIT": "0",
            },
        )

        coverage = report["coverage"]
        self.assertEqual(coverage["max_pages"], 60)
        self.assertEqual(coverage["list_pages_fetched"], 60)
        self.assertTrue(coverage["max_pages_reached"])
        self.assertEqual(coverage["coverage_status"], "partial")
        self.assertIn("max_pages", coverage["warnings"])

    def test_max_pages_counter_does_not_exceed_guard_or_is_split(self):
        report = collect_report_with_env(
            MultiTypeMaxPagesDailyReportClient(),
            date(2026, 6, 20),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_REQUESTS_LIMIT": "1",
                "SKLADBOT_DAILY_REPORT_MAX_PAGES": "1",
            },
        )

        coverage = report["coverage"]
        self.assertEqual(coverage["max_pages"], 1)
        self.assertEqual(coverage["max_pages_per_request_type"], 1)
        self.assertEqual(coverage["list_page_guard_max_total"], 2)
        self.assertEqual(coverage["list_pages_fetched"], 2)
        self.assertEqual(coverage["pages_fetched"], coverage["list_pages_fetched"])
        self.assertLessEqual(coverage["list_pages_fetched"], coverage["list_page_guard_max_total"])
        self.assertEqual(coverage["total_http_pages_fetched"], coverage["list_pages_fetched"] + coverage["detail_pages_fetched"])
        self.assertTrue(coverage["max_pages_reached"])
        self.assertEqual(report["summary"]["requests_total"], 2)

    def test_daily_report_stops_on_repeated_page_ids_with_partial_coverage(self):
        report = collect_report_with_env(
            RepeatedPageDailyReportClient(),
            date(2026, 6, 20),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_REQUESTS_LIMIT": "2",
                "SKLADBOT_DAILY_REPORT_MAX_PAGES": "5",
            },
        )

        coverage = report["coverage"]
        self.assertEqual(coverage["coverage_status"], "partial")
        self.assertEqual(coverage["pages_fetched"], 2)
        self.assertEqual(coverage["duplicate_request_ids"], 2)
        self.assertIn("repeated_page_ids", coverage["warnings"])

    def test_daily_report_records_list_and_detail_api_errors_in_coverage(self):
        list_report = collect_report_with_env(
            ListFailureDailyReportClient(),
            date(2026, 6, 20),
            {"SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0"},
        )
        self.assertEqual(list_report["coverage"]["coverage_status"], "failed")
        self.assertEqual(list_report["coverage"]["api_error_count"], 1)
        self.assertEqual(list_report["summary"]["requests_total"], 0)

        detail_report = collect_report_with_env(
            DetailFailureDailyReportClient(),
            date(2026, 6, 20),
            {"SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0"},
        )
        self.assertEqual(detail_report["coverage"]["coverage_status"], "partial")
        self.assertEqual(detail_report["coverage"]["detail_attempted"], 1)
        self.assertEqual(detail_report["coverage"]["detail_errors"], 1)
        self.assertEqual(detail_report["coverage"]["api_error_count"], 1)
        self.assertEqual(detail_report["excluded_requests"][0]["request_id"], 501)
        self.assertEqual(detail_report["excluded_requests"][0]["detail_loaded"], False)
        self.assertIn("detail 500", detail_report["excluded_requests"][0]["error_message"])

    def test_daily_report_xlsx_handles_partial_data_and_errors(self):
        report = {
            "report_date": date(2026, 6, 8),
            "generated_at": datetime(2026, 6, 8, 22, 0),
            "customer_id": 6211,
            "requests": [],
            "movements": [],
            "stock": {"total": 0, "rows": []},
            "errors": ["Не удалось получить остаток SkladBot"],
            "summary": {
                "requests_total": 0,
                "category_counts": {"Отгрузка": 0, "Отгрузка в браке": 0, "Возврат": 0, "Приемка": 0, "Прочее": 0},
                "type_counts": {},
                "request_blocks_by_category": {"Отгрузка": 0, "Отгрузка в браке": 0, "Возврат": 0, "Приемка": 0, "Прочее": 0},
                "movements_total": 0,
                "movement_in_rows": 0,
                "movement_out_rows": 0,
                "movement_in_amount": 0,
                "movement_out_amount": 0,
                "stock_total": 0,
                "stock_rows": 0,
                "errors": 1,
            },
        }

        content, _filename = build_skladbot_daily_report_xlsx(report)
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)

        self.assertEqual(workbook["Сводка"]["B2"].value, 0)
        self.assertEqual(workbook["Сводка"]["B3"].value, 0)
        self.assertEqual(workbook["Сводка"]["B4"].value, 0)
        self.assertEqual(workbook["Сводка"]["B5"].value, 0)
        self.assertEqual(workbook["Сводка"]["B6"].value, 0)
        self.assertEqual(workbook["Остатки"].max_row, 2)
        self.assertEqual(workbook["Остатки"]["E2"].value, 0)
        self.assertEqual(report["errors"], ["Не удалось получить остаток SkladBot"])

    def test_daily_report_retries_request_detail_after_rate_limit(self):
        client = TransientRateLimitDailyReportClient()
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        original_retry_seconds = os.environ.get("SKLADBOT_DAILY_REPORT_429_RETRY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            os.environ["SKLADBOT_DAILY_REPORT_429_RETRY_SECONDS"] = "0"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 8),
                client=client,
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay
            if original_retry_seconds is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_429_RETRY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_429_RETRY_SECONDS"] = original_retry_seconds

        self.assertEqual(client.detail_calls[101], 2)
        self.assertEqual(report["errors"], [])
        self.assertEqual(report["summary"]["requests_total"], 3)

    def test_daily_report_skips_completed_request_when_created_date_is_stale(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 20),
                client=PreviousDayReceivingCompletedTodayClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        self.assertEqual(report["summary"]["requests_total"], 0)
        self.assertEqual(report["summary"]["category_counts"]["Приемка"], 0)
        self.assertEqual(report["summary"]["request_blocks_by_category"]["Приемка"], 0)

    def test_daily_report_details_stale_created_requests_for_diagnostics(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 20),
                client=PreviousDayReceivingCompletedTodayClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        self.assertEqual(report["summary"]["requests_total"], 0)
        self.assertEqual(report["errors"], [])
        self.assertEqual(report["coverage"]["detail_attempted"], 1)
        self.assertEqual(report["excluded_requests"][0]["exclusion_reason"], "out_of_scope")

    def test_daily_report_skips_completed_request_created_before_report_date(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 21),
                client=PreviousDayReceivingCompletedTodayClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        self.assertEqual(report["summary"]["requests_total"], 0)
        self.assertEqual(report["summary"]["request_blocks_by_category"]["Приемка"], 0)

    def test_daily_report_includes_created_and_completed_today_request(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report_20 = collect_skladbot_daily_report(
                report_date=date(2026, 6, 20),
                client=CompletedAfterCutoffReceivingClient(),
            )
            report_21 = collect_skladbot_daily_report(
                report_date=date(2026, 6, 21),
                client=CompletedAfterCutoffReceivingClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        self.assertEqual(report_20["summary"]["requests_total"], 1)
        self.assertEqual(report_20["summary"]["request_blocks_by_category"]["Приемка"], 500)
        self.assertEqual(report_20["requests"][0]["include_reasons"], ["Дата выгрузки"])
        self.assertEqual(report_20["requests"][0]["date_field_used"], "unloading_date")
        self.assertEqual(report_21["summary"]["requests_total"], 0)
        self.assertEqual(report_21["summary"]["request_blocks_by_category"]["Приемка"], 0)

    def test_daily_report_skips_completed_today_request_when_created_date_is_stale(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 20),
                client=OldListDateReceivingCompletedTodayClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        self.assertEqual(report["summary"]["requests_total"], 0)
        self.assertEqual(report["summary"]["request_blocks_by_category"]["Приемка"], 0)

    def test_daily_report_ignores_archive_cutoff_when_created_on_report_date(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report_20 = collect_skladbot_daily_report(
                report_date=date(2026, 6, 20),
                client=ArchivedAfterCutoffReceivingClient(),
            )
            report_21 = collect_skladbot_daily_report(
                report_date=date(2026, 6, 21),
                client=ArchivedAfterCutoffReceivingClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        self.assertEqual(report_20["summary"]["requests_total"], 1)
        self.assertEqual(report_20["summary"]["request_blocks_by_category"]["Приемка"], 900)
        self.assertEqual(report_20["requests"][0]["include_reasons"], ["Дата выгрузки"])
        self.assertEqual(report_20["requests"][0]["date_field_used"], "unloading_date")
        self.assertEqual(report_21["summary"]["requests_total"], 0)
        self.assertEqual(report_21["summary"]["request_blocks_by_category"]["Приемка"], 0)

    def test_daily_report_includes_request_by_unloading_date_when_created_before_report_date(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 20),
                client=OldCreatedUnloadingTodayClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        self.assertEqual(report["summary"]["requests_total"], 1)
        self.assertEqual(report["summary"]["category_counts"]["Отгрузка"], 1)
        self.assertEqual(report["summary"]["request_blocks_by_category"]["Отгрузка"], 11)
        self.assertEqual(report["requests"][0]["include_reasons"], ["Дата выгрузки"])

    def test_daily_report_records_warehouse_movement_without_including_stale_request(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 20),
                client=MovementTodayRequestClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        self.assertEqual(report["summary"]["requests_total"], 0)
        self.assertEqual(report["summary"]["category_counts"]["Отгрузка"], 0)
        self.assertEqual(report["summary"]["request_blocks_by_category"]["Отгрузка"], 0)
        self.assertEqual(report["summary"]["movement_out_amount"], 5)

    def test_daily_report_includes_completed_old_request_by_movement_without_double_counting(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 20),
                client=CompletedArchivedMovementTodayRequestClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        self.assertEqual(report["summary"]["requests_total"], 1)
        self.assertEqual(report["summary"]["category_counts"]["Отгрузка"], 1)
        self.assertEqual(report["summary"]["request_blocks_by_category"]["Отгрузка"], 5)
        self.assertEqual(report["requests"][0]["date_field_used"], "movement_date")
        self.assertEqual(report["summary"]["movement_out_amount"], 5)
        self.assertEqual(report["coverage"]["coverage_status"], "partial")
        self.assertEqual(report["coverage"]["included_date_conflict_count"], 1)
        self.assertEqual(report["coverage"]["unloading_movement_conflict_count"], 1)
        self.assertIn("date_conflict_unloading_vs_movement", report["coverage"]["warnings"])

    def test_included_unloading_movement_date_conflict_marks_coverage_partial(self):
        report = collect_report_without_delay(
            CompletedArchivedMovementTodayRequestClient(),
            date(2026, 6, 20),
        )

        self.assertEqual(report["summary"]["requests_total"], 1)
        self.assertEqual(report["requests"][0]["diagnostic_reason"], "conflicting_date_fields")
        self.assertEqual(report["coverage"]["coverage_status"], "partial")
        self.assertIn("date_conflict_unloading_vs_movement", report["coverage"]["warnings"])

    def test_date_conflict_stays_in_internal_coverage_and_diagnostics(self):
        report = collect_report_without_delay(
            CompletedArchivedMovementTodayRequestClient(),
            date(2026, 6, 20),
        )
        coverage_values = report["coverage"]
        diagnostics_rows = report["date_diagnostics"]

        self.assertEqual(coverage_values["coverage_status"], "partial")
        self.assertEqual(coverage_values["included_date_conflict_count"], 1)
        self.assertEqual(diagnostics_rows[0]["diagnostic_reason"], "conflicting_date_fields")

    def test_included_unloading_movement_date_conflict_blocks_scheduled_send(self):
        report = collect_report_without_delay(
            CompletedArchivedMovementTodayRequestClient(),
            date(2026, 6, 20),
        )

        blocker = telegram_worker_module.scheduled_skladbot_daily_report_blocker(report)

        self.assertIn("coverage_status=partial", blocker)

    def test_daily_report_keeps_only_report_date_movements(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 20),
                client=MixedDateMovementsClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay

        self.assertEqual(report["summary"]["movement_out_amount"], 5)
        self.assertEqual([item["request_number"] for item in report["movements"]], ["WH-R-409"])

    def test_daily_report_prioritizes_today_created_request_before_stale_list_items(self):
        original_delay = os.environ.get("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS")
        original_detail_limit = os.environ.get("SKLADBOT_DAILY_REPORT_DETAIL_LIMIT")
        try:
            os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = "0"
            os.environ["SKLADBOT_DAILY_REPORT_DETAIL_LIMIT"] = "1"
            report = collect_skladbot_daily_report(
                report_date=date(2026, 6, 20),
                client=MovementTodayRequestAfterStaleListItemsClient(),
            )
        finally:
            if original_delay is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS"] = original_delay
            if original_detail_limit is None:
                os.environ.pop("SKLADBOT_DAILY_REPORT_DETAIL_LIMIT", None)
            else:
                os.environ["SKLADBOT_DAILY_REPORT_DETAIL_LIMIT"] = original_detail_limit

        self.assertEqual(report["summary"]["requests_total"], 1)
        self.assertEqual(report["requests"][0]["number"], "WH-R-409")
        self.assertEqual(report["requests"][0]["include_reasons"], ["Дата выгрузки"])
        self.assertEqual(report["errors"], [])
        self.assertEqual(report["coverage"]["coverage_status"], "partial")
        self.assertEqual(report["coverage"]["api_error_count"], 0)
        self.assertIn("detail_limit", report["coverage"]["warnings"])
        self.assertEqual(report["summary"]["movement_out_amount"], 5)

    def test_detail_limit_not_consumed_by_known_out_of_scope_rows(self):
        report = collect_report_with_env(
            MovementTodayRequestAfterStaleListItemsClient(),
            date(2026, 6, 20),
            {
                "SKLADBOT_DAILY_REPORT_REQUEST_DELAY_SECONDS": "0",
                "SKLADBOT_DAILY_REPORT_DETAIL_LIMIT": "1",
            },
        )

        self.assertEqual(report["summary"]["requests_total"], 1)
        self.assertEqual(report["requests"][0]["number"], "WH-R-409")
        self.assertEqual(report["coverage"]["detail_attempted"], 1)
        self.assertEqual(report["coverage"]["detail_attempted_in_scope"], 1)
        self.assertEqual(report["coverage"]["detail_attempted_out_of_scope_sample"], 0)
        self.assertEqual(report["coverage"]["out_of_scope_skipped_without_detail"], 1)
        excluded_by_id = {item["request_id"]: item for item in report["excluded_requests"]}
        self.assertEqual(excluded_by_id[999]["detail_loaded"], False)
        self.assertNotIn(999, [item.get("id") for item in report["requests"]])

    def test_daily_report_product_breakdown_merges_stock_and_request_aliases(self):
        report = {
            "stock": {
                "rows": [{
                    "product": "",
                    "vendor_code": "130400353",
                    "barcode": "4006396053978",
                    "stock": 42,
                }],
            },
            "requests": [{
                "category": "Отгрузка",
                "products": [{
                    "name": "Chapman Brown OP 20",
                    "vendor_code": "130400353",
                    "barcode": "4006396053978",
                    "amount": 4,
                    "accepted_amount": 0,
                }],
            }],
        }

        rows = product_breakdown_for_summary(report)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["ending_stock"], 42)
        self.assertEqual(rows[0]["outbound"], 4)

    def test_daily_report_uses_accepted_amount_for_receiving(self):
        report = {
            "stock": {"rows": []},
            "requests": [{
                "category": "Приемка",
                "products": [{
                    "name": "Chapman Brown OP 20",
                    "vendor_code": "130400353",
                    "barcode": "4006396053978",
                    "amount": 1,
                    "accepted_amount": 1750,
                }],
            }],
        }

        rows = product_breakdown_for_summary(report)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["inbound"], 1750)

    def test_daily_report_uses_zero_accepted_amount_for_receiving_when_skladbot_sent_it(self):
        report = {
            "stock": {"rows": []},
            "requests": [{
                "category": "Приемка",
                "products": [{
                    "name": "Chapman Brown OP 20",
                    "vendor_code": "130400353",
                    "barcode": "4006396053978",
                    "amount": 1,
                    "accepted_amount": 0,
                    "accepted_amount_present": True,
                }],
            }],
        }

        rows = product_breakdown_for_summary(report)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["inbound"], 0)

    def test_daily_report_separates_defect_shipments_in_message_and_summary_sheet(self):
        self.assertEqual(categorize_request_type("Отгрузка 3PL"), "Отгрузка")
        self.assertEqual(categorize_request_type("Отгрузка в браке"), "Отгрузка в браке")

        report = {
            "report_date": date(2026, 6, 23),
            "generated_at": datetime(2026, 6, 23, 22, 0),
            "customer_id": 6211,
            "requests": [
                {
                    "id": 101,
                    "number": "WH-R-101",
                    "category": categorize_request_type("Отгрузка 3PL"),
                    "type": "Отгрузка 3PL",
                    "products": [{
                        "name": "Chapman Brown OP 20",
                        "vendor_code": "130400353",
                        "barcode": "4006396053978",
                        "amount": 10,
                    }],
                },
                {
                    "id": 102,
                    "number": "WH-R-102",
                    "category": categorize_request_type("Отгрузка в браке"),
                    "type": "Отгрузка в браке",
                    "products": [{
                        "name": "Chapman Brown OP 20",
                        "vendor_code": "130400353",
                        "barcode": "4006396053978",
                        "amount": 3,
                    }],
                },
            ],
            "movements": [],
            "stock": {
                "total": 42,
                "rows": [{
                    "product": "Chapman Brown OP 20",
                    "vendor_code": "130400353",
                    "barcode": "4006396053978",
                    "stock": 42,
                }],
            },
            "errors": [],
        }
        report["summary"] = summarize_daily_report(report)

        summary = report["summary"]
        self.assertEqual(summary["category_counts"]["Отгрузка"], 1)
        self.assertEqual(summary["category_counts"]["Отгрузка в браке"], 1)
        self.assertEqual(summary["request_blocks_by_category"]["Отгрузка"], 10)
        self.assertEqual(summary["request_blocks_by_category"]["Отгрузка в браке"], 3)

        message = build_skladbot_daily_report_message(report)
        self.assertIn("Отгрузка: 1 заявок, 10 блоков", message)
        self.assertIn("Отгрузка в браке: 1 заявок, 3 блоков", message)
        self.assertNotIn("Отгрузка: 2 заявок", message)

        content, _filename = build_skladbot_daily_report_xlsx(report)
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
        summary_sheet = workbook["Сводка"]

        self.assertEqual(summary_sheet["A2"].value, "Отгрузка")
        self.assertEqual(summary_sheet["B2"].value, 10)
        self.assertEqual(summary_sheet["C2"].value, 1)
        self.assertEqual(summary_sheet["A3"].value, "Отгрузка в браке")
        self.assertEqual(summary_sheet["B3"].value, 3)
        self.assertEqual(summary_sheet["C3"].value, 1)

    def test_workbook_summary_contains_only_requested_metrics(self):
        report = collect_report_without_delay(FakeSkladBotDailyReportClient(), date(2026, 6, 8))
        content, _filename = build_skladbot_daily_report_xlsx(report)
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
        summary_sheet = workbook["Сводка"]
        summary_values = [cell.value for row in summary_sheet.iter_rows() for cell in row if cell.value]

        self.assertEqual(summary_sheet.max_row, 6)
        self.assertEqual([summary_sheet.cell(row=row, column=1).value for row in range(2, 7)], [
            "Отгрузка",
            "Отгрузка в браке",
            "Возврат",
            "Приемка",
            "Актуальный остаток",
        ])
        self.assertNotIn("Расчетный начальный остаток", summary_values)
        self.assertNotIn("Примечание", summary_values)

    def test_telegram_manual_command_sends_skladbot_daily_report(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = set()
        worker.admin_chat_ids = set()
        messages = []
        documents = []
        captured = {}

        def fake_collect(report_date=None):
            captured["report_date"] = report_date
            return {
                "report_date": report_date,
                "generated_at": datetime(2026, 6, 8, 22, 0),
                "customer_id": 6211,
                "requests": [],
                "movements": [],
                "stock": {"total": 0, "rows": []},
                "errors": [],
                "summary": {
                    "requests_total": 0,
                    "category_counts": {"Отгрузка": 0, "Отгрузка в браке": 0, "Возврат": 0, "Приемка": 0, "Прочее": 0},
                    "request_blocks_by_category": {"Отгрузка": 0, "Отгрузка в браке": 0, "Возврат": 0, "Приемка": 0, "Прочее": 0},
                    "movement_in_amount": 0,
                    "movement_out_amount": 0,
                    "stock_total": 0,
                },
            }

        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        original_build_xlsx = telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx
        try:
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = fake_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = lambda report: (b"xlsx", "daily.xlsx")
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text))
            worker.safe_send_document = lambda chat_id, content, filename, caption="": documents.append((chat_id, content, filename, caption)) or {"message_id": 1}

            worker.handle_update({
                "update_id": 1,
                "message": {"chat": {"id": 123}, "text": "/skladbot_daily 08.06.2026"},
            })
        finally:
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = original_build_xlsx

        self.assertEqual(captured["report_date"], date(2026, 6, 8))
        self.assertEqual(messages[0], ("123", "Собираю SkladBot отчет за 08.06.2026."))
        self.assertEqual(documents[0][2], "daily.xlsx")
        self.assertEqual(documents[0][3], "SkladBot отчет за 08.06.2026")

    def test_manual_skladbot_daily_partial_blocked_by_default(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = set()
        worker.admin_chat_ids = set()
        messages = []
        documents = []

        def fake_collect(report_date=None):
            return {
                "report_date": report_date,
                "requests": [],
                "excluded_requests": [{"request_id": 1}],
                "movements": [],
                "stock": {"total": 0, "rows": []},
                "errors": [],
                "coverage": {
                    "coverage_status": "partial",
                    "warnings": "movements_possible_truncation",
                    "included_operational_requests": 0,
                    "excluded_diagnostic_requests": 1,
                },
                "summary": {
                    "requests_total": 0,
                    "category_counts": {},
                    "request_blocks_by_category": {},
                    "movement_in_amount": 0,
                    "movement_out_amount": 0,
                    "stock_total": 0,
                },
            }

        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        original_build_xlsx = telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx
        try:
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = fake_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = (
                lambda _report: (_ for _ in ()).throw(AssertionError("partial manual xlsx is forbidden by default"))
            )
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text))
            worker.safe_send_document = lambda *args, **kwargs: documents.append((args, kwargs))

            worker.handle_update({
                "update_id": 1,
                "message": {"chat": {"id": 123}, "text": "/skladbot_daily 08.06.2026"},
            })
        finally:
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = original_build_xlsx

        self.assertEqual(documents, [])
        self.assertIn("coverage_status=PARTIAL", messages[-1][1])
        self.assertIn("не отправлен", messages[-1][1])
        self.assertIn("--allow-partial", messages[-1][1])

    def test_manual_skladbot_daily_partial_requires_explicit_allow_flag(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = set()
        worker.admin_chat_ids = set()
        messages = []
        documents = []

        def fake_collect(report_date=None):
            return {
                "report_date": report_date,
                "requests": [],
                "excluded_requests": [{"request_id": 1}],
                "movements": [],
                "stock": {"total": 0, "rows": []},
                "errors": [],
                "coverage": {
                    "coverage_status": "partial",
                    "warnings": "date_conflict_unloading_vs_movement",
                    "included_operational_requests": 1,
                    "excluded_diagnostic_requests": 0,
                },
                "summary": {
                    "requests_total": 1,
                    "category_counts": {},
                    "request_blocks_by_category": {},
                    "movement_in_amount": 0,
                    "movement_out_amount": 0,
                    "stock_total": 0,
                },
            }

        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        original_build_xlsx = telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx
        try:
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = fake_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = lambda _report: (b"xlsx", "daily.xlsx")
            worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text))
            worker.safe_send_document = lambda chat_id, content, filename, caption="": documents.append((chat_id, content, filename, caption)) or {"message_id": 1}

            worker.handle_update({
                "update_id": 1,
                "message": {"chat": {"id": 123}, "text": "/skladbot_daily --allow-partial 08.06.2026"},
            })
        finally:
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = original_build_xlsx

        self.assertEqual(documents[0][2], "daily.xlsx")
        self.assertTrue(any("НЕПОЛНЫЙ ОТЧЕТ" in text for _chat_id, text in messages))

    def test_manual_partial_warning_text_unmistakable_if_override_enabled(self):
        report = {
            "coverage": {"coverage_status": "partial", "warnings": "api_error"},
            "errors": [],
        }

        text = telegram_worker_module.manual_skladbot_daily_partial_override_warning(
            report,
            "SKLADBOT_DAILY_REPORT_COVERAGE_NOT_COMPLETE: coverage_status=partial",
        )

        self.assertIn("НЕПОЛНЫЙ ОТЧЕТ", text)
        self.assertIn("explicit override", text)

    def test_scheduled_report_marks_reported_requests_after_success(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        original_build_xlsx = telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx
        try:
            telegram_worker_module.SessionLocal = SessionLocal

            def fake_collect(report_date=None):
                return {
                    "report_date": report_date,
                    "generated_at": datetime(2026, 6, 20, 22, 0),
                    "customer_id": 6211,
                    "requests": [{
                        "id": 404,
                        "number": "WH-R-404",
                        "category": "Приемка",
                        "include_reasons": ["создана"],
                        "products": [],
                    }],
                    "movements": [],
                    "stock": {"total": 0, "rows": []},
                    "errors": [],
                    "summary": {
                        "requests_total": 1,
                        "category_counts": {"Отгрузка": 0, "Отгрузка в браке": 0, "Возврат": 0, "Приемка": 1, "Прочее": 0},
                        "request_blocks_by_category": {"Отгрузка": 0, "Отгрузка в браке": 0, "Возврат": 0, "Приемка": 0, "Прочее": 0},
                        "movement_in_amount": 0,
                        "movement_out_amount": 0,
                        "stock_total": 0,
                    },
                }

            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = fake_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = lambda report: (b"xlsx", "daily.xlsx")

            worker = TelegramWorker.__new__(TelegramWorker)
            worker.send_message = lambda chat_id, text, reply_markup=None: {"message_id": 1}
            worker.send_document = lambda chat_id, content, filename, caption="": {"message_id": 1}

            self.assertTrue(worker.send_skladbot_daily_report("123", report_date=date(2026, 6, 20), scheduled=True))

            with SessionLocal() as db:
                events = db.execute(select(PendingEvent)).scalars().all()

            self.assertEqual(len(events), 1)
            self.assertEqual(events[0].event_type, telegram_worker_module.SKLADBOT_DAILY_REPORTED_REQUEST_EVENT_TYPE)
            self.assertEqual((events[0].payload or {}).get("request_id"), 404)
            self.assertIn("2026-06-20", events[0].idempotency_key)
            self.assertIn("123", events[0].idempotency_key)
            self.assertIn("scheduled", events[0].idempotency_key)
            self.assertIn("daily_skladbot", events[0].idempotency_key)
            self.assertEqual((events[0].payload or {}).get("mode"), "scheduled")
            self.assertEqual((events[0].payload or {}).get("report_kind"), "daily_skladbot")
            self.assertTrue((events[0].payload or {}).get("report_version"))
        finally:
            telegram_worker_module.SessionLocal = original_session_local
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = original_build_xlsx

    def test_successful_complete_report_still_sends_and_marks_reported(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        original_build_xlsx = telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx
        original_reconciliation = telegram_worker_module.run_daily_reconciliation
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            messages = []
            documents = []
            reconciliations = []

            def fake_collect(report_date=None):
                return {
                    "report_date": report_date,
                    "requests": [{
                        "id": 404,
                        "number": "WH-R-404",
                        "category": "Приемка",
                        "include_reasons": ["Дата выгрузки"],
                        "products": [],
                    }],
                    "excluded_requests": [],
                    "movements": [],
                    "stock": {"total": 0, "rows": []},
                    "errors": [],
                    "coverage": {
                        "coverage_status": "complete",
                        "included_operational_requests": 1,
                        "excluded_diagnostic_requests": 0,
                    },
                    "summary": {
                        "requests_total": 1,
                        "category_counts": {"Приемка": 1},
                        "request_blocks_by_category": {"Приемка": 0},
                        "movement_in_amount": 0,
                        "movement_out_amount": 0,
                        "stock_total": 0,
                    },
                }

            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = fake_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = lambda _report: (b"xlsx", "daily.xlsx")
            telegram_worker_module.run_daily_reconciliation = lambda **kwargs: reconciliations.append(kwargs)

            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"123"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.daily_reconciliation_enabled = True
            worker.daily_reconciliation_chat_ids = set()
            worker.send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text)) or {"message_id": 1}
            worker.send_document = lambda chat_id, content, filename, caption="": documents.append((chat_id, filename, caption)) or {"message_id": 1}

            now = datetime(2026, 6, 20, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 1)
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 0)

            with SessionLocal() as db:
                events = db.execute(select(PendingEvent).order_by(PendingEvent.created_at)).scalars().all()

            self.assertEqual(len(messages), 1)
            self.assertEqual(len(documents), 1)
            self.assertEqual(documents[0][1], "daily.xlsx")
            self.assertEqual(len(reconciliations), 1)
            self.assertEqual(reconciliations[0]["report_date"], date(2026, 6, 20))
            send_events = [event for event in events if event.event_type == SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE]
            reported_events = [
                event for event in events
                if event.event_type == telegram_worker_module.SKLADBOT_DAILY_REPORTED_REQUEST_EVENT_TYPE
            ]
            self.assertEqual(len(send_events), 1)
            self.assertEqual(send_events[0].status, "completed")
            self.assertEqual(len(reported_events), 1)
            self.assertEqual((reported_events[0].payload or {}).get("request_id"), 404)
        finally:
            telegram_worker_module.SessionLocal = original_session_local
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = original_build_xlsx
            telegram_worker_module.run_daily_reconciliation = original_reconciliation

    def test_telegram_manual_skladbot_daily_rejects_invalid_date(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        worker.allowed_chat_ids = set()
        worker.admin_chat_ids = set()
        messages = []
        sends = []

        worker.safe_send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text))
        worker.send_skladbot_daily_report = lambda *args, **kwargs: sends.append((args, kwargs))

        worker.handle_update({
            "update_id": 1,
            "message": {"chat": {"id": 123}, "text": "/skladbot_daily bad-date"},
        })

        self.assertEqual(sends, [])
        self.assertEqual(messages[0][0], "123")
        self.assertIn("Неверная дата отчета", messages[0][1])

    def test_scheduled_report_is_sent_once_per_chat_and_date(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"123"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.skladbot_daily_report_retry_minutes = 15
            sends = []
            worker.send_skladbot_daily_report = lambda chat_id, report_date=None, scheduled=False, **_kwargs: sends.append((chat_id, report_date, scheduled)) or True

            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 1)
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 0)

            with SessionLocal() as db:
                events = db.execute(select(PendingEvent)).scalars().all()

            self.assertEqual(len(events), 1)
            self.assertEqual(events[0].event_type, SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE)
            self.assertEqual(events[0].status, "completed")
            self.assertEqual(sends, [("123", date(2026, 6, 8), True)])
        finally:
            telegram_worker_module.SessionLocal = original_session_local

    def test_scheduled_report_runs_reconciliation_for_configured_chat(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        original_reconciliation = telegram_worker_module.run_daily_reconciliation
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            reconciliations = []
            telegram_worker_module.run_daily_reconciliation = (
                lambda report_date=None, alert_chat_ids=None: reconciliations.append((report_date, list(alert_chat_ids or []))) or {"status": "ok"}
            )
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"-5271267499"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.skladbot_daily_report_retry_minutes = 15
            worker.daily_reconciliation_enabled = True
            worker.daily_reconciliation_chat_ids = set()
            sends = []
            worker.send_skladbot_daily_report = lambda chat_id, report_date=None, scheduled=False, **_kwargs: sends.append((chat_id, report_date, scheduled)) or True

            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 1)

            self.assertEqual(sends, [("-5271267499", date(2026, 6, 8), True)])
            self.assertEqual(reconciliations, [(date(2026, 6, 8), ["-5271267499"])])
        finally:
            telegram_worker_module.SessionLocal = original_session_local
            telegram_worker_module.run_daily_reconciliation = original_reconciliation

    def test_scheduled_report_sends_reconciliation_to_configured_private_chat(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        original_reconciliation = telegram_worker_module.run_daily_reconciliation
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            reconciliations = []
            telegram_worker_module.run_daily_reconciliation = (
                lambda report_date=None, alert_chat_ids=None: reconciliations.append((report_date, list(alert_chat_ids or []))) or {"status": "ok"}
            )
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"-5271267499"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.skladbot_daily_report_retry_minutes = 15
            worker.daily_reconciliation_enabled = True
            worker.daily_reconciliation_chat_ids = {"999"}
            sends = []
            worker.send_skladbot_daily_report = lambda chat_id, report_date=None, scheduled=False, **_kwargs: sends.append((chat_id, report_date, scheduled)) or True

            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 1)

            self.assertEqual(sends, [("-5271267499", date(2026, 6, 8), True)])
            self.assertEqual(reconciliations, [(date(2026, 6, 8), ["999"])])
        finally:
            telegram_worker_module.SessionLocal = original_session_local
            telegram_worker_module.run_daily_reconciliation = original_reconciliation

    def test_scheduled_report_marks_stale_processing_failed_without_same_day_retry(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.skladbot_daily_report_retry_minutes = 15
            report_date = date(2026, 6, 8)
            key = worker.skladbot_daily_report_idempotency_key("123", report_date)
            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            now_utc = now.astimezone(timezone.utc)
            with SessionLocal() as db:
                db.add(PendingEvent(
                    event_type=SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE,
                    idempotency_key=key,
                    status="processing",
                    attempts=1,
                    payload={"chat_id": "123", "report_date": report_date.isoformat()},
                    updated_at=now_utc - timedelta(minutes=30),
                    last_error="worker died",
                ))
                db.commit()

            event_id = worker.claim_scheduled_skladbot_daily_report("123", report_date, now=now)

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()

            self.assertEqual(event_id, "")
            self.assertEqual(event.status, "failed")
            self.assertEqual(event.attempts, 1)
            self.assertEqual(event.last_error, "STUCK_PROCESSING_AFTER_TTL")
            self.assertEqual((event.payload or {}).get("error"), "STUCK_PROCESSING_AFTER_TTL")
            self.assertEqual((event.payload or {}).get("success"), False)
            self.assertTrue((event.payload or {}).get("finished_at"))
            self.assertNotIn("chat_id", event.payload or {})
        finally:
            telegram_worker_module.SessionLocal = original_session_local

    def test_scheduled_report_claim_payload_does_not_store_chat_id(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            report_date = date(2026, 6, 8)
            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))

            event_id = worker.claim_scheduled_skladbot_daily_report("123", report_date, now=now)

            self.assertTrue(event_id)
            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()
            payload = event.payload or {}
            self.assertEqual(event.status, "processing")
            self.assertEqual(payload.get("report_date"), "2026-06-08")
            self.assertEqual(payload.get("mode"), "scheduled")
            self.assertNotIn("chat_id", payload)
        finally:
            telegram_worker_module.SessionLocal = original_session_local

    def test_same_day_corrected_report_does_not_auto_send(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"123"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.send_skladbot_daily_report = lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("same-day corrected report must not auto-send"))
            report_date = date(2026, 6, 8)
            key = worker.skladbot_daily_report_idempotency_key("123", report_date)
            with SessionLocal() as db:
                db.add(PendingEvent(
                    event_type=SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE,
                    idempotency_key=key,
                    status="completed",
                    attempts=1,
                    payload={"report_date": report_date.isoformat(), "result_status": "completed_sent"},
                    last_error="",
                ))
                db.commit()

            now = datetime(2026, 6, 8, 22, 20, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 0)
        finally:
            telegram_worker_module.SessionLocal = original_session_local

    def test_same_day_corrected_report_creates_manual_recovery_needed_status(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            report_date = date(2026, 6, 8)
            key = worker.skladbot_daily_report_idempotency_key("123", report_date)
            with SessionLocal() as db:
                db.add(PendingEvent(
                    event_type=SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE,
                    idempotency_key=key,
                    status="failed",
                    attempts=1,
                    payload={"report_date": report_date.isoformat()},
                    last_error="telegram_send_failed",
                ))
                db.commit()

            now = datetime(2026, 6, 8, 22, 20, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.claim_scheduled_skladbot_daily_report("123", report_date, now=now), "")
            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()

            self.assertEqual(event.status, "failed")
            self.assertEqual((event.payload or {}).get("result_status"), "manual_recovery_required")
            self.assertEqual((event.payload or {}).get("manual_recovery_required"), True)
            self.assertEqual((event.payload or {}).get("same_day_existing_event_status"), "failed")

        finally:
            telegram_worker_module.SessionLocal = original_session_local

    def test_same_day_corrected_report_version_key_possible_but_operator_only(self):
        worker = TelegramWorker.__new__(TelegramWorker)
        report_date = date(2026, 6, 8)

        default_key = worker.skladbot_daily_report_idempotency_key("123", report_date)
        recovery_key = worker.skladbot_daily_report_idempotency_key("123", report_date, report_version="v3")

        self.assertNotEqual(default_key, recovery_key)
        self.assertIn(":v2", default_key)
        self.assertIn(":v3", recovery_key)

    def test_scheduled_report_failed_claim_does_not_auto_retry_same_day(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.skladbot_daily_report_retry_minutes = 15
            report_date = date(2026, 6, 8)
            key = worker.skladbot_daily_report_idempotency_key("123", report_date)
            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            now_utc = now.astimezone(timezone.utc)
            with SessionLocal() as db:
                db.add(PendingEvent(
                    event_type=SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE,
                    idempotency_key=key,
                    status="failed",
                    attempts=1,
                    payload={"chat_id": "123", "report_date": report_date.isoformat()},
                    updated_at=now_utc - timedelta(minutes=5),
                    last_error="telegram_send_failed",
                ))
                db.commit()

            self.assertEqual(worker.claim_scheduled_skladbot_daily_report("123", report_date, now=now), "")
            retry_now = now + timedelta(minutes=16)
            self.assertEqual(worker.claim_scheduled_skladbot_daily_report("123", report_date, now=retry_now), "")
            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()
            self.assertEqual(event.status, "failed")
            self.assertEqual(event.attempts, 1)
            self.assertEqual(event.last_error, "telegram_send_failed")
        finally:
            telegram_worker_module.SessionLocal = original_session_local

    def test_stale_same_day_event_not_retried_automatically(self):
        self.test_scheduled_report_failed_claim_does_not_auto_retry_same_day()

    def test_ready_not_hard_degraded_when_failed_daily_resolved_by_successful_catchup(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        failed_at = datetime(2026, 7, 6, 17, 7, tzinfo=timezone.utc)
        catchup_at = datetime(2026, 7, 7, 7, 13, tzinfo=timezone.utc)
        try:
            with SessionLocal() as db:
                db.execute(text("CREATE TABLE alembic_version (version_num VARCHAR(32) NOT NULL)"))
                db.execute(text("INSERT INTO alembic_version (version_num) VALUES ('20260710_0010')"))
                db.add(PendingEvent(
                    event_type=SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE,
                    idempotency_key="skladbot_daily_report:2026-07-06:chat-1:scheduled:daily_skladbot:v2",
                    status="failed",
                    attempts=1,
                    payload={
                        "report_date": "2026-07-06",
                        "mode": "scheduled",
                        "kind": "daily_skladbot",
                        "success": False,
                        "coverage_status": "partial",
                    },
                    last_error="SKLADBOT_DAILY_REPORT_COVERAGE_NOT_COMPLETE: coverage_status=partial",
                    created_at=failed_at,
                    updated_at=failed_at,
                ))
                db.add(PendingEvent(
                    event_type=SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE,
                    idempotency_key="skladbot_daily_report:2026-07-06:chat-1:manual_catchup:daily_skladbot:v3",
                    status="completed",
                    attempts=1,
                    payload={
                        "report_date": "2026-07-06",
                        "mode": "manual_catchup",
                        "kind": "daily_skladbot",
                        "success": True,
                        "result_status": "CATCHUP_SENT_COMPLETE_ONCE",
                    },
                    last_error="",
                    created_at=catchup_at,
                    updated_at=catchup_at,
                ))
                db.commit()

                readiness = build_readiness_report(
                    db,
                    SimpleNamespace(service_name="taksklad-backend", environment="test"),
                )

            self.assertEqual(readiness["status"], "ok")
            self.assertEqual(readiness["queue"]["hot_path_last_errors"], [])
            self.assertEqual(len(readiness["queue"]["resolved_historical_errors"]), 1)
            self.assertEqual(
                readiness["queue"]["resolved_historical_errors"][0]["resolved_by"],
                "later_successful_daily_report",
            )
        finally:
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_ready_is_unhealthy_for_unresolved_failed_or_stale_processing_daily(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        failed_at = datetime(2026, 7, 6, 17, 7, tzinfo=timezone.utc)
        try:
            with SessionLocal() as db:
                db.execute(text("CREATE TABLE alembic_version (version_num VARCHAR(32) NOT NULL)"))
                db.execute(text("INSERT INTO alembic_version (version_num) VALUES ('20260710_0010')"))
                db.add(PendingEvent(
                    event_type=SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE,
                    idempotency_key="skladbot_daily_report:2026-07-06:chat-1:scheduled:daily_skladbot:v2",
                    status="failed",
                    attempts=1,
                    payload={
                        "report_date": "2026-07-06",
                        "mode": "scheduled",
                        "kind": "daily_skladbot",
                        "success": False,
                    },
                    last_error="SKLADBOT_DAILY_REPORT_COVERAGE_NOT_COMPLETE: coverage_status=partial",
                    created_at=failed_at,
                    updated_at=failed_at,
                ))
                db.commit()

                readiness = build_readiness_report(
                    db,
                    SimpleNamespace(service_name="taksklad-backend", environment="test"),
                )

            self.assertFalse(readiness["ready"])
            self.assertEqual(readiness["status"], "unhealthy")
            self.assertEqual(len(readiness["queue"]["hot_path_last_errors"]), 1)
            self.assertEqual(readiness["queue"]["resolved_historical_errors"], [])
        finally:
            Base.metadata.drop_all(engine)
            engine.dispose()

    def test_scheduled_report_coverage_failure_marks_event_failed_without_telegram_or_reconciliation(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        original_build_xlsx = telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx
        original_reconciliation = telegram_worker_module.run_daily_reconciliation
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            calls = []
            reconciliations = []

            def fake_collect(report_date=None):
                return {
                    "report_date": report_date,
                    "requests": [{"id": 501, "number": "WH-R-501", "category": "Отгрузка", "products": []}],
                    "movements": [],
                    "stock": {"total": 0, "rows": []},
                    "errors": ["SkladBot list timeout"],
                    "coverage": {"coverage_status": "failed"},
                    "summary": {
                        "requests_total": 0,
                        "category_counts": {},
                        "request_blocks_by_category": {},
                        "movement_in_amount": 0,
                        "movement_out_amount": 0,
                        "stock_total": 0,
                    },
                }

            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = fake_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = lambda _report: (_ for _ in ()).throw(AssertionError("xlsx must not be built"))
            telegram_worker_module.run_daily_reconciliation = lambda **kwargs: reconciliations.append(kwargs)

            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"123"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.send_message = lambda *args, **kwargs: calls.append(("message", args, kwargs))
            worker.send_document = lambda *args, **kwargs: calls.append(("document", args, kwargs))

            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 0)

            with SessionLocal() as db:
                events = db.execute(select(PendingEvent).order_by(PendingEvent.created_at)).scalars().all()

            self.assertEqual(calls, [])
            self.assertEqual(reconciliations, [])
            self.assertEqual(len(events), 1)
            self.assertEqual(events[0].event_type, SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE)
            self.assertEqual(events[0].status, "failed")
            self.assertIn("SKLADBOT_DAILY_REPORT_COVERAGE_NOT_COMPLETE", events[0].last_error)
            self.assertTrue((events[0].payload or {}).get("finished_at"))
            self.assertEqual((events[0].payload or {}).get("success"), False)
        finally:
            telegram_worker_module.SessionLocal = original_session_local
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = original_build_xlsx
            telegram_worker_module.run_daily_reconciliation = original_reconciliation

    def test_scheduled_blocks_partial_report_no_telegram_send(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        original_build_xlsx = telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx
        original_reconciliation = telegram_worker_module.run_daily_reconciliation
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            calls = []
            reconciliations = []

            def fake_collect(report_date=None):
                return {
                    "report_date": report_date,
                    "requests": [],
                    "excluded_requests": [{"request_id": 999, "exclusion_reason": "out_of_scope"}],
                    "movements": [],
                    "stock": {"total": 0, "rows": []},
                    "errors": [],
                    "coverage": {
                        "coverage_status": "complete",
                        "included_operational_requests": 0,
                        "excluded_diagnostic_requests": 1,
                    },
                    "summary": {
                        "requests_total": 0,
                        "category_counts": {},
                        "request_blocks_by_category": {},
                        "movement_in_amount": 0,
                        "movement_out_amount": 0,
                        "stock_total": 0,
                    },
                }

            telegram_worker_module.SessionLocal = SessionLocal
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = fake_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = lambda _report: (_ for _ in ()).throw(AssertionError("xlsx must not be built"))
            telegram_worker_module.run_daily_reconciliation = lambda **kwargs: reconciliations.append(kwargs)

            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"123"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.send_message = lambda *args, **kwargs: calls.append(("message", args, kwargs))
            worker.send_document = lambda *args, **kwargs: calls.append(("document", args, kwargs))

            now = datetime(2026, 6, 20, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 0)

            with SessionLocal() as db:
                events = db.execute(select(PendingEvent).order_by(PendingEvent.created_at)).scalars().all()

            self.assertEqual(calls, [])
            self.assertEqual(reconciliations, [])
            self.assertEqual(len(events), 1)
            self.assertEqual(events[0].event_type, SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE)
            self.assertEqual(events[0].status, "failed")
            self.assertIn("included=0 excluded=1", events[0].last_error)
            self.assertTrue((events[0].payload or {}).get("finished_at"))
        finally:
            telegram_worker_module.SessionLocal = original_session_local
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = original_build_xlsx
            telegram_worker_module.run_daily_reconciliation = original_reconciliation

    def test_scheduled_report_partial_detail_coverage_marks_failed_without_send(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        try:
            telegram_worker_module.SessionLocal = SessionLocal

            def fake_collect(report_date=None):
                return {
                    "report_date": report_date,
                    "requests": [],
                    "movements": [],
                    "stock": {"total": 0, "rows": []},
                    "errors": [],
                    "coverage": {"coverage_status": "partial", "detail_errors": 1},
                    "summary": {
                        "requests_total": 0,
                        "category_counts": {},
                        "request_blocks_by_category": {},
                        "movement_in_amount": 0,
                        "movement_out_amount": 0,
                        "stock_total": 0,
                    },
                }

            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = fake_collect
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"123"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.send_message = lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("telegram send is forbidden"))
            worker.send_document = lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("telegram document is forbidden"))

            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 0)

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()

            self.assertEqual(event.status, "failed")
            self.assertIn("coverage_status=partial", event.last_error)
            self.assertTrue((event.payload or {}).get("finished_at"))
        finally:
            telegram_worker_module.SessionLocal = original_session_local
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect

    def test_scheduled_report_generation_exception_marks_failed_without_send(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        try:
            telegram_worker_module.SessionLocal = SessionLocal

            def fail_collect(report_date=None):
                raise RuntimeError("report generation failed")

            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = fail_collect
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"123"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.send_message = lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("telegram send is forbidden"))
            worker.send_document = lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("telegram document is forbidden"))

            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 0)

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()

            self.assertEqual(event.status, "failed")
            self.assertIn("report generation failed", event.last_error)
            self.assertTrue((event.payload or {}).get("finished_at"))
        finally:
            telegram_worker_module.SessionLocal = original_session_local
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect

    def test_scheduled_report_runtime_timeout_marks_failed_without_send(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        try:
            telegram_worker_module.SessionLocal = SessionLocal

            def timeout_collect(report_date=None):
                raise telegram_worker_module.skladbot_daily_report.SkladBotDailyReportTimeout(
                    "SkladBot daily report runtime exceeded at request_detail: 1501.0s > 1500s"
                )

            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = timeout_collect
            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"123"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.send_message = lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("telegram send is forbidden"))
            worker.send_document = lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("telegram document is forbidden"))

            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 0)

            with SessionLocal() as db:
                event = db.execute(select(PendingEvent)).scalar_one()

            self.assertEqual(event.status, "failed")
            self.assertIn("runtime exceeded", event.last_error)
            self.assertEqual((event.payload or {}).get("success"), False)
            self.assertNotIn("chat_id", event.payload or {})
        finally:
            telegram_worker_module.SessionLocal = original_session_local
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect

    def test_runtime_timeout_marks_event_failed_no_reported(self):
        self.test_scheduled_report_runtime_timeout_marks_failed_without_send()

    def test_scheduled_report_send_document_failure_marks_failed_without_reported_or_reconciliation(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        original_collect = telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report
        original_build_xlsx = telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx
        original_reconciliation = telegram_worker_module.run_daily_reconciliation
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            messages = []
            documents = []
            reconciliations = []

            def fake_collect(report_date=None):
                return {
                    "report_date": report_date,
                    "requests": [{"id": 404, "number": "WH-R-404", "category": "Приемка", "products": []}],
                    "movements": [],
                    "stock": {"total": 0, "rows": []},
                    "errors": [],
                    "coverage": {"coverage_status": "complete"},
                    "summary": {
                        "requests_total": 1,
                        "category_counts": {"Приемка": 1},
                        "request_blocks_by_category": {"Приемка": 0},
                        "movement_in_amount": 0,
                        "movement_out_amount": 0,
                        "stock_total": 0,
                    },
                }

            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = fake_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = lambda _report: (b"xlsx", "daily.xlsx")
            telegram_worker_module.run_daily_reconciliation = lambda **kwargs: reconciliations.append(kwargs)

            worker = TelegramWorker.__new__(TelegramWorker)
            worker.skladbot_daily_report_enabled = True
            worker.skladbot_daily_report_chat_ids = {"123"}
            worker.skladbot_daily_report_hour = 22
            worker.skladbot_daily_report_minute = 0
            worker.send_message = lambda chat_id, text, reply_markup=None: messages.append((chat_id, text)) or {"message_id": 1}

            def fail_document(chat_id, content, filename, caption=""):
                documents.append((chat_id, filename, caption))
                raise RuntimeError("sendDocument timeout")

            worker.send_document = fail_document

            now = datetime(2026, 6, 8, 22, 5, tzinfo=ZoneInfo("Asia/Tashkent"))
            self.assertEqual(worker.send_due_skladbot_daily_reports(now=now), 0)

            with SessionLocal() as db:
                events = db.execute(select(PendingEvent).order_by(PendingEvent.created_at)).scalars().all()

            self.assertEqual(len(messages), 1)
            self.assertEqual(len(documents), 1)
            self.assertEqual(reconciliations, [])
            self.assertEqual(len(events), 1)
            self.assertEqual(events[0].event_type, SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE)
            self.assertEqual(events[0].status, "failed")
            self.assertIn("sendDocument timeout", events[0].last_error)
        finally:
            telegram_worker_module.SessionLocal = original_session_local
            telegram_worker_module.skladbot_daily_report.collect_skladbot_daily_report = original_collect
            telegram_worker_module.skladbot_daily_report.build_skladbot_daily_report_xlsx = original_build_xlsx
            telegram_worker_module.run_daily_reconciliation = original_reconciliation

    def test_send_document_fail_marks_event_failed_no_reported(self):
        self.test_scheduled_report_send_document_failure_marks_failed_without_reported_or_reconciliation()

    def test_scheduled_report_progress_payload_does_not_store_chat_or_secret_fields(self):
        engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        original_session_local = telegram_worker_module.SessionLocal
        try:
            telegram_worker_module.SessionLocal = SessionLocal
            worker = TelegramWorker.__new__(TelegramWorker)
            event_uuid = uuid.uuid4()
            event_id = str(event_uuid)
            with SessionLocal() as db:
                db.add(PendingEvent(
                    id=event_uuid,
                    event_type=SKLADBOT_DAILY_REPORT_SEND_EVENT_TYPE,
                    status="processing",
                    attempts=1,
                    payload={
                        "chat_id": "123",
                        "raw_payload": {"token": "secret"},
                        "report_date": "2026-06-08",
                    },
                ))
                db.commit()

            worker.update_scheduled_skladbot_daily_report_progress(
                event_id,
                "telegram sendDocument started",
                chat_id="123",
                token="secret",
                raw_payload="forbidden",
                payload_summary="forbidden",
                jwt="forbidden",
                api_key="forbidden",
                report_date="2026-06-08",
                detail_attempted=4,
            )

            with SessionLocal() as db:
                event = db.get(PendingEvent, event_uuid)

            payload = event.payload or {}
            self.assertEqual(payload.get("stage"), "telegram sendDocument started")
            self.assertEqual(payload.get("detail_attempted"), 4)
            self.assertNotIn("chat_id", payload)
            self.assertNotIn("token", payload)
            self.assertNotIn("raw_payload", payload)
            self.assertNotIn("payload_summary", payload)
            self.assertNotIn("jwt", payload)
            self.assertNotIn("api_key", payload)
            self.assertNotIn("progress_chat_id", payload)
        finally:
            telegram_worker_module.SessionLocal = original_session_local

    def test_no_sensitive_progress_payload(self):
        self.test_scheduled_report_progress_payload_does_not_store_chat_or_secret_fields()


if __name__ == "__main__":
    unittest.main()
