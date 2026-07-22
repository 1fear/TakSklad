import hashlib
import importlib
import sys
import tempfile
import types
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from taksklad.geocoding import clean_geocoded_address
from taksklad.excel_normalizer import build_source_columns, detect_excel_source, is_summary_row
from taksklad import spreadsheet_safety


def import_excel_import():
    try:
        return importlib.import_module("taksklad.excel_import")
    except ModuleNotFoundError as exc:
        if exc.name != "gspread":
            raise

    fake_sheets = types.ModuleType("taksklad.sheets")
    for name in [
        "build_import_record_row",
        "ensure_import_sheet_layout",
        "get_existing_import_keys",
        "get_existing_order_duplicate_keys",
        "get_google_client",
    ]:
        setattr(fake_sheets, name, lambda *args, **kwargs: None)
    sys.modules["taksklad.sheets"] = fake_sheets
    return importlib.import_module("taksklad.excel_import")


class ExcelNormalizerTests(unittest.TestCase):
    def test_coordinate_header_semantic_contains_match_precedes_content_inference(self):
        columns, missing, _score = build_source_columns([
            "Клиент",
            "Тип оплаты",
            "Товары",
            "Кол-во ШТ",
            "Служебная GPS позиция клиента",
        ])

        self.assertEqual(missing, [])
        self.assertEqual(columns["coords"], 4)
        self.assertEqual(columns["coords_candidates"], [4])

    def test_explicit_coordinate_headers_reject_adjacent_decoys_and_out_of_range_values(self):
        excel_import = import_excel_import()
        for coordinate_header in ("Координаты", "Служебная GPS позиция клиента"):
            with self.subTest(coordinate_header=coordinate_header):
                columns, missing, _score = build_source_columns([
                    "Клиент",
                    "Тип оплаты",
                    "Товары",
                    "Кол-во ШТ",
                    coordinate_header,
                    "Создан",
                    "Резерв",
                ])

                self.assertEqual(missing, [])
                self.assertEqual(
                    excel_import.get_coordinates_from_row(
                        ["Client", "Терминал", "Product", 20, "", "21.07.2026", "91.0,69.2"],
                        columns,
                    ),
                    "",
                )
                self.assertEqual(
                    excel_import.get_coordinates_from_row(
                        ["Client", "Терминал", "Product", 20, "91.0,69.2", "", ""],
                        columns,
                    ),
                    "",
                )

        self.assertEqual(
            excel_import.normalize_coordinates("GPS: 41.311081,69.240562,15"),
            "41.311081, 69.240562",
        )
        self.assertEqual(
            excel_import.normalize_coordinates(
                "https://maps.google.com/?q=41.311081,69.240562"
            ),
            "41.311081, 69.240562",
        )

    def test_clean_geocoded_address_removes_country_prefix(self):
        self.assertEqual(
            clean_geocoded_address("Узбекистан, Ташкент, улица Укчи, 3"),
            "Ташкент, улица Укчи, 3",
        )

    def save_constructor_report(self, path):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Конструктор отчетов"
        worksheet.append(["Статус:  В обработке"])
        worksheet.append(["Дата заказа:  21.05.2026-22.05.2026"])
        worksheet.append(["", "", "", "", "", "Дата заказа", "22.05.2026", "", "ИТОГО", ""])
        worksheet.append([
            "Торговый представитель",
            "Клиент",
            "Координаты клиента",
            "ТМЦ",
            "Тип оплаты",
            "Статус",
            "Количество заказа",
            "Сумма с переоценкой",
            "Количество заказа",
            "Сумма с переоценкой",
        ])
        worksheet.append([
            "ТП1 Суюнбеков Умид Бахрдирович",
            '"BUSINESS MURODOV TRADE" MCHJ (1 филиал)',
            "41.373879,69.322741",
            "Chapman Brown OP 20",
            "Терминал",
            "В обработке",
            20,
            480000,
            20,
            480000,
        ])
        worksheet.append(["ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", 20, 480000, 20, 480000])
        workbook.save(path)

    def test_detects_constructor_report_header_and_first_quantity_column(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "Шаблон_отправки_заказов_на_склад_25_05_2026.xlsx"
            self.save_constructor_report(path)

            workbook = load_workbook(path, data_only=True, read_only=True)
            source = detect_excel_source(workbook, str(path))

        self.assertEqual(source["sheet_name"], "Конструктор отчетов")
        self.assertEqual(source["header_row"], 4)
        self.assertEqual(source["columns"]["quantity"], 6)
        self.assertEqual(source["default_date"], "25.05.2026")

    def test_parses_constructor_report_as_import_records(self):
        excel_import = import_excel_import()
        excel_import.reverse_geocode_yandex = lambda coords, cache=None: (f"Адрес {coords}", "")

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "Шаблон_отправки_заказов_на_склад_25_05_2026.xlsx"
            self.save_constructor_report(path)
            result = excel_import.parse_excel_order_files([str(path)])

        self.assertEqual(result["errors"], [])
        self.assertEqual(result["source_rows_count"], 1)
        self.assertEqual(len(result["records"]), 1)
        record = result["records"][0]
        self.assertEqual(record["Дата отгрузки"], "25.05.2026")
        self.assertEqual(record["Тип оплаты"], "Терминал")
        self.assertEqual(record["Товары"], "Chapman Brown OP 20")
        self.assertEqual(record["Кол-во ШТ"], 20)
        self.assertEqual(record["Адрес"], "Адрес 41.373879, 69.322741")
        self.assertEqual(record["Координаты"], "41.373879, 69.322741")

    def test_desktop_parsed_record_satisfies_typed_backend_dto(self):
        from backend.app.schemas import ImportCreate
        from taksklad.backend_client import backend_import_records

        excel_import = import_excel_import()
        excel_import.reverse_geocode_yandex = lambda coords, cache=None: (f"Адрес {coords}", "")
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "constructor.xlsx"
            self.save_constructor_report(path)
            parsed = excel_import.parse_excel_order_files([str(path)])

        self.assertTrue(parsed["records"])
        self.assertIn("_source_file_sha256", parsed["records"][0])
        sanitized = backend_import_records(parsed["records"])
        payload = ImportCreate(filename="constructor.xlsx", rows=sanitized)
        self.assertEqual(len(payload.rows), 1)
        self.assertNotIn("_source_file_sha256", payload.rows[0])

    def test_desktop_import_rejects_unsupported_file_extension_before_openpyxl(self):
        excel_import = import_excel_import()

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "legacy_orders.xls"
            path.write_text("not a supported xlsx file", encoding="utf-8")
            result = excel_import.parse_excel_order_files([str(path)])

        self.assertEqual(result["records"], [])
        self.assertEqual(result["source_rows_count"], 0)
        self.assertEqual(result["files_count"], 1)
        self.assertEqual(result["errors"], ["spreadsheet_rejected:filename_extension"])
        self.assertNotIn("legacy_orders.xls", result["errors"][0])

    def test_desktop_import_marks_missing_address_as_pickup(self):
        excel_import = import_excel_import()
        calls = []

        def fake_reverse_geocoder(coords, cache=None):
            calls.append(coords)
            return f"Адрес {coords}", ""

        excel_import.reverse_geocode_yandex = fake_reverse_geocoder

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "pickup_without_address.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Заявки"
            worksheet.append([
                "Клиент",
                "Тип оплаты",
                "Товары",
                "Кол-во ШТ",
                "Адрес",
            ])
            worksheet.append([
                "Pickup Client",
                "Терминал",
                "Chapman Brown OP 20",
                20,
                "",
            ])
            workbook.save(path)

            result = excel_import.parse_excel_order_files([str(path)])

        self.assertEqual(calls, [])
        self.assertEqual(result["errors"], [])
        self.assertEqual(result["records"][0]["Адрес"], "Самовывоз со склада")
        self.assertEqual(result["records"][0]["Координаты"], "")

    def test_parses_delivery_date_from_upper_header(self):
        excel_import = import_excel_import()
        excel_import.reverse_geocode_yandex = lambda coords, cache=None: (f"Адрес {coords}", "")

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "Шаблон_отправки_заказов_на_склад_04_06_2026.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Конструктор отчетов"
            worksheet.append(["", "", "", "", "", "", "", "ИТОГО", "", "ДАТА ДОСТАВКИ"])
            worksheet.append([
                "Торговый представитель",
                "Клиент",
                "Координаты клиента",
                "",
                "",
                "ТМЦ",
                "Тип оплаты",
                "Количество заказа",
                "Сумма с переоценкой",
                "",
            ])
            worksheet.append([
                "ТП1",
                "Client One",
                "41.320075",
                "69.298547",
                "41.320075,69.298547",
                "Chapman Brown OP 20",
                "Перечисление",
                20,
                480000,
                "2026-06-05",
            ])
            workbook.save(path)

            workbook_read = load_workbook(path, data_only=True, read_only=True)
            source = detect_excel_source(workbook_read, str(path))
            result = excel_import.parse_excel_order_files([str(path)])

        self.assertEqual(source["columns"]["date"], 9)
        self.assertEqual(result["errors"], [])
        self.assertEqual(result["records"][0]["Дата отгрузки"], "05.06.2026")

    def test_parses_constructor_report_with_repeated_split_coordinates(self):
        excel_import = import_excel_import()
        calls = []

        def fake_reverse_geocoder(coords, cache=None):
            calls.append(coords)
            return "Ташкент, Юнусабадский район", ""

        excel_import.reverse_geocode_yandex = fake_reverse_geocoder

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "Шаблон_отправки_заказов_на_склад_01_06_2026.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Конструктор отчетов"
            worksheet.append([
                "Торговый представитель",
                "Клиент",
                "Координаты клиента",
                "Координаты клиента",
                "Координаты клиента",
                "ТМЦ",
                "Тип оплаты",
                "Статус",
                "Количество заказа",
                "Сумма с переоценкой",
                "Дата отгрузки",
            ])
            worksheet.append([
                "ТП1",
                "Client One",
                "41.325658539017745",
                "69.23166364431383",
                "41.325658539017745,69.23166364431383",
                "Chapman Brown OP 20",
                "Терминал",
                "В обработке",
                20,
                480000,
                "2026-06-03",
            ])
            workbook.save(path)
            result = excel_import.parse_excel_order_files([str(path)])

        self.assertEqual(calls, ["41.325658539017745, 69.23166364431383"])
        self.assertEqual(result["errors"], [])
        self.assertEqual(result["records"][0]["Адрес"], "Ташкент, Юнусабадский район")
        self.assertEqual(result["records"][0]["Координаты"], "41.325658539017745, 69.23166364431383")

    def test_desktop_import_reads_exact_gps_client_coordinates_alias(self):
        excel_import = import_excel_import()
        excel_import.reverse_geocode_yandex = lambda coords, cache=None: (f"Адрес {coords}", "")

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "orders_23_07_2026.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Заявки"
            worksheet.append([
                "Клиент",
                "Тип оплаты",
                "Товары",
                "Кол-во ШТ",
                "Адрес",
                "GPS-координаты клиента",
            ])
            worksheet.append([
                "Delivery Client",
                "Терминал",
                "Chapman Brown OP 20",
                20,
                "Адрес не найден",
                "41.311081,69.240562",
            ])
            workbook.save(path)

            workbook_read = load_workbook(path, data_only=True, read_only=True)
            source = detect_excel_source(workbook_read, str(path))
            workbook_read.close()
            result = excel_import.parse_excel_order_files([str(path)])

        self.assertEqual(source["columns"]["coords"], 5)
        self.assertEqual(result["errors"], [])
        self.assertEqual(result["records"][0]["Адрес"], "Адрес 41.311081, 69.240562")
        self.assertEqual(result["records"][0]["Координаты"], "41.311081, 69.240562")

    def test_desktop_import_infers_unique_coordinate_column_from_contents(self):
        excel_import = import_excel_import()
        calls = []

        def fake_reverse_geocoder(coords, cache=None):
            calls.append(coords)
            return f"Адрес {coords}", ""

        excel_import.reverse_geocode_yandex = fake_reverse_geocoder

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "content_coordinates.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Заявки"
            worksheet.append([
                "Клиент",
                "Тип оплаты",
                "Товары",
                "Кол-во ШТ",
                "Адрес",
                "GPS ID",
                "Создан",
                "Бюджет",
                "Телефон",
                "Точка на карте",
            ])
            worksheet.append([
                "Delivery One",
                "Терминал",
                "Chapman Brown OP 20",
                20,
                "Адрес не найден",
                "WH-R-209244",
                "21.07.2026",
                480000,
                "+998 90 123 45 67",
                "41.311081,69.240562,15",
            ])
            worksheet.append([
                "Delivery Two",
                "Терминал",
                "Chapman Brown OP 20",
                20,
                "Адрес не найден",
                "WH-R-209245",
                "22.07.2026",
                960000,
                "+998 91 765 43 21",
                "41,300000;69,200000;500",
            ])
            workbook.save(path)

            workbook_read = load_workbook(path, data_only=True, read_only=True)
            source = detect_excel_source(workbook_read, str(path))
            workbook_read.close()
            result = excel_import.parse_excel_order_files([str(path)])

        self.assertEqual(source["columns"]["coords"], 9)
        self.assertEqual(calls, ["41.311081, 69.240562", "41.300000, 69.200000"])
        self.assertEqual(
            [record["Координаты"] for record in result["records"]],
            ["41.311081, 69.240562", "41.300000, 69.200000"],
        )

    def test_content_inferred_coordinates_do_not_scan_adjacent_cells_and_validate_range(self):
        excel_import = import_excel_import()
        calls = []

        def fake_reverse_geocoder(coords, cache=None):
            calls.append(coords)
            return "", "timeout"

        excel_import.reverse_geocode_yandex = fake_reverse_geocoder

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "content_coordinates_with_gaps.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Заявки"
            worksheet.append([
                "Клиент",
                "Тип оплаты",
                "Товары",
                "Кол-во ШТ",
                "Адрес",
                "Точка на карте",
                "Соседняя ячейка 1",
                "Соседняя ячейка 2",
            ])
            rows = [
                ("Valid One", "41.311081,69.240562", "21.07.2026", "+998 90 123 45 67"),
                ("Blank Near Date", "", "21.07.2026", "+998 91 765 43 21"),
                ("Blank Near Number", "", 480000, "+998 93 111 22 33"),
                ("Invalid Latitude", "91.000000,69.200000", "22.07.2026", 960000),
                ("Valid Two", "41.300000,69.200000", "22.07.2026", "+998 94 444 55 66"),
                ("Valid Three", "41.320000,69.250000", "23.07.2026", 1_440_000),
                ("Valid Four", "41.330000,69.260000", "23.07.2026", "+998 95 777 88 99"),
            ]
            for client, coordinates, adjacent_one, adjacent_two in rows:
                worksheet.append([
                    client,
                    "Терминал",
                    "Chapman Brown OP 20",
                    20,
                    "Адрес не найден",
                    coordinates,
                    adjacent_one,
                    adjacent_two,
                ])
            workbook.save(path)

            workbook_read = load_workbook(path, data_only=True, read_only=True)
            source = detect_excel_source(workbook_read, str(path))
            workbook_read.close()
            result = excel_import.parse_excel_order_files([str(path)])

        self.assertTrue(source["columns"]["coords_inferred_from_content"])
        self.assertEqual(
            [record["Координаты"] for record in result["records"]],
            [
                "41.311081, 69.240562",
                "",
                "",
                "",
                "41.300000, 69.200000",
                "41.320000, 69.250000",
                "41.330000, 69.260000",
            ],
        )
        self.assertEqual(
            calls,
            [
                "41.311081, 69.240562",
                "41.300000, 69.200000",
                "41.320000, 69.250000",
                "41.330000, 69.260000",
            ],
        )

    def test_desktop_import_does_not_infer_numeric_decoy_columns(self):
        excel_import = import_excel_import()
        calls = []
        excel_import.reverse_geocode_yandex = lambda coords, cache=None: calls.append(coords)

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "numeric_decoys.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Заявки"
            worksheet.append([
                "Клиент",
                "Тип оплаты",
                "Товары",
                "Кол-во ШТ",
                "Адрес",
                "Номер маршрута",
                "Создан",
                "Бюджет",
                "Телефон",
                "Коэффициент",
            ])
            for suffix, created, budget, phone in (
                ("244", "21.07.2026", 480000, "+998 90 123 45 67"),
                ("245", "22.07.2026", 960000, "+998 91 765 43 21"),
            ):
                worksheet.append([
                    f"Pickup {suffix}",
                    "Терминал",
                    "Chapman Brown OP 20",
                    20,
                    "Адрес не найден",
                    f"WH-R-209{suffix}",
                    created,
                    budget,
                    phone,
                    "41,69",
                ])
            workbook.save(path)

            workbook_read = load_workbook(path, data_only=True, read_only=True)
            source = detect_excel_source(workbook_read, str(path))
            workbook_read.close()
            result = excel_import.parse_excel_order_files([str(path)])

        self.assertIsNone(source["columns"]["coords"])
        self.assertEqual(calls, [])
        self.assertEqual([record["Координаты"] for record in result["records"]], ["", ""])
        self.assertEqual(
            [record["Адрес"] for record in result["records"]],
            ["Самовывоз со склада"] * 2,
        )

    def test_desktop_import_fails_closed_on_ambiguous_coordinate_columns(self):
        excel_import = import_excel_import()
        calls = []
        excel_import.reverse_geocode_yandex = lambda coords, cache=None: calls.append(coords)

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "ambiguous_coordinates.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Заявки"
            worksheet.append([
                "Клиент",
                "Тип оплаты",
                "Товары",
                "Кол-во ШТ",
                "Адрес",
                "Точка A",
                "Точка B",
            ])
            worksheet.append([
                "Delivery One",
                "Терминал",
                "Chapman Brown OP 20",
                20,
                "Адрес не найден",
                "41.311081,69.240562",
                "41.320000,69.250000",
            ])
            worksheet.append([
                "Delivery Two",
                "Терминал",
                "Chapman Brown OP 20",
                20,
                "Адрес не найден",
                "41.300000,69.200000",
                "41.330000,69.260000",
            ])
            workbook.save(path)

            workbook_read = load_workbook(path, data_only=True, read_only=True)
            try:
                with self.assertRaisesRegex(
                    ValueError,
                    "^Неоднозначные координаты: найдено несколько подходящих колонок$",
                ):
                    detect_excel_source(workbook_read, str(path))
            finally:
                workbook_read.close()

        self.assertEqual(calls, [])

    def test_desktop_import_keeps_coordinates_when_reverse_geocode_fails(self):
        excel_import = import_excel_import()
        excel_import.reverse_geocode_yandex = lambda coords, cache=None: ("", "timeout")

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "coordinates_without_address.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Заявки"
            worksheet.append([
                "Клиент",
                "Тип оплаты",
                "Товары",
                "Кол-во ШТ",
                "Адрес",
                "Координаты",
            ])
            worksheet.append([
                "Delivery Client",
                "Терминал",
                "Chapman Brown OP 20",
                20,
                "",
                "41.31, 69.27",
            ])
            workbook.save(path)

            result = excel_import.parse_excel_order_files([str(path)])

        self.assertEqual(result["geocode_failed_count"], 1)
        self.assertIn("адрес по координатам не получен", result["warnings"][0])
        self.assertEqual(result["records"][0]["Адрес"], "Координаты: 41.31, 69.27")
        self.assertEqual(result["records"][0]["Координаты"], "41.31, 69.27")

    def test_summary_rows_are_skipped(self):
        row = ["ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", 20]
        columns = {"client": 1, "payment": 4, "product": 3, "quantity": 6}

        self.assertTrue(is_summary_row(row, columns))


class SpreadsheetInputSafetyTests(unittest.TestCase):
    def save_rows(self, path, row_count=2, column_count=4, cell_value=None):
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet("Заявки")
        for row_number in range(1, row_count + 1):
            values = [f"value-{row_number}-{column}" for column in range(column_count)]
            if row_number == 2 and cell_value is not None:
                values[0] = cell_value
            worksheet.append(values)
        workbook.save(path)
        workbook.close()

    def write_minimal_archive(self, path, extra_members=None):
        content_types = b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"/>'
        workbook = b'<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"/>'
        worksheet = b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData/></worksheet>'
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", content_types)
            archive.writestr("xl/workbook.xml", workbook)
            archive.writestr("xl/worksheets/sheet1.xml", worksheet)
            for member_name, content in extra_members or []:
                archive.writestr(member_name, content)

    def assert_safety_code(self, expected_code, callback):
        with self.assertRaises(spreadsheet_safety.SpreadsheetSafetyError) as raised:
            callback()
        self.assertEqual(raised.exception.code, expected_code)
        self.assertEqual(str(raised.exception), f"spreadsheet_rejected:{expected_code}")

    def test_filename_rejections_are_fixed_and_redacted(self):
        invalid_names = {
            "filename_traversal": ["../customer-secret.xlsx", "folder\\customer-secret.xlsx", "C:customer.xlsx"],
            "filename_control_character": ["customer\x00secret.xlsx"],
            "filename_too_long": [f"{'x' * 124}.xlsx"],
            "filename_extension": ["customer-secret.xls"],
        }
        for expected_code, names in invalid_names.items():
            for name in names:
                with self.subTest(name=name):
                    self.assert_safety_code(
                        expected_code,
                        lambda value=name: spreadsheet_safety.normalize_spreadsheet_filename(value),
                    )

    def test_desktop_import_rejects_source_name_traversal_without_echoing_it(self):
        excel_import = import_excel_import()
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "safe.xlsx"
            self.save_rows(path)
            unsafe_name = "../customer-secret.xlsx"
            result = excel_import.parse_excel_order_files(
                [str(path)],
                source_names={str(path): unsafe_name},
            )

        self.assertEqual(result["errors"], ["spreadsheet_rejected:filename_traversal"])
        self.assertNotIn("customer-secret", result["errors"][0])

    def test_archive_member_path_traversal_is_rejected_before_parser(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "unsafe.xlsx"
            self.write_minimal_archive(path, [("../customer-secret.txt", b"secret")])
            self.assert_safety_code(
                "archive_path_traversal",
                lambda: spreadsheet_safety.inspect_xlsx_archive(path),
            )

    def test_excessive_archive_compression_ratio_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "ratio.xlsx"
            self.write_minimal_archive(path, [("xl/media/padding.bin", b"A" * (1024 * 1024))])
            with patch("openpyxl.load_workbook") as parser:
                self.assert_safety_code(
                    "compression_ratio_exceeded",
                    lambda: spreadsheet_safety.load_safe_workbook(path),
                )
                parser.assert_not_called()

    def test_archive_entry_and_size_limits_are_deterministic(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "bounded.xlsx"
            self.save_rows(path)
            cases = [
                ("archive_entries_exceeded", "MAX_XLSX_ENTRIES", 2),
                ("compressed_size_exceeded", "MAX_XLSX_COMPRESSED_BYTES", 64),
                ("uncompressed_size_exceeded", "MAX_XLSX_UNCOMPRESSED_BYTES", 64),
            ]
            for expected_code, constant_name, limit in cases:
                with self.subTest(expected_code=expected_code), patch.object(spreadsheet_safety, constant_name, limit):
                    self.assert_safety_code(
                        expected_code,
                        lambda: spreadsheet_safety.inspect_xlsx_archive(path),
                    )

    def test_encrypted_member_flag_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "encrypted.xlsx"
            self.write_minimal_archive(path)
            content = bytearray(path.read_bytes())
            offset = 0
            while True:
                offset = content.find(b"PK\x03\x04", offset)
                if offset < 0:
                    break
                flags = int.from_bytes(content[offset + 6:offset + 8], "little") | 0x1
                content[offset + 6:offset + 8] = flags.to_bytes(2, "little")
                offset += 4
            offset = 0
            while True:
                offset = content.find(b"PK\x01\x02", offset)
                if offset < 0:
                    break
                flags = int.from_bytes(content[offset + 8:offset + 10], "little") | 0x1
                content[offset + 8:offset + 10] = flags.to_bytes(2, "little")
                offset += 4
            path.write_bytes(content)

            self.assert_safety_code(
                "archive_encrypted",
                lambda: spreadsheet_safety.inspect_xlsx_archive(path),
            )

    def test_row_column_and_cell_limits_are_checked_before_openpyxl(self):
        long_value = "".join(hashlib.sha256(str(index).encode()).hexdigest() for index in range(257))
        cases = [
            ("rows_exceeded", 5002, 1, None),
            ("columns_exceeded", 2, 129, None),
            ("cell_length_exceeded", 2, 1, long_value),
        ]
        with tempfile.TemporaryDirectory() as tmp_dir:
            for index, (expected_code, rows, columns, value) in enumerate(cases):
                with self.subTest(expected_code=expected_code):
                    path = Path(tmp_dir) / f"limit-{index}.xlsx"
                    self.save_rows(path, row_count=rows, column_count=columns, cell_value=value)
                    self.assert_safety_code(
                        expected_code,
                        lambda current_path=path: spreadsheet_safety.load_safe_workbook(current_path),
                    )

    def test_exact_row_boundary_passes_preflight(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "boundary.xlsx"
            self.save_rows(path, row_count=spreadsheet_safety.MAX_XLSX_ROWS, column_count=1)
            evidence = spreadsheet_safety.inspect_xlsx_archive(path)

        self.assertGreater(evidence["entries"], 0)
        self.assertLessEqual(evidence["compression_ratio"], spreadsheet_safety.MAX_XLSX_COMPRESSION_RATIO)

    def test_desktop_report_writer_keeps_formula_prefixes_as_text(self):
        from taksklad.reports import write_day_report_workbook

        values = ["=1+1", "+1+1", "-1+1", "@SUM(A1:A2)"]
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "report.xlsx"
            write_day_report_workbook(
                path,
                [{"Клиент": value, "Количество": index + 1} for index, value in enumerate(values)],
                [],
                [],
            )
            workbook = load_workbook(path, data_only=False)
            try:
                cells = [workbook["Терминал"].cell(row=index + 2, column=1) for index in range(len(values))]
                self.assertEqual([cell.value for cell in cells], values)
                self.assertEqual([cell.data_type for cell in cells], ["s"] * len(values))
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
