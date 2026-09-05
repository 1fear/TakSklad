import unittest
from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.client_points_service import ClientPointApiError
from backend.app.logistics_calendar_orders_service import list_logistics_calendar_day_orders
from backend.app.logistics_manual_stops_service import (
    delete_logistics_manual_stop,
    save_logistics_manual_stop,
)
from backend.app.logistics_service import build_logistics_report_xlsx
from backend.app.models import (
    AuditLog,
    Base,
    ClientPoint,
    LogisticsManualStop,
    LogisticsRegionPoint,
    Order,
    OrderItem,
    PendingEvent,
)
from backend.app.orders_service import ApiError
from backend.app.schemas import LogisticsManualStopUpsert


SHIPMENT_DATE = date(2030, 3, 4)
CITY_COORDINATES = "41.311081, 69.240562"
REGION_COORDINATES = "41.018778, 70.083423"


class LogisticsManualStopTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        self.db = self.SessionLocal()
        # Непустой справочник областных точек: иначе вся доставка считается
        # городской и разделение зон в тесте ничего бы не проверяло
        self.db.add(LogisticsRegionPoint(
            client_name="Тест Клиент Область",
            normalized_client="тестклиентобласть",
            latitude=41.018778,
            longitude=70.083423,
        ))
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def add_order(self, client="Тест Клиент Город", coordinates=CITY_COORDINATES, blocks=5):
        order = Order(
            payment_type="Наличные",
            client=client,
            address="Тестовый адрес заказа",
            order_date=SHIPMENT_DATE,
            status="not_completed",
            raw_payload={"coordinates": coordinates, "skladbot_request_number": "WH-R-1"},
        )
        order.items = [OrderItem(product="Тестовый товар", quantity_blocks=blocks)]
        self.db.add(order)
        self.db.commit()
        return order

    def save_stop(self, **overrides):
        payload = {
            "service_date": SHIPMENT_DATE,
            "client_name": "Тест Ручная Точка",
            "address": "Ташкент, ручной адрес 1",
            "coordinates": CITY_COORDINATES,
            "blocks": 12,
        }
        payload.update(overrides)
        return save_logistics_manual_stop(self.db, LogisticsManualStopUpsert(**payload))

    def zone_rows(self, zone):
        content, _filename = build_logistics_report_xlsx(self.db, SHIPMENT_DATE.isoformat(), zone)
        sheet = load_workbook(BytesIO(content))["Orders"]
        return [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]

    def test_manual_stop_shows_in_the_day_card_with_its_own_zone(self):
        self.add_order()
        self.save_stop()
        self.save_stop(
            client_name="Тест Клиент Область",
            address="Область, ручной адрес 2",
            coordinates=REGION_COORDINATES,
            blocks=0,
        )

        payload = list_logistics_calendar_day_orders(self.db, SHIPMENT_DATE)
        by_client = {row["client"]: row for row in payload["manual_stops"]}

        self.assertEqual(len(payload["orders"]), 1)
        self.assertEqual(len(payload["manual_stops"]), 2)
        self.assertEqual(by_client["Тест Ручная Точка"]["zone"], "city")
        self.assertEqual(by_client["Тест Ручная Точка"]["blocks"], 12)
        self.assertEqual(by_client["Тест Ручная Точка"]["delivery_from"], "10:00")
        self.assertEqual(by_client["Тест Ручная Точка"]["delivery_to"], "18:00")
        self.assertEqual(by_client["Тест Клиент Область"]["zone"], "region")
        self.assertEqual(by_client["Тест Клиент Область"]["blocks"], 0)

    def test_manual_stop_creates_no_order_and_no_queued_event(self):
        self.save_stop()

        self.assertEqual(self.db.execute(select(Order)).scalars().all(), [])
        self.assertEqual(self.db.execute(select(OrderItem)).scalars().all(), [])
        self.assertEqual(self.db.execute(select(PendingEvent)).scalars().all(), [])

    def test_manual_stop_row_carries_empty_external_id_and_empty_product(self):
        self.add_order()
        self.save_stop(representative="Тест Представитель")

        rows = self.zone_rows("city")
        manual = [row for row in rows if row[3] == "Тест Ручная Точка"]

        self.assertEqual(len(manual), 1)
        row = manual[0]
        self.assertEqual(row[0], "delivery")
        self.assertIn(row[1], ("", None))
        self.assertEqual(row[3], "Тест Ручная Точка")
        self.assertEqual(row[6], "Тест Представитель")
        self.assertEqual(row[16], "41.311081")
        self.assertEqual(row[17], "69.240562")
        self.assertEqual(row[18], "Ташкент, ручной адрес 1")
        self.assertEqual(str(row[19]), "2030-03-04 10:00:00")
        self.assertEqual(str(row[20]), "2030-03-04 18:00:00")
        self.assertIn(row[28], ("", None))   # Название товара
        self.assertEqual(row[31], 0)         # Вес (кг)
        self.assertEqual(row[32], 0)         # Объем (m3)
        self.assertEqual(row[33], 12)        # Короба

    def test_manual_stop_joins_line_id_sequence_with_quantity_one(self):
        # «Айди товара» сквозной по всему файлу, ручная точка получает свой
        # номер после товарных строк, «Количество товара» у неё тоже 1
        self.add_order()
        self.save_stop()

        rows = self.zone_rows("city")

        self.assertEqual(len(rows), 2)
        self.assertEqual([row[29] for row in rows], [1, 2])
        self.assertEqual([row[30] for row in rows], [1, 1])
        self.assertEqual(rows[1][3], "Тест Ручная Точка")

    def test_zero_block_stop_still_goes_into_the_report(self):
        self.add_order()
        self.save_stop(blocks=0)

        rows = self.zone_rows("city")
        manual = [row for row in rows if row[3] == "Тест Ручная Точка"]

        self.assertEqual(len(manual), 1)
        self.assertEqual(manual[0][33], 0)

    def test_day_with_manual_stops_only_still_builds_the_report(self):
        self.save_stop()

        rows = self.zone_rows("city")

        self.assertEqual([row[3] for row in rows], ["Тест Ручная Точка"])
        with self.assertRaises(ApiError):
            build_logistics_report_xlsx(self.db, SHIPMENT_DATE.isoformat(), "region")

    def test_deleted_stop_leaves_the_day_card_and_the_report(self):
        self.add_order()
        stop = self.save_stop()

        delete_logistics_manual_stop(self.db, stop["id"], actor="test")

        payload = list_logistics_calendar_day_orders(self.db, SHIPMENT_DATE)
        rows = self.zone_rows("city")

        self.assertEqual(payload["manual_stops"], [])
        self.assertEqual([row[3] for row in rows], ["Тест Клиент Город"])
        self.assertEqual(
            [log.action for log in self.db.execute(select(AuditLog)).scalars().all()],
            ["logistics_manual_stop_created", "logistics_manual_stop_deleted"],
        )

    def test_editing_a_stop_replaces_it_instead_of_adding_a_second_one(self):
        stop = self.save_stop()

        self.save_stop(id=stop["id"], client_name="Тест Ручная Точка", blocks=3)

        payload = list_logistics_calendar_day_orders(self.db, SHIPMENT_DATE)
        stored = self.db.execute(select(LogisticsManualStop)).scalars().all()

        self.assertEqual(len(stored), 1)
        self.assertEqual(len(payload["manual_stops"]), 1)
        self.assertEqual(payload["manual_stops"][0]["blocks"], 3)

    def test_broken_coordinates_and_negative_blocks_are_refused(self):
        with self.assertRaises(ClientPointApiError) as broken:
            self.save_stop(coordinates="где-то рядом")
        self.assertEqual(broken.exception.status_code, 422)

        with self.assertRaises(ValueError):
            self.save_stop(blocks=-1)

    def test_saving_to_the_directory_never_overwrites_an_existing_point(self):
        self.db.add(ClientPoint(
            client_name="Тест Ручная Точка",
            address="Ташкент, ручной адрес 1",
            normalized_client="тестручнаяточка",
            normalized_address="ташкентручнойадрес1",
            coordinates="",
            representative=None,
            delivery_from="09:00",
            delivery_to="12:00",
        ))
        self.db.commit()

        self.save_stop(representative="Тест Представитель", delivery_from="14:00", delivery_to="16:00")

        point = self.db.execute(select(ClientPoint)).scalars().one()
        self.assertEqual(point.delivery_from, "09:00")
        self.assertEqual(point.delivery_to, "12:00")
        self.assertEqual(point.coordinates, CITY_COORDINATES)
        self.assertEqual(point.representative, "Тест Представитель")

    def test_new_point_lands_in_the_directory_only_when_asked(self):
        self.save_stop(save_to_directory=False)
        self.assertEqual(self.db.execute(select(ClientPoint)).scalars().all(), [])

        self.save_stop(client_name="Тест Вторая Точка", address="Ташкент, ручной адрес 2")
        saved = self.db.execute(select(ClientPoint)).scalars().all()

        self.assertEqual([point.client_name for point in saved], ["Тест Вторая Точка"])
        self.assertEqual(saved[0].coordinates, CITY_COORDINATES)


if __name__ == "__main__":
    unittest.main()
